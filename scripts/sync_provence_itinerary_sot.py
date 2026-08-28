#!/usr/bin/env python3
"""Synchronize the active tracker with the Provence itinerary source of truth."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TRACKER = ROOT / "source/OPERATIONS/TP_Europe_Travel_Master_Tracker_v1.2.xlsx"
DAILY_ROUTES = ROOT / "source/ASSETS/maps/daily-routes.json"
UPDATED = datetime(2026, 8, 28)


def d(month: int, day: int) -> datetime:
    return datetime(2026, month, day)


MASTER = {
    d(9, 9): ("Moustiers (Verdon)", 1, "이동", "Nice역 렌터카 인수 → Route Napoléon → Point Sublime → Moustiers",
              "Nice→Saint-Paul→Grasse→Castellane→Moustiers", 5, "Moustiers 1박 숙소(미정) 확보", "Grasse→Point Sublime 순으로 삭제"),
    d(9, 10): ("Aix-en-Provence", 1, "이동", "Moustiers 아침 → Route des Crêtes → Galetas → Valensole → Aix 체크인",
               "Moustiers→La Palud→Galetas→Valensole→Aix", 4, "Aix 4박(확정 9/10~9/14)·Crêtes 통제 확인", "Crêtes 축소→호수·고원 정차 생략"),
    d(9, 11): ("Aix-en-Provence", 2, "체류", "Marseille 전일 당일치기 (TER)",
               "Aix↔Marseille TER", 4, "TER·Mucem(화요일만 휴관)", "Vallon des Auffes 생략 후 조기 복귀"),
    d(9, 12): ("Aix-en-Provence", 3, "체류", "Aix 토요 대형시장 · Vieil Aix · 세잔 아틀리에 · Granet",
               "Aix 시내 도보", 3, "Atelier 예약(매일 09–18 확인)", "Granet 압축·카페 휴식"),
    d(9, 13): ("Aix-en-Provence", 4, "체류", "Cassis & Calanques 유람선",
               "Aix↔Cassis 차량 왕복", 3, "유람선 48시간 전 예약(7일 운항·기상)", "결항 시 Cap Canaille 드라이브"),
    d(9, 14): ("Gordes", 1, "이동", "Aix 체크아웃 → Lourmarin → Lacoste 성 → Gordes 1박차",
               "Aix→Lourmarin→Lacoste→Gordes", 3, "Gordes 2박(미정)·Lacoste 월 13-17 확정", "Bonnieux→Lacoste 관람 축소"),
    d(9, 15): ("Gordes", 2, "체류", "Gordes 화요시장 · Roussillon · Sénanque · [오후 L'Isle]",
               "Gordes↔Roussillon·Sénanque·L'Isle", 4, "Sénanque 9/15 회차 재확인", "L'Isle 왕복 삭제"),
    d(9, 16): ("Avignon", 1, "이동", "Gordes 체크아웃 → Saint-Rémy 수요시장 → Les Baux → Avignon 체크인",
               "Gordes→Saint-Rémy→Les Baux→Avignon", 4, "Avignon 4박(미정)·주차", "Les Baux 축소, Orange 제외"),
    d(9, 17): ("Avignon", 2, "체류", "Uzès · Pont du Gard · Nîmes & 렌터카 최종 반납",
               "Avignon→Uzès→Pont du Gard→Nîmes→Avignon TGV", 4,
               "18:30 이전 Hertz 조기 반납", "Nîmes 체류부터 축소"),
    d(9, 18): ("Avignon", 3, "체류", "Arles 철도 당일치기",
               "Avignon Centre↔Arles TER", 3, "TER·JEP 재확인", "선택시설 삭제"),
    d(9, 19): ("Avignon", 4, "체류", "교황도시 핵심 (Avignon 시내 도보일)",
               "성벽 안 도보", 3, "Palais·식당", "추가 미술관 삭제"),
    d(9, 20): ("Lyon", 1, "이동", "TGV 이동 & Lyon 적응",
               "Avignon TGV→Lyon", 3, "TGV·짐 보관", "Lyon 도착 후 숙소권 휴식"),
}


def rows_by_id(ws) -> dict[str, int]:
    return {ws.cell(row, 1).value: row for row in range(4, ws.max_row + 1)
            if ws.cell(row, 1).value}


def update_master(ws) -> None:
    ws.cell(1, 1).value = "TP Europe Travel — Master Itinerary v1.2 (43일·42박 · Paris 15박 + 기내 1박)"
    for row in range(4, ws.max_row + 1):
        day = ws.cell(row, 1).value
        if day not in MASTER:
            continue
        base, night, kind, theme, movement, fatigue, lock, alternative = MASTER[day]
        ws.cell(row, 3).value = base
        ws.cell(row, 4).value = night
        ws.cell(row, 5).value = kind
        ws.cell(row, 6).value = theme
        ws.cell(row, 11).value = movement
        ws.cell(row, 14).value = fatigue
        ws.cell(row, 15).value = "확정 일정 반영"
        ws.cell(row, 16).value = lock
        ws.cell(row, 17).value = alternative
        ws.cell(row, 18).value = (
            "06B_Verdon_Moustiers_v1.0.md" if day == d(9, 9)
            else "07_Aix_en_Provence_v2.0.md" if day <= d(9, 13)
            else "08_Luberon_Farmhouse_v2.0.md" if day <= d(9, 15)
            else "10_Lyon_v2.0.md" if day == d(9, 20)
            else "09_Avignon_Alpilles_Pont_du_Gard_v2.0.md"
        )
        ws.cell(row, 19).value = UPDATED


def update_reservations(ws) -> None:
    rows = rows_by_id(ws)
    changes = {
        "R005": {4: "Gordes 숙소 2박", 5: d(9, 14),
                 20: "9/14 체크인·9/16 체크아웃·주차·야간 출입 확인",
                 21: UPDATED, 22: "Gordes 2박 숙소 후보 확정 필요 (RS01)"},
        "R006": {4: "Avignon 숙소 4박", 5: d(9, 16),
                 20: "9/16 체크인·9/20 체크아웃·성벽 안 생활권·주차 확인",
                 21: UPDATED, 22: "Avignon 4박 숙소 후보 확정 필요 (RS01)"},
        "R011": {6: "계약 9/20 09:00 · 운영계획 9/17 18:30 이전 반납(변경 필요)",
                 20: "기존 9/20 계약을 9/17 저녁 조기 반납으로 변경 필요",
                 21: UPDATED,
                 22: "확약서 계약 반납은 9/20 09:00 · 일정 정본은 Day 20(9/17) 18:30 이전 반납 · 변경 완료 전 기존 바우처 조건 유효"},
        "R018": {5: d(9, 19), 20: "Day 22 Palais 결합권·입장시간 확인",
                 21: UPDATED, 22: "Avignon 핵심 도보일은 9/19"},
    }
    for ident, cells in changes.items():
        for column, value in cells.items():
            ws.cell(rows[ident], column).value = value


def update_transport(ws) -> None:
    rows = rows_by_id(ws)
    changes = {
        "T005": {2: d(9, 14), 3: "Aix-en-Provence", 4: "Gordes", 10: "재확인",
                 14: "약 1시간 10분", 15: "체크인·주차·차내수하물·Lacoste 경유", 16: "Bonnieux 삭제", 18: UPDATED},
        "T006": {2: d(9, 16), 3: "Gordes", 4: "Avignon", 10: "재확인",
                 14: "약 1시간 20분 (Saint-Rémy·Les Baux 경유)", 15: "Gordes 체크아웃·수요시장·Avignon 4박 체크인", 16: "Les Baux 축소 후 직행", 18: UPDATED},
        "T007": {5: "TGV", 13: "차량 절차 없이 Avignon TGV에서 승차",
                 14: "역 도착 40분 이상 여유", 15: "수하물·플랫폼·짐 보관", 16: "도착 후 숙소권 휴식", 18: UPDATED},
    }
    for ident, cells in changes.items():
        for column, value in cells.items():
            ws.cell(rows[ident], column).value = value


def update_dashboard(ws) -> None:
    ws["D17"], ws["E17"], ws["F17"] = "Luberon allocation", "완료", "9/14~9/16 · Gordes 2박"
    ws["D18"], ws["E18"], ws["F18"] = "Avignon allocation", "완료", "9/16~9/20 · 4박 · 9/17 차량 반납"


def update_phase8(ws) -> None:
    for row in range(4, ws.max_row + 1):
        category, item = ws.cell(row, 1).value, ws.cell(row, 2).value
        if category == "일정" and item == "43일·42박":
            ws.cell(row, 7).value = "Barcelona 3·Bàscara 3·Nice 5·Moustiers 1·Aix 4·Gordes 2·Avignon 4·Lyon 4·Paris 15박 + 기내 1박"
        elif category == "숙박" and item == "8개 거점 숙박배분":
            ws.cell(row, 4).value = "3/3/5/1/4/2/4/4/15박"
            ws.cell(row, 7).value = "Moustiers·Gordes·Avignon 숙소 확정 필요 · 마지막 밤은 기내"
        elif category == "렌터카" and item == "NCE→Avignon TGV":
            ws.cell(row, 4).value = "9/9 Nice-Ville 인수, 9/17 Avignon TGV 반납"
            ws.cell(row, 5).value = "기존 9/20 계약을 9/17 18:30 이전 반납으로 변경"
        elif category == "숙소" and item == "Luberon":
            ws.cell(row, 2).value = "Gordes"
            ws.cell(row, 4).value = "2박 · 9/14~9/16"
            ws.cell(row, 5).value = "숙소 후보·주차·체크인 확정"
        elif category == "숙소" and item == "Avignon":
            ws.cell(row, 4).value = "4박 · 9/16~9/20"
            ws.cell(row, 5).value = "숙소 후보·주차·체크인 확정"


def update_field_shots(ws) -> None:
    changes = {
        "9/9": ("Verdon·Moustiers", "협곡 전망대·절벽 마을 저녁", "오후·저녁", "협곡 진입·1박 거점"),
        "9/10": ("Route des Crêtes·Galetas", "능선 절벽·호수 청록 수면", "오전·낮", "협곡 본편·Aix 이동"),
        "9/15": ("Gordes·Roussillon·L'Isle", "화요시장·오크르·수로", "전일", "장날·오크르·물의 마을"),
        "9/16": ("Saint-Rémy·Les Baux·Avignon", "수요시장·석회암 마을·체크인", "오전·오후", "시장·Alpilles·거점 전환"),
        "9/18": ("Arles", "Arènes·Théâtre·Saint-Trophime", "전일", "JEP·로마유산"),
        "9/19": ("Avignon", "Les Halles·Palais·Rocher·Pont", "전일", "교황도시 핵심"),
    }
    for row in range(1, ws.max_row + 1):
        key = str(ws.cell(row, 1).value or "")
        if key in changes:
            place, subject, time, purpose = changes[key]
            ws.cell(row, 2).value = place
            ws.cell(row, 3).value = subject
            ws.cell(row, 4).value = time
            ws.cell(row, 6).value = purpose


def update_assets(ws) -> None:
    row = rows_by_id(ws).get("A005")
    if row:
        ws.cell(row, 8).value = "Moustiers 1박·Aix 4박·Gordes 2박·Avignon 4박·9/17 차량 반납"


def update_daily_routes() -> None:
    """Keep the map projection aligned with the daily-card source of truth."""
    replacements = {
        "2026-09-17": {
            "date": "2026-09-17",
            "city": "Avignon · Uzès · Pont du Gard · Nîmes",
            "title": "Uzès · Pont du Gard · Nîmes & 렌터카 최종 반납",
            "center": [43.92, 4.52],
            "zoom": 9,
            "defaultMode": "driving",
            "stops": [
                {"placeId": "avignon-stay-candidate", "order": 0, "plannedTime": "07:45 출발", "note": "Avignon 2박차(4박 기준). 렌터카 마지막 운행일."},
                {"placeId": "uzes", "order": 1, "plannedTime": "08:30–10:30", "note": "구시가지와 Place aux Herbes."},
                {"placeId": "pont-du-gard", "order": 2, "plannedTime": "10:50–13:10", "note": "수로교 관람과 간단한 점심."},
                {"placeId": "arenes-de-nimes", "order": 3, "plannedTime": "13:45–14:45", "note": "Nîmes 원형경기장."},
                {"placeId": "maison-carree", "order": 4, "plannedTime": "14:45–15:45", "note": "15:45 출발 Hard Stop."},
                {"placeId": "avignon-tgv", "order": 5, "plannedTime": "16:30–18:30", "note": "주유·차량 촬영 후 Hertz 조기 반납. 기존 9/20 계약은 9/17로 변경 필요."},
                {"placeId": "avignon-stay-candidate", "order": 6, "plannedTime": "19:00 이후", "note": "TER 또는 택시로 숙소 복귀."},
            ],
            "segments": [
                {"from": "avignon-stay-candidate", "to": "uzes", "mode": "driving"},
                {"from": "uzes", "to": "pont-du-gard", "mode": "driving"},
                {"from": "pont-du-gard", "to": "arenes-de-nimes", "mode": "driving"},
                {"from": "arenes-de-nimes", "to": "maison-carree", "mode": "walking"},
                {"from": "maison-carree", "to": "avignon-tgv", "mode": "driving"},
                {"from": "avignon-tgv", "to": "avignon-stay-candidate", "mode": "transit"},
            ],
        },
        "2026-09-18": {
            "date": "2026-09-18",
            "city": "Avignon · Arles",
            "title": "Arles 철도 당일치기",
            "center": [43.677, 4.63],
            "zoom": 13,
            "defaultMode": "transit",
            "stops": [
                {"placeId": "avignon-centre", "order": 0, "plannedTime": "08:30", "note": "Avignon Centre에서 Arles행 TER 탑승."},
                {"placeId": "arles-arenes", "order": 1, "plannedTime": "09:20–10:45", "note": "Arènes d'Arles."},
                {"placeId": "arles-theatre-antique", "order": 2, "plannedTime": "10:45–11:45", "note": "Théâtre antique와 공화국 광장."},
                {"placeId": "cloitre-saint-trophime", "order": 3, "plannedTime": "13:30–14:45", "note": "Saint-Trophime 회랑 뒤 La Roquette 산책."},
                {"placeId": "avignon-centre", "order": 4, "plannedTime": "17:00 이후", "note": "Arles에서 TER로 Avignon 복귀."},
            ],
            "segments": [
                {"from": "avignon-centre", "to": "arles-arenes", "mode": "transit"},
                {"from": "arles-arenes", "to": "arles-theatre-antique", "mode": "walking"},
                {"from": "arles-theatre-antique", "to": "cloitre-saint-trophime", "mode": "walking"},
                {"from": "cloitre-saint-trophime", "to": "avignon-centre", "mode": "transit"},
            ],
        },
        "2026-09-19": {
            "date": "2026-09-19",
            "city": "Avignon",
            "title": "교황도시 핵심 (Avignon 시내 도보일)",
            "center": [43.951, 4.806],
            "zoom": 15,
            "defaultMode": "walking",
            "stops": [
                {"placeId": "les-halles-avignon", "order": 0, "plannedTime": "08:30", "note": "Les Halles 아침 시장."},
                {"placeId": "palais-des-papes", "order": 1, "plannedTime": "09:45–12:15", "note": "Palais des Papes 시간지정 관람."},
                {"placeId": "rocher-des-doms", "order": 2, "plannedTime": "14:00–15:00", "note": "정원과 Rhône 조망."},
                {"placeId": "pont-saint-benezet", "order": 3, "plannedTime": "15:15–16:45", "note": "교황궁 결합권으로 관람."},
            ],
            "segments": [
                {"from": "les-halles-avignon", "to": "palais-des-papes", "mode": "walking"},
                {"from": "palais-des-papes", "to": "rocher-des-doms", "mode": "walking"},
                {"from": "rocher-des-doms", "to": "pont-saint-benezet", "mode": "walking"},
            ],
        },
        "2026-09-20": {
            "date": "2026-09-20",
            "city": "Avignon · Lyon",
            "title": "TGV 이동 & Lyon 적응",
            "center": [44.85, 4.82],
            "zoom": 8,
            "defaultMode": "transit",
            "stops": [
                {"placeId": "avignon-stay-candidate", "order": 0, "plannedTime": "08:30 체크아웃", "note": "차량 반납은 9/17 완료. 렌터카 절차 없이 역으로 이동."},
                {"placeId": "avignon-tgv", "order": 1, "plannedTime": "09:30 이전", "note": "10:22 TGV 탑승 준비."},
                {"placeId": "lyon-part-dieu", "order": 2, "plannedTime": "11:28 도착", "note": "TGV INOUI 12176 확정."},
                {"placeId": "lyon-stay-candidate", "order": 3, "plannedTime": "12:00 이후", "note": "택시 이동·짐 보관 후 15:00 체크인."},
            ],
            "segments": [
                {"from": "avignon-stay-candidate", "to": "avignon-tgv", "mode": "transit"},
                {"from": "avignon-tgv", "to": "lyon-part-dieu", "mode": "transit"},
                {"from": "lyon-part-dieu", "to": "lyon-stay-candidate", "mode": "driving"},
            ],
        },
    }
    payload = json.loads(DAILY_ROUTES.read_text(encoding="utf-8"))
    payload["days"] = [replacements.get(day["date"], day) for day in payload["days"]]
    DAILY_ROUTES.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    wb = load_workbook(TRACKER)
    update_master(wb["Master Itinerary"])
    update_reservations(wb["Reservations"])
    update_transport(wb["Transport"])
    update_dashboard(wb["Dashboard"])
    update_phase8(wb["Phase8 Lock Status"])
    update_field_shots(wb["Field Shots"])
    update_assets(wb["Assets"])
    wb.save(TRACKER)
    update_daily_routes()
    print("Provence itinerary SOT synchronized to tracker and daily routes")


if __name__ == "__main__":
    main()
