#!/usr/bin/env python3
"""Apply the 2026-08-04 confirmed itinerary to the operational XLSX tracker."""

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PATH = ROOT / "source" / "OPERATIONS" / "TP_Europe_Travel_Master_Tracker_v1.2.xlsx"
UPDATED = datetime(2026, 8, 4)


def d(month: int, day: int) -> datetime:
    return datetime(2026, month, day)


MASTER = {
    d(9, 11): ("Aix-en-Provence", 3, "체류", "Marseille 대중교통 당일치기", "Aix↔Marseille L50·도보", "Mucem·L50", "Cassis·Calanques 하루 전체 대체", "07_Aix_en_Provence_v2.0.md", 4),
    d(9, 13): ("Luberon", 1, "이동", "Lourmarin·장보기·농가 체크인", "Aix→Lourmarin→농가", "농가 3박 변경", "경유지 축소", "08_Luberon_Farmhouse_v2.0.md", 4),
    d(9, 14): ("Luberon", 2, "체류", "Roussillon + Goult 또는 Bonnieux", "농가↔Roussillon·선택마을", "주차·오커길", "두 번째 마을 삭제", "08_Luberon_Farmhouse_v2.0.md", 3),
    d(9, 15): ("Luberon", 3, "체류", "Gordes·Village des Bories + 선택마을", "농가↔Gordes·Bories", "시장·주차", "Ménerbes/Oppède 삭제", "08_Luberon_Farmhouse_v2.0.md", 3),
    d(9, 16): ("Avignon", 1, "이동", "농가 체크아웃·Avignon 이동·정착", "Luberon→Avignon", "농가·Avignon 숙소 변경", "추가마을·저녁관람 삭제", "09_Avignon_Alpilles_Pont_du_Gard_v2.0.md", 3),
    d(9, 17): ("Avignon", 2, "체류", "Avignon 핵심 관광", "성벽 안 도보", "Palais des Papes", "추가 미술관 삭제", "09_Avignon_Alpilles_Pont_du_Gard_v2.0.md", 3),
    d(9, 18): ("Avignon", 3, "체류", "Uzès 구시가지·Pont du Gard", "Avignon↔Uzès·Pont du Gard", "주차·Pont", "Pont 실내전시 삭제", "09_Avignon_Alpilles_Pont_du_Gard_v2.0.md", 4),
    d(9, 19): ("Avignon", 4, "체류", "Arles 당일치기", "Avignon Centre↔Arles 철도", "TER·JEP 재확인", "LUMA/Alyscamps/고대박물관 삭제", "09_Avignon_Alpilles_Pont_du_Gard_v2.0.md", 4),
    d(9, 20): ("Lyon", 1, "이동", "렌터카 반납·TGV·Lyon 도착", "Avignon TGV→Lyon", "렌터카·TGV 변경", "Lyon 도착 후 체크인만", "10_Lyon_v2.0.md", 5),
    d(9, 21): ("Lyon", 2, "체류", "Fourvière·Vieux Lyon", "푸니쿨라·도보", "월요일 운영 재확인", "실내관람 삭제", "10_Lyon_v2.0.md", 4),
    d(9, 22): ("Lyon", 3, "체류", "Croix-Rousse·Halles·공원", "TCL·도보", "시장·식당 휴무", "공원 삭제", "10_Lyon_v2.0.md", 3),
    d(9, 23): ("Lyon", 4, "체류", "Annecy 당일치기", "Lyon↔Annecy 철도", "TER·크루즈", "크루즈 삭제·Lyon 생활일", "10_Lyon_v2.0.md", 4),
    d(9, 24): ("Paris", 1, "이동", "TGV·Paris 체크인·짐정리·최소 장보기", "Lyon→Paris Gare de Lyon", "TGV·Paris 숙소 변경", "대형 미술관·공연·근교 금지", "11_Paris_Long_Stay_v2.0.md", 3),
    d(9, 25): ("Paris", 2, "체류", "세탁·생활권 적응 완충일", "숙소 반경 800m", "없음", "장거리 이동 삭제", "11_Paris_Long_Stay_v2.0.md", 2),
}


def update_master(ws) -> None:
    ws["A1"] = "TP Europe Travel — Master Itinerary v1.2 (43일·42박 · Paris 16박)"
    for row in range(4, ws.max_row + 1):
        day = ws.cell(row, 1).value
        if day in MASTER:
            base, night, kind, theme, movement, booking, alt, chapter, fatigue = MASTER[day]
            ws.cell(row, 3).value = base
            ws.cell(row, 4).value = night
            ws.cell(row, 5).value = kind
            ws.cell(row, 6).value = theme
            ws.cell(row, 11).value = movement
            ws.cell(row, 14).value = fatigue
            ws.cell(row, 15).value = "확정 일정 반영"
            ws.cell(row, 16).value = booking
            ws.cell(row, 17).value = alt
            ws.cell(row, 18).value = chapter
            ws.cell(row, 19).value = UPDATED
        elif isinstance(day, datetime) and day >= d(9, 26):
            ws.cell(row, 3).value = "Paris"
            ws.cell(row, 4).value = (day.date() - d(9, 24).date()).days + 1
            ws.cell(row, 18).value = "11_Paris_Long_Stay_v2.0.md"
            ws.cell(row, 19).value = UPDATED


def rows_by_id(ws) -> dict[str, int]:
    return {ws.cell(r, 1).value: r for r in range(4, ws.max_row + 1) if ws.cell(r, 1).value}


def update_reservations(ws) -> None:
    rows = rows_by_id(ws)
    changes = {
        "R004": ("Aix 숙소 4박", d(9, 9), "미조사", "4박 총액·주차·도심진입·9/9 늦은 체크인", "9/11 Marseille 기본·Cassis 완전 대체안"),
        "R005": ("농가 숙소 3박", d(9, 13), "재확인", "기존 4박→3박 예약 변경 필요; 입지·주방·세탁·주차", "9/16 체크아웃으로 변경 완료 여부 미확인"),
        "R006": ("숙소 4박", d(9, 16), "재확인", "기존 9/17 체크인→9/16 변경 필요; 주차·TGV 반납 동선", "예약 변경 필요"),
        "R007": ("숙소 4박", d(9, 20), "재확인", "기존 9/21 체크인→9/20 변경 필요; Part-Dieu 접근·야간 귀가", "예약 변경 필요"),
        "R008": ("숙소 16박", d(9, 24), "재확인", "기존 15박→16박 예약 변경 필요; 총액·세탁·생활권", "9/24 추가 1박 변경 완료 여부 미확인"),
        "R011": ("NCE 인수→Avignon TGV 반납", d(9, 9), "재확인", "9/9 인수·9/20 반납으로 예약 변경 필요", "자동변속·편도수수료·보험·짐"),
        "R012": ("Avignon TGV→Lyon", d(9, 20), "재확인", "기존 9/21→9/20 열차 변경 필요", "반납·주유·플랫폼 버퍼"),
        "R013": ("Lyon→Paris", d(9, 24), "재확인", "기존 9/25→9/24 열차 변경 필요", "체크아웃·Paris 체크인"),
        "R018": ("Palais des Papes 등", d(9, 17), "미조사", "결합권·입장시간", "기존 예약이 있으면 날짜 고정 여부 확인"),
        "R019": ("Annecy 왕복", d(9, 23), "미조사", "왕복시간·수요일 시장 없음·크루즈 대체", "날짜는 유지, 운영 재확인"),
    }
    for ident, (item, date, status, risk, note) in changes.items():
        row = rows[ident]
        ws.cell(row, 4).value = item
        ws.cell(row, 5).value = date
        ws.cell(row, 8).value = status
        ws.cell(row, 19).value = "itinerary.json / v2.0 지역 챕터"
        ws.cell(row, 20).value = risk
        ws.cell(row, 21).value = UPDATED
        ws.cell(row, 22).value = note


def update_transport(ws) -> None:
    rows = rows_by_id(ws)
    changes = {
        "T006": (d(9, 16), "Luberon", "Avignon", "농가 3박·Avignon 9/16 체크인 변경", "Avignon 직행"),
        "T007": (d(9, 20), "Avignon", "Lyon", "렌터카 반납·TGV 날짜 변경 필요", "더 이른 열차 또는 택시"),
        "T008": (d(9, 23), "Lyon", "Annecy", "왕복열차·수요일 시장 없음", "Lyon 완충일로 전환"),
        "T009": (d(9, 24), "Lyon", "Paris", "체크아웃·도착 후 짐·열차 변경 필요", "Paris 도착 후 생활 완충만"),
    }
    for ident, (date, origin, dest, risk, alt) in changes.items():
        row = rows[ident]
        ws.cell(row, 2).value = date
        ws.cell(row, 3).value = origin
        ws.cell(row, 4).value = dest
        ws.cell(row, 10).value = "재확인"
        ws.cell(row, 15).value = risk
        ws.cell(row, 16).value = alt
        ws.cell(row, 18).value = UPDATED


def update_accommodation(ws) -> None:
    changes = {
        "Luberon": (d(9, 13), d(9, 16), 3, "기존 4박→3박 예약 변경 필요"),
        "Avignon": (d(9, 16), d(9, 20), 4, "기존 9/17 체크인 예약 변경 필요"),
        "Lyon": (d(9, 20), d(9, 24), 4, "기존 9/21 체크인 예약 변경 필요"),
        "Paris": (d(9, 24), d(10, 10), 16, "기존 15박→16박 예약 변경 필요; 9/24 생활 완충일"),
    }
    for row in range(4, ws.max_row + 1):
        base = ws.cell(row, 1).value
        if base in changes:
            checkin, checkout, nights, note = changes[base]
            ws.cell(row, 2).value = checkin
            ws.cell(row, 3).value = checkout
            ws.cell(row, 4).value = nights
            ws.cell(row, 5).value = "재확인"
            ws.cell(row, 20).value = note
        elif base == "Aix-en-Provence":
            ws.cell(row, 20).value = "4박·주차·주방·9/11 Marseille 기본"


def update_dashboard(ws) -> None:
    ws["A1"] = "TP Europe Travel Master Tracker v1.2 — 2026-08-04 Itinerary Lock"
    ws["D17"], ws["E17"], ws["F17"] = "Luberon allocation", "재확인", "9/13~9/16 · 3박 · 예약 변경 필요"
    ws["D18"], ws["E18"], ws["F18"] = "Marseille / Arles", "완료", "9/11·9/19 기본 일정"
    ws["D19"], ws["E19"], ws["F19"] = "Paris allocation", "재확인", "9/24~10/10 · 16박 · 예약 변경 필요"
    ws["D20"], ws["E20"], ws["F20"] = "Phase 8 actual bookings", "차단", "변경 예약번호·주소·금액 미입력"


def update_phase8(ws) -> None:
    for row in range(4, ws.max_row + 1):
        category, item = ws.cell(row, 1).value, ws.cell(row, 2).value
        if category == "일정" and item == "43일·42박":
            ws.cell(row, 7).value = "Nice 5·Aix 4·Luberon 3·Avignon 4·Lyon 4·Paris 16박"
        elif category == "숙박" and item == "8개 거점 숙박배분":
            ws.cell(row, 4).value = "3/3/5/4/3/4/4/16박"
            ws.cell(row, 7).value = "Luberon·Avignon·Lyon·Paris 기존 예약 변경 필요"
        elif category == "렌터카" and item == "NCE→Avignon TGV":
            ws.cell(row, 4).value = "9/9 NCE 인수, 9/20 Avignon 반납"
            ws.cell(row, 5).value = "업체·예약번호·9/20 반납 변경 확인"
        elif category == "철도" and item == "Avignon TGV→Lyon":
            ws.cell(row, 4).value = "9/20"
            ws.cell(row, 5).value = "기존 9/21 표 변경·열차번호·좌석"
        elif category == "철도" and item == "Lyon→Paris":
            ws.cell(row, 4).value = "9/24"
            ws.cell(row, 5).value = "기존 9/25 표 변경·열차번호·좌석"
        elif category == "숙소" and item == "Luberon":
            ws.cell(row, 4).value = "3박 · 예약 변경 필요"
            ws.cell(row, 5).value = "기존 4박 예약을 9/13~9/16으로 변경"
        elif category == "숙소" and item == "Avignon":
            ws.cell(row, 4).value = "4박 · 9/16~9/20"
            ws.cell(row, 5).value = "기존 체크인일 변경 확인"
        elif category == "숙소" and item == "Lyon":
            ws.cell(row, 4).value = "4박 · 9/20~9/24"
            ws.cell(row, 5).value = "기존 체크인일 변경 확인"
        elif category == "숙소" and item == "Paris":
            ws.cell(row, 4).value = "16박 · 예약 변경 필요"
            ws.cell(row, 5).value = "기존 15박 예약에 9/24 한 박 추가"


def main() -> None:
    wb = load_workbook(PATH)
    update_master(wb["Master Itinerary"])
    update_reservations(wb["Reservations"])
    update_transport(wb["Transport"])
    update_accommodation(wb["Accommodation"])
    update_dashboard(wb["Dashboard"])
    update_phase8(wb["Phase8 Lock Status"])
    wb.save(PATH)


if __name__ == "__main__":
    main()
