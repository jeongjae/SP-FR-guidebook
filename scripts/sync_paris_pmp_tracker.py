#!/usr/bin/env python3
"""Synchronize PMP itinerary and confirmed Paris bookings in the XLSX tracker."""
from copy import copy
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
PATH=ROOT/"source/OPERATIONS/TP_Europe_Travel_Master_Tracker_v1.2.xlsx"

def main():
    wb=load_workbook(PATH); master=wb["Master Itinerary"]
    # Sheet row = Day + 3.
    updates={
      32:{6:"PMP 시작 — Sainte-Chapelle & Conciergerie",7:"오전 생활 일정·숙소 휴식",9:"Sainte-Chapelle 15:00(PMP 최초 사용)→Conciergerie→Île de la Cité",10:"Bouillon Racine",16:"PMP 6일/144시간 구매완료"},
      33:{6:"Picasso·Notre-Dame & Classical Paris",7:"Marché Convention·Pichard",8:"Le Marais",9:"Picasso 10:30(PMP)→Notre-Dame(무료)→Tuileries→Palais Royal→Garnier",16:"Notre-Dame PMP 불필요"},
      34:{9:"Moreau→Fashion Week·Marais→Arc de Triomphe Optional(PMP)",16:"Arc de Triomphe 선택"},
      36:{9:"Orangerie 10:00→Chez Savy→Fashion Week→Guimet 14:40→Palais de Tokyo",16:"Orangerie 예약완료·Guimet PMP"},
      38:{6:"Louvre 11:00 — PMP 마지막 핵심 사용",7:"10:15 출발",9:"Louvre 11:00~15:00→Cour Carrée·Seine",16:"Louvre 11:00·PMP last planned use"},
      40:{16:"Qatar Prix de l'Arc de Triomphe General Entry 예약완료"},
      42:{6:"Orsay Mary Cassatt 특별전 & 마레",9:"Orsay Mary Cassatt→Carnavalet·Place des Vosges",16:"10/1 상설 방문과 독립 유지"},
      44:{6:"MAM & Trocadéro 고별 일몰",9:"Musée d'Art Moderne→Trocadéro",16:"Guimet는 9/30으로 이동"},
    }
    for row,cells in updates.items():
        for col,val in cells.items(): master.cell(row,col).value=val

    r=wb["Reservations"]
    # Existing race row: preserve identity, update exact ticket type and status.
    r.cell(26,6).value="General Entry"; r.cell(26,8).value="확정"
    r.cell(26,22).value="사용자 예약 완료 확인 · General Entry · 지정석/VIP 아님"

    template=33
    records={
      "R034":("Château de Versailles — Passport",datetime(2026,9,29),"10:00 전후","https://www.chateauversailles.fr/","사용자 예약 완료 확인"),
      "R035":("Paris Museum Pass — 6-day / 144-hour",datetime(2026,9,26),"15:00 첫 사용","https://www.parismuseumpass.fr/","구매완료 · Sainte-Chapelle 첫 사용 → 10/2 Louvre 11:00 마지막 계획"),
    }
    for ident,(name,date,time,url,note) in records.items():
        row=next((i for i in range(4,r.max_row+1) if r.cell(i,1).value==ident),r.max_row+1)
        existing=r.cell(row,4).value
        if existing and existing!=name: raise RuntimeError(f"reservation ID collision: {ident}")
        for c in range(1,23): r.cell(row,c)._style=copy(r.cell(template,c)._style)
        values={1:ident,2:"입장권",3:"Paris",4:name,5:date,6:time,7:"P1",8:"확정",18:url,19:"11_Paris_Long_Stay_v2.0.md",21:datetime(2026,9,6),22:note}
        for c,v in values.items(): r.cell(row,c).value=v
    wb.save(PATH)
    print("tracker synchronized: PMP window + five confirmed Paris reservations")

if __name__=="__main__": main()
