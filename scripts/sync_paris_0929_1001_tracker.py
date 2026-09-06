#!/usr/bin/env python3
"""Synchronize the 9/29 Versailles and 10/1 Orsay/Rodin tracker rows."""

from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TRACKER = ROOT / "source" / "OPERATIONS" / "TP_Europe_Travel_Master_Tracker_v1.2.xlsx"


def expect_date(sheet, row: int, expected: str) -> None:
    value = sheet.cell(row, 1).value
    actual = value.date().isoformat() if isinstance(value, datetime) else str(value)
    if actual != expected:
        raise RuntimeError(f"Master Itinerary row {row}: expected {expected}, found {actual}")


def main() -> None:
    workbook = load_workbook(TRACKER)
    master = workbook["Master Itinerary"]
    expect_date(master, 35, "2026-09-29")
    expect_date(master, 36, "2026-09-30")
    expect_date(master, 37, "2026-10-01")
    expect_date(master, 38, "2026-10-02")

    updates = {
        35: {
            5: "근교·왕실문화",
            6: "Versailles Palace & Gardens",
            7: "08:30 숙소 출발",
            8: "La Flottille 12:45",
            9: "RER C→Versailles 본관 10:00 전후→Gardens→Trianon 선택",
            10: "Le Grand Pan 20:00",
            11: "RER C·영지 도보",
            12: "Trianon·왕비의 촌락",
            13: "왕실 정원 산책",
            14: "4",
            15: "검토중",
            16: "Versailles Passport 10:00 전후·La Flottille·Le Grand Pan",
            17: "RER C·Passport 슬롯 확인, 피로 시 미니트레인",
        },
        37: {
            5: "미술·조각",
            6: "Impressionism & Sculpture — Orsay & Rodin",
            7: "09:30 숙소 출발",
            8: "Café Varenne 13:00",
            9: "Musée d'Orsay 10:30(확정)→Musée Rodin 14:15→Invalides 선택",
            10: "Café du Commerce 18:30",
            11: "메트로 8·12호선·도보",
            12: "Invalides 외관",
            13: "인상주의·조각 스케치",
            14: "4",
            15: "확정",
            16: "Orsay 10:30 예약 확정·Rodin 14:15 권장",
            17: "10/1 상설 중심, 10/6 Mary Cassatt 특별전 재방문과 분리",
        },
    }
    for row, cells in updates.items():
        for column, value in cells.items():
            master.cell(row, column).value = value

    reservations = workbook["Reservations"]
    ident = "R033"
    row = next(
        (index for index in range(4, reservations.max_row + 1)
         if reservations.cell(index, 1).value == ident),
        reservations.max_row + 1,
    )
    existing = reservations.cell(row, 4).value
    if existing and existing != "Musée d'Orsay — Permanent Collection":
        raise RuntimeError(f"Reservation ID collision: {ident}")
    for column in range(1, 23):
        reservations.cell(row, column)._style = copy(reservations.cell(33, column)._style)
    values = {
        1: ident, 2: "입장권", 3: "Paris",
        4: "Musée d'Orsay — Permanent Collection",
        5: datetime(2026, 10, 1), 6: "10:30", 7: "P1", 8: "확정",
        18: "https://www.musee-orsay.fr/",
        19: "11_Paris_Long_Stay_v2.0.md",
        21: datetime(2026, 9, 6),
        22: "사용자 예약 완료 확인 · 10/1 상설 컬렉션 · 예약번호·금액·취소조건 미제공",
    }
    for column, value in values.items():
        reservations.cell(row, column).value = value

    # Existing bundled museum action row remains useful for unconfirmed sites,
    # but the three confirmed bookings are now represented by individual rows.
    reservations.cell(23, 22).value = (
        "Grand Palais 9/25 17:00·Orangerie 9/30 10:00·Orsay 10/1 10:30은 "
        "별도 확정 행 참조; 이 행은 나머지 미술관 예약"
    )
    workbook.save(TRACKER)
    print("tracker synchronized: 9/29 Versailles · 10/1 Orsay 10:30 + Rodin")


if __name__ == "__main__":
    main()
