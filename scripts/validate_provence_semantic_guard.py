#!/usr/bin/env python3
"""Fail when Provence follow-up work changes protected Day 20+ itinerary semantics."""

from __future__ import annotations

import argparse
import io
import json
import re
import subprocess
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
MASTER = "source/CURRENT/10_Core/03_Whole_Trip_Master_Itinerary_v1.2.md"
AVIGNON = "source/CURRENT/20_Regional_Chapters/09_Avignon_Alpilles_Pont_du_Gard_v2.0.md"
TRACKER = "source/OPERATIONS/TP_Europe_Travel_Master_Tracker_v1.2.xlsx"
DAY_FIELDS = ("day", "date", "city", "title", "startTime", "endTime")
STOP_FIELDS = ("id", "order", "start", "end", "name", "category", "optional", "place_ref")
LEG_FIELDS = ("from", "to", "mode", "duration", "distance", "line")


def git_bytes(ref: str, path: str) -> bytes:
    return subprocess.check_output(["git", "show", f"{ref}:{path}"], cwd=ROOT)


def current_bytes(path: str) -> bytes:
    return (ROOT / path).read_bytes()


def daily_semantics(raw: bytes) -> dict:
    payload = json.loads(raw)
    result = {key: payload.get(key) for key in DAY_FIELDS}
    result["stops"] = [
        {key: stop.get(key) for key in STOP_FIELDS} for stop in payload.get("stops", [])
    ]
    result["legs"] = [
        {key: leg.get(key) for key in LEG_FIELDS} for leg in payload.get("legs", [])
    ]
    for key in ("transport", "highlights", "backup"):
        result[key] = payload.get(key)
    return result


def master_rows(raw: bytes) -> list[str]:
    text = raw.decode("utf-8")
    return [
        re.sub(r"\s+", " ", line.strip())
        for line in text.splitlines()
        if re.match(r"^\|\s*(?:2[0-9]|3[0-9]|4[0-3])\s*\|", line)
    ]


def avignon_day20_body(raw: bytes) -> str:
    text = raw.decode("utf-8")
    marker = "## Day 20 —"
    return text[text.index(marker) :]


def tracker_rows(raw: bytes) -> list[list[str]]:
    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=False)
    sheet = workbook["Master Itinerary"]
    rows = []
    for row in sheet.iter_rows(values_only=True):
        value = row[0]
        if isinstance(value, datetime):
            value = value.date()
        if isinstance(value, date) and value >= date(2026, 9, 17):
            rows.append(["" if cell is None else str(cell) for cell in row])
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--base", default="origin/main", help="Git ref used as protected baseline")
    args = parser.parse_args()

    errors = []
    for day in range(20, 44):
        path = f"data/daily-cards/day-{day:02d}.json"
        if daily_semantics(git_bytes(args.base, path)) != daily_semantics(current_bytes(path)):
            errors.append(f"Day {day:02d} Daily Card schedule semantics changed")

    if master_rows(git_bytes(args.base, MASTER)) != master_rows(current_bytes(MASTER)):
        errors.append("Master Day 20–43 rows changed")
    if avignon_day20_body(git_bytes(args.base, AVIGNON)) != avignon_day20_body(current_bytes(AVIGNON)):
        errors.append("Avignon Day 20+ body changed")
    if tracker_rows(git_bytes(args.base, TRACKER)) != tracker_rows(current_bytes(TRACKER)):
        errors.append("Canonical route schedule from 2026-09-17 changed")

    if errors:
        for error in errors:
            print(f"ERROR {error}")
        return 1
    print(f"PASS protected itinerary semantics match {args.base}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
