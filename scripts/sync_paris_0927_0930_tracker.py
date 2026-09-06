#!/usr/bin/env python3
"""Synchronize the 9/27 and 9/30 Paris swap into the operations tracker.

Only those two dates and their confirmed reservations are touched. Date
preconditions make the binary edit fail closed if the workbook layout changes.
"""

from datetime import datetime
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TRACKER = ROOT / "source" / "OPERATIONS" / "TP_Europe_Travel_Master_Tracker_v1.2.xlsx"


def expect_date(ws, row: int, expected: str) -> None:
    value = ws.cell(row, 1).value
    actual = value.date().isoformat() if isinstance(value, datetime) else str(value)
    if actual != expected:
        raise RuntimeError(f"Master Itinerary row {row}: expected {expected}, found {actual}")


def main() -> None:
    workbook = load_workbook(TRACKER)
    sheet = workbook["Master Itinerary"]
    expect_date(sheet, 31, "2026-09-25")
    expect_date(sheet, 33, "2026-09-27")
    expect_date(sheet, 36, "2026-09-30")

    updates = {
        31: {16: "Grand Palais 세잔전 9/25 17:00 예약 확정"},
        33: {
            5: "고전 도시공간",
            6: "Belle Époque & Classical Paris",
            7: "Marché Convention·Pichard 후 숙소 점심·휴식",
            8: "숙소 점심",
            9: "Petit Palais 13:00→Tuileries 15:00→Palais Royal 16:10→Opéra Garnier 17:30",
            10: "Bouillon Chartier Montparnasse 저녁",
            11: "메트로 8+1호선·도보",
            16: "",
            17: "야외 구간 단축",
        },
        36: {
            5: "미술·패션",
            6: "Monet & Contemporary Paris / Fashion",
            7: "09:15 숙소 출발",
            8: "Chez Savy 12:15",
            9: "Orangerie 10:00–11:30→Avenue Montaigne→Grand Palais 공개 동선→Président Wilson→Palais de Tokyo·Alma/Seine",
            10: "Stéphane Martin 19:30",
            11: "메트로 8·9호선·도보",
            16: "Orangerie 10:00 예약 확정·Fashion Week 공개행사 재확인",
            17: "수련 우선·Walter-Guillaume 주요작 후 11:30 종료",
        },
    }
    for row, cells in updates.items():
        for column, value in cells.items():
            sheet.cell(row, column).value = value

    reservations = workbook["Reservations"]
    for ident, name, day, time, url in [
        ("R031", "Grand Palais — Cézanne et nous", 25, "17:00", "https://www.grandpalais.fr/"),
        ("R032", "Musée de l'Orangerie", 30, "10:00", "https://www.musee-orangerie.fr/"),
    ]:
        row = next((i for i in range(4, reservations.max_row + 1)
                    if reservations.cell(i, 1).value == ident), reservations.max_row + 1)
        existing = reservations.cell(row, 4).value
        if existing and existing != name:
            raise RuntimeError(f"Reservation ID collision: {ident}")
        values = {1: ident, 2: "입장권", 3: "Paris", 4: name,
                  5: datetime(2026, 9, day), 6: time, 7: "P1", 8: "확정",
                  18: url, 19: "11_Paris_Long_Stay_v2.0.md",
                  21: datetime(2026, 9, 6),
                  22: "사용자 예약 완료 확인 · 예약번호·금액·취소조건 미제공"}
        for col in range(1, 23):
            reservations.cell(row, col)._style = copy(reservations.cell(33, col)._style)
        for col, value in values.items():
            reservations.cell(row, col).value = value
    reservations.cell(23, 22).value = "Grand Palais 9/25 17:00·Orangerie 9/30 10:00은 별도 확정 행 참조; 이 행은 나머지 미술관 예약"
    workbook.save(TRACKER)
    print("legacy tracker synchronization completed; final PMP itinerary is applied by sync_paris_pmp_tracker.py")


if __name__ == "__main__":
    main()
