#!/usr/bin/env python3
"""Validate the structured itinerary and its tracker/daily-page projections."""

import json
import re
import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "build"))

from itinerary import load_itinerary  # noqa: E402


def main():
    payload = load_itinerary(ROOT)
    trip = payload["trip"]
    stays = payload["stays"]
    errors = []

    expected_by_date = {}
    for stay in stays:
        cursor = date.fromisoformat(stay["checkin"])
        checkout = date.fromisoformat(stay["checkout"])
        while cursor < checkout:
            if cursor.isoformat() in expected_by_date:
                errors.append(f"숙박 날짜 중복: {cursor}")
            expected_by_date[cursor.isoformat()] = stay["base"]
            cursor += timedelta(days=1)

    # 마지막 밤은 숙소가 아니라 기내다 (OZ502, 10/9 CDG → 10/10 인천).
    # 그 박을 거점 숙박으로 세면 파리 체크아웃이 하루 밀린다.
    lodging_nights = trip["nights"] - trip.get("inflightNights", 0)
    expected_dates = {
        (date.fromisoformat(trip["start"]) + timedelta(days=i)).isoformat()
        for i in range(lodging_nights)
    }
    missing = sorted(expected_dates - set(expected_by_date))
    extra = sorted(set(expected_by_date) - expected_dates)
    if missing:
        errors.append("숙박 날짜 누락: " + ", ".join(missing))
    if extra:
        errors.append("숙박 날짜 범위 초과: " + ", ".join(extra))

    tracker = ROOT / "source" / "OPERATIONS" / "TP_Europe_Travel_Master_Tracker_v1.2.xlsx"
    wb = load_workbook(tracker, data_only=True, read_only=True)
    ws = wb["Accommodation"]
    headers = [cell.value for cell in ws[3]]
    rows = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=4, values_only=True)
            if row[0] is not None]
    tracker_by_base = {row["거점"]: row for row in rows}
    for stay in stays:
        row = tracker_by_base.get(stay["base"])
        if not row:
            errors.append(f"트래커 숙소 누락: {stay['base']}")
            continue
        actual = (row["체크인"].date().isoformat(), row["체크아웃"].date().isoformat(), row["박수"])
        expected = (stay["checkin"], stay["checkout"], stay["nights"])
        if actual != expected:
            errors.append(f"트래커 숙박 불일치 {stay['base']}: {actual} != {expected}")

    # ---- 사이트 투영 검사 --------------------------------------------
    # 데이터가 맞아도 화면에 안 나오면 소용이 없다. 아래는 "일정 사실이
    # 실제로 렌더된 페이지에 있는가" 를 본다. URL 은 새 IA 기준이다.
    site = ROOT / "site"

    index_js = site / "assets" / "search-index.js"
    if not index_js.exists():
        errors.append("검색 색인 누락: site/assets/search-index.js")
    else:
        raw = index_js.read_text(encoding="utf-8")
        match = re.fullmatch(r"window\.SEARCH_INDEX\s*=\s*(\[.*\]);\s*", raw, re.S)
        if not match:
            errors.append("search-index.js 해석 실패")
        else:
            entries = json.loads(match.group(1))
            searchable = " ".join(
                f"{e.get('t','')} {e.get('x','')} {e.get('u','')}" for e in entries
            ).casefold()
            for term in ("marseille", "mucem", "fort saint-jean", "arles",
                         "saint-trophime", "la roquette"):
                if term.casefold() not in searchable:
                    errors.append(f"검색 색인 누락: {term}")
            for e in entries:
                url = e.get("u", "")
                if not url or url.startswith(("http://", "https://", "mailto:")):
                    continue
                if not (site / url.split("#", 1)[0]).exists():
                    errors.append(f"검색 링크 대상 없음: {url}")

    daily_dir = site / "daily"
    if daily_dir.exists():
        pages = sorted(daily_dir.glob("day-*.html"))
        if len(pages) != trip["days"]:
            errors.append(f"데일리 페이지 {len(pages)}개 (기대 {trip['days']}개)")

    # 그날 원고의 핵심 장소가 실제로 그 날 화면에 있는가
    required_page_terms = {
        # Day 14 는 2026-08-19 에 Marseille → Cassis·Calanques 로 바뀌었다
        "daily/day-14.html": ("Cassis", "Calanques", "Port-Miou"),
        "daily/day-21.html": ("Arles", "Saint-Trophime", "La Roquette"),
        "daily/day-22.html": ("Palais", "Rocher des Doms", "Pont Saint-Bénézet"),
    }
    # 지역 페이지에는 그 거점의 박수와 날짜가 나와야 한다
    for stay in stays:
        ci = date.fromisoformat(stay["checkin"])
        co = date.fromisoformat(stay["checkout"])
        required_page_terms[f"guide/{stay['key']}.html"] = (
            f"{stay['nights']}박", f"{ci.month}/{ci.day}", f"{co.month}/{co.day}")
    for relative, terms in required_page_terms.items():
        path = site / relative
        if not path.exists():
            errors.append(f"필수 생성 페이지 누락: {relative}")
            continue
        rendered = path.read_text(encoding="utf-8")
        for term in terms:
            if term not in rendered:
                errors.append(f"필수 생성 콘텐츠 누락: {relative} → {term}")

    # 선택안(대체 일정)이 기본 일정의 장소처럼 보이면 안 된다. 현장에서
    # "오늘 가는 곳" 으로 읽고 움직이게 된다.
    forbidden_day_place_links = {
        # Marseille 는 이제 Day 14 의 대체안이다. 대체안이 기본 시간표에 섞이면
        # 현장에서 "오늘 가는 곳" 으로 읽는다 — backup 문구에만 남아야 한다.
        "daily/day-14.html": ("places/arles.html", "places/marseille.html",
                              "places/mucem.html"),
        "daily/day-21.html": ("places/les-baux-de-provence.html",
                              "places/saint-remy-de-provence.html"),
        "daily/day-22.html": ("places/les-baux-de-provence.html",
                              "places/saint-remy-de-provence.html"),
    }
    for relative, links in forbidden_day_place_links.items():
        path = site / relative
        if not path.exists():
            continue
        rendered = path.read_text(encoding="utf-8")
        timeline = rendered.split('class="timeline"', 1)[-1].split("</ol>", 1)[0]
        for link in links:
            if link in timeline:
                errors.append(f"선택안이 기본 일정에 노출됨: {relative} → {link}")

    home = site / "index.html"
    if home.exists():
        rendered = home.read_text(encoding="utf-8")
        for stay in stays:
            marker = f'data-region="{stay["key"]}"'
            if marker not in rendered:
                errors.append(f"홈 거점 누락: {stay['base']}")
                continue
            card = rendered.split(marker, 1)[1].split("</article>", 1)[0]
            if f'{stay["nights"]}박' not in card:
                errors.append(f"홈 숙박 요약 불일치: {stay['base']} → {stay['nights']}박")

    if errors:
        print("일정 검증 실패:")
        for error in errors:
            print("  " + error)
        raise SystemExit(1)

    print(f"일정 검증 통과: {trip['days']}일 · {trip['nights']}박 · 날짜 누락 0 · 중복 0 · 거점 연결 {len(stays)-1}건 일치")


if __name__ == "__main__":
    main()
