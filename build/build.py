#!/usr/bin/env python3
"""TP Europe Travel Guidebook — 정적 사이트 빌드 스크립트.

source/ 의 통합 패키지 v1.37(Phase 10 LatestOnly)을
site/ 아래의 순수 정적 HTML 사이트로 변환한다.

콘텐츠 기준 (CURRENT/00_Governance/00_Current_Source_of_Truth_Index_v1.9.md):
 - 본문: 정식 지역 챕터(20_Regional_Chapters) + Core 문서(10_Core)
 - 지도: ASSETS/75_Execution_Maps 8개 지역
 - 데일리 카드: ASSETS/80_Daily_Mobile_Guide_Images 43장 (Day 12–24는 Phase 4 카드 우선)
 - 트래커: OPERATIONS/TP_Europe_Travel_Master_Tracker_v1.2.xlsx

UI/UX 설계: docs/UIUX_Design_v1.0.md
 - '오늘' 버튼 → 당일 데일리 카드 페이지(카드 이미지 + 챕터 일정 + 지역 지도 링크)

사용법:
    python3 build/build.py

필요 패키지: pip install markdown openpyxl
"""

import html
import itertools
import json
import re
import shutil
import sys
import unicodedata
import urllib.parse
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import icons                     # noqa: E402  — 아이콘 마스크 생성기
import content_model             # noqa: E402  — stable-ID content graph
from xml.etree import ElementTree

try:
    import markdown
    from markdown.extensions.toc import TocExtension, slugify_unicode
except ImportError:
    sys.exit("markdown 패키지가 필요합니다: pip install markdown")

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl 패키지가 필요합니다: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
SOURCE = ROOT / "source"
SITE = ROOT / "site"
ASSETS = ROOT / "build" / "assets"

SITE_TITLE = "2026 유럽 여행 가이드북"
SITE_SHORT = "2026 유럽 여행 가이드북"
TRIP_PERIOD = "2026-08-29 ~ 2026-10-10 · 43일 42박"
TRIP_START = date(2026, 8, 29)
TRIP_END = date(2026, 10, 10)
WEEKDAY_KO = ["월", "화", "수", "목", "금", "토", "일"]

CORE = "CURRENT/10_Core"
REGIONAL = "CURRENT/20_Regional_Chapters"

# 챕터 매니페스트 — kind: intro(안내) / schedule(전체일정) / region(지역 가이드)
CHAPTERS = [
    dict(path=f"{CORE}/01_How_to_Use_This_Guidebook_v1.0.md", slug="01", name="how-to-use",
         kind="intro", title="가이드북 사용법", sub="읽는 법과 기준 문서"),
    dict(path=f"{CORE}/02_Whole_Trip_Experience_Highlights_v1.0.md", slug="02", name="highlights",
         kind="intro", title="전체 여행 하이라이트", sub="43일의 경험 설계"),
    dict(path=f"{CORE}/03_Whole_Trip_Master_Itinerary_v1.2.md", slug="03", name="itinerary",
         kind="schedule", title="43일 Master Itinerary", sub="전체 일정·실행성 감사 반영"),
    dict(path=f"{REGIONAL}/04_Barcelona_Sitges_v2.0.md", slug="04", name="barcelona", kind="region",
         title="Barcelona · Sitges", start=date(2026, 8, 29), end=date(2026, 9, 1),
         nights=3, map="barcelona.html", region="Barcelona"),
    dict(path=f"{REGIONAL}/05_Girona_Collioure_Emporda_v2.1.md", slug="05", name="girona", kind="region",
         title="Girona · Collioure · Empordà", start=date(2026, 9, 1), end=date(2026, 9, 4),
         nights=3, map="girona.html", region="Girona"),
    dict(path=f"{REGIONAL}/06_Nice_Cote_d_Azur_v2.0.md", slug="06", name="nice", kind="region",
         title="Nice · Côte d’Azur", start=date(2026, 9, 4), end=date(2026, 9, 9),
         nights=5, map="nice.html", region="Nice"),
    dict(path=f"{REGIONAL}/07_Aix_en_Provence_v2.0.md", slug="07", name="aix", kind="region",
         title="Aix-en-Provence", start=date(2026, 9, 9), end=date(2026, 9, 13),
         nights=4, map="aix.html", region="Aix"),
    dict(path=f"{REGIONAL}/08_Luberon_Farmhouse_v2.0.md", slug="08", name="luberon", kind="region",
         title="Luberon Farmhouse", start=date(2026, 9, 13), end=date(2026, 9, 17),
         nights=4, map="luberon.html", region="Luberon"),
    dict(path=f"{REGIONAL}/09_Avignon_Alpilles_Pont_du_Gard_v2.0.md", slug="09", name="avignon", kind="region",
         title="Avignon · Alpilles · Pont du Gard", start=date(2026, 9, 17), end=date(2026, 9, 21),
         nights=4, map="avignon.html", region="Avignon"),
    dict(path=f"{REGIONAL}/10_Lyon_v2.0.md", slug="10", name="lyon", kind="region",
         title="Lyon · Annecy", start=date(2026, 9, 21), end=date(2026, 9, 25),
         nights=4, map="lyon.html", region="Lyon"),
    dict(path=f"{REGIONAL}/11_Paris_Long_Stay_v2.0.md", slug="11", name="paris", kind="region",
         title="Paris Long Stay", start=date(2026, 9, 25), end=date(2026, 10, 10),
         nights=15, map="paris.html", region="Paris"),
]

# 실행지도 8종 (ASSETS/75_Execution_Map_Index_v1.0.md 기준)
MAP_DIR = SOURCE / "ASSETS" / "75_Execution_Maps"
MAPS = [
    ("Barcelona_Execution_Map_v0.2.html", "barcelona.html", "Barcelona 실행지도"),
    ("Girona_Execution_Map_v0.2.html", "girona.html", "Girona 실행지도"),
    ("Nice_Execution_Map_v0.2.html", "nice.html", "Nice 실행지도"),
    ("Aix_Execution_Map_v0.2.html", "aix.html", "Aix 실행지도"),
    ("Luberon_Execution_Map_v0.2.html", "luberon.html", "Luberon 실행지도"),
    ("Avignon_Execution_Map_v0.2.html", "avignon.html", "Avignon 실행지도"),
    ("Lyon_Execution_Map_v0.2.html", "lyon.html", "Lyon 실행지도"),
    ("Paris_Execution_Map_v0.2.html", "paris.html", "Paris 실행지도"),
]

# Phase 6 정본 라우팅. 데일리 페이지·지도 목록·회귀 가드가 같은 범위를 쓴다.
MAP_META = {
    "barcelona.html": ("Day 1–4", "Barcelona·Sitges"),
    "girona.html": ("Day 4–7", "Girona·Empordà"),
    "nice.html": ("Day 7–12", "Nice·Côte d’Azur"),
    "aix.html": ("Day 12–16", "Aix-en-Provence"),
    "luberon.html": ("Day 16–20", "Luberon"),
    "avignon.html": ("Day 20–24", "Avignon·Alpilles"),
    "lyon.html": ("Day 24–28", "Lyon·Annecy"),
    "paris.html": ("Day 28–43", "Paris"),
}
MAP_DAY_SPANS = {
    "barcelona.html": range(1, 5), "girona.html": range(4, 8),
    "nice.html": range(7, 13), "aix.html": range(12, 17),
    "luberon.html": range(16, 21), "avignon.html": range(20, 25),
    "lyon.html": range(24, 29), "paris.html": range(28, 44),
}

# 데일리 모바일 가이드 (ASSETS/80_Daily_Mobile_Guide_Image_Index_v1.1.md 기준)
DAILY_IMG_DIR = SOURCE / "ASSETS" / "80_Daily_Mobile_Guide_Images"
PHASE4_DIR = DAILY_IMG_DIR / "Phase4_Provence_Final"
PHASE4_DAYS = set(range(12, 25))  # Day 12–24는 Phase 4 카드 우선

TRACKER_XLSX = SOURCE / "OPERATIONS" / "TP_Europe_Travel_Master_Tracker_v1.2.xlsx"
COMMERCIAL_CARDS = SOURCE / "ASSETS" / "89_Commercial_City_Experience_Cards_v1.0.md"
PLACE_DOSSIERS = SOURCE / "ASSETS" / "90_Regional_Context_and_Place_Dossier_Compendium_v1.0.md"
COMMERCIAL_STANDARD = SOURCE / "CURRENT" / "00_Governance" / "89_Commercial_Guidebook_Editorial_and_Layout_Standard_v1.0.md"
DOSSIER_STANDARD = SOURCE / "CURRENT" / "00_Governance" / "90_Regional_and_Place_Dossier_Editorial_Standard_v1.0.md"
TRACKER_SHEETS = [
    ("Master Itinerary", "itinerary", "43일 전체 일정표"),
    ("Reservations", "reservations", "예약 현황"),
    ("Transport", "transport", "이동·교통"),
    ("Accommodation", "accommodation", "숙소 후보·확정"),
    ("Dashboard", "dashboard", "진행 대시보드"),
    ("Phase8 Lock Status", "locks", "예약·운영 잠금"),
]

DAY_RE = re.compile(r"Day\s*(\d+)\s*[—\-–·]\s*(\d+)월\s*(\d+)일")


def chapter_url(c):
    """챕터의 사이트 URL. 번호가 아니라 지명으로 짓는다 (명명규칙 v1.0).

    지역 챕터는 디렉터리를 쓴다 — 분할된 챕터(T11)와 아직 단일 페이지인
    챕터가 같은 URL 모양을 갖게 해서, 나중에 하나씩 쪼개도 링크가 안 바뀐다.
    """
    if c["kind"] == "region":
        return f'chapters/{c["name"]}/index.html'
    return f'chapters/{c["name"]}.html'


def chapter_rel(c):
    """해당 챕터 페이지에서 사이트 루트로 올라가는 상대경로."""
    return "../.." if c["kind"] == "region" else ".."


ITINERARY_URL = "chapters/itinerary.html"

# 원본의 Day 섹션 헤딩. 레벨(#~######)과 섹션번호 접두어("5. ")가 챕터마다 다르다.
#   `### Day 1 — 8월 29일 토요일`   (04·07)
#   `## 5. Day 1 — 9월 25일 금요일` (06·08·09·10·11)
#   `# 6. Day 1 — 9월 1일 화요일`   (05, regroup 단계에서 h2로 내려옴)
SRC_DAY_HEADING_RE = re.compile(
    r"^#{1,6}[ \t]*(?:\d+[A-Z]?[.)][ \t]*)*"
    r"Day[ \t]*(\d+)[ \t]*[—–-][ \t]*(\d+)월[ \t]*(\d+)일[ \t]*([월화수목금토일])요일[ \t]*$")

# 날짜가 없는 Day 참조 헤딩 (`## Day 1 비` · `### Day 16 — Lourmarin` 등).
# 번호가 로컬인지 전역인지는 정규식으로 못 가른다. resolve_day_ref() 가 판단한다.
SRC_DAY_REF_RE = re.compile(r"^(#{1,6})[ \t]*Day[ \t]*(\d+)[ \t]*(.*)$")

# 인라인 VISUAL 토큰. 원고에는 이미지 자리표시자로 남아 있고 화면에 노출되면 안 된다.
VISUAL_TOKEN_RE = re.compile(r"\{\{VISUAL:[A-Z0-9-]+\|[^}]*\}\}[ \t]*")
FENCE_RE = re.compile(r"^[ \t]*(?:```|~~~)")

# ---------------------------------------------------------------- 분류 체계
# 지역 챕터의 h2 섹션을 카테고리로 재분류·재배치한다.
# 소스는 수정하지 않고 빌드 시 매핑한다 (소스는 상류 콘텐츠 워크플로가 관리).

CATEGORIES = [
    ("intro", "지역소개"),
    ("schedule", "일정"),
    ("info", "여행정보"),
    ("food", "먹거리"),
    ("transport", "교통"),
    ("stay", "숙박"),
    ("booking", "예약"),
    ("cost", "경비"),
    ("tips", "여행팁"),
    ("appendix", "부록"),
]
CAT_LABEL = dict(CATEGORIES)

# 제목 키워드 규칙 — 먼저 맞는 규칙이 이긴다 (번호 접두어 제거 후 적용)
CAT_RULES = [
    ("appendix", r"공식자료|검증 상태|검증 기록|검증 범위|검증 출처|참고 출처|참고자료|편집 메모|최종 결론"
                 r"|최종 편집 판단|시각요소|공식정보 원칙"),
    ("cost", r"예상 현지비용|예상 경비|^경비"),
    ("booking", r"^예약|예약카드|예약 게이트"),
    ("tips", r"대체안|대안 루트|확인목록|현장 메모"),
    ("schedule", r"Day \d|날짜별|일정표|일정 요약|일정 교체|피로도|한눈에 보는|운영 원칙|동선 도식"
                 r"|Quick Reference|실행성 감사|의사결정 게이트|세 사이클|삭제 우선순위|중요 정정"),
    ("transport", r"교통|렌터카|주차|공항|문전 이동|대중교통|자동차|철도"),
    ("stay", r"숙소|생활권|농가"),
    ("intro", r"이해하|어떻게 볼 것인가|도시층|읽는 법|지역 이해|편집자 큐레이션|열쇠|다섯 개의 층"
              r"|Editor’s Verdict|Editor's Verdict"),
    ("food", r"레스토랑|카페|시장|먹어야|식당|장보기|음식|빵|식사체계|먹거리"),
    ("info", r"방문지|관광지|주요 장소|핵심 장소|추천등급|추천 등급|미술관|박물관|도서관|서점|공연|축구"
             r"|근교|체험할|행사|특별전|특별운영|이벤트|선택표|전시|장소별|Top 10|놓치면 아쉬운|하루를 완성"),
    ("tips", r"운동|수영|안전|치안|스케치|지속가능|현장 선택|출발 전|휴식"),
]

# 규칙으로 판별이 어려운 제목의 명시적 지정 (챕터 슬러그, 원문 h2 제목) -> 카테고리
CAT_OVERRIDES = {
    ("05", "시체스에서 지로나 도착, 대성당과 성벽"): "schedule",
    ("05", "콜리우르 시장·왕궁·야수파 산책과 페랄라다"): "schedule",
    ("05", "Pals·Peratallada·Calella de Palafrugell"): "schedule",
    ("05", "지로나에서 니스로 이동"): "schedule",
    ("05", "모듈 A — 동일 차량으로 니스 이동"): "schedule",
    ("05", "모듈 B — 지로나 반납 후 철도 이동"): "schedule",
    ("05", "출발 전 30분 옵션"): "schedule",
    ("05", "5.1 추천 생활권"): "stay",
    ("05", "10.1 로마와 중세 성곽도시"): "intro",
    ("05", "13.1 지로나에서 살 것"): "food",
    ("05", "Jason"): "tips",
    ("05", "Julia"): "tips",
    ("05", "스케치"): "tips",
    ("05", "확정적으로 겹치는 전시"): "info",
    ("05", "9월 2일 행사"): "info",
    ("05", "체류기간 이후"): "info",
    ("05", "지로나"): "transport",
    ("05", "콜리우르"): "transport",
    ("05", "Peratallada·Pals"): "transport",
    ("05", "Calella de Palafrugell"): "transport",
    ("05", "Day 1 비"): "tips",
    ("05", "Day 2 비"): "tips",
    ("05", "Day 3 비"): "tips",
    ("05", "교통지연 기준"): "tips",
    ("05", "Girona Cathedral"): "booking",
    ("05", "Château Royal de Collioure"): "booking",
    ("05", "Museu de Peralada"): "booking",
    ("05", "Day 3 점심"): "booking",
    ("05", "확정"): "booking",
    ("05", "미결정"): "booking",
    ("05", "편집자가 고른 핵심 장소 7곳"): "info",
    ("05", "꼭 체험할 것"): "info",
    ("05", "식당 큐레이션"): "food",
    ("05", "현장 선택 규칙"): "tips",
    ("05", "4. 이 일정이 Jason·Julia에게 맞는 이유"): "schedule",
    ("05", "20. 확정사항과 남은 미결정"): "booking",
    # Girona v2.1 보강본 신설 섹션 — 키워드 규칙으로는 엉뚱하게 갈린다
    ("05", "이 3박이 여행 전체에서 하는 일"): "intro",
    ("05", "엠포르다 요리를 이해하는 한 문장"): "food",
    ("05", "주문할 때 찾을 것"): "food",
    ("05", "날짜별 성격"): "food",
    ("05", "시장"): "food",
    ("05", "이 구간이 가장 복잡한 이유"): "transport",
    ("05", "국경 통과"): "transport",
    ("05", "통행료 — 표기가 반대다"): "transport",
    ("05", "연료"): "transport",
    ("05", "운전 규칙"): "transport",
    ("05", "주차 색 구분"): "transport",
    ("05", "도시별 (기존 원고 유지)"): "transport",
    ("05", "3박 4일 구조 (2인)"): "cost",
    ("05", "입장료"): "cost",
    ("05", "절감 여지"): "cost",
    ("05", "결제"): "cost",
    ("05", "우선순위 요약"): "tips",
    # 검증 등록부에서 생성되는 섹션 — 출발 전에 잠그는 정보라 예약으로 보낸다
    **{(s, "공식 확인 정보와 재확인 대상"): "booking"
       for s in ("04", "05", "06", "07", "08", "09", "10", "11")},
    **{(s, "검증 상태"): "appendix"
       for s in ("04", "05", "06", "07", "08", "09", "10", "11")},
    # Barcelona v2.0 보강본 신설 섹션
    ("04", "카탈루냐 요리의 기본 문법"): "food",
    ("04", "시체스의 xató"): "food",
    ("04", "이 구간의 식사 전략"): "food",
    ("04", "시간대"): "food",
    ("04", "이 구간의 성격 — 아직 차가 없다"): "transport",
    ("04", "지하철·버스"): "transport",
    ("04", "Barcelona Sants — Day 4의 기점"): "transport",
    ("04", "시체스까지"): "transport",
    ("04", "저배출구역 (ZBE)"): "transport",
    ("04", "3박 4일 구조 (2인)"): "cost",
    ("04", "입장료 — 이 구간의 특징"): "cost",
    ("04", "절감 여지"): "cost",
    ("04", "결제"): "cost",
    ("04", "우선순위"): "tips",
    ("04", "이 3박이 여행 전체에서 하는 일"): "intro",
    # Nice v2.0 보강본 신설 섹션
    ("06", "이 5박이 여행 전체에서 하는 일"): "intro",
    ("06", "이 도시를 이해하는 축"): "intro",
    ("06", "니스 요리를 이해하는 한 문장"): "food",
    ("06", "반드시 먹을 것"): "food",
    ("06", "진짜와 가짜를 가르는 세 가지 규칙"): "food",
    ("06", "Cuisine Nissarde 라벨"): "food",
    ("06", "날짜별 전략"): "food",
    ("06", "이 구간의 성격 — 차가 없다가 마지막에 생긴다"): "transport",
    ("06", "니스 시내"): "transport",
    ("06", "칸·모나코 — 기차가 정답이다"): "transport",
    ("06", "Day 12 — 니스 공항 렌터카 인수"): "transport",
    ("06", "운전 (Day 12부터)"): "transport",
    ("06", "5박 6일 구조 (2인)"): "cost",
    ("06", "입장료 — 이 구간의 특징"): "cost",
    ("06", "모나코 주의"): "cost",
    ("06", "절감 여지"): "cost",
    ("06", "우선순위"): "tips",
    ("06", "Èze"): "tips",
    ("06", "Villefranche-sur-Mer"): "tips",
    ("06", "마티스 미술관 · 시미에"): "tips",
    ("06", "샤갈 미술관"): "tips",
    ("06", "Antibes"): "tips",
    ("06", "Menton"): "tips",
    # Aix v2.0 보강본 신설 섹션
    ("07", "이 4박이 여행 전체에서 하는 일"): "intro",
    ("07", "이 도시를 이해하는 축"): "intro",
    ("07", "프로방스 요리를 이해하는 한 문장"): "food",
    ("07", "먹어야 할 것"): "food",
    ("07", "카시스에서는 다르게 먹는다"): "food",
    ("07", "날짜별 배치"): "food",
    ("07", "시장 장보기 — 9월"): "food",
    ("07", "이 구간의 성격 — 차가 생기고 시내에서는 못 쓴다"): "transport",
    ("07", "Day 12 렌터카 인수"): "transport",
    ("07", "프랑스 도로 — 요약"): "transport",
    ("07", "엑상 시내 주차 — 이 구간의 실질 과제"): "transport",
    ("07", "카시스 주차 — Day 14의 병목"): "transport",
    ("07", "저배출구역 (ZFE)"): "transport",
    ("07", "4박 5일 구조 (2인)"): "cost",
    ("07", "입장료 — 확인된 값"): "cost",
    ("07", "절감 여지"): "cost",
    ("07", "예약이 필요한 것 — 우선순위"): "booking",
    ("07", "우선순위"): "tips",
    # Luberon v2.0 보강본 신설 섹션
    ("08", "이 4박이 여행 전체에서 하는 일"): "intro",
    ("08", "이 지역을 이해하는 축"): "intro",
    ("08", "이 구간이 다른 이유 — 사 먹는 곳이 아니라 해 먹는 곳이다"): "food",
    ("08", "4박 5일 구조 (2인)"): "cost",
    ("08", "입장료 — 확인된 값"): "cost",
    ("08", "절감과 증가 요인"): "cost",
    ("08", "우선순위"): "tips",
    ("08", "9/14 Roussillon + Goult 또는 Bonnieux"): "schedule",
    ("08", "숙소 평가 최종 기준"): "stay",
    # Avignon v2.0 보강본 신설 섹션
    ("09", "이 4박이 여행 전체에서 하는 일"): "intro",
    ("09", "이 지역을 이해하는 축"): "intro",
    ("09", "9월 19–20일은 유럽 문화유산의 날이다"): "info",
    ("09", "이 구간의 성격 — 다시 사 먹는 구간이다"): "food",
    ("09", "아비뇽·가르 지역에서 먹어야 할 것"): "food",
    ("09", "와인 — 이 구간이 이 여행의 와인 중심지다"): "food",
    ("09", "날짜별 배치"): "food",
    ("09", "이 구간의 성격 — 차를 쓰다가 놓는다"): "transport",
    ("09", "아비뇽 주차 — 성벽 안에 넣지 마라"): "transport",
    ("09", "Day 22·23 주차"): "transport",
    # 렌터카 반납 섹션 — 제목에 Day 가 들어가 일정으로 빠진다. 교통에 둔다.
    ("09", "Day 24 — 렌터카 반납과 TGV {{badge:p0|P0 연결}}"): "transport",
    ("09", "4박 5일 구조 (2인)"): "cost",
    ("09", "입장료 — 확인된 값"): "cost",
    ("09", "절감 여지"): "cost",
    ("09", "예약 우선순위"): "booking",
    ("09", "우선순위"): "tips",
    ("10", "치안 판단과 여행 설계 반영"): "tips",
    # Lyon v2.0 보강본 신설 섹션
    ("10", "이 4박이 여행 전체에서 하는 일"): "intro",
    ("10", "이 도시를 이해하는 축"): "intro",
    ("10", "리옹 요리를 이해하는 한 문장"): "food",
    ("10", "부숑(bouchon)이란 무엇인가"): "food",
    ("10", "이 구간의 성격 — 차가 없다"): "transport",
    ("10", "Day 24 — 아비뇽에서 리옹으로 {{badge:p0|P0 연결}}"): "transport",
    ("10", "리옹 시내 — TCL"): "transport",
    ("10", "역이 둘이다 — Part-Dieu와 Perrache"): "transport",
    ("10", "Day 27 — 안시 왕복"): "transport",
    ("10", "Day 28 — 파리로 {{badge:p0|P0 연결}}"): "transport",
    ("10", "안전"): "transport",
    ("10", "4박 5일 구조 (2인)"): "cost",
    ("10", "입장료"): "cost",
    ("10", "절감 여지"): "cost",
    ("10", "예약 우선순위"): "booking",
    # 대안 루트에 딸린 후보들 — 박물관·마을 이름이라 여행정보로 빠진다
    ("10", "우선순위"): "tips",
    ("10", "Musée des Tissus (직물박물관)"): "tips",
    ("10", "Pérouges"): "tips",
    ("10", "Vienne"): "tips",
    ("10", "Beaujolais"): "tips",
    ("10", "넣지 말아야 할 것"): "tips",
    # Paris v2.0 보강본 신설 섹션
    ("11", "이 16일이 여행 전체에서 하는 일"): "intro",
    ("11", "16일 개요"): "schedule",
    ("11", "요일이 일정을 만들었다"): "schedule",
    ("11", "2026년이 만든 세 개의 고정점"): "info",
    ("11", "15박의 운영 원칙"): "tips",
    ("11", "예약 게이트 — 출발 전에 끝내야 할 것"): "booking",
    # 대안 루트 한 덩어리 — D·E·F 는 '일정'·'추가' 낱말 때문에 흩어진다
    ("11", "먼저 — 건드리면 안 되는 것"): "tips",
    ("11", "대안 D — 일정 교체 매트릭스"): "tips",
    ("11", "대안 E — 추가하고 싶어질 것들"): "tips",
    ("11", "대안 F — 이월된 항목"): "tips",
    ("11", "마지막 이틀"): "tips",
}

# 모듈 h1 — h1은 제거하고 하위 h2를 개별 분류한다 (직속 도입부의 기본 카테고리)
MODULE_H1 = {
    "Commercial Guide Module": "intro",
    "Regional Context & Scheduled Place Dossiers": "intro",
}
TITLE_H1_RE = re.compile(r"^(Chapter\s*)?\d+\.\s")

# 지역 대표 사진 (ASSETS/88_Representative_Public_Photo_Credits_v1.0.md 기준, CC 라이선스)
HERO_DIR = SOURCE / "ASSETS" / "88_Representative_Public_Photos"
HERO_PHOTOS = {
    "04": ("01_Barcelona_Sagrada_Familia_CC_BY_SA_4_0.jpg", "Sagrada Família", "Lolo7433",
           "CC BY-SA 4.0", "https://creativecommons.org/licenses/by-sa/4.0/",
           "https://commons.wikimedia.org/wiki/File:Sagrada_Familia_Ext%C3%A9rieur.jpg"),
    "05": ("02_Girona_Onyar_Houses_CC_BY_SA_4_0.jpg", "Onyar 강변 주택", "Guyw4444",
           "CC BY-SA 4.0", "https://creativecommons.org/licenses/by-sa/4.0/",
           "https://commons.wikimedia.org/wiki/File:The_houses_of_the_Onyar.jpg"),
    "06": ("03_Nice_Promenade_des_Anglais_CC_BY_SA_2_5.jpg", "Promenade des Anglais", "Floflo",
           "CC BY-SA 2.5", "https://creativecommons.org/licenses/by-sa/2.5/",
           "https://commons.wikimedia.org/wiki/File:Nice_promenade-anglais.jpg"),
    "07": ("04_Aix_Cours_Mirabeau_CC_BY_2_0.jpg", "Cours Mirabeau", "Andrea Schaffer",
           "CC BY 2.0", "https://creativecommons.org/licenses/by/2.0/",
           "https://commons.wikimedia.org/wiki/File:Cours_Mirabeau,_Aix-en-Provence.jpg"),
    "08": ("05_Luberon_Gordes_CC_BY_SA_4_0.jpg", "Gordes", "Einaz80",
           "CC BY-SA 4.0", "https://creativecommons.org/licenses/by-sa/4.0/",
           "https://commons.wikimedia.org/wiki/File:Village_of_Gordes.jpg"),
    "09": ("06_Avignon_Palais_des_Papes_CC_BY_SA_4_0.jpg", "Palais des Papes", "Gzen92",
           "CC BY-SA 4.0", "https://creativecommons.org/licenses/by-sa/4.0/",
           "https://commons.wikimedia.org/wiki/File:Palais_des_papes_(Avignon)_(9).jpg"),
    "10": ("07_Lyon_Saone_Fourviere_CC_BY_SA_4_0.jpg", "Saône와 Fourvière", "Gloumouth1",
           "CC BY-SA 4.0", "https://creativecommons.org/licenses/by-sa/4.0/",
           "https://commons.wikimedia.org/wiki/File:Panoramics_Lyon_Sa%C3%B4ne_Fourvi%C3%A8re.jpg"),
    "11": ("08_Paris_Seine_Sunset_CC_BY_SA_4_0.jpg", "Seine 일몰", "DiscoA340",
           "CC BY-SA 4.0", "https://creativecommons.org/licenses/by-sa/4.0/",
           "https://commons.wikimedia.org/wiki/File:River_Seine_at_Sunset_in_Paris.jpg"),
}

# 편집 도식 (ASSETS/85_Editorial_Visuals_Index_v1.0.md 권장 위치 기준, 내부 제작 자산)
VISUALS_DIR = SOURCE / "ASSETS" / "85_Editorial_Visuals"
VISUALS = {
    "route": "01_Whole_Trip_Route_and_Stay_Structure_v1.0.png",
    "rhythm": "02_Living_Travel_Daily_Rhythm_v1.0.png",
    "fatigue": "03_Fatigue_and_Deletion_Hierarchy_v1.0.png",
    "risk": "04_Booking_and_Operational_Risk_Matrix_v1.0.png",
    "cardays": "05_Provence_Car_Day_Operating_Logic_v1.0.png",
    "cycles": "06_Paris_Long_Stay_Three_Cycles_v1.0.png",
}


def hero_figure(slug, rel=".."):
    """지역소개 첫머리의 대표 사진 (CC 저작자·라이선스 표시)."""
    if slug not in HERO_PHOTOS:
        return ""
    fname, subject, author, lic, lic_url, src_url = HERO_PHOTOS[slug]
    return f"""<figure class="hero-photo">
<img src="{rel}/assets/heroes/{slug}.jpg" alt="{html.escape(subject)}" loading="lazy">
<figcaption>{html.escape(subject)} — 사진: {html.escape(author)},
<a href="{lic_url}" target="_blank" rel="noopener">{lic}</a>
(<a href="{src_url}" target="_blank" rel="noopener">Wikimedia Commons</a>, 크롭·리사이즈)</figcaption>
</figure>"""


def net_note(detail=""):
    """오프라인일 때만 보이는 안내. nav.js 가 hidden 을 토글한다.

    기본이 hidden 이라 스크립트가 죽으면 안 보인다 — 잘못된 안내가 떠 있는
    것보다 낫다 (K4 점진적 향상).
    """
    tail = f" {html.escape(detail)}" if detail else ""
    return (f'<p class="offline-note net-note" hidden>'
            f'<b>오프라인입니다.</b>{tail}</p>')


def check_hero_credits():
    """HERO_PHOTOS 가 크레딧 원본과 어긋나지 않았는지 확인한다.

    저작자 표시는 CC 라이선스의 배포 조건이라 조용히 틀리면 안 된다.
    원본 표가 갱신됐는데 build.py 상수가 그대로면 빌드를 중단시킨다.
    """
    src = SOURCE / "ASSETS" / "88_Representative_Public_Photo_Credits_v1.0.md"
    text = src.read_text(encoding="utf-8")
    rows = {}
    for line in text.splitlines():
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) != 8 or not cells[0].isdigit():
            continue
        fname = re.search(r"\[([^\]]+\.jpg)\]", cells[3])
        author = cells[4]
        lic = re.match(r"\[([^\]]+)\]", cells[5])
        if fname and lic:
            rows[fname.group(1)] = (author, lic.group(1))
    problems = []
    for slug, (fname, _subject, author, lic, *_rest) in sorted(HERO_PHOTOS.items()):
        if fname not in rows:
            problems.append(f"{slug}: {fname} 이 크레딧 표에 없음")
            continue
        src_author, src_lic = rows[fname]
        if (src_author, src_lic) != (author, lic):
            problems.append(f"{slug}: 원본은 {src_author}/{src_lic}, build.py 는 {author}/{lic}")
    if len(rows) != len(HERO_PHOTOS):
        problems.append(f"크레딧 표 {len(rows)}행 vs HERO_PHOTOS {len(HERO_PHOTOS)}개")
    if problems:
        print(f"대표사진 크레딧 불일치 ({src.name}):")
        for p in problems:
            print("  " + p)
        sys.exit(1)


def build_regions():
    """지역 인덱스 — 하단탭 '지역' 의 도착지.

    지역 챕터가 드로어 2뎁스에만 있어 3탭 이내 도달이 안 됐다.
    라벨에 챕터 번호를 쓰지 않는다 (명명규칙 v1.0).
    """
    cards = []
    for c in CHAPTERS:
        if c["kind"] != "region":
            continue
        first, last = day_no(c["start"]), day_no(c["end"])
        # a 안에 a 를 넣을 수 없다. 카드는 div 로 두고 제목만 링크로 만든다.
        cards.append(f"""<div class="card rg-card">
<a class="card-title" href="{chapter_url(c)}">{html.escape(c["title"])}</a>
<span class="card-sub">Day {first}–{last} · {date_label(c["start"])}–{date_label(c["end"])}
 · {c["nights"]}박</span>
<span class="pl-links"><a href="maps/{c["map"]}">실행지도</a>
<a href="daily/day-{first:02d}.html">첫날 카드</a></span>
</div>""")
    body = f"""<h1>지역</h1>
<p class="meta">8개 거점을 이동 순서대로 놓았다. 각 지역 챕터로 바로 들어간다.</p>
<div class="grid">{"".join(cards)}</div>

<div class="related"><a href="{ITINERARY_URL}"><b class="ic ic-list" aria-hidden="true"></b>43일 전체 일정표</a>
<a href="daily/index.html"><b class="ic ic-today" aria-hidden="true"></b>데일리 카드 43일</a>
<a href="maps/offline.html"><b class="ic ic-map" aria-hidden="true"></b>오프라인 지도 준비</a></div>"""
    (SITE / "regions.html").write_text(
        page("지역", body, rel=".", back=crumbs_for(("지역", None)),
             coords=coords_bar(".", day=("43일", "daily/index.html"),
                               region=("8곳", None),
                               topic=("분류·상태", "topics/index.html"))),
        encoding="utf-8")
    SEARCH_INDEX.append({"t": "지역", "c": "목록", "u": "regions.html"})
    print(f"  지역 인덱스: {len(cards)}개 거점 → regions.html")


def build_credits():
    """사진 저작자 표시 페이지. CC BY / CC BY-SA 의 배포 조건이다."""
    check_hero_credits()
    region_name = {c["slug"]: c["region"] for c in CHAPTERS if c["kind"] == "region"}
    rows = []
    for slug, (fname, subject, author, lic, lic_url, src_url) in sorted(HERO_PHOTOS.items()):
        rows.append(f"""<figure class="credit-item">
<img src="assets/heroes/{slug}.jpg" alt="{html.escape(subject)}" loading="lazy">
<figcaption>
  <b>{html.escape(region_name.get(slug, slug))} — {html.escape(subject)}</b>
  <span>저작자 {html.escape(author)}</span>
  <span>라이선스 <a class="needs-net" target="_blank" rel="noopener license"
    href="{lic_url}">{lic}</a></span>
  <span>원본 <a class="needs-net" target="_blank" rel="noopener"
    href="{src_url}">Wikimedia Commons</a></span>
  <span>수정 가이드북용 크롭·리사이즈 (파생본)</span>
</figcaption>
</figure>""")
    body = f"""<h1>사진 저작자 표시</h1>
<p class="meta">이 가이드북의 지역 대표사진 {len(HERO_PHOTOS)}장은 Wikimedia Commons 의
공개 라이선스 사진을 가이드북용으로 크롭·리사이즈한 파생본이다.</p>

{net_note("저작자와 라이선스는 그대로 읽힙니다. 라이선스 전문과 원본 링크만 연결이 필요합니다.")}

<div class="credit-list">{"".join(rows)}</div>

<h2>사용 조건</h2>
<ul>
<li>CC BY · CC BY-SA 의 저작자 표시와 라이선스 링크를 유지한다.</li>
<li>CC BY-SA 사진을 수정한 파생본은 동일하거나 호환되는 조건으로 배포한다.</li>
<li>원본을 대체하지 않고 파생본임을 명시한다.</li>
</ul>

<h2>그 밖의 자료</h2>
<ul>
<li>본문 서체는 <b>나눔고딕</b> (NHN Corporation · 디자인 Sandoll Communications) 이며
  <a class="needs-net" target="_blank" rel="noopener license"
  href="https://scripts.sil.org/OFL">SIL Open Font License 1.1</a> 로 배포된다.
  CDN 을 쓰지 않고 <code>assets/vendor/nanum/</code> 에 번들했다 — 오프라인에서도 뜬다.
  라이선스 전문은 같은 폴더의 <code>LICENSE.txt</code> 다.</li>
<li>색은 스페인·프랑스 국기색에서 뽑았다. 국기 원색을 그대로 쓰면 야외에서
  읽히지 않는 값이 있어(스페인 노랑은 흰 바탕에 1.7:1) 글자용은 명암비를 맞춰 낮췄다.</li>
<li>편집 도식과 데일리 카드는 이 여행을 위해 직접 제작했다.</li>
<li>실행지도의 배경 타일은 <a class="needs-net" target="_blank" rel="noopener"
  href="https://www.openstreetmap.org/copyright">OpenStreetMap</a> 기여자들의 것이며
  ODbL 조건으로 제공된다. 지도 라이브러리는 Leaflet (BSD-2-Clause) 을 로컬 번들로 쓴다.</li>
<li>주요 방문지 카드의 사진은 온라인일 때 Wikipedia 에서 불러온다. 각 사진의 출처는
  사진 위 링크에 표시된다.</li>
</ul>

<p class="offline-note">원본 표 — <code>source/ASSETS/88_Representative_Public_Photo_Credits_v1.0.md</code>.
빌드가 이 표와 대조해 저작자·라이선스가 어긋나면 중단한다.</p>"""
    (SITE / "credits.html").write_text(
        page("사진 저작자 표시", body, rel=".", back=crumbs_for(("저작자 표시", None))), encoding="utf-8")
    SEARCH_INDEX.append({"t": "사진 저작자 표시", "c": "라이선스", "u": "credits.html"})
    print(f"  저작자 표시: 대표사진 {len(HERO_PHOTOS)}장 → credits.html")


def visual_figure(key, caption, rel="../assets"):
    return (f'<figure class="ed-visual"><img src="{rel}/visuals/{VISUALS[key]}" '
            f'alt="{html.escape(caption)}" loading="lazy">'
            f'<figcaption>{html.escape(caption)}</figcaption></figure>')

NUM_PREFIX_RE = re.compile(r"^\d+[A-Z]?[\.\)]\s*")
SUBNUM_RE = re.compile(r"^\d+\.\d+\s")


def classify(slug, title, prev_cat):
    """h2 섹션 제목을 카테고리로 분류한다."""
    if (slug, title) in CAT_OVERRIDES:
        return CAT_OVERRIDES[(slug, title)]
    if SUBNUM_RE.match(title):        # 7.3 같은 하위번호는 직전 섹션을 따라간다
        return prev_cat or "appendix"
    norm = NUM_PREFIX_RE.sub("", title)
    for cat, pattern in CAT_RULES:
        if re.search(pattern, norm):
            return cat
    return prev_cat or "appendix"

# 빌드 중 수집되는 전역 데이터
CHAPTER_DATE_URL = {}  # 'YYYY-MM-DD' -> 챕터(#Day 앵커) URL — 데일리 페이지에서 사용
DAY_OVERRIDES = {}     # 'YYYY-MM-DD' -> Day 섹션 앵커 URL (범위 매핑보다 우선)
# Day 번호 -> [(지역, 제목, 원고 마크다운)]. 챕터 빌드가 채우고 데일리 카드가 렌더한다.
# 거점 이동일은 출발지·도착지 원고가 둘 다 들어와 항목이 둘이다.
DAY_MD = {}
SEARCH_INDEX = []      # {t: 제목, c: 위치, u: URL}


def regroup_regional(slug, body_md, rel=".."):
    """지역 챕터 본문을 카테고리 순서로 재편성한 마크다운을 만든다.

    h1 유형별 처리:
    - 챕터 제목 h1 (`04.` / `Chapter 05.` 형태의 첫 등장) → 헤더 영역 (+ 부제 h2)
    - Layer h1 → 제거, 하위 h2 개별 분류
    - Pass B/C h1 → h2로 낮춰 하위와 함께 '일정' (예외는 CAT_OVERRIDES)
    - 모듈 h1 (Commercial Guide Module 등) → 제거, 하위 h2 개별 분류
    - 그 외 번호 h1 (재구조화된 챕터의 섹션) → h2로 낮춰 일반 분류
    각 카테고리는 `# 카테고리명` h1로 시작한다 (서브내비 앵커).
    """
    lines = body_md.splitlines()
    header, sections = [], []   # sections: [title, cat, [lines]]
    cur = None                  # 현재 수집 중인 섹션
    mode = "start"              # start | header | sections
    pass_group = False
    seen_title = False
    prev_cat = None

    def close():
        nonlocal cur
        if cur:
            sections.append(cur)
            cur = None

    for line in lines:
        h1 = re.match(r"^# (.+)", line)
        h2 = re.match(r"^## (.+)", line)
        if h1:
            close()
            title = h1.group(1).strip()
            pass_group = False
            if re.match(r"Layer\s*\d", title):
                mode = "sections"
                cur = ["브리프", "appendix", []]   # h1 직속 내용 임시 수집
                continue
            if re.match(r"Pass\s*[B-Z]", title):
                mode, pass_group = "sections", True
                prev_cat = "schedule"
                cur = [title, "schedule", [f"## {title}"]]
                continue
            if title in MODULE_H1:
                mode = "sections"
                prev_cat = MODULE_H1[title]
                cur = ["브리프", MODULE_H1[title], []]
                continue
            if not seen_title and TITLE_H1_RE.match(title):
                seen_title = True
                mode = "header"
                header.append(line)
                continue
            # 재구조화된 챕터의 번호 h1 섹션 — h2로 낮춰 일반 분류
            mode = "sections"
            cat = classify(slug, title, prev_cat)
            cur = [title, cat, [f"## {title}"]]
            prev_cat = cat
            continue
        if h2:
            title = h2.group(1).strip()
            if mode == "header":
                close()
                cat = None
                if "편집 메모" in title:
                    cat = "appendix"
                elif re.search(CAT_RULES[0][1], NUM_PREFIX_RE.sub("", title)):
                    cat = "appendix"   # 부록성 h2 (공식정보 원칙 등)
                if cat:
                    cur = [title, cat, [line]]
                else:
                    header.append(line)   # 부제 h2 — cur가 None이므로 후속 줄도 헤더로
                continue
            if pass_group:
                cat = CAT_OVERRIDES.get((slug, title))
                if cat is None:
                    close()
                    cur = [title, "schedule", [f"### {title}"]]
                    prev_cat = "schedule"
                    continue
                close()
                cur = [title, cat, [line]]
                prev_cat = cat
                continue
            cat = classify(slug, title, prev_cat)
            close()
            cur = [title, cat, [line]]
            prev_cat = cat
            continue
        if cur is not None:
            cur[2].append(line)
        elif mode in ("header", "start"):
            header.append(line)
    close()

    # 빈 '브리프' 의사섹션 제거, 내용 있으면 부록으로
    sections = [s for s in sections
                if not (s[0] == "브리프" and not any(x.strip() for x in s[2]))]

    # 소스에 경비 섹션이 없는 챕터에는 트래커로 안내하는 스텁을 넣어 메뉴를 일관되게 유지
    if not any(s[1] == "cost" for s in sections):
        sections.append(["경비 안내", "cost", [
            "## 경비 참고처",
            "",
            "이 지역 챕터의 소스에는 아직 별도 경비 정리가 없다. 다음을 참고한다.",
            "",
            "- 숙박 예산: 위 **숙박** 카테고리의 숙소 전략·예산 항목",
            f"- 전체 예산과 예약 지출: [진행 대시보드]({rel}/tracker/dashboard.html)"
            f" · [예약 현황]({rel}/tracker/reservations.html)",
            "- 입장료·식비 기준: 본문 각 장소·식당 항목의 가격 표기",
        ]])

    out = header[:]
    counts = {}
    for key, label in CATEGORIES:
        matched = [s for s in sections if s[1] == key]
        if not matched:
            continue
        counts[label] = len(matched)
        out += ["", f"# {label}", ""]
        for s in matched:
            out += s[2] + [""]
    unknown = [s[0] for s in sections if s[1] not in CAT_LABEL]
    if unknown:
        print(f"  경고: 미분류 섹션({slug}): {unknown}")
        sys.exit(1)
    return "\n".join(out), counts


# ---------------------------------------------------------------- 장소 큐레이션
# 여행정보 카테고리 상단의 '주요 방문지' 카드. 이름·설명은 큐레이션, 좌표·Google Maps
# 링크는 실행지도 geojson/html에서 가져온다. 사진은 열람 시 브라우저가 Wikipedia
# REST API에서 불러오는 점진적 향상 방식 (오프라인·실패 시 자동 숨김).
PLACES = {
    "04": [
        ("Sagrada Família", "Sagrada Família", "en", "가우디 필생의 성당 — 1882년 착공해 지금도 건축 중"),
        ("Sant Pau", "Hospital de Sant Pau", "en", "세계문화유산 모데르니스메 병원 단지"),
        ("Gòtic", "Gothic Quarter, Barcelona", "en", "로마 성벽 위에 쌓인 중세 고딕 지구"),
        ("Biblioteca de Catalunya", "Biblioteca de Catalunya", "en", "15세기 병원 건물에 들어선 카탈루냐 도서관"),
        ("MACBA", "Museu d'Art Contemporani de Barcelona", "en", "라발 지구의 현대미술관"),
        ("Barcelona Sants", "Barcelona Sants railway station", "en", "고속철·근교선이 모이는 중앙역"),
        ("Sitges", "Sitges", "en", "해변·구시가·미술관의 휴양 소도시"),
    ],
    "05": [
        ("Girona Cathedral", "Girona Cathedral", "en", "세계에서 가장 넓은 고딕 신랑을 가진 대성당"),
        ("Onyar Houses", "Onyar", "en", "오냐르 강변의 색색 파사드와 붉은 철교"),
        ("Collioure", "Collioure", "en", "야수파 화가들이 사랑한 프랑스 카탈루냐 항구마을"),
        ("Peralada", "Peralada", "en", "성과 와이너리의 엠포르다 귀족 마을"),
        ("Pals", "Pals", "en", "엠포르다 평야를 내려다보는 중세 석조마을"),
        ("Peratallada", "Peratallada", "en", "돌을 깎아 만든 해자와 요새의 마을"),
        ("Calella de Palafrugell", "Calella de Palafrugell", "en", "코스타브라바의 어촌 해변마을"),
    ],
    "06": [
        ("Cours Saleya", "Cours Saleya", "en", "니스 구시가의 식품·꽃 시장 거리"),
        ("Castle Hill", "Castle Hill, Nice", "en", "천사의 만과 항구를 내려다보는 전망 언덕"),
        ("Nice-Ville", "Gare de Nice-Ville", "en", "칸·모나코행 TER이 출발하는 중앙역"),
        ("Cannes", "Cannes", "en", "영화제의 도시 — 르 쉬케 언덕과 크루아제트"),
        ("Monaco", "Monaco", "en", "왕궁과 몬테카를로의 도시국가"),
        ("Libération Market", None, "en", "현지 생활형 아침시장 — 9/8 회복일의 장보기"),
        ("NCE T2", "Nice Côte d'Azur Airport", "en", "9/9 렌터카 인수 지점 (Terminal 2)"),
    ],
    "07": [
        ("Rotonde", "Fontaine de la Rotonde", "en", "미라보 대로 초입의 대분수 로터리"),
        ("Cours Mirabeau", "Cours Mirabeau", "en", "플라타너스 그늘이 덮는 엑상의 중심 산책로"),
        ("Musée Granet", "Musée Granet", "en", "세잔과 유럽 회화의 미술관"),
        ("Cassis", "Cassis", "en", "칼랑크 석회암 절벽 아래의 항구마을"),
        ("Atelier Cézanne", "Atelier de Cézanne", "fr", "세잔이 말년을 보낸 아틀리에"),
        ("Lourmarin", "Lourmarin", "en", "카뮈가 잠든 뤼베롱 초입 마을 — 9/13 경유"),
    ],
    "08": [
        ("Coustellet", None, "en", "농산물 직판장이 서는 교차 마을"),
        ("Roussillon", "Roussillon, Vaucluse", "en", "오커 채석장의 붉은 절벽 마을"),
        ("Goult", "Goult", "en", "조용한 언덕 위 와인 마을"),
        ("Gordes", "Gordes", "en", "절벽에 쌓아 올린 뤼베롱의 대표 석조마을"),
        ("Village des Bories", "Village des Bories", "fr", "돌로만 쌓은 옛 농경 오두막 마을"),
        ("Ménerbes", "Ménerbes", "en", "『프로방스에서의 1년』의 무대"),
        ("L’Isle-sur-la-Sorgue", "L'Isle-sur-la-Sorgue", "en", "물레방아와 골동품 시장의 수상 마을"),
    ],
    "09": [
        ("Les Halles", "Les Halles d'Avignon", "fr", "아비뇽의 실내 중앙시장"),
        ("Palais des Papes", "Palais des Papes", "en", "14세기 교황들이 머문 거대한 궁전"),
        ("Pont Saint-Bénézet", "Pont Saint-Bénézet", "en", "노래로 남은 론 강의 끊어진 다리"),
        ("Uzès", "Uzès", "en", "토요시장이 유명한 공작령 도시"),
        ("Pont du Gard", "Pont du Gard", "en", "로마 수도교 — 세계문화유산"),
        ("Les Baux", "Les Baux-de-Provence", "en", "석회암 바위산 위의 요새 마을"),
        ("Saint-Rémy", "Saint-Rémy-de-Provence", "en", "고흐가 요양하며 그림을 그린 마을"),
    ],
    "10": [
        ("Bellecour", "Place Bellecour", "en", "유럽 최대급 광장 — 리옹의 중심"),
        ("Fourvière", "Basilica of Notre-Dame de Fourvière", "en", "도시를 내려다보는 언덕 위 대성당"),
        ("Vieux Lyon", "Vieux Lyon", "en", "르네상스 골목과 트라불의 구시가"),
        ("Croix-Rousse", "La Croix-Rousse", "en", "비단직공의 언덕 동네"),
        ("Halles Paul Bocuse", "Les Halles de Lyon-Paul Bocuse", "en", "미식도시 리옹의 실내 시장"),
        ("Parc Tête d’Or", "Parc de la Tête d'or", "en", "호수가 있는 대공원 — 러닝 코스"),
        ("Annecy", "Annecy", "en", "알프스 호수와 운하의 도시 — 당일치기"),
    ],
    "11": [
        ("Notre-Dame", "Notre-Dame de Paris", "en", "복원을 마친 시테섬의 대성당"),
        ("Louvre", "Louvre", "en", "세계 최대의 미술관"),
        ("Montmartre", "Montmartre", "en", "사크레쾨르와 화가들의 언덕"),
        ("BnF Richelieu", "Bibliothèque nationale de France", "en", "리슐리외 열람실의 국립도서관"),
        ("Grand Palais", "Grand Palais", "en", "특별전이 열리는 대전시장"),
        ("Orsay", "Musée d'Orsay", "en", "기차역을 개조한 인상파 미술관"),
        ("Bourse de Commerce", "Bourse de Commerce", "en", "피노 컬렉션의 현대미술관"),
        ("Versailles", "Palace of Versailles", "en", "절대왕정의 궁전과 정원 — 근교 옵션"),
        ("Giverny", "Giverny", "en", "모네의 정원 마을 — 근교 옵션"),
    ],
}


def load_map_links():
    """실행지도 HTML에서 장소별 Google Maps 링크를 추출한다."""
    links = {}
    for src_name, _, _ in MAPS:
        text = (MAP_DIR / src_name).read_text(encoding="utf-8")
        m = re.search(r"const pts=(\[.*?\]);", text, re.S)
        if m:
            for pt in json.loads(m.group(1)):
                links[pt["name"]] = pt.get("url", "")
    return links


def places_block(chapter, map_links, rel=".."):
    """여행정보 카테고리 상단의 주요 방문지 카드 HTML."""
    places = PLACES.get(chapter["slug"])
    if not places:
        return ""
    cards = []
    for name, wiki, lang, desc in places:
        gmaps = map_links.get(name) or (
            "https://www.google.com/maps/search/?api=1&query=" + name.replace(" ", "+"))
        wiki_attr = (f' data-wiki="{html.escape(wiki, quote=True)}" data-wlang="{lang}"'
                     if wiki else "")
        cards.append(f"""<div class="pl-card"{wiki_attr}>
  <div class="pl-photo" hidden><img alt="{html.escape(name)}" loading="lazy"><a class="pl-credit"
    target="_blank" rel="noopener" href="#">사진: Wikipedia</a></div>
  <div class="pl-body">
    <b>{html.escape(name)}</b>
    <p>{html.escape(desc)}</p>
    <div class="pl-links"><a target="_blank" rel="noopener" href="{html.escape(gmaps)}">Google Maps</a>
    <a href="{rel}/maps/{chapter["map"]}">실행지도</a></div>
  </div>
</div>""")
    return ('<section class="places"><h3 class="ic ic-pin">주요 방문지</h3>'
            '<p class="note">사진은 온라인 상태에서 Wikipedia로부터 불러옵니다.</p>'
            f'{net_note("Google Maps 링크와 방문지 사진은 연결되면 다시 동작합니다.")}'
            f'<div class="pl-grid">{"".join(cards)}</div></section>')


# ---------------------------------------------------------------- utilities

# 본문 중간에 남은 frontmatter. 원고 여럿이 대표사진과 Commercial Guide Module
# 로 시작해 정작 챕터 frontmatter 가 파일 머리가 아니다. 앞머리만 훑으면
# 이 블록이 본문 문단으로 렌더돼 화면에 YAML 이 그대로 나온다.
# 키를 ASCII 로 제한해 `---` 수평선 뒤의 한국어 문장과 섞이지 않게 한다.
EMBEDDED_FM_RE = re.compile(
    r"^---[ \t]*\n((?:[A-Za-z_][A-Za-z0-9_-]*:[^\n]*\n)+)---[ \t]*\n", re.M)


def _read_fm(chunk, meta):
    for line in chunk.strip().splitlines():
        if ":" in line:
            key, _, value = line.partition(":")
            meta[key.strip()] = value.strip().strip('"')


def parse_frontmatter(text):
    """단순 YAML frontmatter를 dict로 파싱하고 본문을 돌려준다.

    파일 머리의 블록과, 본문 중간에 남은 블록 하나를 모두 걷어낸다.
    """
    meta = {}
    if text.startswith("---"):
        end = text.find("\n---", 3)
        if end != -1:
            _read_fm(text[3:end], meta)
            text = text[end + 4:]
    m = EMBEDDED_FM_RE.search(text)
    if m:
        _read_fm(m.group(1), meta)
        text = text[:m.start()] + text[m.end():]
    return meta, text.lstrip("\n")


def strip_visual_tokens(md_text):
    """인라인 `{{VISUAL:...}}` 토큰을 본문에서 걷어낸다.

    토큰은 대부분 헤딩 앞머리에 붙어 있고 뒤에 실제 제목이 이어진다
    (`### {{VISUAL:VIS-MAP-055|...}} 리옹 4박 전체 생활권과 동선 지도`).
    토큰만 지우면 제목이 그대로 남으므로 내용 손실이 없다.

    **펜스 코드블록 안은 건드리지 않는다.** 챕터 01의 '시각자료 표기' 절은
    토큰 문법 자체를 예시로 보여주는 곳이라 지우면 설명이 무너진다.
    """
    out, in_fence, removed = [], False, 0
    for line in md_text.splitlines():
        if FENCE_RE.match(line):
            in_fence = not in_fence
            out.append(line)
            continue
        if not in_fence and "{{VISUAL:" in line:
            line, n = VISUAL_TOKEN_RE.subn("", line)
            removed += n
        out.append(line)
    text = "\n".join(out)
    # 토큰이 헤딩 전체를 차지했던 경우의 안전장치 — 빈 헤딩은 지운다
    text = re.sub(r"^#{1,6}[ \t]*$\n?", "", text, flags=re.M)
    return text, removed


BADGE_RE = re.compile(r"\{\{badge:([a-z0-9]+)\|([^}|]+)\}\}")
GRADE_RE = re.compile(r"\{\{grade:([a-z]+)\|([^}|]+)\}\}")
BADGE_KINDS = {"p0", "pending", "done", "rest"}
GRADE_KINDS = {"essential", "priority", "optional", "alternative", "excluded"}

# 원고의 추천등급 표기 → 등급 슬러그. 모양(■●○◇▨)으로 구분되므로 색만으로
# 정보를 전달하지 않는다. 여기에 없는 변형(`필수에 가까운 우선 추천` 등)은
# 그대로 둔다 — 다섯 칸에 억지로 밀어넣지 않는다.
GRADE_WORDS = {
    "필수": "essential", "Essential": "essential",
    "우선 추천": "priority", "Priority": "priority",
    "선택": "optional", "Optional": "optional",
    "대체": "alternative", "Alternative": "alternative",
    "비추천": "excluded", "Not recommended": "excluded",
}

# 값이 확정이 아닌 열 — 원고가 변동정보로 규정한 요금·예약 계열.
# 셀마다 배지를 달면 표가 배지밭이 되므로 열 머리에 한 번만 단다.
VOLATILE_COL_RE = re.compile(r"요금|예약|가격|예산|비용|입장료")


def render_inline_tokens(text):
    """`{{badge:...}}` · `{{grade:...}}` 인라인 문법을 span 으로 바꾼다.

    md_convert 이전에 돌리므로 결과 span 이 마크다운 표·목록 안에서도 산다.
    알 수 없는 종류는 바꾸지 않고 남겨 둔다 — 빌드 검사가 잡는다.
    """
    def badge(m):
        kind, label = m.group(1), m.group(2).strip()
        if kind not in BADGE_KINDS:
            return m.group(0)
        return f'<span class="badge badge-{kind}">{html.escape(label)}</span>'

    def grade(m):
        kind, label = m.group(1), m.group(2).strip()
        if kind not in GRADE_KINDS:
            return m.group(0)
        return f'<span class="grade grade-{kind}">{html.escape(label)}</span>'

    return GRADE_RE.sub(grade, BADGE_RE.sub(badge, text))


def annotate_tables(md_text):
    """원고의 표에 등급 모양과 미확정 표시를 넣는다.

    - `등급` 열의 정규 5종 → 모양이 붙은 등급 span
    - 요금·예약 계열 열 머리 → `재확인` 배지 (열 전체가 변동정보라는 뜻)
    원본 MD 는 그대로 두고 빌드에서만 붙인다.
    """
    lines = md_text.splitlines()
    out = list(lines)
    graded = volatile = 0
    i = 0
    while i < len(lines) - 1:
        row, sep = lines[i].strip(), lines[i + 1].strip()
        if not (row.startswith("|") and re.fullmatch(r"\|[\s:|-]+\|", sep)):
            i += 1
            continue
        header = [c.strip() for c in row.strip("|").split("|")]
        grade_col = next((k for k, h in enumerate(header) if h.strip("*") == "등급"), None)
        vol_cols = [k for k, h in enumerate(header) if VOLATILE_COL_RE.search(h)]
        if vol_cols:
            for k in vol_cols:
                header[k] = f"{header[k]} {{{{badge:pending|재확인}}}}"
            out[i] = "| " + " | ".join(header) + " |"
            volatile += len(vol_cols)
        if grade_col is not None:
            j = i + 2
            while j < len(lines) and lines[j].strip().startswith("|"):
                cells = [c.strip() for c in lines[j].strip().strip("|").split("|")]
                if len(cells) > grade_col:
                    word = cells[grade_col].strip("*").strip()
                    slug = GRADE_WORDS.get(word)
                    if slug:
                        cells[grade_col] = f"{{{{grade:{slug}|{word}}}}}"
                        out[j] = "| " + " | ".join(cells) + " |"
                        graded += 1
                j += 1
            i = j
            continue
        i += 1
    return "\n".join(out), graded, volatile


# 헤딩 앞머리의 섹션 번호 — `## 5. 숙소 전략`, `### 5.1 추천 생활권`, `## 13A) …`
HEADING_NUM_RE = re.compile(
    r"^(#{1,6})\s+(?:Chapter\s*)?"
    r"(?:\d+(?:\.\d+)+|\d+(?:\.\d+)*[A-Z]?[.)])\s+(?=\S)")
# `① ~ ⑳` — 순서가 정보인 인라인 열거에 쓰인다. 뜻은 살리고 글리프만 없앤다.
CIRCLED = {chr(0x2460 + k): f"{k + 1})" for k in range(20)}
CIRCLED_RE = re.compile("[" + "".join(CIRCLED) + "]")


def strip_naming_noise(md_text):
    """헤딩의 섹션 번호와 원문자를 걷어낸다 (명명규칙 v1.0).

    번호는 재배치 뒤 순서와 어긋나 노이즈가 됐다 — Girona 의 h2 등장 순서가
    10 → 1 → 2 … 14 → 13 처럼 뒤섞인다. `Day N` 은 전체 여행 중 위치를
    알려주므로 남긴다 (normalize_day_headings 가 이미 다시 쓴다).

    원문자는 `평가 순서는 ① … ⑧` 처럼 순서가 내용인 자리에만 쓰인다.
    항목을 지우면 뜻이 사라지므로 평문 숫자로 바꾼다.
    """
    out, in_fence, n_num, n_circ = [], False, 0, 0
    for line in md_text.splitlines():
        if FENCE_RE.match(line):
            in_fence = not in_fence
            out.append(line)
            continue
        if not in_fence:
            m = HEADING_NUM_RE.match(line)
            if m and not re.match(r"^#{1,6}\s+\d+[.)]\s*Day\s", line):
                line = HEADING_NUM_RE.sub(r"\1 ", line)
                n_num += 1
            if CIRCLED_RE.search(line):
                line, k = CIRCLED_RE.subn(lambda x: CIRCLED[x.group(0)], line)
                n_circ += k
        out.append(line)
    return "\n".join(out), n_num, n_circ


HERO_MD_IMG_RE = re.compile(
    r"^!\[[^\]]*\]\(\.\./\.\./ASSETS/88_Representative_Public_Photos/[^)]+\)\s*$"
    r"(?:\n^\*[^\n]*\*\s*$)?", re.M)


def drop_source_hero(md_text):
    """원고 앞머리의 대표사진과 그 캡션 줄을 걷어낸다.

    같은 사진을 `hero_figure()` 가 지역소개 첫머리에 저작자·라이선스와 함께
    다시 넣는다. 두 장이 한 페이지에 겹쳐 보였다 — 분할 챕터 허브에서는
    두 장이 나란히 붙어 특히 두드러진다. 크레딧을 담은 쪽을 남긴다.
    """
    return HERO_MD_IMG_RE.subn("", md_text)


def strip_title_number(title):
    """`05. 지로나·콜리우르` · `Chapter 05. …` 의 번호 접두어를 없앤다."""
    return TITLE_H1_RE.sub("", title).strip()


def resolve_day_ref(n, chapter):
    """날짜 없는 Day 참조 헤딩의 번호를 전역 번호로 해석한다.

    원고에 두 가지 표기가 섞여 있다.
    - v1.x 원본 섹션은 챕터 로컬 번호다 — Girona 의 `## Day 1 비` 는 전체 Day 4다.
    - v2.0 보강본 섹션은 이미 전역 번호다 — Luberon 의 `### Day 16 — Lourmarin`.
    표기만 봐서는 구분되지 않으므로 챕터의 실제 날짜 범위로 판정한다.

    번호가 챕터 구간 안에 있으면 전역으로 보고 그대로 둔다. 그렇지 않고
    로컬 범위(1..박수+1) 안이면 전역으로 옮긴다. 둘 다 아니면 다른 챕터를
    가리키는 상호참조로 보고 **손대지 않는다** — 해석할 수 없는 번호를
    고쳐 쓰면 현장에서 틀린 날짜를 읽게 된다.

    반환: (전역번호, 경고문 또는 None)
    """
    base = day_no(chapter["start"])
    last = base + chapter["nights"]
    local_ok = 1 <= n <= chapter["nights"] + 1
    if base <= n <= last:
        # 겹침 구간 — 로컬로도 읽히는 번호다. 지금 원고에는 해당 사례가 없다.
        if local_ok and base > 1:
            return n, f"전역·로컬 양쪽으로 읽힌다. 전역 Day {n} 으로 봤다"
        return n, None
    if local_ok:
        return base + n - 1, None
    return n, f"챕터 구간(Day {base}–{last}) 밖이다. 상호참조로 보고 그대로 둔다"



# ---------------------------------------------------------------- 작성 흔적 제거

# 통째로 지우는 섹션 — 제작 공정 문서이지 여행 정보가 아니다
DROP_SECTION_RE = re.compile(
    r"^(#{1,6})[ \t]*(?:\d+[A-Z]?[.)][ \t]*)*"
    r"(?:Phase\s*\d+[^\n]*원칙|시각요소 자리표시자|사진 에셋 브리프[^\n]*|"
    r"시각자료 설계[^\n]*)[ \t]*$", re.M)

# 제목만 바꾸는 것 — 내용은 여행 정보다
RENAME_HEADING = [
    (re.compile(r"^(#{1,6}[ \t]*)(?:\d+[A-Z]?[.)][ \t]*)*Pass\s*[B-Z][.．]?\s*", re.M), r"\1"),
    (re.compile(r"^(#{1,6}[ \t]*)편집자 큐레이션[ \t]*[—–-][ \t]*", re.M), r"\1"),
    (re.compile(r"^(#{1,6}[ \t]*)검증 상태[ \t]*[—–-][ \t]*보강본 근거[ \t]*$", re.M),
     r"\1확인이 필요한 것"),
]

# 문장 단위로 지우는 것 — 이 문서가 어떻게 쓰였는지에 대한 말
DROP_SENTENCE = [
    re.compile(r"기존 원고[가의는를에]?[^.\n]*?(?:판단|정리|배치|구성|서술|표현)이"
               r"[^.\n]*?(?:정확하다|옳다|옳았[^.\n]*|맞다)[^.\n]*\.\s*"),
    re.compile(r"기존 원고의 (?:판단|지시|정리)(?:다|이다)\.\s*"),
    re.compile(r"보강본(?:은|이|에서)[^.\n]*?(?:넣었다|명시했다|추가했다|반영했다)\.\s*"),
    re.compile(r"[^.\n]*(?:현재 작업본|현재 Markdown 작업본|최종 Word|최종본에는|자리표시자)"
               r"[^.\n]*\.\s*"),
    re.compile(r"[^.\n]*보강본(?:은|이|의)[^.\n]*\.\s*"),
    re.compile(r"[^.\n]*이번 개정의 핵심[^.\n]*\.\s*"),
]

# 문장은 살리고 귀속만 떼는 것
STRIP_ATTRIB = [
    (re.compile(r"기존 원고[가의][ \t]*"), ""),
    (re.compile(r"기존 원고에도[ \t]*"), ""),
    (re.compile(r"기존 원고에만[ \t]*"), "여기에만 "),
    (re.compile(r"기존 원고에[ \t]*"), ""),
    (re.compile(r"기존 원고는[ \t]*"), ""),
    (re.compile(r"기존 원고[ \t]+(Day\s*\d+)"), r"\1"),
    (re.compile(r"기존 원고[ \t]*(기준|기반|지침대로|유지)"), r"\1"),
    # 검증 상태표의 근거 표기 — 독자에겐 '확인 안 됨' 이라는 뜻이다
    (re.compile(r"근거가 기존 원고뿐"), "교차 확인된 출처 없음"),
    (re.compile(r"기존 원고[ \t]*\+[ \t]*일반 지식"), "교차 확인 안 됨"),
    (re.compile(r"\([ \t]*기존 원고[^)]*\)"), ""),
    (re.compile(r"기존 원고[ \t]*"), ""),
    (re.compile(r"[ \t]*\((?:기존 )?원고 유지\)"), ""),
    # 번호를 지운 뒤라 `32.1 저녁 배분` 같은 참조는 가리킬 곳이 없다. 제목만 남긴다.
    (re.compile(r"`\d+[A-Z]?(?:\.\d+)?[.．]?[ \t]+([^`]{1,40})`"), r"**\1**"),
]

DESIGN_NOTE_RE = re.compile(r"^>[ \t]*\*\*디자인 메모\*\*[^\n]*\n?", re.M)


def strip_process_notes(md_text):
    """가이드북이 어떻게 만들어졌는지에 대한 서술을 최종본에서 걷는다.

    현장에서 읽는 사람에게 '기존 원고가 …로 잡았다' 는 정보가 아니다.
    원본 MD 는 손대지 않는다 — 작성 경위는 저장소에 남고 출력에서만 사라진다.
    """
    n = 0
    lines, out, drop_level = md_text.splitlines(), [], None
    for line in lines:
        h = re.match(r"^(#{1,6})[ \t]", line)
        if drop_level is not None:
            if h and len(h.group(1)) <= drop_level:
                drop_level = None
            else:
                n += 1
                continue
        if DROP_SECTION_RE.match(line):
            drop_level = len(re.match(r"^(#{1,6})", line).group(1))
            n += 1
            continue
        out.append(line)
    text = "\n".join(out)
    for pat, rep in RENAME_HEADING:
        text, k = pat.subn(rep, text)
        n += k
    text, k = DESIGN_NOTE_RE.subn("", text)
    n += k
    for pat in DROP_SENTENCE:
        text, k = pat.subn("", text)
        n += k
    for pat, rep in STRIP_ATTRIB:
        text, k = pat.subn(rep, text)
        n += k
    # 문장을 지운 뒤 남는 빈 인용줄·빈 문단 정리
    text = re.sub(r"^>[ \t]*$\n", "", text, flags=re.M)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text, n




def section_end(marks, idx, total):
    """marks[idx] 절이 끝나는 줄. 같은 레벨 이상 헤딩을 만나면 끝난다.

    다음 헤딩까지로 자르면 하위 절이 남아 본문이 반토막 난다.
    """
    lvl = marks[idx][1]
    for j in range(idx + 1, len(marks)):
        if marks[j][1] <= lvl:
            return marks[j][0]
    return total


def place_title_key(name):
    """장소 절 제목에서 장소명만 남긴다. 등급 토큰·섹션번호·부제를 뗀다."""
    n = re.sub(r"\{\{grade:[^}]*\}\}", "", name)
    n = NUM_PREFIX_RE.sub("", SUBNUM_RE.sub("", n.strip()))
    n = re.split(r"\s+[—–]\s+", n)[0]
    n = re.sub(r"\s*\(.*$", "", n).strip()
    return place_key(n)


def merge_place_sections(md_text, slug):
    """같은 장소를 두 번 쓴 절을 하나로 합친다.

    보강본이 장소 상세를 새로 쓰면서 원본 원고의 같은 장소 절이 그대로 남았다.
    `places.html` 에 Palais des Papes 가 두 번 나오는 식이다. 읽는 사람에게는
    스크롤만 늘어난다.

    **내용은 버리지 않는다.** 뒤 절의 본문을 앞 절 아래로 옮기고 제목만 지운다.
    합칠지 말지는 장소 레지스트리에 있는 이름과 정확히 맞을 때만 정한다 —
    제목이 비슷하다는 이유로 합치면 다른 내용을 뭉갠다.
    """
    known = {place_key(r["name"]) for r in load_place_registry()
             if r["chapter"] == slug and r["type"] == "spot"}
    if not known:
        return md_text, 0
    lines = md_text.splitlines()
    # 카테고리를 넘어 합치면 안 된다. 일정(Day)·교통에 있는 같은 이름의 절은
    # 그 날·그 맥락의 정보이지 장소 상세의 사본이 아니다. 여행정보 안에서만 본다.
    info_label = CAT_LABEL["info"]
    marks, in_info = [], False
    for i, line in enumerate(lines):
        h1 = re.match(r"^#[ \t]+(.+)$", line)
        if h1:
            in_info = h1.group(1).strip() == info_label
            marks.append((i, 1, None))       # 카테고리 경계 — 절이 여기서 끝난다
            continue
        m = re.match(r"^(#{2,4})[ \t]+(.+)$", line)
        if m:
            k = place_title_key(m.group(2)) if in_info else None
            marks.append((i, len(m.group(1)), k if k in known else None))
    # 절 범위를 먼저 구한다
    spans = {}
    for idx, (i, lvl, k) in enumerate(marks):
        if k:
            spans.setdefault(k, []).append((i, section_end(marks, idx, len(lines))))
    # 기준 절은 등급이 붙은 것 — 레지스트리가 그 제목을 가리킨다.
    # 순서로 정하면 원본 절이 기준이 되어 등급 절이 사라진다.
    merged, drop = 0, []
    for k, group in spans.items():
        if len(group) < 2:
            continue
        host = next((g for g in group if "{{grade:" in lines[g[0]]), group[0])
        for g in group:
            if g == host:
                continue
            sub = re.split(r"\s+[—–]\s+", lines[g[0]].lstrip("# ").strip(), 1)
            lead = f"**{sub[1].strip()}**" if len(sub) > 1 and sub[1].strip() else ""
            drop.append((g[0], g[1], host[1], lead, lines[g[0] + 1:g[1]]))
            merged += 1
    if not merged:
        return md_text, 0
    insert, skip = {}, set()
    for start, end, host_end_line, lead, body in drop:
        insert.setdefault(host_end_line, []).append(([lead] if lead else []) + body)
        skip.update(range(start, end))
    out = []
    for i, line in enumerate(lines):
        if i not in skip:
            out.append(line)
        if i + 1 in insert:
            for ch in insert[i + 1]:
                out += [""] + ch
    text = re.sub(r"\n{3,}", "\n\n", "\n".join(out))
    return text, merged



PLACE_BODY = {}   # 장소 슬러그 -> 상세 마크다운. 챕터 빌드가 채우고 장소 페이지가 쓴다.


def extract_place_bodies(md_text, slug, rel):
    """장소 상세 본문을 챕터에서 떼어 장소 카드로 옮긴다.

    최소 단위를 장소로 정한 이상 상세의 집은 장소 카드다. 챕터의 여행정보
    페이지가 그 본문을 다 안고 있으면 파리는 한 장이 16,000자가 된다.

    **복제가 아니라 이사다.** 챕터에는 제목과 등급, 그리고 카드로 가는 줄만
    남는다. 제목을 남기는 것은 레지스트리 가드가 그 문자열로 대조하기 때문이고,
    챕터를 훑을 때 무엇이 있는지 보이게 하려는 것이기도 하다.
    """
    by_key = {place_key(r["name"]): r for r in load_place_registry()
              if r["chapter"] == slug and r["type"] == "spot"}
    if not by_key:
        return md_text, 0
    lines = md_text.splitlines()
    info_label = CAT_LABEL["info"]
    marks, in_info = [], False
    for i, line in enumerate(lines):
        h1 = re.match(r"^#[ \t]+(.+)$", line)
        if h1:
            in_info = h1.group(1).strip() == info_label
            marks.append((i, 1, None))       # 카테고리 경계
            continue
        m = re.match(r"^(#{2,4})[ \t]+(.+)$", line)
        if m:
            r = by_key.get(place_title_key(m.group(2))) if in_info else None
            marks.append((i, len(m.group(1)), r))
    moved, cut = 0, []
    for idx, (i, lvl, r) in enumerate(marks):
        end = section_end(marks, idx, len(lines))
        if not r:
            continue
        # 이 절 안에 다른 장소 절이 들어 있으면 옮기지 않는다. 상위 묶음 헤딩을
        # 통째로 들어내면 그 아래 장소들이 챕터에서 사라진다.
        if any(marks[j][0] < end and marks[j][2] for j in range(idx + 1, len(marks))):
            continue
        body = [x for x in lines[i + 1:end] if x.strip()]
        if len(body) < 3:          # 이미 짧으면 그냥 둔다
            continue
        # 한 장소의 상세라기엔 너무 크면 형제 절을 삼킨 것이다. 손대지 않는다.
        # 정상 범위는 1,500자 안쪽이다 (가장 긴 것이 크루아루스 1,339자).
        if len(re.sub(r"\s", "", " ".join(body))) > 2500:
            continue
        PLACE_BODY.setdefault(r["slug"], []).extend(lines[i + 1:end])
        cut.append((i + 1, end, r))
        moved += 1
    if not moved:
        return md_text, 0
    skip, note = set(), {}
    for s, e, r in cut:
        skip.update(range(s, e))
        note[s] = (f'[◈ {r["name"]} — 장소 카드에서 보기]'
                   f'({rel}/places/{r["slug"]}.html)')
    out = []
    for i, line in enumerate(lines):
        if i in note:
            out += ["", note[i], ""]
        if i not in skip:
            out.append(line)
    return re.sub(r"\n{3,}", "\n\n", "\n".join(out)), moved


def normalize_day_headings(md_text, chapter):
    """Day 섹션 헤딩을 h2로 통일하고 Day 번호를 전체 여행 기준으로 바꾼다.

    원본은 챕터 로컬 번호를 쓴다 — Paris의 `Day 16`이 실제로는 전체 Day 43이다.
    데일리 페이지(`daily/day-43.html`)와 어긋나 현장에서 혼란을 만든다.
    헤딩에 이미 날짜가 있으므로 날짜를 정본으로 삼아 번호를 역산한다.
    원본 MD는 수정하지 않는다.
    """
    if chapter["kind"] != "region":
        return md_text, 0
    base = day_no(chapter["start"])
    out, changed, warn = [], 0, []
    for line in md_text.splitlines():
        m = SRC_DAY_HEADING_RE.match(line)
        if m:
            local, month, dom, weekday = int(m[1]), int(m[2]), int(m[3]), m[4]
            d = date(TRIP_START.year, month, dom)
            expected = WEEKDAY_KO[d.weekday()]
            if expected != weekday:
                sys.exit(f"Day 헤딩 요일 불일치({chapter['slug']}): {line.strip()}"
                         f" — {d} 는 {expected}요일")
            if not TRIP_START <= d <= TRIP_END:
                sys.exit(f"Day 헤딩 날짜가 여행 기간 밖({chapter['slug']}): {line.strip()}")
            if base + local - 1 != day_no(d):
                sys.exit(f"Day 로컬번호와 날짜 불일치({chapter['slug']}): {line.strip()}"
                         f" — 로컬 {local}은 전체 Day {base + local - 1}이어야 하는데"
                         f" 날짜는 전체 Day {day_no(d)}")
            out.append(f"## Day {day_no(d)} · {month}월 {dom}일 {expected}")
            changed += 1
            continue
        m = SRC_DAY_REF_RE.match(line)
        if m and m[3].strip():
            n, rest = int(m[2]), m[3].strip()
            g, note = resolve_day_ref(n, chapter)
            if note:
                warn.append(f"{note}: {line.strip()}")
            if g != n:
                out.append(f"{m[1]} Day {g} {rest}")
                changed += 1
                continue
        out.append(line)
    for w in warn:
        print(f"  주의: Day 참조 헤딩({chapter['slug']}) {w}")
    return "\n".join(out), changed


def md_convert(text):
    md = markdown.Markdown(
        extensions=["tables", "fenced_code",
                    TocExtension(slugify=slugify_unicode, toc_depth="1-3")],
        output_format="html5",
    )
    body = md.convert(text)
    return body, md.toc_tokens


def flatten_tokens(tokens):
    flat = []
    for tok in tokens:
        flat.append(tok)
        flat.extend(flatten_tokens(tok.get("children", [])))
    return flat


def wrap_tables(body):
    body = body.replace("<table>", '<div class="table-wrap"><table>')
    return body.replace("</table>", "</table></div>")


CAT_ICON = {
    "지역소개": "book", "일정": "clock", "여행정보": "pin", "먹거리": "food",
    "교통": "train", "숙박": "stay", "예약": "lock", "경비": "cost",
    "여행팁": "tip", "부록": "source",
}


def first_heading_icon(body_html, icon):
    """페이지 첫 헤딩에 아이콘을 단다.

    원고 제목은 라벨과 정확히 같지 않다 — `먹거리 — 니스 요리의 문법`. 문자열
    대조로는 못 잡는다. 그 페이지가 어느 카테고리인지는 빌드가 알고 있다.
    """
    if not icon:
        return body_html
    return re.sub(r"<(h1|h2)(?![^>]*\bclass=)", f'<\\1 class="ic ic-{icon}"',
                  body_html, count=1)


def mark_category_icons(body_html):
    """카테고리 헤딩에 아이콘을 단다.

    원고가 만든 헤딩이라 빌드가 클래스를 붙일 자리가 없다. 렌더된 뒤에
    제목 문자열로 찾는다 — 카테고리 라벨 10개는 CATEGORIES 가 정하는
    닫힌 목록이라 문자열 대조가 흔들리지 않는다.
    """
    def sub(m):
        icon = CAT_ICON.get(m.group(3).strip())
        if not icon:
            return m.group(0)
        return f'<{m.group(1)}{m.group(2)} class="ic ic-{icon}">{m.group(3)}</{m.group(1)}>'
    return re.sub(r"<(h1|h2)([^>]*?)>([^<]+)</\1>", sub, body_html)


def mark_layer_headings(body):
    """카테고리 경계 h1에 시각 구분용 클래스를 부여한다."""
    labels = "|".join(label for _, label in CATEGORIES)
    return re.sub(rf'<h1 id="([^"]*)">({labels})</h1>',
                  r'<h1 id="\1" class="layer-h">\2</h1>', body)


def rewrite_md_links(body, by_file, rel):
    def repl(match):
        name = Path(match.group(1).split("#")[0]).name
        if name in by_file:
            return f'href="{rel}/{chapter_url(by_file[name])}"'
        return match.group(0)
    return re.sub(r'href="([^"]+\.md)"', repl, body)


def rewrite_asset_links(body, rel=".."):
    """소스 내부 상대경로의 자산 링크를 사이트 경로로 재작성한다."""
    for src_name, out_name, _ in MAPS:
        body = body.replace(f"../../ASSETS/75_Execution_Maps/{src_name}",
                            f"{rel}/maps/{out_name}")
    body = re.sub(r"\.\./\.\./ASSETS/75_Execution_Maps/([^\"']+\.(?:geojson|kml))",
                  rf"{rel}/maps/data/\1", body)
    for slug, (fname, *_) in HERO_PHOTOS.items():
        body = body.replace(f"../../ASSETS/88_Representative_Public_Photos/{fname}",
                            f"{rel}/assets/heroes/{slug}.jpg")
    body = body.replace("../../ASSETS/85_Editorial_Visuals/", f"{rel}/assets/visuals/")
    return body


def toc_html(tokens):
    items = []
    for tok in tokens:
        items.append(f'<li><a href="#{tok["id"]}">{html.escape(tok["name"])}</a></li>')
        for child in tok.get("children", []):
            items.append(
                f'<li class="toc-sub"><a href="#{child["id"]}">{html.escape(child["name"])}</a></li>')
    if not items:
        return ""
    return ('<details class="toc"><summary>전체 목차</summary><ul>'
            + "".join(items) + "</ul></details>")


def date_label(d):
    return f"{d.month}/{d.day}"


def day_no(d):
    return (d - TRIP_START).days + 1


def date_of_day(n):
    return TRIP_START + timedelta(days=n - 1)


def chapter_for_date(d):
    """해당 날짜의 지역 챕터 (경계일은 도착 챕터)."""
    for c in reversed([c for c in CHAPTERS if c["kind"] == "region"]):
        if c["start"] <= d <= c["end"]:
            return c
    return None


# ---------------------------------------------------------------- page shell

def search_sheet_html(rel):
    """검색 시트. 전체 메뉴는 여기 없다 — 있었지만 지웠다.

    햄버거 메뉴가 담던 것은 하나도 빠짐없이 홈·하단탭·꼬리말에도 있었다.
    같은 목록을 세 번째로 늘어놓으면 메뉴가 사이트의 사본이 되고, 어느
    쪽이 최신인지 알 수 없게 된다. 검색만 남긴다 — 검색은 다른 어디에도
    없는 기능이라 중복이 아니다.
    """
    return f"""<div id="overlay"></div>
<aside id="search-sheet" aria-label="검색">
  <div class="ss-head">
    <button id="sheet-close" aria-label="닫기"><b class="ic ic-only ic-close" aria-hidden="true"></b></button>
    <span>검색</span>
  </div>
  <input id="search-input" type="search" placeholder="장소·섹션 검색" autocomplete="off">
  <div id="search-results"></div>
</aside>"""


# 주제 축 인덱스. 지역 챕터를 훑어 카테고리별·상태별로 모은다.
# 본문을 복제하지 않고 제목·발췌·앵커만 담는다 — 복제하면 갱신 때 어긋난다.
TOPIC_INDEX = defaultdict(list)    # 카테고리 라벨 -> [항목]
STATUS_INDEX = defaultdict(list)   # 'badge:pending' 등 -> [항목]

HEADING_SCAN_RE = re.compile(r'<h([12]) id="([^"]+)"[^>]*>(.*?)</h\1>', re.S)
TAG_RE = re.compile(r"<[^>]+>")
BADGE_SPAN_RE = re.compile(r'<span class="(badge|grade) \1-(\w+)">(.*?)</span>', re.S)
STATUS_WANTED = {"badge:pending", "badge:p0", "grade:essential"}


def plain(fragment, limit=0):
    t = html.unescape(TAG_RE.sub(" ", fragment))
    t = re.sub(r"\s+", " ", t).strip()
    return (t[:limit].rstrip() + "…") if limit and len(t) > limit else t


def scan_sections(c, body, url, fixed_cat=None):
    """렌더된 지역 챕터 본문에서 주제·상태 인덱스에 넣을 것을 한 번에 뽑는다.

    단일 페이지 챕터는 h1 이 카테고리고 그 아래 h2 가 섹션이다. 분할 챕터는
    페이지 하나가 곧 카테고리라 fixed_cat 으로 받는다.
    """
    if c["kind"] != "region":
        return
    labels = {label for _, label in CATEGORIES}
    marks = list(HEADING_SCAN_RE.finditer(body))
    cat = fixed_cat
    for k, m in enumerate(marks):
        name = plain(m.group(3))
        chunk = body[m.end():marks[k + 1].start() if k + 1 < len(marks) else len(body)]
        if m.group(1) == "1":
            if fixed_cat is None:
                cat = name if name in labels else None
            continue
        if not cat:
            continue
        item = {"region": c["region"], "slug": c["slug"], "title": name,
                "url": f'{url}#{m.group(2)}', "excerpt": plain(chunk, 130)}
        TOPIC_INDEX[cat].append(item)
        for kind, sub, label in BADGE_SPAN_RE.findall(chunk):
            key = f"{kind}:{sub}"
            if key in STATUS_WANTED:
                STATUS_INDEX[key].append(dict(item, note=plain(label), cat=cat))



def chapter_coords(c, rel):
    """지역 챕터의 좌표 — 일자 축은 첫날 카드로, 주제 축은 주제 허브로 연다."""
    if c["kind"] != "region":
        return ""
    first, last = day_no(c["start"]), day_no(c["end"])
    return coords_bar(rel,
                      day=(f"Day {first}–{last}", f"daily/day-{first:02d}.html"),
                      region=(c["region"], None),
                      topic=("전체 주제", "topics/index.html"))


DAY_H2_RE = re.compile(r'(<h2 id="[^"]*">Day (\d+) · [^<]*</h2>)')


PLACE_BY_HEAD = {}   # 등급 헤딩 원문 -> 장소 슬러그 (레지스트리에서 채운다)


def link_place_cards(body, rel):
    """챕터 본문의 등급 헤딩마다 그 장소 카드로 가는 링크를 붙인다.

    장소가 최소 단위이므로 본문에서 좌표로 나가는 문이 있어야 한다.
    """
    def add(m):
        name = plain(re.sub(r'<span class="grade.*?</span>', "", m.group(0), flags=re.S))
        slug = PLACE_BY_HEAD.get(name)
        if not slug:
            return m.group(0)
        return (f'{m.group(0)}<p class="day-jump">'
                f'<a href="{rel}/places/{slug}.html"><b class="ic ic-download" aria-hidden="true"></b>{html.escape(name)} 장소 카드</a></p>')
    return re.sub(r'<h[234][^>]*>(?:(?!</h[234]>).)*?class="grade grade-.*?</h[234]>',
                  add, body, flags=re.S)


def link_day_cards(body, rel):
    """챕터 본문의 Day 섹션마다 그 날 데일리 카드로 가는 링크를 붙인다.

    데일리 → 챕터 방향은 있었지만 역방향이 없었다. 현장에서는 챕터를 읽다가
    '오늘 카드' 로 돌아가는 동선이 더 잦다.
    """
    def add(m):
        n = int(m.group(2))
        return (f'{m.group(1)}<p class="day-jump">'
                f'<a href="{rel}/daily/day-{n:02d}.html"><b class="ic ic-today" aria-hidden="true"></b>Day {n} 카드 · 시간표</a></p>')
    return DAY_H2_RE.sub(add, body)


def coords_bar(rel, *, day=None, region=None, topic=None):
    """일자·지역·주제 세 축의 현재 좌표. 현재 축은 굳고 나머지는 링크가 된다.

    각 인자는 (라벨, URL 또는 None) 이다. URL 이 None 이면 현재 페이지다.
    """
    cells = []
    for axis, item in (("일자", day), ("지역", region), ("주제", topic)):
        if item is None:
            continue
        label, url = item
        inner = f'<b>{axis}</b><span>{html.escape(label)}</span>'
        cells.append(f'<a href="{rel}/{url}">{inner}</a>' if url
                     else f'<span class="cd-here">{inner}</span>')
    if not cells:
        return ""
    return (f'<nav class="coords" aria-label="일자·지역·주제로 이동">{"".join(cells)}</nav>')



def crumbs_for(*items):
    """홈에서 시작하는 위치 경로. 마지막 항목은 링크가 아니라 현재 위치다."""
    out = [("홈", "index.html")]
    for lab, url in items:
        out.append((lab, url))
    out[-1] = (out[-1][0], None)
    return out


def region_crumbs(*items):
    """지역을 지나는 경로. 홈 다음에 지역 목록을 끼운다.

    `홈 › 지역 › Nice › 먹거리`. 장소·일자 페이지에서 지역 하나를 건너뛰고
    8개 거점 목록으로 바로 올라갈 수 있어야 한다 — 이동 중에 다음 거점을
    확인하는 동선이 그 길이다.
    """
    return crumbs_for(("지역", "regions.html"), *items)


def regions_of_date(d):
    """그 날짜를 품는 지역 전부. 이동일은 두 지역에 걸친다."""
    return [c for c in CHAPTERS if c["kind"] == "region" and c["start"] <= d <= c["end"]]


# 지역 슬러그 → 나라. 스페인 두 곳, 나머지는 프랑스. Day 7 에 국경을 넘는다.
ES_REGIONS = {"barcelona", "girona"}


def country_of(*names):
    """이 페이지가 어느 나라인가. 이동일은 둘 다다.

    통화는 같지만 시장 요일·공휴일·긴급번호·언어가 다르다. 지금 어느
    나라에 있는지는 장식이 아니라 현장 정보다.
    """
    seen = []
    for n in names:
        if not n:
            continue
        c = "es" if n in ES_REGIONS else "fr"
        if c not in seen:
            seen.append(c)
    if not seen:
        return ""
    return "-".join(sorted(seen))


def page(title, body, *, rel="..", topbar_title=None, meta_line="", subnav="",
         coords="", back=None, country=""):
    """back=(라벨, URL). HIG 네비게이션 바의 뒤로가기다.

    상단바는 둘로 고정한다 — 앞: 위치 경로(홈까지 올라간다), 뒤: 검색.
    전체 메뉴 버튼은 없다. 축 진입은 하단탭이, 목록은 홈이 맡는다.
    """
    meta_html = f'<p class="meta">{meta_line}</p>' if meta_line else ""
    country_attr = f' data-country="{country}"' if country else ""
    tb_title = html.escape(topbar_title or title)
    if back:
        crumbs = back if isinstance(back, list) else [("홈", "index.html"), back]
        parts = []
        for k, (lab, url) in enumerate(crumbs):
            sep = '<span class="tb-sep" aria-hidden="true">›</span>' if k else ""
            if url:
                parts.append(f'{sep}<a href="{rel}/{url}">{html.escape(lab)}</a>')
            else:
                parts.append(f'{sep}<b>{html.escape(lab)}</b>')
        lead = f'<nav class="tb-crumbs" aria-label="위치">{"".join(parts)}</nav>'
    else:
        lead = '<span class="tb-crumbs tb-crumbs-none"></span>'
    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<title>{html.escape(title)} — {SITE_TITLE}</title>
<link rel="stylesheet" href="{rel}/assets/style.css">
</head>
<body data-rel="{rel}"{country_attr}>
<header class="topbar">
  {lead}
  <button id="search-btn" aria-label="검색 열기"><b class="ic ic-only ic-search" aria-hidden="true"></b></button>
</header>
{subnav}
{search_sheet_html(rel)}
<main>
{meta_html}
{coords}
{body}
</main>
<footer>
  <p>{SITE_TITLE} · {TRIP_PERIOD}</p>
  <p><a href="{rel}/credits.html">사진 저작자 표시 · 라이선스</a> ·
     <a href="{rel}/maps/offline.html">오프라인 지도 준비</a></p>
</footer>
<nav class="bottomnav" aria-label="주요 메뉴">
  <a href="#" class="nav-today" data-tab="today"><b class="ic ic-only ic-today" aria-hidden="true"></b><span>오늘</span></a>
  <a href="{rel}/{ITINERARY_URL}" data-tab="itinerary"><b class="ic ic-only ic-list" aria-hidden="true"></b><span>일정</span></a>
  <a href="{rel}/regions.html" data-tab="regions"><b class="ic ic-only ic-region" aria-hidden="true"></b><span>지역</span></a>
  <a href="{rel}/tracker/index.html" data-tab="prepare"><b class="ic ic-only ic-check" aria-hidden="true"></b><span>준비</span></a>
  <a href="#search" class="nav-search" data-tab="search"><b class="ic ic-only ic-search" aria-hidden="true"></b><span>검색</span></a>
</nav>
<button id="back-top" aria-label="맨 위로"><b class="ic ic-only ic-up" aria-hidden="true"></b></button>
<script src="{rel}/assets/data.js" defer></script>
<script src="{rel}/assets/nav.js" defer></script>
</body>
</html>
"""


# ---------------------------------------------------------------- chapters

def chapter_subnav(chapter, flat_tokens):
    """카테고리 점프 + Day 칩 서브내비 (지역 챕터 전용)."""
    labels = {label for _, label in CATEGORIES}
    cats, days, seen_days = [], [], set()
    for tok in flat_tokens:
        if tok["level"] == 1 and tok["name"] in labels:
            cats.append(f'<a href="#{tok["id"]}">{tok["name"]}</a>')
        dm = DAY_RE.search(tok["name"])
        if dm and dm.group(1) not in seen_days:
            seen_days.add(dm.group(1))
            days.append(f'<a href="#{tok["id"]}" title="{html.escape(tok["name"])}">'
                        f'{int(dm.group(2))}/{int(dm.group(3))}</a>')
    if not cats and not days:
        return ""
    cats_html = f'<div class="sn-layers">{"".join(cats)}</div>' if cats else ""
    days_html = f'<div class="sn-days">{"".join(days)}</div>' if days else ""
    return f'<nav class="subnav">{cats_html}{days_html}</nav>'


def related_box(chapter):
    """지역 챕터 상단의 관련 리소스 링크."""
    if chapter["kind"] != "region":
        return ""
    rel = chapter_rel(chapter)
    links = [f'<a href="{rel}/maps/{chapter["map"]}"><b class="ic ic-map" aria-hidden="true"></b>{chapter["region"]} 실행지도</a>',
             f'<a href="{rel}/daily/day-{day_no(chapter["start"]):02d}.html"><b class="ic ic-today" aria-hidden="true"></b>데일리 카드</a>',
             f'<a href="{rel}/tracker/reservations.html"><b class="ic ic-table" aria-hidden="true"></b>예약 현황</a>',
             f'<a href="{rel}/{ITINERARY_URL}"><b class="ic ic-list" aria-hidden="true"></b>43일 일정표</a>']
    return f'<div class="related">{"".join(links)}</div>'


def collect_search(chapter, flat_tokens):
    label = f'{chapter["slug"]} {chapter["title"]}'
    SEARCH_INDEX.append({"t": chapter["title"], "c": f'챕터 {chapter["slug"]}',
                         "u": chapter_url(chapter)})
    for tok in flat_tokens:
        name = tok["name"].strip()
        if not name or name.startswith(("Layer", "Pass")):
            continue
        SEARCH_INDEX.append({"t": name, "c": label,
                             "u": f'{chapter_url(chapter)}#{tok["id"]}'})


def collect_chapter_dates(chapter, flat_tokens):
    """날짜 → 챕터 URL 매핑. 지역 범위로 채우고 Day 섹션 앵커는 별도 수집."""
    if chapter["kind"] != "region":
        return
    url = chapter_url(chapter)
    d = chapter["start"]
    last = chapter["end"] if chapter["slug"] == "11" else chapter["end"] - timedelta(days=1)
    while d <= last:
        CHAPTER_DATE_URL[d.isoformat()] = url
        d += timedelta(days=1)
    split = chapter.get("name") in SPLIT_CHAPTERS
    for tok in flat_tokens:
        dm = DAY_RE.search(tok["name"])
        if dm:
            day_date = date(2026, int(dm.group(2)), int(dm.group(3)))
            # 그 날의 본문은 데일리 카드에 있다. 여기 담는 것은 '그 날이 속한
            # 지역의 일정 한눈에' 로, 데일리 카드에서 지역 축으로 나가는 문이다.
            DAY_OVERRIDES[day_date.isoformat()] = (
                f'chapters/{chapter["name"]}/schedule.html' if split
                else f'{url}#{tok["id"]}')


# 분할 대상 챕터. Girona 파일럿이 자리를 잡은 뒤 8개 거점 전부로 넓혔다.
# 원고는 여전히 챕터당 파일 하나다 — 분할은 출력 단계에서만 일어난다.
SPLIT_CHAPTERS = {"barcelona", "girona", "nice", "aix",
                  "luberon", "avignon", "lyon", "paris"}

# 카테고리 → 주제 페이지 슬러그. 일정은 일자 페이지로 따로 나간다.
TOPIC_SLUG = {
    "intro": "about", "info": "places", "food": "food", "transport": "transport",
    "stay": "stay", "booking": "booking", "cost": "cost", "tips": "tips",
    "appendix": "sources",
}
TOPIC_SUB = {
    "about": "이 지역을 어떻게 읽을 것인가",
    "places": "무엇을 보고 어디를 갈 것인가",
    "food": "무엇을 먹고 어디서 살 것인가",
    "transport": "어떻게 움직일 것인가",
    "stay": "어디에 묵을 것인가",
    "booking": "무엇을 미리 잠글 것인가",
    "cost": "얼마가 드는가",
    "tips": "현장에서 무엇을 조정할 것인가",
    "sources": "무엇을 근거로 썼는가",
}


def split_sections(body_md):
    """재편성된 마크다운을 `# 카테고리` 경계로 잘라 (카테고리, 본문) 목록을 만든다."""
    out, cur, buf, head = [], None, [], []
    labels = {label: key for key, label in CATEGORIES}
    for line in body_md.splitlines():
        m = re.match(r"^# (.+)$", line)
        if m and m.group(1).strip() in labels:
            if cur:
                out.append((cur, "\n".join(buf).strip()))
            cur, buf = labels[m.group(1).strip()], []
            continue
        (buf if cur else head).append(line)
    if cur:
        out.append((cur, "\n".join(buf).strip()))
    return "\n".join(head).strip(), out


def split_day_sections(schedule_md):
    """'일정' 본문을 Day 헤딩 기준으로 잘라 일자별 조각을 만든다.

    첫 Day 헤딩 앞의 내용(전체 일정표·동선 도식 등)은 허브에 남긴다.
    """
    parts = re.split(r"^(## Day (\d+) · [^\n]+)$", schedule_md, flags=re.M)
    head = parts[0].strip()
    days, tails = [], []
    for k in range(1, len(parts), 3):
        heading, n, body = parts[k], int(parts[k + 1]), parts[k + 3 - 1]
        title = heading[3:].strip()
        # Day 헤딩은 h2 다. 다음 h2 를 만나면 그 날이 끝난 것이다. 이걸 안 자르면
        # 마지막 날이 뒤에 오는 구간 전체(피로도표·실행성 감사·게이트)를 삼킨다.
        m = re.search(r"^## ", body, re.M)
        if m:
            tails.append(body[m.start():].rstrip())
            body = body[:m.start()]
        days.append((n, title, (heading + "\n" + body).strip()))
    if tails:
        head = (head + "\n\n" + "\n\n".join(tails)).strip()
    return head, days


def render_split_page(c, title, sub, body_md, crumbs, prev_nx, map_links, extra="",
                      coords="", subnav=""):
    """분할 페이지 하나를 기존 페이지 셸로 렌더한다.

    별도 마크다운 변환기를 두지 않는다 — 드로어·하단탭·검색·오프라인 처리와
    배지·등급 문법이 다른 페이지와 어긋나면 안 된다.
    """
    rel = "../.."
    body, toc_tokens = md_convert(body_md)
    body = first_heading_icon(
        mark_category_icons(
            link_place_cards(wrap_tables(rewrite_asset_links(body, rel)), rel)),
        CAT_ICON.get(title) or ("clock" if title.startswith("일정") else None))
    prev_link, next_link = prev_nx
    pager = f'<nav class="pager">{prev_link}<span></span>{next_link}</nav>'
    crumb_html = ('<nav class="crumbs" aria-label="위치">'
                  + " › ".join(crumbs) + "</nav>")
    sub_html = f'<p class="page-sub">{html.escape(sub)}</p>' if sub else ""
    content = crumb_html + sub_html + extra + body + pager
    return (page(title, content, rel=rel, topbar_title=title, coords=coords, subnav=subnav,
                 back=region_crumbs((c["region"], f'chapters/{c["name"]}/index.html'),
                                    (title, None)),
                 country=country_of(c["name"]),
                 meta_line=f'{c["title"]} · {date_label(c["start"])}–{date_label(c["end"])}'),
            flatten_tokens(toc_tokens), body)


def day_card_url(n, rel):
    """하루의 페이지는 daily/day-NN.html 하나다.

    전에는 지역 챕터에도 같은 날의 페이지가 있었다. 둘 다 원고를 렌더하는데
    한쪽은 정규식으로 다시 긁어 만든 사본이라 시간표가 30일에서 어긋났고,
    20일은 원고에 있는 피로도를 '없다'고 표시했다. 사본을 없앴다.
    """
    return f"{rel}/daily/day-{n:02d}.html"


def siblings_nav(groups):
    """L2 형제 내비 — 지금 묶음 안에서 옆으로 움직이는 한 가지 일만 한다.

    L0 하단탭이 축을 고르고, L1 좌표 바가 축을 건너뛰고, 여기가 형제를 훑는다.
    챕터·주제·데일리가 모두 같은 컴포넌트를 쓴다. 묶음마다 내용만 다르다.

    groups: [(축 라벨, 스타일, [(표시, url, 현재인가), ...]), ...]
    """
    out = []
    for label, style, items in groups:
        if not items:
            continue
        def link(text, u, here, tip):
            cls = ' class="here"' if here else ""
            ttl = f' title="{html.escape(tip)}"' if tip else ""
            return f'<a href="{u}"{cls}{ttl}>{html.escape(text)}</a>'

        links = "".join(link(*it) for it in items)
        out.append(f'<div class="sn-group"><span class="sn-label">{label}</span>'
                   f'<div class="{style}">{links}</div></div>')
    if not out:
        return ""
    return f'<nav class="subnav" aria-label="이 묶음 안에서 이동">{"".join(out)}</nav>'


def chapter_siblings(pages, current, with_days=True):
    """분할 챕터의 형제 — 주제 9개와 일자 N개. 두 축을 제 줄에 갈라 놓는다.

    주제 페이지에서는 일자 행을 감춘다. 세 줄짜리 고정 크롬은 작은 화면에서
    본문을 덮는다. 일자가 필요한 곳은 허브와 일자 페이지다.
    """
    cats = [("허브", "index.html", current == "index.html", "")]
    days = []
    for fname, title, _sub, _md, _cat in pages:
        dm = re.match(r"day-(\d+)\.html$", fname)
        if fname == "schedule.html":
            days.append(("한눈에", fname, fname == current, title))
        elif dm:
            md = re.search(r"(\d+)월\s*(\d+)일", title)
            label = f"{int(md.group(1))}/{int(md.group(2))}" if md else f"D{int(dm.group(1))}"
            # 하루의 페이지는 하나뿐이다. 지역 축은 목록만 맡고 본문은 갖지 않는다.
            days.append((label, day_card_url(int(dm.group(1)), "../.."),
                         False, title))
        else:
            cats.append((title, fname, fname == current, ""))
    return siblings_nav([("주제", "sn-layers", cats)]
                        + ([("일자", "sn-days", days)] if with_days else []))


def build_split_chapter(c, body_md, map_links):
    """지역 챕터를 허브 + 일자 + 주제 페이지로 나눠 낸다.

    authoring 파일 수와 output 페이지 수를 분리한다 — 원고 하나를 고치면
    해당 페이지들이 함께 갱신된다. 기존 단일 페이지 챕터와 공존한다.
    """
    out_dir = SITE / "chapters" / c["name"]
    out_dir.mkdir(parents=True, exist_ok=True)
    rel = "../.."
    hub_url = f'chapters/{c["name"]}/index.html'
    hub_crumb = f'<a href="index.html">{html.escape(c["title"])}</a>'

    header_md, sections = split_sections(body_md)
    by_cat = {k: v for k, v in sections}
    sched_head, days = split_day_sections(by_cat.get("schedule", ""))

    for n, title, md_body in days:
        DAY_MD.setdefault(n, []).append((c["region"], title, md_body))

    pages = []      # (파일명, 제목, 부제, 마크다운, 카테고리 라벨)
    if sched_head:
        # 일자에 속하지 않는 일정 내용(전체 일정표·운영 원칙 등). 허브에 쏟으면
        # 허브가 목차가 아니라 본문이 된다. 제 페이지를 준다.
        pages.append(("schedule.html", "일정 한눈에", "이 구간을 어떻게 굴리는가",
                      sched_head, CAT_LABEL["schedule"]))
    for n, title, md_body in days:
        pages.append((f"day-{n:02d}.html", title, "", md_body, CAT_LABEL["schedule"]))
    for key, label in CATEGORIES:
        if key == "schedule" or key not in by_cat:
            continue
        slug = TOPIC_SLUG[key]
        pages.append((f"{slug}.html", label, TOPIC_SUB[slug], by_cat[key], label))

    def href(fname):
        """일자 페이지의 실체는 데일리 카드다. 챕터 쪽 주소는 리다이렉트만 남는다."""
        dm = re.match(r"day-(\d+)\.html$", fname)
        return day_card_url(int(dm.group(1)), rel) if dm else fname

    made = []
    for idx, (fname, title, sub, md_body, page_cat) in enumerate(pages):
        dm = re.match(r"day-(\d+)\.html$", fname)
        if dm:
            # 하루에 페이지 하나. 옛 주소는 리다이렉트로 남긴다.
            write_redirect(out_dir / fname, href(fname), title)
            made.append((href(fname), title, sub, True))
            continue
        prev_link = next_link = ""
        if idx > 0:
            p = pages[idx - 1]
            prev_link = f'<a href="{href(p[0])}">← {html.escape(p[1])}</a>'
        if idx < len(pages) - 1:
            nx = pages[idx + 1]
            next_link = f'<a href="{href(nx[0])}">{html.escape(nx[1])} →</a>'
        crumbs = [f'<a href="{rel}/regions.html">지역</a>', hub_crumb,
                  f'<span>{html.escape(title)}</span>']
        fig = ""
        if fname == "transport.html" and c["slug"] in ("07", "08", "09"):
            fig = visual_figure("cardays", "Provence 차량일 운영 논리", "../../assets")
        rendered, flat, page_body = render_split_page(
            c, title, sub, md_body, crumbs, (prev_link, next_link), map_links,
            extra=fig + (places_block(c, map_links, "../..") if fname == "places.html" else ""),
            subnav=chapter_siblings(pages, fname, with_days=fname == "schedule.html"))
        (out_dir / fname).write_text(rendered, encoding="utf-8")
        url = f'chapters/{c["name"]}/{fname}'
        scan_sections(c, page_body, url, fixed_cat=page_cat)
        SEARCH_INDEX.append({"t": f'{c["region"]} · {title}', "c": c["title"], "u": url})
        for tok in flat:
            name = tok["name"].strip()
            if name and not name.startswith(("Layer", "Pass")):
                SEARCH_INDEX.append({"t": name, "c": f'{c["title"]} · {title}',
                                     "u": f'{url}#{tok["id"]}'})
        made.append((fname, title, sub, fname == "schedule.html"))

    # 허브
    day_cards = "".join(
        f'<a class="card" href="{f}"><span class="card-title">{html.escape(t)}</span>'
        + (f'<span class="card-sub">{html.escape(s)}</span>' if s else "") + '</a>'
        for f, t, s, is_day in made if is_day)
    topic_cards = "".join(
        f'<a class="card" href="{f}"><span class="card-title">{html.escape(t)}</span>'
        f'<span class="card-sub">{html.escape(s)}</span></a>'
        for f, t, s, is_day in made if not is_day)
    intro_html, _ = md_convert(header_md) if header_md else ("", None)
    intro_html = wrap_tables(rewrite_asset_links(
        mark_layer_headings(intro_html), rel)) if header_md else ""
    hub_body = (
        related_box(c)
        + hero_figure(c["slug"], rel)
        + intro_html
        + (visual_figure("cycles", "Paris 15박의 세 사이클", "../../assets")
           if c["slug"] == "11" else "")
        + f'<h2 class="ic ic-clock">일자</h2><div class="grid">{day_cards}</div>'
        + f'<h2 class="ic ic-topic">주제</h2><div class="grid">{topic_cards}</div>'
    )
    (out_dir / "index.html").write_text(
        page(c["title"], hub_body, rel=rel, topbar_title=c["title"],
             back=region_crumbs((c["region"], None)),
             country=country_of(c["name"]),
             subnav=chapter_siblings(pages, "index.html"),
             meta_line=f'{date_label(c["start"])} ~ {date_label(c["end"])} · {c["nights"]}박'),
        encoding="utf-8")
    SEARCH_INDEX.append({"t": c["title"], "c": "지역", "u": hub_url})
    print(f'  {c["name"]}: 분할 {len(made) + 1}페이지 '
          f'(허브 1 · 일자 {len(days)} · 주제 {len(made) - len(days)})')
    return len(made) + 1


def cat_summary(counts):
    return "  [" + " ".join(f"{k}:{v}" for k, v in counts.items()) + "]" if counts else ""


def fix_summary(n_tokens, n_days, n_grade, n_vol, n_num, n_circ, n_hero=0, n_proc=0,
                n_merge=0, n_move=0):
    bits = []
    for label, n in (("VISUAL토큰 -", n_tokens), ("작성흔적 -", n_proc), ("장소절병합 ", n_merge), ("장소이사 ", n_move), ("Day헤딩 ", n_days), ("등급 ", n_grade),
                     ("재확인 ", n_vol), ("섹션번호 -", n_num), ("원문자 -", n_circ),
                     ("중복사진 -", n_hero)):
        if n:
            bits.append(f"{label}{n}")
    return ("  (" + " · ".join(bits) + ")") if bits else ""


VERIFY_MD = SOURCE / "OPERATIONS" / "116_Phase10_Official_Source_Fact_Verification_Register_v1.0.md"
REVERIFY_MD = SOURCE / "OPERATIONS" / "41_Operational_Variables_and_Reverification_Register_v1.0.md"
GATE_MD = SOURCE / "OPERATIONS" / "117_Departure_and_Daily_Reverification_Calendar_v1.0.md"

# 검증 상태 → 화면 표기. 재확인이 필요한 상태에는 배지를 단다.
VERIFY_STATUS = {
    "VERIFIED": ("확인됨", False),
    "CORRECTED": ("정정됨", False),
    "VERIFIED_CURRENT": ("현재 기준 확인", True),
    "CONFLICT_RECHECK": ("출처 간 불일치", True),
    "DATE_GATE": ("당일 확정", True),
}
# 116·41 이 쓰는 지역 이름 → 챕터 슬러그
VERIFY_REGION = {
    "Barcelona": "04", "Girona": "05", "Nice": "06", "Aix": "07", "Aix/Cassis": "07",
    "Luberon": "08", "Avignon": "09", "Lyon": "10", "Paris": "11",
}


def md_table_rows(text, header_has):
    """머리글에 header_has 가 모두 들어간 표의 데이터 행을 돌려준다."""
    lines, out, cols = text.splitlines(), [], None
    for i, line in enumerate(lines):
        s = line.strip()
        if not (s.startswith("|") and s.endswith("|")):
            cols = None
            continue
        cells = [c.strip() for c in s.strip("|").split("|")]
        if all(re.fullmatch(r":?-{2,}:?", c) for c in cells):
            continue
        if cols is None:
            if all(any(h in c for c in cells) for h in header_has):
                cols = cells
            continue
        out.append(dict(zip(cols, cells)))
    return out


def load_verification():
    """공식자료 검증표(116)를 챕터 슬러그별로 묶는다."""
    rows = md_table_rows(VERIFY_MD.read_text(encoding="utf-8"),
                         ["ID", "지역", "장소", "상태", "확인내용", "공식출처"])
    by_slug = {}
    for r in rows:
        slug = VERIFY_REGION.get(r["지역"])
        if slug:
            by_slug.setdefault(slug, []).append(r)
    return by_slug


def load_reverify():
    """지역별 재확인 항목과 최종 확인시점(41 §3)."""
    rows = md_table_rows(REVERIFY_MD.read_text(encoding="utf-8"),
                         ["지역", "재확인 항목", "최종 확인시점"])
    return {VERIFY_REGION[r["지역"]]: r for r in rows if r["지역"] in VERIFY_REGION}


def load_gates():
    """날짜별 핵심 게이트(117)를 날짜로 푼다."""
    rows = md_table_rows(GATE_MD.read_text(encoding="utf-8"), ["날짜", "확인"])
    out = []
    for r in rows:
        for m in re.finditer(r"(\d+)/(\d+)", r["날짜"]):
            d = date(TRIP_START.year, int(m[1]), int(m[2]))
            if TRIP_START <= d <= TRIP_END:
                out.append((d, r["확인"]))
    return out


def verification_block(c, verified, reverify, gates):
    """공식 출처로 확인된 운영정보와 재확인 대상을 챕터에 싣는다.

    저장소의 검증 등록부에서 그대로 뽑는다 — 새로 지어낸 사실이 없다.
    확정이 아닌 값에는 재확인 배지를 달아 확정처럼 보이지 않게 한다.
    """
    rows = verified.get(c["slug"], [])
    rv = reverify.get(c["slug"])
    mine = [(d, x) for d, x in gates if c["start"] <= d <= c["end"]]
    if not (rows or rv or mine):
        return ""

    out = ["## 공식 확인 정보와 재확인 대상", "",
           "출발 전 공식 출처로 확인한 항목이다. 확정이 아닌 값에는 재확인 표시를",
           "달았다. 계절·행사에 따라 바뀌므로 방문 전 공식 페이지를 다시 본다.", ""]

    if rows:
        out += ["### 공식 출처 확인 항목", "",
                "| 장소 | 항목 | 상태 | 확인 내용 | 출처 |", "|---|---|---|---|---|"]
        for r in rows:
            label, pending = VERIFY_STATUS.get(r["상태"], (r["상태"], True))
            badge = " {{badge:pending|재확인}}" if pending else ""
            src = re.sub(r"\[[^\]]*\]\(([^)]+)\)", r"[공식](\1)", r["공식출처"])
            out.append(f'| {r["장소"]} | {r["항목"]} | {label}{badge} | '
                       f'{r["확인내용"]} | {src} |')
        out.append("")
        acts = [r for r in rows if r.get("조치")]
        if acts:
            out += ["**출발 전 조치**", ""]
            out += [f'- {r["장소"]} — {r["조치"]}' for r in acts]
            out.append("")

    if rv:
        out += ["### 이 지역에서 다시 확인할 것", "",
                f'{rv["재확인 항목"]} {{{{badge:pending|재확인}}}}', "",
                f'**최종 확인시점** {rv["최종 확인시점"]}', ""]

    if mine:
        out += ["### 날짜에 걸린 확인", "",
                "| 날짜 | 확인 |", "|---|---|"]
        for d, what in mine:
            out.append(f'| Day {day_no(d)} · {d.month}월 {d.day}일 '
                       f'{WEEKDAY_KO[d.weekday()]} | {what} |')
        out.append("")

    src_names = []
    if rows:
        src_names.append("공식자료 검증표(116)")
    if rv:
        src_names.append("변동정보·재확인 등록부(41)")
    if mine:
        src_names.append("출발 전·날짜별 재검증 달력(117)")
    out += [f'> 출처 — {" · ".join(src_names)}. '
            f'본문은 이 등록부를 그대로 옮긴 것이고 새로 추가한 사실은 없다.', ""]
    return "\n".join(out)


def verification_status_block(c, verified, reverify, gates):
    """챕터 말미의 검증 상태표. Girona 보강본의 형식을 따른다."""
    rows = verified.get(c["slug"], [])
    rv = reverify.get(c["slug"])
    mine = [(d, x) for d, x in gates if c["start"] <= d <= c["end"]]
    if not (rows or rv or mine):
        return ""
    out = ["## 검증 상태", "", "| 항목 | 상태 |", "|---|---|"]
    for r in rows:
        label, pending = VERIFY_STATUS.get(r["상태"], (r["상태"], True))
        mark = f"**{label}** — 방문 전 공식 페이지 재확인" if pending else label
        out.append(f'| {r["장소"]} · {r["항목"]} | {mark} |')
    if rv:
        out.append(f'| 지역 재확인 항목 | **미확정** — {rv["최종 확인시점"]} |')
    if mine:
        out.append(f'| 날짜 게이트 {len(mine)}건 | **당일 확정** |')
    out += [f'| 방문지 해설 | Phase 9D 공식링크 정리분 |',
            f'| 숙소·식당·공연 가격 | **미확정** — Phase 8B 예약 완료 전 |', ""]
    out += ["> 이 표는 저장소의 검증 등록부에서 생성된다. 등록부가 갱신되면 함께 바뀐다.", ""]
    return "\n".join(out)


def write_redirect(dest, target, label):
    """옮겨간 주소에 리다이렉트를 남긴다.

    즐겨찾기·외부 링크가 걸려 있을 수 있다. JS 없이도 동작하도록
    meta refresh 와 본문 링크를 함께 둔다.
    """
    css = "../" * (len(Path(dest).relative_to(SITE).parts) - 1) + "assets/style.css"
    dest.write_text(f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<meta name="robots" content="noindex">
<meta http-equiv="refresh" content="0; url={target}">
<link rel="canonical" href="{target}">
<title>{html.escape(label)} — 주소가 바뀌었습니다</title>
<link rel="stylesheet" href="{css}">
</head>
<body>
<main>
<h1>주소가 바뀌었습니다</h1>
<p>이 페이지는 <a href="{target}">{html.escape(label)}</a> 로 옮겨졌습니다.
자동으로 넘어가지 않으면 링크를 누르세요.</p>
</main>
</body>
</html>
""", encoding="utf-8")


def write_legacy_redirect(c):
    """기존 번호 URL(`chapters/05.html`) → 지명 URL."""
    write_redirect(SITE / "chapters" / f'{c["slug"]}.html',
                   chapter_url(c).replace("chapters/", "", 1), c["title"])


VERIFIED_FACTS = {}
REVERIFY_ITEMS = {}
DATE_GATES = []


def build_chapters():
    global VERIFIED_FACTS, REVERIFY_ITEMS, DATE_GATES
    for r in load_place_registry():
        if r["type"] == "spot" and r["head"]:
            PLACE_BY_HEAD.setdefault(r["head"], r["slug"])
    VERIFIED_FACTS = load_verification()
    REVERIFY_ITEMS = load_reverify()
    DATE_GATES = load_gates()
    out_dir = SITE / "chapters"
    out_dir.mkdir(parents=True, exist_ok=True)
    by_file = {Path(c["path"]).name: c for c in CHAPTERS}
    map_links = load_map_links()

    for i, c in enumerate(CHAPTERS):
        text = (SOURCE / c["path"]).read_text(encoding="utf-8")
        meta, body_md = parse_frontmatter(text)
        # 토큰 제거가 먼저다 — 헤딩 텍스트가 목차·검색 인덱스·앵커의 원천이라
        # md_convert 이후에 손대면 토큰이 data.js 로 새어 나간다.
        body_md, n_tokens = strip_visual_tokens(body_md)
        body_md, n_proc = strip_process_notes(body_md)
        n_verify = 0
        if c["kind"] == "region":
            # 검증 등록부 블록은 분류 이전에 붙인다. 그래야 카테고리로 배분된다.
            vb = verification_block(c, VERIFIED_FACTS, REVERIFY_ITEMS, DATE_GATES)
            sb = verification_status_block(c, VERIFIED_FACTS, REVERIFY_ITEMS, DATE_GATES)
            if vb or sb:
                body_md = body_md.rstrip() + "\n\n" + vb + "\n" + sb
                n_verify = 1
            body_md, counts = regroup_regional(c["slug"], body_md, chapter_rel(c))
        # 분류는 원본 제목으로 끝난 뒤에 Day 헤딩을 정규화한다 (CAT_OVERRIDES 보존)
        body_md, n_days = normalize_day_headings(body_md, c)
        body_md, n_num, n_circ = strip_naming_noise(body_md)
        n_merge = n_move = 0
        if c["kind"] == "region":
            body_md, n_merge = merge_place_sections(body_md, c["slug"])
            body_md, n_move = extract_place_bodies(body_md, c["slug"], chapter_rel(c))
        body_md, n_hero = drop_source_hero(body_md)
        body_md, n_grade, n_vol = annotate_tables(body_md)
        body_md = render_inline_tokens(body_md)
        body, toc_tokens = md_convert(body_md)
        flat = flatten_tokens(toc_tokens)
        body = mark_category_icons(mark_layer_headings(wrap_tables(
            rewrite_asset_links(rewrite_md_links(body, by_file, chapter_rel(c)),
                                chapter_rel(c)))))
        if c["kind"] == "region":
            def insert_after(pattern, html_block, body=None):
                m = re.search(pattern, body)
                return (body[:m.end()] + html_block + body[m.end():]) if m else body
            # 여행정보 카테고리 첫머리에 주요 방문지 카드 삽입
            body = insert_after(r'<h1 id="[^"]*" class="layer-h">여행정보</h1>',
                                places_block(c, map_links, chapter_rel(c)), body=body)
            # 지역소개 첫머리에 대표 사진 (CC 크레딧)
            body = insert_after(r'<h1 id="[^"]*" class="layer-h">지역소개</h1>',
                                hero_figure(c["slug"], chapter_rel(c)), body=body)
            # 편집 도식: 권장 위치 기준 삽입
            if c["slug"] in ("07", "08", "09"):
                body = insert_after(r'<h1 id="[^"]*" class="layer-h">교통</h1>',
                                    visual_figure("cardays", "Provence 차량일 운영 논리", chapter_rel(c) + "/assets"), body=body)
            if c["slug"] == "11":
                body = insert_after(r'<h1 id="[^"]*" class="layer-h">일정</h1>',
                                    visual_figure("cycles", "Paris 15박의 세 사이클", chapter_rel(c) + "/assets"), body=body)

        collect_chapter_dates(c, flat)
        if c.get("name") in SPLIT_CHAPTERS:
            build_split_chapter(c, body_md, map_links)
            write_legacy_redirect(c)
            print(f'  {c["name"]}: {Path(c["path"]).name} → chapters/{c["name"]}/'
                  f'{cat_summary(counts)}{fix_summary(n_tokens, n_days, n_grade, n_vol, n_num, n_circ, n_hero, n_proc, n_merge, n_move)}')
            continue
        collect_search(c, flat)

        if c["slug"] == "01":
            body = re.sub(r"(</h1>)", r"\1" + visual_figure("rhythm", "생활형 여행의 하루 리듬"),
                          body, count=1)
        if c["slug"] == "03":
            body = re.sub(r"(</h1>)", r"\1" + visual_figure("route", "43일 전체 루트와 숙박구조"),
                          body, count=1)
            body += visual_figure("fatigue", "피로도와 일정 삭제순서")

        meta_bits = []
        if c["kind"] == "region":
            meta_bits.append(f'{date_label(c["start"])} ~ {date_label(c["end"])} · {c["nights"]}박')
        if meta.get("version"):
            meta_bits.append(f'v{meta["version"]}')

        rel = chapter_rel(c)
        prev_link = next_link = ""
        if i > 0:
            p = CHAPTERS[i - 1]
            prev_link = f'<a href="{rel}/{chapter_url(p)}">← {p["title"]}</a>'
        if i < len(CHAPTERS) - 1:
            nx = CHAPTERS[i + 1]
            next_link = f'<a href="{rel}/{chapter_url(nx)}">{nx["title"]} →</a>'
        pager = f'<nav class="pager">{prev_link}<span></span>{next_link}</nav>'

        scan_sections(c, body, chapter_url(c))
        content = related_box(c) + toc_html(toc_tokens) + link_day_cards(body, rel) + pager
        dest = SITE / chapter_url(c)
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_text(
            page(c["title"], content, rel=rel,
                 topbar_title=c["title"], back=crumbs_for((c["title"], None)),
                 meta_line=" · ".join(meta_bits),
                 coords=chapter_coords(c, rel),
                 subnav=chapter_subnav(c, flat)),
            encoding="utf-8")
        write_legacy_redirect(c)
        print(f'  {c["name"]}: {Path(c["path"]).name} → {chapter_url(c)}'
              f'{cat_summary(counts if c["kind"] == "region" else {})}'
              f'{fix_summary(n_tokens, n_days, n_grade, n_vol, n_num, n_circ, n_hero, n_proc, n_merge, n_move)}')


# ---------------------------------------------------------------- daily cards

def find_daily_images():
    """Day 번호 → 카드 이미지 경로. Day 12–24는 Phase 4 카드를 우선한다."""
    images = {}
    for f in sorted(DAILY_IMG_DIR.glob("Day_*.png")):
        m = re.match(r"Day_(\d+)_(.+)\.png", f.name)
        if m:
            images[int(m.group(1))] = {"src": f, "region": m.group(2).replace("_", " ")}
    for f in sorted(PHASE4_DIR.glob("Day_*_Phase4_v1.0.png")):
        m = re.match(r"Day_(\d+)_(.+)_Phase4_v1\.0\.png", f.name)
        if m and int(m.group(1)) in PHASE4_DAYS:
            images[int(m.group(1))] = {"src": f, "region": m.group(2).replace("_", " "),
                                       "phase4": True}
    return images


# ⚠ 세 플래그는 뜻이 다르다. 뭉치면 Day 43(귀국 항공)에 경고가 안 뜬다.
P0_CONNECTION = [4, 7, 12, 24, 28, 43]        # 놓치면 대안이 없는 교통 연결
MAP_TRANSITION = [4, 7, 12, 16, 20, 24, 28]   # 실행지도 2장이 필요한 날
DUAL_CHAPTER = [12, 16, 20, 24, 28]           # 양쪽 챕터에 원고가 있는 날

AUDIT_MD = SOURCE / "OPERATIONS" / "100_Whole_Trip_43_Day_Execution_Audit_v1.0.md"
AUDIT_FIELDS = ["day", "date", "base", "core", "depart", "buffer",
                "meals", "risk", "cut", "alt", "lock"]
AUDIT_LABELS = {
    "core": "핵심 실행", "depart": "권장 출발", "buffer": "완충",
    "meals": "식사·휴식", "risk": "최고 리스크", "cut": "우선 삭제",
    "alt": "대체안", "lock": "잠금 필요",
}
TIME_CELL_RE = re.compile(r"^\d{1,2}:\d{2}\s*[–\-~]\s*\d{1,2}:\d{2}$|^\d{1,2}:\d{2}$")
FATIGUE_RE = re.compile(r"([0-9]+(?:[–\-~][0-9]+)?)\s*/\s*5")


def md_inline(text):
    """표 셀 하나 분량의 최소 마크다운 (굵게·기울임·코드·링크)."""
    t = html.escape(text, quote=False)
    t = re.sub(r"\[([^\]]+)\]\(([^)\s]+)\)",
               lambda m: f'<a href="{html.escape(m.group(2), quote=True)}">{m.group(1)}</a>', t)
    t = re.sub(r"\*\*([^*]+)\*\*", r"<strong>\1</strong>", t)
    t = re.sub(r"(?<!\*)\*([^*]+)\*(?!\*)", r"<em>\1</em>", t)
    t = re.sub(r"`([^`]+)`", r"<code>\1</code>", t)
    return render_inline_tokens(t)


def load_audit():
    """43일 실행 감사표를 Day 번호 → dict 로 읽는다. 결측이 있으면 중단한다."""
    rows = {}
    for line in AUDIT_MD.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if not (s.startswith("|") and re.match(r"^\|\s*\d+\s*\|", s)):
            continue
        cells = [c.strip() for c in s.strip("|").split("|")]
        if len(cells) != len(AUDIT_FIELDS):
            sys.exit(f"감사표 열 수 불일치: {cells[:2]} ({len(cells)}열)")
        row = dict(zip(AUDIT_FIELDS, cells))
        n = int(row["day"])
        empty = [k for k in AUDIT_LABELS if not row[k] or row[k] in ("—", "-")]
        if empty:
            sys.exit(f"감사표 Day {n} 결측 필드: {empty}")
        rows[n] = row
    if sorted(rows) != list(range(1, 44)):
        sys.exit(f"감사표가 43일이 아니다: {len(rows)}행")
    return rows


_DAY_DETAILS = None


def load_day_details():
    """지역 챕터에서 날짜별 피로도와 시간표를 긁는다.

    피로도는 원고에 있는 날만 담는다. 없는 날을 추정해 채우지 않는다 —
    현장에서 틀린 수치는 없는 것보다 나쁘다.
    """
    global _DAY_DETAILS
    if _DAY_DETAILS is not None:
        return _DAY_DETAILS
    fatigue, timetable, conflicts = {}, {}, []
    for c in CHAPTERS:
        if c["kind"] != "region":
            continue
        text = (SOURCE / c["path"]).read_text(encoding="utf-8")
        lines = text.splitlines()

        # (1) `날짜별 … 요약` 표의 피로도 열
        i = 0
        while i < len(lines) - 1:
            row, sep = lines[i].strip(), lines[i + 1].strip()
            if row.startswith("|") and re.fullmatch(r"\|[\s:|-]+\|", sep):
                header = [x.strip() for x in row.strip("|").split("|")]
                fi = next((k for k, h in enumerate(header) if "피로도" in h), None)
                di = next((k for k, h in enumerate(header) if "날짜" in h or "일자" in h), None)
                if fi is not None and di is not None:
                    j = i + 2
                    while j < len(lines) and lines[j].strip().startswith("|"):
                        cells = [x.strip() for x in lines[j].strip().strip("|").split("|")]
                        if len(cells) > max(fi, di):
                            dm = re.match(r"(\d+)/(\d+)", cells[di])
                            fm = FATIGUE_RE.search(cells[fi])
                            if dm and fm:
                                key = date(TRIP_START.year, int(dm[1]), int(dm[2])).isoformat()
                                fatigue.setdefault(key, {})[c["slug"]] = fm[1]
                        j += 1
                    i = j
                    continue
            i += 1

        # (2) Day 섹션 안의 `오늘의 피로도` 와 시각표
        parts = re.split(r"^#{1,6}[^\n]*Day\s*\d+\s*[—–-]\s*(\d+)월\s*(\d+)일[^\n]*$",
                         text, flags=re.M)
        for k in range(1, len(parts), 3):
            d = date(TRIP_START.year, int(parts[k]), int(parts[k + 1]))
            key, body = d.isoformat(), parts[k + 2]
            fm = re.search(r"오늘의 피로도[:\s]*\**\s*" + FATIGUE_RE.pattern, body)
            if fm:
                fatigue.setdefault(key, {})[c["slug"]] = fm[1]
            rows = []
            for line in body.splitlines():
                s = line.strip()
                if not (s.startswith("|") and s.endswith("|")):
                    continue
                cells = [x.strip() for x in s.strip("|").split("|")]
                if cells and TIME_CELL_RE.match(cells[0].replace("**", "")):
                    rows.append(cells)
            if rows:
                timetable.setdefault(key, []).append((c["slug"], c["region"], rows))

    for key, per in fatigue.items():
        if len(set(per.values())) > 1:
            conflicts.append(f"{key} {per}")
    _DAY_DETAILS = (fatigue, timetable, conflicts)
    return _DAY_DETAILS


def fatigue_html(value):
    """피로도 막대 + 숫자. 색을 빼도 막대 개수와 숫자로 읽힌다."""
    lead = int(re.match(r"\d+", value).group(0))
    bars = "".join(f'<i class="{"on" if k < lead else ""}"></i>' for k in range(5))
    return (f'<span class="fatigue" data-v="{lead}" '
            f'aria-label="피로도 {value} / 5">{bars}<b>{value}/5</b></span>')


def day_flags(n):
    """Day 번호에 붙는 플래그 배지. 세 목록을 각각 본다."""
    out = []
    if n in P0_CONNECTION:
        out.append('<span class="badge badge-p0">P0 연결</span>')
    if n in MAP_TRANSITION:
        out.append('<span class="badge badge-pending">거점 이동</span>')
    if n in DUAL_CHAPTER:
        out.append('<span class="badge badge-rest">양쪽 챕터</span>')
    return out


# `**피로도**: 4/5.` `오늘의 피로도 1–2/5` — 문장 앞머리의 표기만 집어낸다.
# 뒤에 붙는 안내("…해변 체류를 조절한다")는 원고에 그대로 남긴다.
FATIGUE_TOKEN_RE = re.compile(
    r"\**\s*(?:오늘의\s*)?피로도\s*\**\s*[:：]?\s*\**\s*"
    + FATIGUE_RE.pattern + r"(?:\s*[.。]|\s*\*{1,2})*\s*")


def pull_fatigue(md):
    """원고에서 피로도 값을 꺼내고, 그 표기만 본문에서 지운다.

    막대로 보여줄 것을 문장으로 한 번 더 반복하지 않는다. 값을 못 찾으면
    아무것도 건드리지 않고 아무 말도 하지 않는다 — '표기가 없다' 고 단정하던
    문구가 20일에서 틀렸다. 원고에 있는 것을 없다고 말한 것이다.
    """
    for line in md.splitlines():
        s = line.strip()
        if s.startswith("|") or not s:
            continue
        m = FATIGUE_TOKEN_RE.match(s)
        if m:
            rest = s[m.end():].strip()
            return md.replace(line, rest, 1) if rest else md.replace(line + "\n", "", 1), \
                m.group(1)
    return md, None


TT_HEADER_RE = re.compile(r"시간|시각|시간대")


def is_timetable(rows):
    """시간표인가. 머리글에 '시간'이 있거나 첫 열이 시각·상대시간이면 그렇다.

    Day 섹션에는 표가 여럿이다 — 리옹 원고의 첫 표는 시간표가 아니라
    피로도·삭제 우선순위 표다. 순서로 고르면 그 표가 시간표 자리에 앉는다.
    """
    if len(rows) < 3:
        return False
    header = [x.strip() for x in rows[0].strip().strip("|").split("|")]
    if header and TT_HEADER_RE.search(header[0]):
        return True
    firsts = [r.strip().strip("|").split("|")[0].replace("**", "").strip()
              for r in rows[2:]]
    hits = sum(1 for x in firsts
               if TIME_CELL_RE.match(x) or re.search(r"\d{1,2}:\d{2}|기차\s*\d+", x))
    return hits >= max(2, len(firsts) // 2)


def split_day_md(md):
    """그 날 원고를 (시간표, 나머지 본문) 으로 가른다.

    표는 통째로 옮긴다. 예전에는 첫 셀이 시각인 행만 골라 담아서
    `저녁 · 로티세리 치킨` 같은 줄이 사라졌다. 43일 중 30일의 시간표가
    챕터 쪽과 어긋나 있었던 원인이다.
    """
    lines = md.splitlines()
    table, rest, k = [], [], 0
    while k < len(lines):
        s = lines[k].strip()
        if s.startswith("|") and k + 1 < len(lines) \
                and re.fullmatch(r"\|[\s:|-]+\|", lines[k + 1].strip()):
            block, j = [], k
            while j < len(lines) and lines[j].strip().startswith("|"):
                block.append(lines[j])
                j += 1
            if not table and is_timetable(block):
                table = block
            else:
                rest.extend(block)
            k = j
            continue
        rest.append(lines[k])
        k += 1
    return "\n".join(table), "\n".join(rest).strip()


def build_daily():
    out_dir = SITE / "daily"
    img_dir = out_dir / "img"
    img_dir.mkdir(parents=True, exist_ok=True)
    images = find_daily_images()

    missing = [n for n in range(1, 44) if n not in images]
    if missing:
        print("데일리 카드 누락:", missing)
        sys.exit(1)

    # Day 섹션 앵커는 챕터 빌드가 모아 둔다. 여기서 합쳐야 데일리 카드가
    # 챕터 첫머리가 아니라 그 날 섹션으로 바로 간다 — 파리 챕터는 5만 자다.
    CHAPTER_DATE_URL.update(DAY_OVERRIDES)

    audit = load_audit()
    fatigue, timetable, conflicts = load_day_details()
    TIMETABLE.update(timetable)
    for x in conflicts:
        print(f"  주의: 피로도 값이 챕터마다 다름 — {x} (도착 챕터 값을 쓴다)")

    index_items = []
    n_fat = n_tt = 0
    for n in range(1, 44):
        d = date_of_day(n)
        key = d.isoformat()
        info = images[n]
        row = audit[n]
        shutil.copy(info["src"], img_dir / f"day-{n:02d}.png")
        c = chapter_for_date(d)
        wd = WEEKDAY_KO[d.weekday()]
        title = f"Day {n} · {date_label(d)} {wd} · {row['base']}"

        # ── 1. 날짜 · 거점 · 플래그
        flags = day_flags(n)
        flag_html = f'<p class="day-flags">{"".join(flags)}</p>' if flags else ""

        # ── 2. 실행 요약. 첫 화면은 현장에서 바로 쓸 세 가지만 놓고,
        # 삭제·대체·식사·잠금은 필요할 때 펴친다.
        quick_summary = "".join(
            f'<div class="dq-item dq-{k}"><dt>{AUDIT_LABELS[k]}</dt>'
            f'<dd>{html.escape(row[k])}</dd></div>'
            for k in ("core", "depart", "buffer"))
        summary = "".join(
            f"<div class=\"ds-item\"><dt>{AUDIT_LABELS[k]}</dt>"
            f"<dd>{html.escape(row[k])}</dd></div>"
            for k in ("risk", "cut", "alt", "meals", "lock"))
        audit_block = (f'<details class="day-details"><summary>상세 실행 확인</summary>'
                       f'<dl class="day-summary">{summary}</dl></details>')

        # ── 3·4. 시간표 · 피로도 · 메모 — 그 날 원고 하나에서 전부 나온다.
        #   전에는 이 자리에 원고를 정규식으로 다시 긁은 사본이 있었고, 챕터
        #   쪽에는 원본이 따로 렌더됐다. 같은 하루가 두 페이지에서 다른 말을
        #   했다. 사본을 없애고 원고를 여기로 옮겼다.
        blocks = DAY_MD.get(n, [])
        multi = len(blocks) > 1        # 거점 이동일은 출발지·도착지 원고가 둘이다
        tt_parts, note_parts, value = [], [], None
        for region, _dtitle, dmd in blocks:
            body_md, fat = pull_fatigue(dmd)
            value = value or fat
            table_md, rest_md = split_day_md(body_md)
            head = f"<h3>{html.escape(region)}</h3>" if multi else ""
            if table_md:
                n_tt += 1
                tbl, _ = md_convert(table_md)
                tt_parts.append(head + wrap_tables(tbl).replace(
                    "<table>", '<table class="day-time">'))
            # 첫 줄은 `## Day 9 · 9월 6일 일` 헤딩이다. h1 이 이미 같은 말을 한다.
            rest_md = re.sub(r"^#{1,6}\s*Day\s*\d+[^\n]*\n?", "", rest_md).strip()
            if rest_md:
                rh, _ = md_convert(rest_md)
                note_parts.append(head + link_place_cards(
                    wrap_tables(rewrite_asset_links(rh, "..")), ".."))

        if tt_parts:
            tt_block = '<h2 class="ic ic-clock">시간표</h2>' + "".join(tt_parts)
        else:
            # 원고에 그 날의 시각표가 없는 날이 있다. 지어내지 않고, 그 구간의
            # 일정을 한 표로 담은 '일정 한눈에' 로 보낸다.
            where = CHAPTER_DATE_URL.get(key, ITINERARY_URL)
            tt_block = ('<h2 class="ic ic-clock">시간표</h2><p class="note">이 날은 원고에 시각표가 없다. '
                        f'구간 일정은 <a href="../{where}">일정 한눈에</a> 에 있다.</p>')
        if value:
            n_fat += 1
            fat_block = f'<h2 class="ic ic-gauge">피로도</h2><p>{fatigue_html(value)}</p>'
        else:
            fat_block = ""
        note_block = ('<h2 class="ic ic-note">이 날의 메모</h2>' + "".join(note_parts)) if note_parts else ""

        # ── 5·6. 장소 · 지도 (거점 이동일은 두 지역 지도를 모두 건다)
        ch_url = CHAPTER_DATE_URL.get(key, ITINERARY_URL)
        maps_seen, map_links = [], []
        for cand in (chapter_for_date(d), chapter_for_date(d - timedelta(days=1))):
            if cand and cand["map"] not in maps_seen:
                maps_seen.append(cand["map"])
                map_links.append(
                    f'<a href="../maps/{cand["map"]}"><b class="ic ic-map" aria-hidden="true"></b>{cand["region"]} 실행지도</a>')
        if n not in MAP_TRANSITION:
            map_links = map_links[:1]
        links = [f'<a href="../{ch_url}"><b class="ic ic-list" aria-hidden="true"></b>챕터 일정</a>'] + map_links
        more_links = [
            '<a href="../maps/offline.html"><b class="ic ic-download" aria-hidden="true"></b>오프라인 지도</a>',
            '<a href="index.html"><b class="ic ic-today" aria-hidden="true"></b>전체 데일리 목록</a>']

        # ── 6b. 이 날의 장소 — 장소 축에서 되돌아오는 문
        spots = PLACES_BY_DAY.get(n, [])
        if spots:
            def chip(s):
                g = f' <b>{GRADE_LABEL[s["grade"]]}</b>' if s["grade"] else ""
                return (f'<a class="pl-day" href="../places/{s["slug"]}.html">'
                        f'{html.escape(s["name"])}{g}</a>')

            chips = "".join(chip(s) for s in spots)
            place_block = (f'<h2 class="ic ic-pin">이 날의 장소</h2><div class="pl-chips">{chips}</div>'
                           f'<p class="note">일자별 시간표와 본문 표기에서 찾은 것이다. '
                           f'하루의 전부가 아니라 장소 카드가 있는 것만 담는다.</p>')
        else:
            place_block = ""

        # ── 7·8. 원문 · 카드 이미지
        card_block = f"""<details class="day-details day-card-archive"><summary>모바일 카드 이미지</summary>
<figure class="daily-card"><img src="img/day-{n:02d}.png"
  alt="Day {n} 데일리 모바일 가이드 카드"></figure>
<p class="note">카드의 지도영역은 일정 순서를 보여주는 개요이며 내비게이션이 아닙니다.
실제 도보·운전 경로는 Google Maps에서 다시 계산하세요.</p></details>"""

        prev_link = f'<a href="day-{n-1:02d}.html">← Day {n-1}</a>' if n > 1 else ""
        next_link = f'<a href="day-{n+1:02d}.html">Day {n+1} →</a>' if n < 43 else ""
        pager = f'<nav class="pager">{prev_link}<span></span>{next_link}</nav>'
        body = f"""<h1>{title}</h1>
{flag_html}
<section class="day-command" aria-label="오늘 실행 요약">
<dl class="day-quick">{quick_summary}</dl>
<nav class="day-actions" aria-label="오늘 바로가기">{''.join(links)}</nav>
</section>
{tt_block}
{fat_block}
{place_block}
{note_block}
{audit_block}
<div class="related day-more">{''.join(more_links)}</div>
{card_block}
{pager}"""
        region_cell = ((c["region"], CHAPTER_DATE_URL.get(key, ITINERARY_URL))
                       if c else ("이동 중", ITINERARY_URL))
        # 43일 전체를 한 줄에 늘어놓으면 못 쓴다. 현재 날짜 앞뒤 창만 보인다.
        lo, hi = max(1, n - 5), min(43, n + 5)
        window = [(f"{date_label(date_of_day(k))}", f"day-{k:02d}.html", k == n,
                   f"Day {k}") for k in range(lo, hi + 1)]
        day_nav = siblings_nav([
            ("일자", "sn-layers", [("전체 목록", "index.html", False, ""),
                                 ("← 앞", f"day-{max(1, n - 1):02d}.html", False, ""),
                                 ("뒤 →", f"day-{min(43, n + 1):02d}.html", False, "")]),
            ("날짜", "sn-days", window)])
        regs = regions_of_date(d)
        reg_crumb = ((" · ".join(x["region"] for x in regs), chapter_url(regs[0]))
                     if regs else ("이동", ITINERARY_URL))
        (out_dir / f"day-{n:02d}.html").write_text(
            page(title, body, rel="..", topbar_title=title, subnav=day_nav,
                 back=region_crumbs(reg_crumb, (f"Day {n}", None)),
                 country=country_of(*(x["name"] for x in regs)),
                 coords=coords_bar("..",
                                   day=(f"Day {n} · {date_label(d)} {wd}", None),
                                   region=region_cell,
                                   topic=("전체 주제", "topics/index.html"))),
            encoding="utf-8")

        index_items.append((c["region"] if c else "이동",
            f'<a class="daily-item{" is-p0" if n in P0_CONNECTION else ""}" href="day-{n:02d}.html">'
            f'<b>Day {n}</b><span>{date_label(d)} {wd}</span>'
            f'<span class="di-region">{html.escape(row["base"])}</span>'
            f'<span class="di-core">{html.escape(row["core"])}</span></a>'))
        SEARCH_INDEX.append({"t": title, "c": "데일리 가이드", "u": f"daily/day-{n:02d}.html"})
        SEARCH_INDEX.append({"t": f"Day {n} 핵심 실행 — {row['core']}",
                             "c": "데일리 가이드", "u": f"daily/day-{n:02d}.html"})
        for slug, region, rows in timetable.get(key, []):
            for cells in rows:
                if len(cells) >= 2:
                    label = re.sub(r"<[^>]+>", "", md_inline(cells[1]))
                    SEARCH_INDEX.append({
                        "t": f"Day {n} {cells[0]} {label}", "c": "일자별 시간표",
                        "u": f"daily/day-{n:02d}.html"})

    groups = []
    for region, items in itertools.groupby(index_items, key=lambda x: x[0]):
        cards = "".join(x[1] for x in items)
        groups.append(f'<section class="daily-region"><h2>{html.escape(region)}</h2>'
                      f'<div class="daily-grid">{cards}</div></section>')
    body = ('<h1>43일 실행 일정</h1>'
            '<p class="meta">지역별로 날짜·거점·핵심 실행을 확인한다. '
            '빨간 표시는 놓치면 대안이 없는 교통 연결일이다.</p>'
            f'{"".join(groups)}')
    (out_dir / "index.html").write_text(
        page("데일리 가이드", body, rel="..", topbar_title="데일리 가이드",
             back=crumbs_for(("데일리", None)),
             coords=coords_bar("..", day=("43일", None),
                               region=("8곳", "regions.html"),
                               topic=("분류·상태", "topics/index.html"))),
        encoding="utf-8")
    print(f"  데일리 카드: 43일 → daily/day-01~43.html (Phase4 적용 {len(PHASE4_DAYS)}일)"
          f" · 피로도 {n_fat}일 · 시간표 {n_tt}건")


# ---------------------------------------------------------------- home

def build_home():
    """홈 — 여행자가 지금 할 행동부터 고른다.

    편집 분류가 아니라 오늘/일정/지역/준비/검색/비상 순으로 연다. 상세
    여정과 주제·지도는 아래에서 계속 접근할 수 있지만 첫 화면의 선택지를
    압도하지 않는다.
    """
    stops, prev_es = [], False
    for c in CHAPTERS:
        if c["kind"] != "region":
            continue
        first = day_no(c["start"])
        acts = [("일정", f'chapters/{c["name"]}/schedule.html'
                 if (SITE / "chapters" / c["name"] / "schedule.html").exists()
                 else chapter_url(c)),
                ("첫날", f"daily/day-{first:02d}.html"),
                ("지도", f'maps/{c["map"]}')]
        btns = "".join(f'<a class="tl-act" href="{u}">{n}</a>' for n, u in acts)
        # 스페인에서 프랑스로 넘어가는 지점. 여기서 시장 요일·공휴일·긴급번호가
        # 바뀐다. 여정 목록에서 그 자리가 보여야 한다.
        if c["name"] not in ES_REGIONS and prev_es:
            stops.append(f'''<li class="tl-border" aria-label="국경">
  <div class="bc">
    <span class="bc-flag bc-es" aria-hidden="true"></span>
    <span class="bc-label">{date_label(c["start"])} · 국경을 넘는다 — 스페인에서 프랑스로</span>
    <span class="bc-flag bc-fr" aria-hidden="true"></span>
  </div>
</li>''')
        prev_es = c["name"] in ES_REGIONS
        stops.append(f"""<li>
  <div class="tl-body" data-region="{c["name"]}">
    <div class="tl-head">
      <a class="tl-title" href="{chapter_url(c)}">{c["title"]}</a>
      <span class="tl-when">{date_label(c["start"])}–{date_label(c["end"])}
        ({c["nights"]}박)</span>
    </div>
    <div class="tl-acts">{btns}</div>
  </div>
</li>""")

    intro_cards = "".join(
        f'<a class="card" href="{chapter_url(c)}">'
        f'<span class="card-title">{c["title"]}</span>'
        f'<span class="card-sub">{c["sub"]}</span></a>'
        for c in CHAPTERS if c["kind"] != "region")

    primary_rows = "".join(
        f'<a class="home-action{extra}" href="{u}">'
        f'<span class="ha-icon ic ic-only ic-{g}" aria-hidden="true"></span>'
        f'<span class="ha-copy"><b>{n}</b><span>{s}</span></span>'
        f'<span class="lr-go" aria-hidden="true">›</span></a>'
        for g, u, n, s, extra in (
            ("list", ITINERARY_URL, "43일 일정", "날짜별 전체 여정을 확인", ""),
            ("region", "regions.html", "지역별 가이드", "8개 거점의 일정과 생활권", ""),
            ("check", "tracker/index.html", "여행 준비", "예약 · 교통 · 숙소 상태", ""),
            ("search", "#search", "통합 검색", "날짜 · 장소 · 식당 · 교통 찾기", " nav-search"),
            ("flag", "maps/offline.html", "비상 · 오프라인", "연결이 약할 때 필요한 정보", "")))

    # 하단탭(L0)이 없는 보조 진입점은 홈에 남긴다. 본문을 복제하지 않고
    # 원본 목록으로만 연결해 길이 끊기지 않게 한다.
    tool_rows = "".join(
        f'<a class="list-row" href="{u}">'
        f'<span class="lr-icon ic ic-only ic-{g}" aria-hidden="true"></span>'
        f'<span class="lr-text"><b class="lr-title">{n}</b>'
        f'<span class="lr-sub">{s}</span></span>'
        f'<span class="lr-go" aria-hidden="true">›</span></a>'
        for g, u, n, s in (
            ("today", "daily/index.html", "데일리 카드 전체", "43일을 한 줄씩"),
            ("pin", "places/index.html", "장소", "갈 곳 전체 목록"),
            ("topic", "topics/index.html", "주제별 보기", "먹거리 · 교통 · 숙박 등"),
            ("map", "maps/index.html", "지역 지도", "주요 기준점과 외부 길찾기"),
            ("book", "chapters/how-to-use.html", "가이드 사용법", "이 가이드북을 읽는 법"),
            ("license", "credits.html", "사진 저작자 표시", "CC 라이선스와 출처")))

    body = f"""<section class="hero">
  <h1>{SITE_TITLE}</h1>
  <p class="period">{TRIP_PERIOD}</p>
  <div class="today-bar">
    <span class="today-date" id="today-date">{TRIP_START.isoformat()}</span>
    <a href="#" class="nav-today btn-today">오늘 일정 열기</a>
  </div>
</section>
<nav class="home-actions" aria-label="여행 주요 기능">{primary_rows}</nav>
<h2 class="ic ic-list">전체 여정</h2>
<ol class="timeline">{''.join(stops)}</ol>
<h2 class="ic ic-book">Quick Summary</h2>
<div class="grid">{intro_cards}</div>
<h2 class="ic ic-topic">더 보기</h2>
<div class="list-group">{tool_rows}</div>
<p class="note">지도 배경 타일은 인터넷 연결 시 표시됩니다.
본문·데일리 카드·마커 목록은 오프라인에서도 열람됩니다.</p>
"""
    (SITE / "index.html").write_text(
        page("홈", body, rel=".", topbar_title=SITE_SHORT,
             back=[(SITE_SHORT, None)],
             coords=coords_bar(".", day=("43일", "daily/index.html"),
                               region=("8곳", "regions.html"),
                               topic=("분류·상태", "topics/index.html"))),
        encoding="utf-8")
    print("  홈 → index.html")


# ---------------------------------------------------------------- maps

# 엔티티로 시작하지 않는 맨 & — `Pepper & Paper` 같은 장소 이름에서 나온다.
BARE_AMP_RE = re.compile(r"&(?!(?:[A-Za-z][A-Za-z0-9]*|#[0-9]+|#[xX][0-9A-Fa-f]+);)")


def sanitize_kml(text, label):
    """KML 을 well-formed XML 로 만들고 파싱을 확인한다.

    원본 KML 에 이스케이프되지 않은 `&` 가 있어 XML 파서가 거부한다.
    Organic Maps 도 같은 이유로 임포트에 실패한다. 원본을 고치지 않고
    배포 시점에 막는다 — 원본이 다시 생성되면 같은 문자가 또 들어온다.
    """
    fixed, n = BARE_AMP_RE.subn("&amp;", text)
    try:
        root = ElementTree.fromstring(fixed)
    except ElementTree.ParseError as e:
        sys.exit(f"KML 이 well-formed XML 이 아니다 ({label}): {e}")
    ns = {"k": "http://www.opengis.net/kml/2.2"}
    names = [e.text or "" for e in root.iterfind(".//k:Placemark/k:name", ns)]
    if not names:
        sys.exit(f"KML 에 Placemark 이름이 없다 ({label})")
    broken = [x for x in names if "�" in x]
    if broken:
        sys.exit(f"KML 핀 이름에 대체문자(U+FFFD) ({label}): {broken}")
    return fixed, names, n


def build_offline_maps():
    """Organic Maps 북마크 임포트 안내 + KML 직접 다운로드.

    자체 타일 지도는 구현하지 않는다 (용량·라이선스로 폐기된 안이다).
    Organic Maps 는 KML 북마크 임포트와 오프라인 턴바이턴 안내를 제공한다.
    """
    out_dir = SITE / "maps"
    kml_dir = out_dir / "kml"
    kml_dir.mkdir(parents=True, exist_ok=True)
    region_by_map = {c["map"]: c["region"] for c in CHAPTERS if c["kind"] == "region"}
    order = [out_name for _, out_name, _ in MAPS]

    files, total_pins, candidate_pins, repaired = [], 0, 0, 0
    for out_name in order:
        region = region_by_map[out_name]
        src = MAP_DIR / f"{region}_Execution_Map_v0.2.kml"
        if not src.exists():
            sys.exit(f"KML 없음: {src}")
        text, names, fixed_amps = sanitize_kml(
            src.read_text(encoding="utf-8"), src.name)
        pins = len(names)
        cand = sum(1 for n in names if "숙소 후보" in n)
        total_pins += pins
        candidate_pins += cand
        repaired += fixed_amps
        slug = out_name.replace(".html", "")
        dest = kml_dir / f"{slug}.kml"
        dest.write_text(text, encoding="utf-8")
        files.append((slug, region, pins, cand, dest.stat().st_size))

    rows = "".join(
        f"<tr><td>{html.escape(region)}</td>"
        f'<td><a href="kml/{slug}.kml" download>{slug}.kml</a></td>'
        f"<td>{pins}</td><td>{cand or ''}</td><td>{size:,} B</td></tr>"
        for slug, region, pins, cand, size in files)

    body = f"""<h1>오프라인 지도 — Organic Maps</h1>
<p class="meta">로밍이 끊겨도 도보·운전 안내가 되게 하는 준비다.
출발 전 Wi-Fi 에서 한 번만 해두면 된다.</p>

{net_note("KML 파일은 이미 이 기기에 있어 지금도 내려받을 수 있습니다.")}

<p class="offline-note"><b>실기기 검증 전이다.</b>
아래 KML {len(files)}개는 빌드가 매번 XML 파서로 열어보고 핀 {total_pins}개의 이름이
UTF-8 로 온전한지 확인한 파일이다. 한글·프랑스어 이름이 섞여 있다.
<b>다만 Organic Maps 가 기기에서 이 이름들을 어떻게 표시하는지는 확인하지 않았다.</b>
출발 전에 한 지역만 먼저 임포트해 핀 이름이 깨지지 않는지 직접 보고,
깨지면 나머지는 이 앱의 실행지도로 대신한다.</p>

<h2>준비 순서</h2>
<ol>
<li><b>Organic Maps 설치</b> — 무료·오픈소스·광고 없음. 계정이 필요 없다.</li>
<li><b>지도 내려받기 (Wi-Fi 에서)</b> — 앱에서 프랑스와 스페인 지도를 받는다.
  용량이 크므로 반드시 출발 전 Wi-Fi 에서 한다.</li>
<li><b>KML 내려받기</b> — 아래 표의 파일을 이 기기에 저장한다.</li>
<li><b>임포트</b> — 저장한 파일을 열어 <i>Organic Maps 로 열기</i> 를 고른다.
  핀이 앱의 북마크로 들어간다.</li>
<li><b>확인</b> — 앱의 북마크 목록에서 핀 {total_pins}개가 보이고
  이름이 깨지지 않았는지 본다.</li>
</ol>

<h2>지역별 북마크 파일</h2>
<div class="table-wrap"><table>
<thead><tr><th>지역</th><th>파일</th><th>핀</th><th>숙소 후보</th><th>크기</th></tr></thead>
<tbody>{rows}</tbody>
</table></div>
<p class="meta">전체 {total_pins}개 핀. 이동 순서대로 정렬돼 있다.</p>

<h2 id="후보-주의">숙소 핀은 확정이 아니다</h2>
<p class="offline-note"><b>{candidate_pins}개 핀이 <code>[숙소 후보]</code> 로 시작한다.
예약이 확정된 주소가 아니라 검토 중인 후보다.</b>
이 좌표를 목적지로 잡고 이동하면 안 된다. 확정 주소는
<a href="../tracker/accommodation.html">숙소 후보·확정</a> 과
<a href="../tracker/reservations.html">예약 현황</a> 에서 확인한다.
예약이 잠기면 KML 을 다시 만들어 이 페이지에 올린다.</p>

<h2>이 앱의 실행지도와 무엇이 다른가</h2>
<div class="table-wrap"><table>
<thead><tr><th></th><th>이 앱의 실행지도</th><th>Organic Maps</th></tr></thead>
<tbody>
<tr><td>핀 위치</td><td>오프라인 동작</td><td>오프라인 동작</td></tr>
<tr><td>배경 지도</td><td>연결 필요</td><td>오프라인 동작</td></tr>
<tr><td>길찾기</td><td>없음 (Google Maps 로 넘김)</td><td>도보·운전 턴바이턴 음성 안내</td></tr>
<tr><td>검색</td><td>이 가이드북 안에서만</td><td>주변 상점·주유소·화장실</td></tr>
</tbody>
</table></div>
<p>둘 다 쓴다. 계획과 설명은 이 가이드북에서 보고, 실제 길찾기는 Organic Maps 로 한다.</p>

<p class="offline-note">Organic Maps 는 이 가이드북과 무관한 별도 앱이다.
설치 링크를 여기에 걸지 않은 것은 스토어 주소가 바뀔 수 있어서다.
앱스토어·Play 스토어에서 <b>Organic Maps</b> 로 검색한다.</p>

<nav class="pager"><a href="index.html">← 실행지도 목록</a><span></span></nav>"""
    (out_dir / "offline.html").write_text(
        page("오프라인 지도", body, rel="..",
             back=crumbs_for(("지도", "maps/index.html"), ("오프라인", None))), encoding="utf-8")
    SEARCH_INDEX.append({"t": "오프라인 지도 — Organic Maps", "c": "실행지도",
                         "u": "maps/offline.html"})
    amp_note = f" · & 이스케이프 {repaired}건 교정" if repaired else ""
    print(f"  오프라인 지도: KML {len(files)}개 · 핀 {total_pins}개"
          f"(숙소 후보 {candidate_pins}){amp_note} → maps/offline.html")


POPUP_SRC = ("marker.bindPopup(`<b>${i+1}. ${p.name}</b><br>${p.category}"
             "<br><small>${p.status}</small><br>"
             '<a target="_blank" href="${p.url}">Google Maps 열기</a>`);')


def link_map_places(text, out_name):
    """지도 팝업에 장소 카드 링크를 넣는다.

    핀 이름은 레지스트리의 `지도 핀` 열과 대조한다. 장소가 URL 을 가진 이상
    지도에서도 같은 곳을 가리켜야 한다. 매칭이 없으면 팝업은 그대로 둔다.
    """
    slug = next((c["slug"] for c in CHAPTERS
                 if c["kind"] == "region" and c["map"] == out_name), None)
    if slug is None or POPUP_SRC not in text:
        return text
    table = {r["pin"]: r["slug"] for r in load_place_registry()
             if r["chapter"] == slug and r["type"] == "spot" and r["pin"]}
    if not table:
        return text
    repl = ("const PLACE=" + json.dumps(table, ensure_ascii=False) + ";\n "
            "marker.bindPopup(`<b>${i+1}. ${p.name}</b><br>${p.category}"
            "<br><small>${p.status}</small><br>"
            '<a target="_blank" href="${p.url}">Google Maps 열기</a>'
            "${PLACE[p.name]?`<br><a href=\"../places/${PLACE[p.name]}.html\">"
            "◇ 장소 카드</a>`:''}`);")
    return text.replace(POPUP_SRC, repl, 1)


def build_maps():
    out_dir = SITE / "maps"
    out_dir.mkdir(parents=True, exist_ok=True)
    shutil.copytree(ASSETS / "vendor" / "leaflet", out_dir / "vendor" / "leaflet",
                    dirs_exist_ok=True)
    data_dir = out_dir / "data"
    data_dir.mkdir(exist_ok=True)
    for f in MAP_DIR.glob("*.geojson"):
        shutil.copy(f, data_dir / f.name)
    for f in MAP_DIR.glob("*.kml"):
        text, _, _ = sanitize_kml(f.read_text(encoding="utf-8"), f.name)
        (data_dir / f.name).write_text(text, encoding="utf-8")
    cards = []
    total_points = 0
    for src_name, out_name, title in MAPS:
        geo_name = src_name.replace(".html", ".geojson")
        geo = json.loads((MAP_DIR / geo_name).read_text(encoding="utf-8"))
        features = geo.get("features", [])
        if not features:
            sys.exit(f"실행지도 GeoJSON 포인트 없음: {geo_name}")
        point_rows = []
        for i, feature in enumerate(features, 1):
            props = feature.get("properties", {})
            coords = feature.get("geometry", {}).get("coordinates", [])
            if len(coords) != 2 or not all(isinstance(v, (int, float)) for v in coords):
                sys.exit(f"실행지도 좌표 오류: {geo_name} #{i} {coords}")
            lon, lat = coords
            if not (-180 <= lon <= 180 and -90 <= lat <= 90):
                sys.exit(f"실행지도 좌표 범위 오류: {geo_name} #{i} {coords}")
            name = props.get("name") or props.get("Name")
            category = props.get("category") or props.get("Category")
            url = props.get("google_maps") or props.get("url") or props.get("URL")
            if not name or not category or not url:
                sys.exit(f"실행지도 필수속성 누락: {geo_name} #{i}")
            point_rows.append(
                f'<li><span><b>{i}. {html.escape(str(name))}</b>'
                f'<small>{html.escape(str(category))}</small></span>'
                f'<a target="_blank" rel="noopener" href="{html.escape(str(url), quote=True)}">길찾기</a></li>')
        total_points += len(features)
        text = (MAP_DIR / src_name).read_text(encoding="utf-8")
        text = text.replace(
            "https://unpkg.com/leaflet@1.9.4/dist/leaflet.css", "vendor/leaflet/leaflet.css")
        text = text.replace(
            "https://unpkg.com/leaflet@1.9.4/dist/leaflet.js", "vendor/leaflet/leaflet.js")
        back = ('<a href="../maps/index.html" style="position:absolute;z-index:1100;right:12px;top:12px;'
                'background:#1f4e78;color:#fff;padding:7px 12px;border-radius:8px;'
                'font-size:13px;text-decoration:none;box-shadow:0 1px 6px rgba(0,0,0,.3)">← 지도 목록</a>')
        text = text.replace('<div id="map"></div>', f'<div id="map"></div>{back}', 1)
        # 원본 지도는 데스크톱용 고정 패널이라 작은 화면에서 지도를 가린다.
        # 안내 접기와 텍스트 기준점 목록을 빌드 시 주입해 원본 자산은 보존한다.
        map_ui_css = """
<style id="phase6-map-ui">
.map-info-toggle{position:absolute;z-index:1200;left:12px;bottom:12px;border:0;border-radius:999px;background:#1f4e78;color:#fff;padding:10px 14px;font-weight:700;box-shadow:0 2px 10px #0005;min-height:44px}
.point-list{margin-top:9px;border-top:1px solid #d8dee5;padding-top:7px}.point-list summary{cursor:pointer;font-size:12px;font-weight:700}.point-list ol{max-height:32vh;overflow:auto;margin:7px 0 0;padding:0;list-style:none}.point-list li{display:flex;align-items:center;justify-content:space-between;gap:10px;padding:7px 0;border-top:1px solid #e7ebef;font-size:12px}.point-list small{display:block;color:#5d6670;margin-top:2px}.point-list a{color:#1f4e78;font-weight:700;white-space:nowrap}
body.map-info-hidden .panel{display:none}
@media(max-width:600px){.panel{left:max(8px,env(safe-area-inset-left));right:max(8px,env(safe-area-inset-right));top:max(58px,calc(env(safe-area-inset-top) + 50px));max-width:none;max-height:55vh;overflow:auto;padding:10px 12px}.panel h1{font-size:16px}.panel p{font-size:11px}.map-info-toggle{left:max(8px,env(safe-area-inset-left));bottom:max(12px,env(safe-area-inset-bottom))}}
</style>"""
        text = text.replace("</head>", map_ui_css + "</head>", 1)
        point_list = (f'<details class="point-list"><summary>기준점 {len(features)}개 목록</summary>'
                      f'<ol>{"".join(point_rows)}</ol></details>')
        text = text.replace('</div></div>\n<script src="vendor/leaflet/leaflet.js">',
                            f'</div>{point_list}</div>\n<script src="vendor/leaflet/leaflet.js">', 1)
        toggle = ('<button class="map-info-toggle" type="button" aria-expanded="true" '
                  'aria-label="지도 안내 접기">안내 접기</button>')
        text = text.replace('<div id="map"></div>', f'<div id="map"></div>{toggle}', 1)
        toggle_js = """<script>
const mapInfoButton=document.querySelector('.map-info-toggle');
mapInfoButton.addEventListener('click',()=>{
 const hidden=document.body.classList.toggle('map-info-hidden');
 mapInfoButton.textContent=hidden?'안내 보기':'안내 접기';
 mapInfoButton.setAttribute('aria-expanded',String(!hidden));
 mapInfoButton.setAttribute('aria-label',hidden?'지도 안내 보기':'지도 안내 접기');
});
</script>"""
        text = text.replace("</body>", toggle_js + "</body>", 1)
        text = link_map_places(text, out_name)
        # 실행지도는 원본 HTML 이라 페이지 셸을 거치지 않는다. 안전영역을 쓰려면
        # 여기서도 viewport-fit=cover 를 넣어야 한다.
        text = re.sub(r'(<meta name="viewport" content="(?![^"]*viewport-fit)[^"]*)"',
                      r'\1, viewport-fit=cover"', text, count=1)
        (out_dir / out_name).write_text(text, encoding="utf-8")
        day_range, area = MAP_META[out_name]
        cards.append(f'<a class="card card-alt" href="{out_name}">'
                     f'<span class="card-num">🗺️</span><span class="card-title">{title}</span>'
                     f'<span class="card-sub">{day_range} · {area} · 기준점 {len(features)}개</span></a>')
        SEARCH_INDEX.append({"t": title, "c": "실행지도", "u": f"maps/{out_name}"})
    print(f"  지도: {len(MAPS)}개 지역 · 기준점 {total_points}개 → maps/")

    body = ('<h1>실행지도</h1>'
            '<p class="meta">43일 일정에 연결된 8개 지역별 기준점 지도. 카드의 Day 범위를 확인하고, '
            '마커 또는 기준점 목록의 길찾기를 누르면 Google Maps가 열린다. '
            '배경 타일은 인터넷 연결 시 표시된다.</p>'
            '<div class="related"><a href="offline.html">📴 오프라인 지도 준비 — Organic Maps</a></div>'
            + net_note("핀 위치와 목록은 그대로 보입니다. 배경 지도와 Google Maps 링크만 연결이 필요합니다.")
            + f'<div class="grid">{"".join(cards)}</div>')
    (out_dir / "index.html").write_text(
        page("실행지도", body, rel="..", back=crumbs_for(("지도", None))), encoding="utf-8")


# ---------------------------------------------------------------- tracker

def format_cell(v):
    if v is None:
        return ""
    if hasattr(v, "hour"):  # datetime — 시각이 00:00이면 날짜만
        if (v.hour, v.minute, v.second) == (0, 0, 0):
            return v.strftime("%Y-%m-%d")
        return v.strftime("%Y-%m-%d %H:%M")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def sheet_to_table(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = [format_cell(v) for v in row]
        if any(c.strip() for c in cells):
            rows.append(cells)
    if not rows:
        return ""
    # 선두의 제목 행(비어 있지 않은 셀이 1개뿐)은 캡션으로 분리
    captions = []
    while rows and sum(1 for c in rows[0] if c.strip()) == 1:
        captions.append(next(c for c in rows[0] if c.strip()))
        rows.pop(0)
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    keep = [i for i in range(width) if any(r[i].strip() for r in rows)]
    rows = [[r[i] for i in keep] for r in rows]

    head = "".join(f"<th>{html.escape(c)}</th>" for c in rows[0])
    body_rows = []
    for r in rows[1:]:
        body_rows.append("<tr>" + "".join(f"<td>{html.escape(c)}</td>" for c in r) + "</tr>")
    caption_html = "".join(f'<p class="meta">{html.escape(c)}</p>' for c in captions)
    return (f'{caption_html}<div class="table-wrap"><table><thead><tr>{head}</tr></thead>'
            f'<tbody>{"".join(body_rows)}</tbody></table></div>')


def build_tracker():
    out_dir = SITE / "tracker"
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(TRACKER_XLSX, data_only=True)

    def tabs_of(active):
        links = []
        for _, s, label in TRACKER_SHEETS:
            cls = ' class="active"' if s == active else ""
            links.append(f'<a href="{s}.html"{cls}>{label}</a>')
        return '<nav class="tabs">' + "".join(links) + "</nav>"

    cards = []
    for sheet_name, slug, label in TRACKER_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  경고: 시트 없음 — {sheet_name}")
            continue
        table = sheet_to_table(wb[sheet_name])
        visual = (visual_figure("risk", "예약·운영 리스크 매트릭스")
                  if slug == "reservations" else "")
        body = f"<h1>{label}</h1>{tabs_of(slug)}{visual}{table}"
        (out_dir / f"{slug}.html").write_text(
            page(label, body, rel="..", topbar_title=label,
                 back=crumbs_for(("트래커", "tracker/index.html"), (label, None)),
                 meta_line="TP_Europe_Travel_Master_Tracker_v1.2.xlsx 기준"),
            encoding="utf-8")
        cards.append(f'<a class="card card-alt" href="{slug}.html">'
                     f'<span class="card-title">{label}</span>'
                     f'<span class="card-sub">{sheet_name}</span></a>')
        SEARCH_INDEX.append({"t": label, "c": "트래커", "u": f"tracker/{slug}.html"})
    print(f"  트래커: {len(cards)}개 시트 → tracker/")

    body = ('<h1>마스터 트래커</h1>'
            '<p class="meta">TP_Europe_Travel_Master_Tracker_v1.2.xlsx에서 변환</p>'
            f'<div class="grid">{"".join(cards)}</div>')
    (out_dir / "index.html").write_text(
        page("마스터 트래커", body, rel="..", topbar_title="트래커",
             back=crumbs_for(("트래커", None))),
        encoding="utf-8")



# ---------------------------------------------------------------- 주제 축

# 카테고리 슬러그는 분할 챕터와 같은 것을 쓴다 (schedule 은 일자 축이라 따로).
ALL_TOPIC_SLUG = dict(TOPIC_SLUG, schedule="days")
TOPIC_HEAD = dict(TOPIC_SUB, days="언제 무엇을 하는가")

STATUS_PAGES = [
    ("badge:pending", "reverify", "출발 전 확인",
     "재확인 표시가 붙은 항목을 여행 전체에서 모았다. 운영시간·요금·운행처럼 "
     "바뀌는 정보이거나 아직 교차 확인이 안 된 서술이다."),
    ("badge:p0", "p0", "P0 연결",
     "놓치면 다음 일정이 무너지는 연결이다. 항공·렌터카·고속철처럼 대체가 어려운 것들이다."),
    ("grade:essential", "essential", "필수 방문지",
     "각 지역에서 필수 등급이 붙은 장소다. 시간이 없을 때 마지막까지 남기는 것들이다."),
]


def topic_list_html(items, rel):
    """지역별로 묶어 섹션 목록을 낸다. 본문이 아니라 발췌와 앵커만 담는다."""
    out, cur = [], None
    for it in items:
        if it["region"] != cur:
            cur = it["region"]
            out.append(f'<h2 class="tp-region">{html.escape(cur)}</h2>')
        ex = f'<span class="tp-ex">{html.escape(it["excerpt"])}</span>' if it["excerpt"] else ""
        out.append(f'<a class="tp-item" href="{rel}/{it["url"]}">'
                   f'<span class="tp-title">{html.escape(it["title"])}</span>{ex}</a>')
    return "".join(out)


def topic_siblings(current):
    """주제 축의 형제 — 상태 3개와 분류 N개."""
    st = [(label, f"{slug}.html", slug == current, "")
          for _k, slug, label, _lead in STATUS_PAGES]
    cats = [("허브", "index.html", current == "index.html", "")]
    cats += [(label, f"{ALL_TOPIC_SLUG[key]}.html", ALL_TOPIC_SLUG[key] == current, "")
             for key, label in CATEGORIES if TOPIC_INDEX.get(label)]
    return siblings_nav([("분류", "sn-layers", cats), ("상태", "sn-days", st)])


def build_topics():
    """일자·지역과 나란히 서는 세 번째 축. 카테고리 10개 + 상태 3개."""
    out_dir = SITE / "topics"
    out_dir.mkdir(parents=True, exist_ok=True)
    rel = ".."
    order = [c["region"] for c in CHAPTERS if c["kind"] == "region"]
    rank = {r: i for i, r in enumerate(order)}

    cards = []
    for key, label in CATEGORIES:
        items = sorted(TOPIC_INDEX.get(label, []), key=lambda x: rank.get(x["region"], 99))
        if not items:
            continue
        slug = ALL_TOPIC_SLUG[key]
        icon = CAT_ICON.get(label)
        cls = f' class="ic ic-{icon}"' if icon else ""
        body = (f'<h1{cls}>{label}</h1>'
                f'<p class="meta">{TOPIC_HEAD[slug]} · 여행 전체 {len(items)}개 섹션</p>'
                + f'<div class="tp-list">{topic_list_html(items, rel)}</div>')
        (out_dir / f"{slug}.html").write_text(
            page(label, body, rel=rel, topbar_title=label, back=crumbs_for(("주제", "topics/index.html"), (label, None)),
                 subnav=topic_siblings(slug)), encoding="utf-8")
        cards.append(f'<a class="card" href="{slug}.html">'
                     f'<span class="card-title">{label}</span>'
                     f'<span class="card-sub">{TOPIC_HEAD[slug]}</span>'
                     f'<span class="card-n">{len(items)}개 섹션</span></a>')
        SEARCH_INDEX.append({"t": f"{label} — 전체 보기", "c": "주제",
                             "u": f"topics/{slug}.html"})

    st_cards = []
    for key, slug, label, lead in STATUS_PAGES:
        items = sorted(STATUS_INDEX.get(key, []), key=lambda x: rank.get(x["region"], 99))
        rows = "".join(
            f'<tr><td>{html.escape(it["region"])}</td>'
            f'<td><a href="{rel}/{it["url"]}">{html.escape(it["title"])}</a></td>'
            f'<td>{html.escape(it["cat"])}</td>'
            f'<td>{html.escape(it["note"])}</td></tr>' for it in items)
        body = (f'<h1>{label}</h1><p class="meta">{lead}</p>'
                + '<div class="table-wrap"><table><thead><tr>'
                  '<th>지역</th><th>섹션</th><th>분류</th><th>표시</th>'
                  f'</tr></thead><tbody>{rows}</tbody></table></div>')
        (out_dir / f"{slug}.html").write_text(
            page(label, body, rel=rel, topbar_title=label, back=crumbs_for(("주제", "topics/index.html"), (label, None)),
                 subnav=topic_siblings(slug)), encoding="utf-8")
        st_cards.append(f'<a class="card" href="{slug}.html">'
                        f'<span class="card-title">{label}</span>'
                        f'<span class="card-n">{len(items)}건</span></a>')
        SEARCH_INDEX.append({"t": label, "c": "주제", "u": f"topics/{slug}.html"})

    hub = (f'<h1 class="ic ic-topic">주제</h1>'
           '<p class="meta">같은 종류의 내용을 여행 전체에서 한 번에 본다. '
           '일자·지역과 나란히 서는 세 번째 축이다.</p>'
           + '<h2 class="ic ic-check">상태로 보기</h2>'
             '<p class="note">배지 표시를 여행 전체에서 모은 것이다. 출발 전 점검에 쓴다.</p>'
           + f'<div class="rg-grid">{"".join(st_cards)}</div>'
           + '<h2 class="ic ic-topic">분류로 보기</h2>'
           + f'<div class="rg-grid">{"".join(cards)}</div>')
    (SITE / "topics" / "index.html").write_text(
        page("주제", hub, rel=rel, topbar_title="주제", back=crumbs_for(("주제", None)),
             coords=coords_bar("..", day=("43일", "daily/index.html"),
                               region=("8곳", "regions.html"),
                               topic=("분류·상태", None)),
             subnav=topic_siblings("index")), encoding="utf-8")
    SEARCH_INDEX.append({"t": "주제 — 전체 목록", "c": "주제", "u": "topics/index.html"})
    print(f"  주제: 분류 {len(cards)}개 · 상태 {len(st_cards)}개 → topics/")



# ---------------------------------------------------------------- 장소 레지스트리

PLACE_REGISTRY = SOURCE / "ASSETS" / "91_Place_Registry_v1.0.md"
GRADE_KO2SLUG = {"필수": "essential", "우선 추천": "priority", "선택": "optional",
                 "대체": "alternative", "비추천": "excluded"}
GRADE_LABEL = {v: k for k, v in GRADE_KO2SLUG.items()}
TIMETABLE = {}        # 데일리 빌드가 채운다 — 장소 페이지가 Day 매핑에 쓴다
PLACES_BY_DAY = {}    # 장소 빌드가 채운다 — 데일리 카드가 그 날 장소를 싣는다


def wiki_ref(value):
    """레지스트리의 위키 칸을 (제목, 언어) 로 읽는다.

    `fr:Vieux-Nice` 처럼 앞에 언어를 붙일 수 있다. 기본은 영어판이다.
    """
    if not value:
        return {"wiki": None, "wlang": "en"}
    m = re.match(r"([a-z]{2}):(.+)$", value)
    return {"wiki": m.group(2).strip(), "wlang": m.group(1)} if m \
        else {"wiki": value, "wlang": "en"}


def load_place_registry():
    """장소 레지스트리를 읽는다. 지역 h2 아래의 표가 그 지역의 장소 목록이다."""
    if not PLACE_REGISTRY.exists():
        return []
    reg2slug = {c["region"].lower().split()[0]: c["slug"]
                for c in CHAPTERS if c["kind"] == "region"}
    reg2slug.update({c["name"]: c["slug"] for c in CHAPTERS if c["kind"] == "region"})
    out, slug = [], None
    for line in PLACE_REGISTRY.read_text(encoding="utf-8").splitlines():
        h = re.match(r"^## (\S+) \((\d\d)\)\s*$", line)
        if h:
            slug = h.group(2)
            continue
        if not slug or not line.startswith("|") or re.fullmatch(r"\|[\s:|-]+\|", line.strip()):
            continue
        c = [x.strip() for x in line.strip().strip("|").split("|")]
        if len(c) < 7 or c[0].startswith("슬러그"):
            continue
        dash = lambda v: None if v in ("—", "-", "") else v
        out.append({"slug": c[0].strip("`"), "name": c[1], "chapter": slug,
                    "type": c[2], "grade": GRADE_KO2SLUG.get(c[3].rstrip("*")),
                    "grade_raw": c[3], "pin": dash(c[4]),
                    "body": dash(c[5]), "head": dash(c[6]),
                    # 위키백과 문서 제목. 사진과 참고 링크의 근거다. 확인한 것만
                    # 채운다 — 틀린 제목은 남의 사진을 이 장소 것처럼 붙인다.
                    # `fr:제목` 이면 프랑스어판이다. 시장·채석장·구시가지는 영어판에
                    # 문서가 없고 프랑스어판에만 있는 경우가 많다.
                    **wiki_ref(dash(c[7]) if len(c) > 7 else None)})
    return out



def place_key(name):
    """장소 이름을 대조용 열쇠로 줄인다. 부제·괄호를 떼고 발음기호를 없앤다.

    앞 조각이 너무 짧으면 자르지 않는다 — `시장 — Place Richelme` 를 `시장` 으로
    줄이면 본문의 모든 '시장' 절이 그 한 곳으로 몰린다. 실제로 그렇게 됐다.
    """
    def norm(s):
        s = unicodedata.normalize("NFKD", s)
        s = "".join(c for c in s if not unicodedata.combining(c))
        return re.sub(r"[^a-z0-9가-힣]", "", s.lower())

    head = norm(re.split(r"[·—(]", name)[0])
    return head if len(head) >= 4 else norm(name)


def norm_key(s):
    s = "".join(c for c in unicodedata.normalize("NFKD", s.lower())
                if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9가-힣]", "", s)


def place_days(registry, timetable):
    """장소가 어느 Day 에 배정돼 있는지 찾는다.

    근거는 셋이고 우선순위가 있다. 전부 구조화된 기록이거나 저자의 명시다.
      1) 일자별 시간표 — 원고가 그 날 시각에 넣은 일정
      2) 43일 실행 감사표의 `핵심 실행`·`식사·휴식` — 그 날 하기로 한 것
         (`대체안`·`우선 삭제` 는 쓰지 않는다. 그 날 가는 것이 아니다)
      3) 본문 상세의 `Day N` 표기 — 보강본 저자의 서술

    셋 다 없으면 **비운다.** 산문에 이름이 나온다고 그 날 가는 것은 아니다.
    """
    audit_txt = {}
    try:
        for n, row in load_audit().items():
            audit_txt[n] = norm_key(f'{row["core"]} {row["meals"]}')
    except SystemExit:
        audit_txt = {}
    by_day = {}
    for key, blocks in timetable.items():
        n = day_no(date.fromisoformat(key))
        txt = " ".join(plain(" ".join(cells)) for _s, _r, rows in blocks for cells in rows)
        by_day[n] = norm_key(txt)
    out = {}
    for r in registry:
        if r["type"] != "spot":
            continue
        k = place_key(r["name"])
        if len(k) < 3:
            continue
        days = sorted(n for n, txt in by_day.items() if k in txt)
        if days:
            out[r["slug"]] = (days, "시간표")
            continue
        keys = {k} | ({norm_key(r["pin"])} if r["pin"] else set())
        days = sorted(n for n, txt in audit_txt.items()
                      if any(x and len(x) >= 3 and x in txt for x in keys))
        if days:
            out[r["slug"]] = (days, "실행 감사표")
            continue
        if r["body"] and r["head"]:
            f = SITE / r["body"]
            if f.exists():
                body = read_body_for_scan(f)
                m = re.search(r"<h[234][^>]*>\s*" + re.escape(r["head"][:24]), body)
                if m:
                    nxt = re.search(r"<h[234][^>]*>", body[m.end():])
                    seg = body[m.end(): m.end() + (nxt.start() if nxt else 2000)]
                    d = sorted({int(x) for x in re.findall(r"Day\s*(\d+)", plain(seg))})
                    d = [x for x in d if 1 <= x <= 43]
                    if d:
                        out[r["slug"]] = (d, "본문 표기")
    return out


def check_places():
    """레지스트리가 지도 핀·본문 등급 헤딩과 어긋나지 않는지 본다.

    같은 장소 목록이 저장소 세 곳에 따로 있다가 절반이 갈라졌다. 정본을 하나로
    모은 이상, 다시 갈라지면 빌드가 멈춰야 한다.
    """
    reg = load_place_registry()
    if not reg:
        print("장소 레지스트리 검사: 파일 없음 — 건너뜀")
        return
    problems = []
    seen = set()
    for r in reg:
        if r["slug"] in seen:
            problems.append(f'슬러그 중복: {r["slug"]}')
        seen.add(r["slug"])
        if r["type"] not in ("spot", "node"):
            problems.append(f'{r["slug"]}: 타입이 spot·node 가 아니다 ({r["type"]})')
        if r["pin"] and r["pin"] not in [p[0] for p in PLACES.get(r["chapter"], [])]:
            problems.append(f'{r["slug"]}: 지도 핀 "{r["pin"]}" 이 PLACES[{r["chapter"]}] 에 없다')
        if r["body"]:
            f = SITE / r["body"]
            if not f.exists():
                problems.append(f'{r["slug"]}: 본문 페이지 없음 {r["body"]}')
            elif r["head"] and r["head"] not in plain(
                    re.sub(r'<span class="grade.*?</span>', "",
                           f.read_text(encoding="utf-8"), flags=re.S)):
                problems.append(f'{r["slug"]}: 헤딩 "{r["head"]}" 을 {r["body"]} 에서 못 찾았다')

    # 본문에 등급 헤딩이 있는데 레지스트리에 없으면 고아다
    known = {r["head"] for r in reg if r["head"]}
    for c in CHAPTERS:
        if c["kind"] != "region":
            continue
        for f in sorted((SITE / "chapters" / c["name"]).glob("*.html")):
            body = f.read_text(encoding="utf-8")
            for lvl, b in re.findall(r"<h([234])[^>]*>(.*?)</h\1>", body, re.S):
                if 'class="grade grade-' not in b:
                    continue
                name = plain(re.sub(r'<span class="grade.*?</span>', "", b, flags=re.S))
                if name and name not in known and name not in REGISTRY_EXCLUDED:
                    problems.append(f'{c["name"]}/{f.name}: 등급 헤딩 "{name}" 이 레지스트리에 없다')
    if problems:
        print("장소 레지스트리 검사 실패:")
        for p in problems[:20]:
            print("  " + p)
        sys.exit(1)
    spots = [r for r in reg if r["type"] == "spot"]
    undecided = [r for r in spots if not r["grade"]]
    print(f"장소 레지스트리 검사: spot {len(spots)} · node {len(reg) - len(spots)}"
          f" · 등급 미정 {len(undecided)} 이상 없음")


# 등급이 붙어 있지만 갈 곳이 아닌 헤딩 — 하루의 성격이다
REGISTRY_EXCLUDED = {"15구 생활일", "월요일 모듈"}



# ---------------------------------------------------------------- 장소 페이지

def read_body_for_scan(path):
    """대조·발췌용으로 본문을 읽는다. 등급 span 과 빌드가 주입한 점프 링크를 걷는다."""
    s = path.read_text(encoding="utf-8")
    s = re.sub(r'<span class="grade.*?</span>', "", s, flags=re.S)
    return re.sub(r'<p class="day-jump">.*?</p>', "", s, flags=re.S)


def place_excerpt(reg_row):
    """본문 상세의 첫 문단을 발췌한다. 본문을 복제하지 않고 미리보기만 담는다."""
    if not (reg_row["body"] and reg_row["head"]):
        return ""
    f = SITE / reg_row["body"]
    if not f.exists():
        return ""
    body = read_body_for_scan(f)
    m = re.search(r"<h([234])[^>]*>\s*" + re.escape(reg_row["head"][:24]), body)
    if not m:
        return ""
    seg = section_extent(body, m)
    # 하위 헤딩이 먼저 오는 항목이 많다. 헤딩이 아니라 첫 문단을 가져온다.
    for para in re.findall(r"<p>(.*?)</p>", seg, re.S):
        txt = plain(para)
        # `Day 41 (10/8 목) — B안` 같은 표시줄은 설명이 아니다
        if len(txt) < 25 or re.match(r"^Day\s*\d", txt):
            continue
        return plain(para, 220)
    return ""


def section_extent(body, m):
    """헤딩 m 이 여는 섹션의 본문 — 같은 레벨 이상 헤딩을 만나면 끝난다."""
    lvl = int(m.group(1))
    tail = body[m.end():]
    stop = re.search(r"<h([1-%d])[^>]*>" % lvl, tail)
    return tail[: stop.start()] if stop else tail[:4000]


def place_anchor(reg_row):
    """본문 상세로 가는 앵커 URL. 헤딩 id 를 그대로 쓴다."""
    if not (reg_row["body"] and reg_row["head"]):
        return None
    f = SITE / reg_row["body"]
    if not f.exists():
        return None
    body = read_body_for_scan(f)
    m = re.search(r'<h[234] id="([^"]+)"[^>]*>\s*' + re.escape(reg_row["head"][:24]), body)
    return f'{reg_row["body"]}#{m.group(1)}' if m else reg_row["body"]


def build_places(timetable):
    """장소 페이지 — 가이드북의 최소 단위.

    언제(Day) · 어디(지역·지도) · 무엇(본문 발췌와 앵커) · 확인할 것을 한 장에 모은다.
    본문은 복제하지 않는다. 여기는 좌표이지 사본이 아니다.
    """
    reg = load_place_registry()
    if not reg:
        return
    out_dir = SITE / "places"
    out_dir.mkdir(parents=True, exist_ok=True)
    rel = ".."
    days_of = place_days(reg, timetable)
    map_links_all = load_map_links()
    by_ch = {c["slug"]: c for c in CHAPTERS if c["kind"] == "region"}
    order = [c["slug"] for c in CHAPTERS if c["kind"] == "region"]
    spots = [r for r in reg if r["type"] == "spot"]
    spots.sort(key=lambda r: (order.index(r["chapter"]), r["name"]))

    for r in spots:
        c = by_ch[r["chapter"]]
        days, basis = days_of.get(r["slug"], ([], ""))
        grade = (f'<span class="grade grade-{r["grade"]}">{GRADE_LABEL[r["grade"]]}</span>'
                 if r["grade"] else '<span class="badge badge-pending">등급 미정</span>')
        rows = [("지역", f'<a href="{rel}/{chapter_url(c)}">{html.escape(c["region"])}</a>'),
                ("등급", grade)]
        if days:
            chips = "".join(
                f'<a class="pl-day" href="{rel}/daily/day-{n:02d}.html">'
                f'Day {n} · {date_label(date_of_day(n))}</a>' for n in days)
            rows.append(("언제", chips + f'<span class="pl-basis">{basis} 기준</span>'))
        else:
            rows.append(("언제", '<span class="pl-basis">일정에 배정된 날을 '
                                 '시간표에서 찾지 못했다</span>'))
        anchor = place_anchor(r)
        rows.append(("상세", f'<a href="{rel}/{anchor}">챕터 본문에서 보기</a>' if anchor
                     else '<span class="pl-basis">아직 상세 서술이 없다</span>'))
        if r["pin"]:
            rows.append(("지도", f'<a href="{rel}/maps/{c["map"]}">{c["region"]} 실행지도</a>'
                                 f' · <span class="pl-basis">핀 이름 {html.escape(r["pin"])}</span>'))
        detail = ""
        if PLACE_BODY.get(r["slug"]):
            md_body = "\n".join(PLACE_BODY[r["slug"]]).strip()
            # 챕터에서 h4 였던 소제목을 카드에서는 h2 로 올린다
            md_body = re.sub(r"^#{3,6}[ \t]", "## ", md_body, flags=re.M)
            md_body = render_inline_tokens(annotate_tables(md_body)[0])
            html_body, _ = md_convert(md_body)
            detail = wrap_tables(rewrite_asset_links(html_body, rel))
        ex = "" if detail else place_excerpt(r)

        # 사진 — 위키백과에서 온라인일 때만 불러온다. 오프라인이면 자리가 접힌다.
        # 제목이 없으면 아예 걸지 않는다. 비슷한 이름으로 찍으면 남의 사진이
        # 이 장소 것처럼 붙는다 — 현장에서 엉뚱한 곳을 찾게 된다.
        photo = ""
        if r.get("wiki"):
            photo = (f'<figure class="pl-shot" data-wiki="{html.escape(r["wiki"], quote=True)}"'
                     f' data-wlang="{r["wlang"]}" hidden>'
                     f'<img alt="{html.escape(r["name"])}" loading="lazy">'
                     f'<figcaption><a class="pl-credit" target="_blank" rel="noopener"'
                     f' href="#">사진 · 위키백과</a></figcaption></figure>')

        # 참고 링크 — 근거가 있는 것만. 지도 좌표는 실행지도 핀에서, 위키 제목은
        # 레지스트리에서 온다. 둘 다 없으면 그 칩은 나오지 않는다.
        refs = []
        if r.get("wiki"):
            wurl = (f'https://{r["wlang"]}.wikipedia.org/wiki/'
                    + urllib.parse.quote(r["wiki"].replace(" ", "_")))
            refs.append(f'<a class="ref ic ic-link" target="_blank" rel="noopener"'
                        f' href="{html.escape(wurl)}">위키백과</a>')
        gm = map_links_all.get(r["pin"]) if r["pin"] else None
        if gm:
            refs.append(f'<a class="ref ic ic-pin" target="_blank" rel="noopener"'
                        f' href="{html.escape(gm)}">Google Maps</a>')
        refs.append(f'<a class="ref ic ic-map" href="{rel}/maps/{c["map"]}">'
                    f'{html.escape(c["region"])} 실행지도</a>')
        ref_block = (f'<h2 class="ic ic-link">참고</h2><div class="refs">{"".join(refs)}</div>'
                     + net_note("위키백과 링크와 사진은 연결되면 다시 동작합니다."))

        body = (f'<h1>{html.escape(r["name"])}</h1>'
                + photo
                + '<div class="table-wrap"><table><tbody>'
                + "".join(f'<tr><th>{k}</th><td>{v}</td></tr>' for k, v in rows)
                + '</tbody></table></div>'
                # 참고 링크는 본문 앞에 둔다. 길 찾는 손이 먼저 닿아야 하는 것은
                # 서술이 아니라 지도다. 긴 서술 뒤로 밀면 스크롤 끝에서 찾게 된다.
                + ref_block
                + (f'<p class="pl-ex">{html.escape(ex)}</p>' if ex else "")
                + detail
                + ('' if detail else
                   '<p class="note">상세 서술은 챕터 본문에 있다. 여기서 복제하지 않는다.</p>'))
        (out_dir / f'{r["slug"]}.html').write_text(
            page(r["name"], body, rel=rel, topbar_title=r["name"],
                 back=region_crumbs((c["region"], chapter_url(c)), (r["name"], None)),
                 country=country_of(c["name"]),
                 subnav=place_siblings(spots, r["slug"], by_ch)), encoding="utf-8")
        SEARCH_INDEX.append({"t": r["name"], "c": f'장소 · {c["region"]}',
                             "u": f'places/{r["slug"]}.html'})

    # 인덱스 — 지역 순, 등급 표시
    groups = []
    for s in order:
        mine = [r for r in spots if r["chapter"] == s]
        if not mine:
            continue
        cards = "".join(
            f'<a class="card" href="{r["slug"]}.html">'
            f'<span class="card-title">{html.escape(r["name"])}</span>'
            f'<span class="card-sub">'
            f'{GRADE_LABEL[r["grade"]] if r["grade"] else "등급 미정"}</span></a>'
            for r in mine)
        groups.append(f'<h2>{html.escape(by_ch[s]["region"])}</h2>'
                      f'<div class="grid">{cards}</div>')
    idx = (f'<h1>장소</h1><p class="meta">이 가이드북의 최소 단위다. '
           f'여행 전체 {len(spots)}곳 · 갈 수 있는 곳만 담는다. '
           f'역·공항은 지도와 일정에서만 다룬다.</p>'
           + "".join(groups))
    (out_dir / "index.html").write_text(
        page("장소", idx, rel=rel, topbar_title="장소", back=crumbs_for(("장소", None)),
             subnav=place_siblings(spots, "index", by_ch)), encoding="utf-8")
    SEARCH_INDEX.append({"t": "장소 — 전체 목록", "c": "장소", "u": "places/index.html"})
    for r in spots:
        for n in days_of.get(r["slug"], ([], ""))[0]:
            PLACES_BY_DAY.setdefault(n, []).append(
                {"slug": r["slug"], "name": r["name"], "grade": r["grade"]})
    print(f"  장소: {len(spots)}곳 → places/ (Day 확인 {len(days_of)}곳)")


def place_siblings(spots, current, by_ch):
    """장소 페이지의 형제 — 지역 묶음으로 옮겨 다닌다."""
    regions = []
    seen = set()
    for r in spots:
        if r["chapter"] in seen:
            continue
        seen.add(r["chapter"])
        first = next(x for x in spots if x["chapter"] == r["chapter"])
        regions.append((by_ch[r["chapter"]]["region"], f'{first["slug"]}.html',
                        False, ""))
    return siblings_nav([("장소", "sn-layers",
                          [("전체", "index.html", current == "index", "")] + regions)])


# ---------------------------------------------------------------- data.js

def build_data_js():
    CHAPTER_DATE_URL.update(DAY_OVERRIDES)
    today_map = {}
    d = TRIP_START
    while d <= TRIP_END:
        today_map[d.isoformat()] = f"daily/day-{day_no(d):02d}.html"
        d += timedelta(days=1)
    # 검색 결과는 독자 화면이다. 원고의 Markdown 강조·링크 문법을 노출하지 않는다.
    def clean_search_text(value):
        rendered = md_inline(str(value))
        return html.unescape(re.sub(r"<[^>]+>", "", rendered)).strip()

    clean_index = []
    seen = set()
    for entry in SEARCH_INDEX:
        item = dict(entry)
        item["t"] = clean_search_text(item.get("t", ""))
        item["c"] = clean_search_text(item.get("c", ""))
        key = (item["t"], item["c"], item.get("u", ""))
        if not item["t"] or key in seen:
            continue
        seen.add(key)
        clean_index.append(item)

    data = {
        "tripStart": TRIP_START.isoformat(),
        "tripEnd": TRIP_END.isoformat(),
        "today": today_map,
        "search": clean_index,
    }
    js = "window.GUIDE = " + json.dumps(data, ensure_ascii=False) + ";\n"
    (SITE / "assets" / "data.js").write_text(js, encoding="utf-8")
    print(f"  data.js: 날짜 매핑 {len(today_map)}일 · 검색 인덱스 {len(clean_index)}항목")


# ---------------------------------------------------------------- checks

CODE_BLOCK_RE = re.compile(r"<pre\b.*?</pre>|<code\b.*?</code>", re.S)


def check_visual_tokens():
    """산출물에 VISUAL 토큰이 남으면 빌드를 중단한다.

    검색 인덱스(`data.js`)까지 훑는다 — 헤딩 텍스트가 그대로 인덱싱되므로
    HTML 만 검사하면 검색 결과에 토큰이 뜨는 것을 놓친다.
    `<pre>`·`<code>` 안은 챕터 01이 토큰 문법을 예시로 보여주는 자리라 제외한다.
    """
    leftover = []
    for f in sorted(SITE.rglob("*")):
        if not f.is_file() or f.suffix not in (".html", ".js", ".json"):
            continue
        text = f.read_text(encoding="utf-8")
        if f.suffix == ".html":
            text = CODE_BLOCK_RE.sub("", text)
        for tok in ("{{VISUAL:", "{{badge:", "{{grade:"):
            n = text.count(tok)
            if n:
                leftover.append(f"{f.relative_to(SITE)}: {tok}... {n}개")
    empty = []
    for f in sorted(SITE.rglob("*.html")):
        n = len(re.findall(r"<h[1-6][^>]*>\s*</h[1-6]>", f.read_text(encoding="utf-8")))
        if n:
            empty.append(f"{f.relative_to(SITE)}: {n}개")
    if leftover or empty:
        if leftover:
            print("VISUAL 토큰 잔존:")
            for x in leftover:
                print("  " + x)
        if empty:
            print("빈 헤딩 잔존:")
            for x in empty:
                print("  " + x)
        sys.exit(1)
    print("VISUAL 토큰·빈 헤딩 검사: 이상 없음")


CHROME_RE = re.compile(r"<(script|style|nav|header|footer|aside)\b.*?</\1>", re.S)


def check_naming():
    """명명규칙 v1.0 위반을 잡는다. 번호가 아니라 지명으로 이름을 짓는다."""
    hits = []
    for f in sorted(SITE.rglob("*.html")):
        raw = f.read_text(encoding="utf-8")
        if 'http-equiv="refresh"' in raw:
            continue                      # 번호 URL 리다이렉트 스텁은 대상 아님
        text = CHROME_RE.sub("", raw)
        rel = f.relative_to(SITE)
        for m in re.finditer(r"<h[1-4][^>]*>\s*(?:Chapter\s*)?"
                             r"(\d+(?:\.\d+)+|\d+(?:\.\d+)*[.)])\s", text):
            hits.append(f"{rel}: 헤딩 번호 {m.group(1)}")
        for m in re.finditer(r"[①-⑳㉑-㊿]", text):
            hits.append(f"{rel}: 원문자 {m.group(0)}")
        for m in re.finditer(r'class="card-num">(\d+)<', raw):
            hits.append(f"{rel}: 카드 순번 {m.group(1)}")
        for m in re.finditer(r">(\d{2}) (?:Barcelona|Girona|Nice|Aix|Luberon|Avignon|Lyon|Paris)",
                             raw):
            hits.append(f"{rel}: 라벨 챕터번호 {m.group(1)}")
    if hits:
        print("명명규칙 위반:")
        for h in hits[:20]:
            print("  " + h)
        if len(hits) > 20:
            print(f"  … 외 {len(hits) - 20}건")
        sys.exit(1)
    print("명명규칙 검사: 헤딩번호·원문자·챕터번호 라벨 0건")


def check_day_sections():
    """Day 원고가 하나도 빠지지 않고 데일리 카드로 넘어왔는지 검사한다.

    Day 페이지가 챕터에도 있던 시절에는 렌더된 h2 를 셌다. 이제 하루의
    페이지는 하나뿐이라, 세어야 하는 것은 챕터 원고에서 걷어 온 Day 섹션이다.
    거점 전환일 7일은 출발지·도착지 원고가 둘 다 있어 50건이 된다.
    """
    problems, total = [], 0
    for n, blocks in sorted(DAY_MD.items()):
        total += len(blocks)
        for region, title, _md in blocks:
            m = re.search(r"(\d+)월\s*(\d+)일", title)
            if not m:
                problems.append(f"Day {n} ({region}): 제목에 날짜가 없다 — {title}")
                continue
            expected = day_no(date(TRIP_START.year, int(m.group(1)), int(m.group(2))))
            if n != expected:
                problems.append(f"Day {n} ({region}) · {m.group(1)}/{m.group(2)}"
                                f" → 전체 Day {expected}")
    missing = [n for n in range(1, 44) if n not in DAY_MD]
    if missing:
        problems.append(f"원고에서 Day 섹션을 못 찾은 날: {missing}")
    if total != 50:
        problems.append(f"Day 섹션 총수 {total} (전환일 7일 양쪽 포함 50이어야 함)")
    # 챕터 쪽에 Day 본문이 남아 있으면 사본이 되살아난 것이다
    for c in CHAPTERS:
        if c.get("name") not in SPLIT_CHAPTERS:
            continue
        for f in sorted((SITE / "chapters" / c["name"]).glob("day-*.html")):
            if "주소가 바뀌었습니다" not in f.read_text(encoding="utf-8"):
                problems.append(f"{f.relative_to(SITE)}: 리다이렉트가 아니라 본문이다")
    if problems:
        print("Day 섹션 검사 실패:")
        for p in problems:
            print("  " + p)
        sys.exit(1)
    print(f"Day 섹션 검사: 43일 전부 데일리 카드로 · 전환일 포함 {total}건 이상 없음")


def check_phase1_reader_guards():
    """Phase 1에서 고친 독자 화면 결함의 재발을 빌드 단계에서 막는다."""
    problems = []
    first = (SITE / "daily" / "day-01.html").read_text(encoding="utf-8")
    last = (SITE / "daily" / "day-43.html").read_text(encoding="utf-8")
    if 'href="day-00.html"' in first or re.search(r'<nav class="pager">[^<]*<a[^>]*>[^<]*Day 0', first):
        problems.append("Day 1에 활성 이전 링크가 있다")
    if 'href="day-44.html"' in last or re.search(r'<nav class="pager">.*Day 44', last, re.S):
        problems.append("Day 43에 활성 다음 링크가 있다")

    data = (SITE / "assets" / "data.js").read_text(encoding="utf-8")
    for pattern, label in ((r'"t":\s*"[^"]*\*\*', "검색 제목의 ** 강조"),
                           (r'"t":\s*"[^"]*`', "검색 제목의 ` 코드")):
        if re.search(pattern, data):
            problems.append(label + " 문법이 노출된다")

    nav = (SITE / "assets" / "nav.js").read_text(encoding="utf-8")
    if 'u || "chapters/03.html"' in nav:
        problems.append("오늘 버튼의 기간 밖 fallback이 폐기된 번호 URL이다")

    if problems:
        print("Phase 1 독자 화면 가드 실패:")
        for p in problems:
            print("  " + p)
        sys.exit(1)
    print("Phase 1 독자 화면 가드: 경계 탐색 · 검색 표기 · 오늘 fallback 이상 없음")


def check_phase3_navigation_guards():
    """여행 행동 중심의 전역 메뉴와 홈 진입점이 유지되는지 검사한다."""
    problems = []
    home = (SITE / "index.html").read_text(encoding="utf-8")
    tabs = re.findall(r'<a [^>]*data-tab="([^"]+)"[^>]*>.*?<span>([^<]+)</span></a>',
                      home, re.S)
    expected = [("today", "오늘"), ("itinerary", "일정"), ("regions", "지역"),
                ("prepare", "준비"), ("search", "검색")]
    if tabs != expected:
        problems.append(f"전역 메뉴 {tabs} (기대값 {expected})")
    for label in ("43일 일정", "지역별 가이드", "여행 준비", "통합 검색", "비상 · 오프라인"):
        if label not in home:
            problems.append(f"홈 주요 행동 누락: {label}")
    nav = (SITE / "assets" / "nav.js").read_text(encoding="utf-8")
    if 'd < G.tripStart' not in nav or '(G.today || {})[G.tripStart]' not in nav:
        problems.append("출발 전 오늘 버튼이 다음 여행일로 연결되지 않는다")
    if problems:
        print("Phase 3 내비게이션 가드 실패:")
        for p in problems:
            print("  " + p)
        sys.exit(1)
    print("Phase 3 내비게이션 가드: 5탭 · 홈 5행동 · 다음 여행일 이상 없음")


def check_phase4_daily_guards():
    """43일 실행 템플릿의 정보 순서와 보관 카드 접기를 잠근다."""
    problems = []
    for n in range(1, 44):
        path = SITE / "daily" / f"day-{n:02d}.html"
        text = path.read_text(encoding="utf-8")
        required = ('class="day-command"', 'class="day-quick"',
                    'class="day-actions"', '<h2 class="ic ic-clock">시간표</h2>',
                    '<details class="day-details day-card-archive">')
        for token in required:
            if token not in text:
                problems.append(f"Day {n}: {token} 누락")
        order = [text.find(token) for token in required]
        if any(x < 0 for x in order) or order != sorted(order):
            problems.append(f"Day {n}: 실행요약 → 시간표 → 카드 순서 훼손")
        if '<figure class="daily-card">' in text.split('<details class="day-details day-card-archive">', 1)[0]:
            problems.append(f"Day {n}: 카드 이미지가 첫 화면에 노출됨")
    index = (SITE / "daily" / "index.html").read_text(encoding="utf-8")
    if index.count('class="daily-region"') != 8:
        problems.append("일정 목록이 8개 지역 구간으로 묶이지 않음")
    if index.count('class="daily-item') != 43:
        problems.append("일정 목록이 43일이 아님")
    if index.count('class="di-core"') != 43:
        problems.append("일정 목록의 핵심 실행 요약이 43건이 아님")
    if problems:
        print("Phase 4 데일리 템플릿 가드 실패:")
        for p in problems[:30]:
            print("  " + p)
        sys.exit(1)
    print("Phase 4 데일리 템플릿 가드: 43일 공통 순서 · 8지역 목록 · 카드 접기 이상 없음")


def check_phase5_execution_guards():
    """43일 감사표의 날짜·거점·운영일 교정이 다시 갈라지지 않게 잠근다."""
    problems = []
    audit = load_audit()
    for n, row in audit.items():
        d = date_of_day(n)
        expected_date = f"{d.month}/{d.day} {WEEKDAY_KO[d.weekday()]}"
        if row["date"].replace("**", "") != expected_date:
            problems.append(f'Day {n}: 감사표 날짜 "{row["date"]}" (기대값 {expected_date})')
        chapter = chapter_for_date(d)
        if chapter and not row["base"].startswith(chapter["region"]):
            problems.append(
                f'Day {n}: 감사표 거점 "{row["base"]}" (기대값 {chapter["region"]})')

    # 2026-08-03 공식 운영 검증: Matisse·Chagall은 화요일 휴관이다.
    # Day 11에 두 곳을 다시 넣으면 감사표와 실제 데일리 페이지가 동시에 실패한다.
    day11 = (SITE / "daily" / "day-11.html").read_text(encoding="utf-8")
    forbidden = re.compile(r"Matisse Museum\s*(?:또는|or)\s*Chagall Museum", re.I)
    if forbidden.search(day11):
        problems.append("Day 11: 화요일 휴관인 Matisse·Chagall 선택안이 다시 노출됨")
    if "Musée de la Photographie Charles Nègre" not in day11:
        problems.append("Day 11: 화요일 운영 사진미술관 선택안 누락")
    if "화요일 휴관" not in day11:
        problems.append("Day 11: Matisse·Chagall 화요일 휴관 경고 누락")

    if problems:
        print("Phase 5 실행성 감사 가드 실패:")
        for p in problems[:30]:
            print("  " + p)
        sys.exit(1)
    print("Phase 5 실행성 감사 가드: 43일 날짜·거점 · Day 11 휴관 교정 이상 없음")


def check_phase6_map_guards():
    """8개 실행지도·65개 기준점·43일 라우팅을 하나의 계약으로 검사한다."""
    problems, total = [], 0
    map_index = (SITE / "maps" / "index.html").read_text(encoding="utf-8")
    for src_name, out_name, title in MAPS:
        geo_name = src_name.replace(".html", ".geojson")
        geo = json.loads((MAP_DIR / geo_name).read_text(encoding="utf-8"))
        count = len(geo.get("features", []))
        total += count
        day_range, area = MAP_META[out_name]
        if f"{day_range} · {area} · 기준점 {count}개" not in map_index:
            problems.append(f"지도 목록 메타데이터 누락: {out_name}")
        page_text = (SITE / "maps" / out_name).read_text(encoding="utf-8")
        for token in ('id="phase6-map-ui"', 'class="map-info-toggle"',
                      f'기준점 {count}개 목록', 'aria-expanded="true"'):
            if token not in page_text:
                problems.append(f"{out_name}: 모바일 지도 UI 누락 — {token}")
        if page_text.count('>길찾기</a>') != count:
            problems.append(f"{out_name}: 기준점 길찾기 {page_text.count('>길찾기</a>')}건 (기대 {count})")

    if total != 65:
        problems.append(f"GeoJSON 기준점 총 {total}개 (기대 65)")
    for n in range(1, 44):
        text = (SITE / "daily" / f"day-{n:02d}.html").read_text(encoding="utf-8")
        actual = set(re.findall(r'href="\.\./maps/([^"/]+\.html)"', text))
        actual.discard("offline.html")
        expected = {name for name, days in MAP_DAY_SPANS.items() if n in days}
        if actual != expected:
            problems.append(f"Day {n}: 지도 {sorted(actual)} (기대 {sorted(expected)})")
    if problems:
        print("Phase 6 실행지도 가드 실패:")
        for p in problems[:30]:
            print("  " + p)
        sys.exit(1)
    print("Phase 6 실행지도 가드: 8지역 · 65개 기준점 · 43일 라우팅 · 전환일 7일 이상 없음")


def check_phase7_visual_guards():
    """대표사진·편집도식·권리표시를 배포 계약으로 고정한다."""
    problems = []

    # 원본과 배포본이 모두 있어야 한다. 잘못 확장한 파일도 초기에 막는다.
    for slug, (fname, subject, author, lic, _lic_url, _src_url) in sorted(HERO_PHOTOS.items()):
        source = HERO_DIR / fname
        deployed = SITE / "assets" / "heroes" / f"{slug}.jpg"
        for label, path in (("원본", source), ("배포본", deployed)):
            if not path.exists() or path.stat().st_size < 10_000:
                problems.append(f"대표사진 {label} 누락·손상: {path}")
            elif path.read_bytes()[:2] != b"\xff\xd8":
                problems.append(f"대표사진 JPEG 형식 오류: {path}")

        chapter = next(c for c in CHAPTERS if c.get("slug") == slug)
        hub = SITE / chapter_url(chapter)
        page_text = hub.read_text(encoding="utf-8")
        for token in (f'assets/heroes/{slug}.jpg', f'alt="{html.escape(subject)}"',
                      html.escape(author), lic, "Wikimedia Commons", "크롭·리사이즈"):
            if token not in page_text:
                problems.append(f"{chapter['region']} 허브 대표사진·크레딧 누락: {token}")

    for key, fname in VISUALS.items():
        source = VISUALS_DIR / fname
        deployed = SITE / "assets" / "visuals" / fname
        for label, path in (("원본", source), ("배포본", deployed)):
            if not path.exists() or path.stat().st_size < 10_000:
                problems.append(f"편집도식 {label} 누락·손상: {path}")
            elif path.read_bytes()[:8] != b"\x89PNG\r\n\x1a\n":
                problems.append(f"편집도식 PNG 형식 오류: {path}")

    expected_visuals = {
        "chapters/how-to-use.html": ("rhythm",),
        "chapters/itinerary.html": ("route", "fatigue"),
        "tracker/reservations.html": ("risk",),
        "chapters/aix/transport.html": ("cardays",),
        "chapters/luberon/transport.html": ("cardays",),
        "chapters/avignon/transport.html": ("cardays",),
        "chapters/paris/index.html": ("cycles",),
    }
    for relpath, keys in expected_visuals.items():
        page_text = (SITE / relpath).read_text(encoding="utf-8")
        for key in keys:
            if VISUALS[key] not in page_text:
                problems.append(f"{relpath}: 편집도식 {key} 누락")

    credits = (SITE / "credits.html").read_text(encoding="utf-8")
    if credits.count('class="credit-item"') != len(HERO_PHOTOS):
        problems.append("사진 저작자 표시 페이지의 대표사진 항목 수 불일치")
    for _slug, (_fname, _subject, author, lic, lic_url, src_url) in HERO_PHOTOS.items():
        for token in (author, lic, lic_url, src_url):
            if token not in credits:
                problems.append(f"사진 저작자 표시 페이지 누락: {token}")

    leaked = []
    for path in SITE.rglob("*.html"):
        text = path.read_text(encoding="utf-8")
        if "../../ASSETS/88_Representative_Public_Photos/" in text or \
                "../../ASSETS/85_Editorial_Visuals/" in text:
            leaked.append(str(path.relative_to(SITE)))
    if leaked:
        problems.append("배포 HTML에 원본 상대경로 잔존: " + ", ".join(leaked[:10]))

    if problems:
        print("Phase 7 시각자산 가드 실패:")
        for problem in problems[:40]:
            print("  " + problem)
        sys.exit(1)
    print("Phase 7 시각자산 가드: 대표사진 8장 · 편집도식 6장 · 본문 배치 · CC 크레딧 이상 없음")


def check_phase8_operations_guards():
    """Known facts와 실제 예약값의 경계를 배포 계약으로 고정한다."""
    problems = []
    wb = openpyxl.load_workbook(TRACKER_XLSX, data_only=True)
    required_sheets = {name for name, _slug, _label in TRACKER_SHEETS}
    missing_sheets = sorted(required_sheets - set(wb.sheetnames))
    if missing_sheets:
        problems.append("필수 트래커 시트 누락: " + ", ".join(missing_sheets))

    def records(sheet_name):
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[3]]
        return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=4, values_only=True)
                if any(value is not None for value in row)]

    reservations = records("Reservations")
    ids = [row.get("ID") for row in reservations]
    if len(reservations) != 24 or len(set(ids)) != 24 or ids != [f"R{i:03d}" for i in range(1, 25)]:
        problems.append("예약항목은 R001~R024의 고유한 24건이어야 함")
    if sum(row.get("우선순위") == "P0" for row in reservations) != 13:
        problems.append("P0 예약항목은 정확히 13건이어야 함")
    allowed_reservation_states = {"미조사", "예약대기", "재확인", "예약완료", "취소"}
    for row in reservations:
        state = row.get("상태")
        if state not in allowed_reservation_states:
            problems.append(f'{row.get("ID")}: 알 수 없는 예약상태 {state!r}')
        if state == "예약완료":
            required = ("예약번호", "사업자", "소스 URL", "최종확인일")
            absent = [key for key in required if not row.get(key)]
            if absent:
                problems.append(f'{row.get("ID")}: 예약완료인데 필수값 누락 — {", ".join(absent)}')

    expected_stays = {
        "Barcelona": ("2026-08-29", "2026-09-01", 3),
        "Girona": ("2026-09-01", "2026-09-04", 3),
        "Nice": ("2026-09-04", "2026-09-09", 5),
        "Aix-en-Provence": ("2026-09-09", "2026-09-13", 4),
        "Luberon": ("2026-09-13", "2026-09-17", 4),
        "Avignon": ("2026-09-17", "2026-09-21", 4),
        "Lyon": ("2026-09-21", "2026-09-25", 4),
        "Paris": ("2026-09-25", "2026-10-10", 15),
    }
    stays = records("Accommodation")
    if {row.get("거점") for row in stays} != set(expected_stays):
        problems.append("숙소 거점은 Barcelona~Paris의 확정 8개여야 함")
    for row in stays:
        base = row.get("거점")
        if base not in expected_stays:
            continue
        checkin, checkout, nights = expected_stays[base]
        actual = (row.get("체크인").date().isoformat() if row.get("체크인") else None,
                  row.get("체크아웃").date().isoformat() if row.get("체크아웃") else None,
                  row.get("박수"))
        if actual != (checkin, checkout, nights):
            problems.append(f"{base}: 숙박배분 {actual} (기대 {(checkin, checkout, nights)})")
        if row.get("상태") == "예약완료":
            required = ("실제총액", "예약번호", "주소", "소스 URL")
            absent = [key for key in required if not row.get(key)]
            if absent:
                problems.append(f'{base}: 숙소 예약완료인데 필수값 누락 — {", ".join(absent)}')

    locks = records("Phase8 Lock Status")
    allowed_lock_states = {"LOCKED", "PARTIAL", "BLOCKED"}
    for row in locks:
        state = row.get("잠금상태")
        if state not in allowed_lock_states:
            problems.append(f'{row.get("항목")}: 알 수 없는 잠금상태 {state!r}')
        if state != "LOCKED" and not row.get("필요 입력"):
            problems.append(f'{row.get("항목")}: 미잠금 항목의 필요 입력 누락')
    lock_by_item = {row.get("항목"): row for row in locks}
    for item, value in (("43일·42박", "2026-08-29~2026-10-10"),
                        ("8개 거점 숙박배분", "3/3/5/4/4/4/4/15박")):
        row = lock_by_item.get(item, {})
        if row.get("잠금상태") != "LOCKED" or row.get("현재 확정값") != value:
            problems.append(f"Known-Facts Lock 불일치: {item}")

    # 사용자 예약서가 없는 현재 기준에서는 잠금률 0이 정직한 값이다.
    completed = sum(row.get("상태") == "예약완료" for row in reservations)
    dashboard_text = (SITE / "tracker" / "dashboard.html").read_text(encoding="utf-8")
    lock_text = (SITE / "tracker" / "locks.html").read_text(encoding="utf-8")
    if completed == 0 and "실제 예약 잠금률</td><td>0" not in dashboard_text:
        problems.append("Dashboard 실제 예약 잠금률이 예약 데이터와 불일치")
    for token in ("43일·42박", "3/3/5/4/4/4/4/15박", "BLOCKED", "필요 입력"):
        if token not in lock_text:
            problems.append(f"잠금 현황 배포 페이지 누락: {token}")

    if problems:
        print("Phase 8 예약·운영 잠금 가드 실패:")
        for problem in problems[:40]:
            print("  " + problem)
        sys.exit(1)
    print("Phase 8 예약·운영 잠금 가드: 24개 예약 · P0 13개 · 8개 숙소 · Known-Facts/실예약 경계 이상 없음")


def check_phase9_commercial_depth_guards():
    """상용 편집모듈과 51개 장소 dossier가 원고·배포본에서 빠지지 않게 잠근다."""
    problems = []
    required_files = (COMMERCIAL_CARDS, PLACE_DOSSIERS, COMMERCIAL_STANDARD, DOSSIER_STANDARD)
    for path in required_files:
        if not path.exists() or path.stat().st_size < 500:
            problems.append(f"Phase 9 기준파일 누락·손상: {path.relative_to(SOURCE)}")

    cards_text = COMMERCIAL_CARDS.read_text(encoding="utf-8")
    dossier_text = PLACE_DOSSIERS.read_text(encoding="utf-8")
    expected_regions = {
        "barcelona": ("Barcelona", 6), "girona": ("Girona", 7),
        "nice": ("Nice", 6), "aix": ("Aix", 6),
        "luberon": ("Luberon", 6), "avignon": ("Avignon", 7),
        "lyon": ("Lyon", 5), "paris": ("Paris", 8),
    }

    if len(re.findall(r"^## .+$", cards_text, re.M)) != 8:
        problems.append("Commercial City Experience Card는 정확히 8개 지역이어야 함")
    dossier_count = len(re.findall(r"^## .+$", dossier_text, re.M))
    if dossier_count != 51:
        problems.append(f"장소 dossier {dossier_count}개 (기대 51개)")

    # 최신 Regional Chapter만 검사한다. superseded 원고가 통과 근거가 되어서는 안 된다.
    for chapter in (c for c in CHAPTERS if c["kind"] == "region"):
        slug = chapter["name"]
        region_heading, expected_places = expected_regions[slug]
        source_path = SOURCE / chapter["path"]
        source_text = source_path.read_text(encoding="utf-8")
        for token in ("# Commercial Guide Module", "## Editor’s Verdict",
                      "## 놓치면 아쉬운 선택", "## 하루를 완성하는 네 가지 선택",
                      "## 이 지역의 Top 10", "## 현장 메모",
                      "# Regional Context & Scheduled Place Dossiers"):
            if token not in source_text:
                problems.append(f"{source_path.name}: Phase 9 모듈 누락 — {token}")
        if not re.search(r"^## (?:지역을 이해하는 다섯 개의 층|이 (?:도시를|지역을) 이해하는 축)",
                         source_text, re.M):
            problems.append(f"{source_path.name}: 역사·경제·사회·문화 지역맥락 축 누락")

        region_match = re.search(
            rf"^# {re.escape(region_heading)}\s*$([\s\S]*?)(?=^# [^#]|\Z)",
            dossier_text, re.M)
        actual_places = len(re.findall(r"^## .+$", region_match.group(1), re.M)) if region_match else 0
        if actual_places != expected_places:
            problems.append(f"{region_heading}: dossier {actual_places}개 (기대 {expected_places}개)")

        deployed = "\n".join(
            path.read_text(encoding="utf-8")
            for path in sorted((SITE / "chapters" / slug).glob("*.html")))
        for token in ("Editor’s Verdict", "놓치면 아쉬운 선택", "이 지역의 Top 10"):
            if token not in deployed:
                problems.append(f"{slug} 배포 챕터: Phase 9 콘텐츠 누락 — {token}")
        if not ("지역을 이해하는 다섯 개의 층" in deployed or
                "이 도시를 이해하는 축" in deployed or "이 지역을 이해하는 축" in deployed):
            problems.append(f"{slug} 배포 챕터: 역사·경제·사회·문화 지역맥락 축 누락")

    # 모든 dossier는 현장에서 확인할 수 있는 공식 출발점을 가져야 한다.
    official_links = re.findall(r"^- 공식정보:\s+https?://\S+", dossier_text, re.M)
    if len(official_links) != 51:
        problems.append(f"dossier 공식정보 링크 {len(official_links)}개 (기대 51개)")

    if problems:
        print("Phase 9 상용편집·장소심화 가드 실패:")
        for problem in problems[:40]:
            print("  " + problem)
        sys.exit(1)
    print("Phase 9 상용편집·장소심화 가드: 8개 지역 카드 · 최신 챕터 8개 · 장소 dossier 51개 · 공식링크 51개 이상 없음")


def check_phase10_official_fact_guards():
    """공식 사실 레지스터·재확인 달력·배포 노출을 하나의 계약으로 잠근다."""
    problems = []
    report_path = SOURCE / "OPERATIONS" / "118_Phase10_Final_Fact_Verification_Report_v1.0.md"
    required_files = (VERIFY_MD, REVERIFY_MD, GATE_MD, report_path)
    for path in required_files:
        if not path.exists() or path.stat().st_size < 300:
            problems.append(f"Phase 10 기준파일 누락·손상: {path.relative_to(SOURCE)}")

    verify_text = VERIFY_MD.read_text(encoding="utf-8")
    rows = md_table_rows(verify_text,
                         ["ID", "지역", "장소", "상태", "확인내용", "공식출처", "조치"])
    ids = [row.get("ID") for row in rows]
    if ids != [f"F{i:03d}" for i in range(1, 19)]:
        problems.append("공식 검증 레코드는 F001~F018의 연속된 18건이어야 함")

    allowed_states = set(VERIFY_STATUS)
    state_counts = {state: 0 for state in allowed_states}
    official_urls = []
    for row in rows:
        state = row.get("상태")
        if state not in allowed_states:
            problems.append(f'{row.get("ID")}: 알 수 없는 검증상태 {state!r}')
        else:
            state_counts[state] += 1
        for key in ("지역", "장소", "항목", "확인내용", "공식출처", "조치"):
            if not row.get(key):
                problems.append(f'{row.get("ID")}: 필수값 누락 — {key}')
        match = re.fullmatch(r"\[공식\]\((https://[^\s)]+)\)", row.get("공식출처", ""))
        if not match:
            problems.append(f'{row.get("ID")}: HTTPS 공식출처 형식 오류')
        else:
            official_urls.append(match.group(1))

    expected_counts = {
        "VERIFIED": 13, "CORRECTED": 3, "VERIFIED_CURRENT": 1,
        "CONFLICT_RECHECK": 1, "DATE_GATE": 0,
    }
    if state_counts != expected_counts:
        problems.append(f"검증상태 집계 {state_counts} (기대 {expected_counts})")
    if len(official_urls) != 18:
        problems.append(f"공식출처 URL {len(official_urls)}개 (기대 18개)")

    # 검증일은 명시되어야 하고 여행 시작일보다 뒤일 수 없다.
    verified_date = re.search(r"\*\*검증일:\*\*\s*(\d{4}-\d{2}-\d{2})", verify_text)
    if not verified_date:
        problems.append("공식 검증일 누락")
    else:
        try:
            if date.fromisoformat(verified_date.group(1)) > TRIP_START:
                problems.append("공식 검증일이 여행 시작일보다 늦음")
        except ValueError:
            problems.append("공식 검증일 형식 오류")

    gate_text = GATE_MD.read_text(encoding="utf-8")
    gate_rows = md_table_rows(gate_text, ["날짜", "확인"])
    trip_gate_rows = []
    for row in gate_rows:
        dates = [date(TRIP_START.year, int(m[1]), int(m[2]))
                 for m in re.finditer(r"(\d{1,2})/(\d{1,2})", row.get("날짜", ""))]
        if dates and any(TRIP_START <= d <= TRIP_END for d in dates):
            trip_gate_rows.append(row)
        if not row.get("확인"):
            problems.append(f'재확인 게이트 내용 누락: {row.get("날짜")}')
    if len(trip_gate_rows) != 17:
        problems.append(f"여행 중 날짜별 핵심 게이트 {len(trip_gate_rows)}개 (기대 17개)")

    report_text = report_path.read_text(encoding="utf-8")
    for token in ("| 공식 검증·정정 레코드 | 18 |", "| 명시적 정정 | 3 |",
                  "| 공식페이지 충돌·재확인 | 1 |", "| 전날·당일 게이트 | 17개 날짜군 |",
                  "| 임의 생성한 예약값 | 0 |", "Phase 8B"):
        if token not in report_text:
            problems.append(f"Phase 10 완료보고서 집계·경계 누락: {token}")

    # 8개 지역 모두 웹에서 공식 확인 블록과 상태표를 노출해야 한다.
    for chapter in (c for c in CHAPTERS if c["kind"] == "region"):
        pages = "\n".join(path.read_text(encoding="utf-8")
                          for path in sorted((SITE / "chapters" / chapter["name"]).glob("*.html")))
        for token in ("공식 확인 정보와 재확인 대상", "검증 상태", "Phase 8B 예약 완료 전"):
            if token not in pages:
                problems.append(f'{chapter["name"]} 배포 챕터: Phase 10 블록 누락 — {token}')
        expected_rows = len(load_verification().get(chapter["slug"], []))
        if pages.count('>공식</a>') < expected_rows:
            problems.append(f'{chapter["name"]} 배포 챕터: 공식출처 링크 수 부족')

    # 미확정 상태가 확정처럼 보이지 않도록 배지와 문구를 보존한다.
    deployed_all = "\n".join(path.read_text(encoding="utf-8")
                             for path in SITE.rglob("*.html"))
    for token in ("출처 간 불일치", "현재 기준 확인", "재확인", "당일 확정"):
        if token not in deployed_all:
            problems.append(f"배포본 불확실성 표기 누락: {token}")

    if problems:
        print("Phase 10 공식정보 최종 검증 가드 실패:")
        for problem in problems[:50]:
            print("  " + problem)
        sys.exit(1)
    print("Phase 10 공식정보 가드: 18개 공식검증 · 정정 3건 · 충돌 1건 · 날짜게이트 17개 · 8지역 배포 이상 없음")


def check_links():
    broken = []
    for f in SITE.rglob("*.html"):
        text = f.read_text(encoding="utf-8")
        for attr in ("href", "src"):
            for target in re.findall(rf'{attr}="([^"]+)"', text):
                if target.startswith(("http", "#", "mailto:", "data:")) or "${" in target:
                    continue
                path = (f.parent / target.split("#")[0]).resolve()
                if not path.exists():
                    broken.append(f"{f.relative_to(SITE)} → {target}")
    if broken:
        print("깨진 링크:")
        for b in broken:
            print("  " + b)
        sys.exit(1)
    print("링크 검사: 이상 없음")


def check_dates():
    d = TRIP_START
    missing = []
    while d <= TRIP_END:
        if d.isoformat() not in CHAPTER_DATE_URL:
            missing.append(d.isoformat())
        d += timedelta(days=1)
    if missing:
        print("챕터 날짜 매핑 누락:", ", ".join(missing))
        sys.exit(1)
    print(f"날짜 매핑 검사: {TRIP_START} ~ {TRIP_END} 전체 43일 이상 없음")


# ---------------------------------------------------------------- main

def main():
    if SITE.exists():
        shutil.rmtree(SITE)
    SITE.mkdir()
    (SITE / "assets").mkdir()
    # 아이콘은 CSS 마스크로 붙인다. 스프라이트를 페이지마다 인라인하면
    # 314쪽이 각각 무거워진다 — 마스크는 CSS 한 번이고 페이지 무게는 0 이다.
    (SITE / "assets" / "style.css").write_text(
        (ASSETS / "style.css").read_text(encoding="utf-8") + "\n" + icons.css(),
        encoding="utf-8")
    shutil.copy(ASSETS / "nav.js", SITE / "assets" / "nav.js")
    # 나눔고딕 woff2 — CDN 을 쓰지 않고 번들한다 (OFL 1.1)
    shutil.copytree(ASSETS / "vendor" / "nanum", SITE / "assets" / "vendor" / "nanum",
                    dirs_exist_ok=True)
    (SITE / "assets" / "heroes").mkdir()
    for slug, (fname, *_) in HERO_PHOTOS.items():
        shutil.copy(HERO_DIR / fname, SITE / "assets" / "heroes" / f"{slug}.jpg")
    (SITE / "assets" / "visuals").mkdir()
    for fname in VISUALS.values():
        shutil.copy(VISUALS_DIR / fname, SITE / "assets" / "visuals" / fname)

    print("챕터 빌드:")
    build_chapters()
    print("장소 축 빌드:")
    TIMETABLE.update(load_day_details()[1])
    build_places(TIMETABLE)
    print("데일리 카드 빌드:")
    build_daily()
    build_home()
    print("지도 빌드:")
    build_maps()
    build_offline_maps()
    build_regions()
    print("라이선스 빌드:")
    build_credits()
    print("주제 축 빌드:")
    build_topics()
    print("트래커 빌드:")
    build_tracker()
    graph = content_model.build_graph(
        chapters=CHAPTERS, trip_start=TRIP_START, trip_end=TRIP_END,
        places=load_place_registry(), place_days=place_days(load_place_registry(), TIMETABLE),
        tracker_path=TRACKER_XLSX)
    graph_errors = content_model.validate_graph(graph)
    if graph_errors:
        print("콘텐츠 모델 검사 실패:")
        for problem in graph_errors[:20]:
            print("  " + problem)
        sys.exit(1)
    content_model.write_graph(graph, SITE / "assets" / "content-model.json")
    print("콘텐츠 모델: " + " · ".join(
        f'{key} {len(graph[key])}' for key in
        ("regions", "days", "places", "stays", "reservations", "transports")))
    build_data_js()
    check_visual_tokens()
    check_naming()
    check_day_sections()
    check_phase1_reader_guards()
    check_phase3_navigation_guards()
    check_phase4_daily_guards()
    check_phase5_execution_guards()
    check_phase6_map_guards()
    check_phase7_visual_guards()
    check_phase8_operations_guards()
    check_phase9_commercial_depth_guards()
    check_phase10_official_fact_guards()
    check_links()
    check_dates()
    check_places()
    print(f"\n완료: {SITE} ({sum(1 for _ in SITE.rglob('*.html'))}개 HTML 페이지)")


if __name__ == "__main__":
    main()
