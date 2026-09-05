#!/usr/bin/env python3
"""페이지 렌더러 — 콘텐츠 모델에서 사이트를 만든다.

여기 있는 함수는 model.py 의 엔티티만 읽는다. 마크다운을 정규식으로 긁지
않고, 빌드 산출물을 다시 파싱하지도 않는다. 그 두 가지가 이전 파이프라인이
"문서를 웹으로 변환한 것"처럼 보였던 이유였다.

축은 목록을 맡고 본문을 갖지 않는다.

    Day    = 실행의 정본. 시간표·이동·예약·Plan B 가 여기에만 있다.
    Region = 탐색과 이해. Day 시간표를 복제하지 않는다.
    Place  = 장문의 정본. Region 과 Day 는 참조만 한다.
"""
from __future__ import annotations

import hashlib
import html as html_lib
import json
import os
import re
import shutil
import sys
from urllib.parse import quote
from datetime import date
from pathlib import Path

import markdown as md_lib

import md_tidy

import icons
import model
from model import Day, Place, Region, Stop, Trip
from shell import (GRADE_BADGE, SITE_TITLE, alert, badge, esc, ic, page,
                   redirect, sec_head, tabs_strip)

ROOT = Path(__file__).resolve().parent.parent
# 출력 경로. SPFR_SITE_DIR 로 바꿀 수 있다 — 같은 워크트리에서 다른 빌드가
# 동시에 돌 때 서로의 산출물을 지우지 않게 하기 위한 것이다.
SITE = Path(os.environ.get("SPFR_SITE_DIR") or (ROOT / "site"))
ASSETS = ROOT / "build" / "assets"
IMAGE_MANIFEST = ROOT / "data" / "images" / "image-manifest.json"
TRACKER_XLSX = ROOT / "source" / "OPERATIONS" / "TP_Europe_Travel_Master_Tracker_v1.2.xlsx"

SEARCH_INDEX: list[dict] = []

# 지도 키는 환경에서 온다 (CI 시크릿). 없으면 지도 없이 목록만 남는다 —
# 로컬 빌드가 키 때문에 막히지 않게 한다.
MAPS_KEY = os.environ.get("GOOGLE_MAPS_API_KEY", "").strip()

# 숙소를 이름으로 여는 검색어. '바스카라의 B&B' 같은 우리끼리 쓰는 별칭은
# 구글맵에 없다 — 정식 상호(Casa Bascara)로 열어야 찾아진다.
HOTEL_QUERIES: dict[str, str] = {}
MAPS_ID = os.environ.get("GOOGLE_MAPS_MAP_ID", "").strip()
IMAGES: dict = {}

# 어휘표. daily-cards 가 쓰는 값을 화면 말로 옮긴다.
#
# 여기 없는 값이 들어오면 빌드를 세운다 (check_vocabulary). 조용히 넘기면
# 한국어 화면에 영어 코드가 그대로 새고, 더 나쁘게는 'unconfirmed' 같은
# 미확정 표시가 확정처럼 보인다.
CAT_ICON = {
    "culture": "book", "sight": "pin", "food": "food", "cafe": "food",
    "hotel": "stay", "transport": "train", "activity": "gauge",
    "shopping": "pin",
}
MODE_ICON = {
    "walk": "pin", "metro": "train", "tram": "train", "funicular": "train", "bus": "train",
    "train": "train", "drive": "car", "car": "car", "flight": "plane",
    "taxi": "car", "unconfirmed": "alert",
}
MODE_LABEL = {
    "walk": "도보", "metro": "지하철", "tram": "트램", "funicular": "푸니쿨라", "bus": "버스",
    "train": "기차", "drive": "운전", "car": "운전", "flight": "비행",
    "taxi": "택시",
    # 이동수단이 아직 안 정해진 구간. 확정처럼 보이면 안 된다.
    "unconfirmed": "이동수단 미정",
}
FACT_LABEL = {
    "hours": "운영시간", "closed": "휴무", "price_adult": "요금",
    "price_range": "가격대", "booking": "예약", "getting_there": "가는 법",
    "duration": "소요시간", "address": "주소", "phone": "전화", "note": "메모",
    "parking": "주차",
}

EXECUTION_STATUS_UI = {
    "confirmed": ("ok", "CONFIRMED", "check"),
    "book": ("must", "BOOK", "ticket"),
    "ticket": ("caution", "TICKET", "ticket"),
    "check": ("caution", "CHECK", "alert"),
    "caution": ("alert", "CAUTION", "alert"),
    "optional": ("neutral", "OPTIONAL", "gauge"),
    "unavailable": ("alert", "UNAVAILABLE", "alert"),
}

DAY_TYPE_LABEL = {
    "city": ("neutral", "CITY DAY"),
    "transfer": ("caution", "TRANSFER DAY"),
    "driving": ("caution", "DRIVING DAY"),
    "living": ("neutral", "LIVING DAY"),
}


SEP_ROW = re.compile(r"^\s*\|?[\s:|-]*-{2,}[\s:|-]*\|?\s*$")


def _cells(line: str) -> list[str]:
    row = line.strip()
    if row.startswith("|"):
        row = row[1:]
    if row.endswith("|"):
        row = row[:-1]
    return [c.strip() for c in row.split("|")]


def _inline(text: str) -> str:
    """셀 안의 굵게·링크만 변환한다. 문단 태그는 벗긴다."""
    out = md_lib.markdown(text, extensions=["attr_list"]).strip()
    if out.startswith("<p>") and out.endswith("</p>"):
        out = out[3:-4]
    return out


def plain_inline(text: str) -> str:
    """마크다운이 허용되지 않는 UI 필드를 안전한 일반 문자열로 정규화한다.

    Action Card·Timeline 요약은 HTML 조각이 아니라 텍스트 컴포넌트다. 원고의
    `**강조**`가 그대로 보이던 문제를 특정 Day 문자열 치환 없이 여기서 막는다.
    """
    rendered = _inline(text or "")
    return html_lib.unescape(re.sub(r"<[^>]+>", "", rendered)).strip()


def split_stacked_tables(text: str) -> str:
    """빈 줄 없이 이어 붙은 표를 갈라 놓는다.

    원고에는 표 세 개가 줄바꿈 하나로 붙어 있는 곳이 있다. 마크다운은 그것을
    **표 하나**로 읽고, 첫 표의 열 수에 맞춰 뒤 표의 열을 잘라 버린다.
    Barcelona '한눈에 보기' 가 그랬다 — 세 번째 표의 `확정 일정`·`예상 체류`·
    `핵심 이유` 세 열이 화면에서 통째로 사라졌고, 그 값은 사이트 어디에도
    없었다. 열 수가 일정하게 잘리기 때문에 표 검사도 이것을 못 잡는다.

    구분선(`|---|---|`)은 표 하나에 하나뿐이다. 덩어리 안에서 두 번째
    구분선을 만나면 그 위 줄(다음 표의 머리)에서 자른다.
    """
    lines, out, block = text.splitlines(), [], []

    def flush():
        if not block:
            return
        seps = [i for i, ln in enumerate(block) if SEP_ROW.match(ln)]
        cuts = [i - 1 for i in seps[1:] if i >= 2]
        start = 0
        for cut in cuts:
            out.extend(block[start:cut])
            out.append("")
            start = cut
        out.extend(block[start:])
        block.clear()

    for line in lines:
        if line.lstrip().startswith("|"):
            block.append(line)
            continue
        flush()
        out.append(line)
    flush()
    return "\n".join(out)


def headerless_tables(text: str) -> tuple[str, dict[str, str]]:
    """헤더 없는 파이프 표를 직접 HTML 로 만든다.

    마크다운 표는 헤더 행과 구분선이 있어야 표로 읽힌다. 원고에는 둘 없이
    데이터 행만 있는 표가 있고(Barcelona·Paris), Aix 는 구분선만 있고 헤더가
    없다. 그대로 두면 파이프가 글자로 나와 화면이 깨진다.

    첫 행을 헤더로 승격시키지 않는다 — 'Essential' 은 열 이름이 아니라 값이다.
    없는 열 이름을 지어내느니 헤더 없는 표로 둔다.
    """
    lines = text.splitlines()
    parts, holes, i, n = [], {}, 0, 0
    while i < len(lines):
        if not lines[i].lstrip().startswith("|"):
            parts.append(lines[i])
            i += 1
            continue
        j = i
        while j < len(lines) and lines[j].lstrip().startswith("|"):
            j += 1
        block = lines[i:j]
        # 제대로 된 표(2행이 구분선)는 마크다운에게 맡긴다
        if len(block) >= 2 and SEP_ROW.match(block[1]):
            parts.extend(block)
            i = j
            continue
        rows = [b for b in block if not SEP_ROW.match(b)]
        if len(rows) < 2:
            parts.extend(block)
            i = j
            continue
        body = "".join(
            "<tr>" + "".join(f"<td>{_inline(c)}</td>" for c in _cells(r)) + "</tr>"
            for r in rows)
        key = f"@@TABLE{n}@@"
        holes[key] = f'<div class="table-wrap"><table><tbody>{body}</tbody></table></div>'
        # 빈 줄로 감싼다. 표 뒤에 --- 가 오는 원고가 있는데, 그러면 마크다운이
        # 바로 앞 줄(자리표시자)을 제목으로 읽어 표가 <h2> 안에 들어간다.
        parts += ["", key, ""]
        n += 1
        i = j
    return "\n".join(parts), holes


EXTERNAL_ANCHOR = re.compile(r'<a\s+href="(https?://[^"]+)"')


def md(text: str) -> str:
    if not text.strip():
        return ""
    # 표 앞뒤를 빈 줄로 가른다. 원고는 사람이 쓴 것이라 표 바로 뒤에 문장이
    # 붙어 있는 곳이 많고, 그러면 그 문장이 표의 한 행으로 빨려 들어가
    # 열 너비가 문장 길이만큼 늘어난다.
    #
    # 승격 단계가 아니라 여기서 한다. 장소 장문(30_Places)은 이제 손으로
    # 관리하는 정본이라 빌드가 고쳐 쓰지 않는다 — 원고는 쓴 대로 두고
    # 렌더가 방어한다.
    text = md_tidy.tidy(text)
    text = split_stacked_tables(text)
    text, holes = headerless_tables(text)
    html_out = md_lib.markdown(
        text, extensions=["tables", "fenced_code", "sane_lists", "attr_list"])
    # 표는 감싸서 그 안에서만 가로 스크롤시킨다 — 본문이 가로로 흐르면 안 된다
    html_out = re.sub(r"<table>", '<div class="table-wrap"><table>',
                      html_out).replace("</table>", "</table></div>")
    for key, block in holes.items():
        html_out = html_out.replace(f"<p>{key}</p>", block).replace(key, block)
    # 원고가 쓴 외부 링크에도 rel 을 붙인다. 렌더러가 직접 만드는 링크는
    # 이미 붙이고 있었지만 마크다운을 거쳐 나온 것은 빠져 있었다.
    html_out = EXTERNAL_ANCHOR.sub(
        lambda m: m.group(0) if "rel=" in m.group(0)
        else f'<a rel="nofollow noopener" href="{m.group(1)}"',
        html_out)
    return html_out


LAYER_LABEL = {"role": "여행 전체에서의 역할", "rhythm": "추천 체류 리듬"}

FACTS: dict = {}   # slug → {key: Fact}. load_facts() 가 채운다.

FACT_TOKEN = re.compile(r"\{\{fact:([a-z0-9-]+)\.([a-z_]+)\}\}")


def resolve_fact(match: "re.Match") -> str:
    """{{fact:<장소>.<항목>}} 을 실제 값으로 바꾼다.

    값을 감추지 않는다. 확인된 값 76곳을 전부 '확인 필요' 로 덮고 있었는데,
    그건 규칙 3 을 반대 방향으로 어긴 것이다 — 아는 것을 모른다고 쓰면
    현장에서 쓸 수 있는 정보를 버린다.

    다만 확정되지 않은 값에는 반드시 표시를 붙인다. 요금·운영시간을
    확정으로 믿고 움직이는 것이 이 프로젝트 최악의 사고다.
    """
    slug, key = match.group(1), match.group(2)
    fact = (FACTS.get(slug) or {}).get(key)
    if fact is None or not fact.value:
        reason = fact.blocked_reason if fact and fact.blocked_reason else "확인 필요"
        return f"({esc(reason)})"
    if fact.is_confirmed:
        return esc(fact.value)
    return f"{esc(fact.value)} (재확인)"


# 원고에 남아 있는 옛 주소. 승격된 본문이 그대로 들고 오면 링크가 깨진다.
LEGACY_LINK = re.compile(
    r"\((?:\.\./)*(chapters/([a-z]+)/[a-z-]+\.html|topics/[a-z-]+\.html"
    r"|tracker/[a-z-]+\.html|regions\.html|daily/index\.html)([^)]*)\)")


def _relink(m: "re.Match") -> str:
    """옛 주소를 새 IA 로 옮긴다. 리다이렉트가 있긴 하지만 본문 링크가
    한 번 더 튕기게 두지 않는다 — 오프라인에서는 그 왕복이 실패한다."""
    target, region = m.group(1), m.group(2)
    if target.startswith("chapters/") and region:
        return f"(../guide/{region}.html{m.group(3)})"
    if target.startswith("tracker/"):
        return f"(../prepare/index.html{m.group(3)})"
    return f"(../guide/index.html{m.group(3)})"


def strip_tokens(text: str) -> str:
    """원고의 토큰을 사람이 읽는 형태로 바꾼다."""
    text = LEGACY_LINK.sub(_relink, text)
    text = re.sub(r"\{\{grade:[^}]*\}\}", "", text)
    text = re.sub(r"\{\{badge:[^|}]*\|([^}]*)\}\}", r"(\1)", text)
    text = re.sub(r"\{\{badge:([^}]*)\}\}", r"(\1)", text)
    text = FACT_TOKEN.sub(resolve_fact, text)
    return re.sub(r"\{\{[^}]*\}\}", "", text)


# 글자로만 있는 URL 을 누를 수 있게 만든다.
#
# 사실 출처(place-facts 의 source) 286건이 전부 맨 URL 이라 화면에 글자로만
# 나왔다. 현장에서 "운영시간이 맞나" 를 확인하려면 그 주소를 손으로 옮겨
# 적어야 했다는 뜻이다. 확인할 수 없는 근거는 근거가 아니다.
URL_IN_TEXT = re.compile(r'https?://[^\s<>"\')\]]+')


def domain_of(url: str) -> str:
    """보여줄 이름. 전체 주소는 390px 에서 서너 줄을 먹고 읽히지도 않는다."""
    host = url.split("//", 1)[-1].split("/", 1)[0]
    return host[4:] if host.startswith("www.") else host


def linkify(text: str, *, show: str = "domain") -> str:
    """**이스케이프된** 문자열 안의 맨 URL 을 <a> 로 바꾼다.

    이스케이프를 먼저 하고 여기 넣어야 한다. 순서를 뒤집으면 우리가 만든
    태그가 다시 이스케이프돼 글자로 나온다.
    """
    def repl(m):
        raw = m.group(0)
        url = raw.rstrip(".,·;")
        tail = raw[len(url):]
        shown = domain_of(url) if show == "domain" else url
        return (f'<a href="{url}" rel="nofollow noopener">{shown}</a>{tail}')
    return URL_IN_TEXT.sub(repl, text)


def first_source_url(place) -> str:
    """이 장소의 근거 URL 하나. 공식 페이지인 경우가 대부분이라 상단
    행동줄에 '공식 정보' 로 내보낸다 — 현장에서 눌러 지금 값을 확인한다."""
    for key in ("booking", "hours", "price_adult", "closed", "getting_there"):
        f = place.facts.get(key)
        if f and f.source:
            m = URL_IN_TEXT.search(f.source)
            if m:
                return m.group(0).rstrip(".,·;")
    return ""



# ---------------------------------------------------------------- 예약번호 가리기

# 이 사이트는 gh-pages 로 공개 배포된다. 예약번호가 그대로 있으면 누구나
# 그 예약을 조회할 수 있다. 원본(트래커·원고)에는 온전히 남겨 둔다 — 그게
# 기록이고, 확정 사실 토큰 가드도 원고에서 코드가 살아 있는지 본다.
# 가리는 것은 **화면에 나갈 때뿐**이다.
#
# 아는 코드만 정확히 겨냥한다. 정규식으로 '코드처럼 생긴 것' 을 찾으면
# 주소나 전화번호가 함께 가려진다.
BOOKING_CODES: set[str] = set()

SKIP_CODE_WORDS = {"trip.com", "airbnb", "booking.com", "none", "미표기",
                   "발권메일", "확인", "pnr"}


def collect_codes(raw: str) -> None:
    for token in re.findall(r"[A-Za-z0-9][A-Za-z0-9.]{3,}", raw or ""):
        if token.lower() not in SKIP_CODE_WORDS:
            BOOKING_CODES.add(token)


def mask_code(code: str) -> str:
    """프라이버시 완전 보호를 위해 [CONFIRMED] 로 마스킹한다."""
    return "[CONFIRMED]"


def mask_booking_codes(html_text: str) -> str:
    """알려진 예약번호를 화면에서 가린다. 긴 것부터 바꿔야 부분 일치로
    짧은 코드가 긴 코드를 잘라먹지 않는다."""
    for code in sorted(BOOKING_CODES, key=len, reverse=True):
        if code in html_text:
            html_text = html_text.replace(code, mask_code(code))
    return html_text


# ---------------------------------------------------------------- 이미지

NAME_ALIASES = ROOT / "data" / "place-name-aliases.json"


def load_name_aliases() -> dict[str, list[str]]:
    """장소 이름 줄임말. 코드가 아니라 데이터로 둔다."""
    if not NAME_ALIASES.exists():
        return {}
    return json.loads(NAME_ALIASES.read_text(encoding="utf-8")).get("aliases", {})


CONSOLIDATION = ROOT / "data" / "region-consolidation.json"
_CONSOLIDATION: dict | None = None


def consolidation() -> dict:
    """통폐합을 끝낸 지역의 등록부. 지역 전용 분기를 코드에 넣지 않기 위해
    판정과 제목을 전부 데이터로 내린다."""
    global _CONSOLIDATION
    if _CONSOLIDATION is None:
        _CONSOLIDATION = (json.loads(CONSOLIDATION.read_text(encoding="utf-8"))
                          if CONSOLIDATION.exists() else {})
    return _CONSOLIDATION


def is_consolidated(slug: str) -> bool:
    return slug in (consolidation().get("consolidated") or [])


def layer_title(slug: str, key: str, default: str) -> str:
    """접이식 제목. 통폐합을 끝낸 지역은 자기 이름으로 부른다 —
    '지역 교통 심화' 가 아니라 '바르셀로나에서 이동하기' 다."""
    return ((consolidation().get("layerTitles") or {})
            .get(slug, {}).get(key, default))


def load_image_index(trip: Trip | None = None) -> dict:
    """사진 색인. 잇는 규칙은 model.load_images 하나뿐이다 — 여기서 다시
    만들지 않는다. 예전에는 두 곳에 같은 로직이 있었고, 그래서 별칭을 한 곳만
    고치면 다른 쪽이 조용히 옛 규칙으로 돌았다."""
    known = None
    if trip is not None:
        known = set(trip.places) | {r.slug for r in trip.regions}
    return model.load_images(known)


def img_src(img: dict, role: str, rel: str) -> tuple[str, str]:
    """(src, srcset). 카탈로그에 없는 이미지는 애초에 여기 오지 않는다."""
    variants = (img.get("variants") or {}).get(role) or []
    if not variants:
        for other in ("content", "hero", "thumbnail"):
            variants = (img.get("variants") or {}).get(other) or []
            if variants:
                break
    if not variants:
        return "", ""
    ordered = sorted(variants, key=lambda v: v.get("width", 0))
    src = f"{rel}/{ordered[-1]['sitePath']}"
    srcset = ", ".join(f"{rel}/{v['sitePath']} {v['width']}w" for v in ordered)
    return src, srcset


def figure(img: dict | None, rel: str, role: str = "content",
           cls: str = "", sizes: str = "100vw") -> str:
    """사진. 카탈로그에 항목이 없으면 자리를 아예 만들지 않는다.

    저작자 표시가 필요한 라이선스는 화면에 표시한다 — 이 사이트는 공개
    배포되므로 표시 의무가 실제로 발생한다.
    """
    if not img:
        return ""
    src, srcset = img_src(img, role, rel)
    if not src:
        return ""
    alt = esc(img.get("altKo") or img.get("titleKo") or "")
    ss = f' srcset="{srcset}" sizes="{sizes}"' if srcset else ""
    return (f'<img class="{cls}" src="{src}"{ss} alt="{alt}" '
            f'loading="lazy" decoding="async">')


# 사진이 없는 장소는 카드에서 자리를 아예 비웠다. 그러면 `84px 1fr` 그리드의
# 첫 열이 빈 채 남아 제목과 설명이 사진 칸으로 밀려 들어간다 — 니스의 식당
# 세 곳에서 실제로 그렇게 깨졌다. 사진이 없어도 **같은 크기의 자리**를 만든다.
PLACEHOLDER_KIND = {
    "food": ("food", "식당·카페 사진 자리"),
    "cafe": ("food", "카페 사진 자리"),
}


def photo_placeholder(p: "Place", cls: str = "") -> str:
    """사진이 없을 때 같은 비율의 자리를 채운다. 아이콘만 두지 않고
    접근성 이름을 함께 준다 — 아이콘은 스크린리더에 아무 말도 하지 않는다."""
    icon, label = PLACEHOLDER_KIND.get(place_visual_kind(p), ("pin", "장소 사진 자리"))
    return (f'<div class="{cls} thumb-empty" role="img" aria-label="{esc(label)}">'
            f'{ic(icon)}</div>')


def place_visual_kind(p: "Place") -> str:
    """카드에서 쓰는 시각 분류. 새 taxonomy 를 만들지 않고 정본 값을 읽는다."""
    if getattr(p, "food_kind", None):
        return "cafe" if str(p.food_kind).upper() in ("CAFE", "BAKERY") else "food"
    return "place"


def credit_line(img: dict | None) -> str:
    if not img:
        return ""
    who = esc(img.get("creator") or "")
    lic = esc(img.get("license") or "")
    src = img.get("sourcePage") or ""
    if not (who or lic):
        return ""
    link = f'<a href="{esc(src)}" rel="nofollow noopener">{who}</a>' if src else who
    return f'<p class="meta">사진 {link} · {lic}</p>'


# ---------------------------------------------------------------- 컴포넌트

def place_card(p: Place, rel: str, large: bool = False) -> str:
    """PlaceCard. 목록에서는 이 형태 하나만 쓴다 — 페이지마다 카드를 새로
    만들면 사이트가 조각난다."""
    img = IMAGES["by_place"].get(p.slug)
    grade = GRADE_BADGE.get(p.grade or "")
    grade_html = badge(*grade) if grade else ""
    url = f"{rel}/places/{p.slug}.html"

    type_badge = (
        f'<span class="badge" title="식당·미식" aria-label="식당·미식">{ic("food")}식당·미식</span>'
        if p.is_food else
        f'<span class="badge" title="명소·관광" aria-label="명소·관광">{ic("pin")}명소</span>'
    )

    if large:
        thumb = figure(img, rel, "content", "thumb", "(min-width:600px) 50vw, 100vw")
        thumb = thumb or photo_placeholder(p, "thumb")
        return f"""<article class="card place-card-lg">
  {thumb}
  <div class="card-body">
    <div class="metarow">{grade_html}{type_badge}</div>
    <h3 class="card-title"><a class="card-link" href="{url}">{esc(p.name)}</a></h3>
    <p class="card-dek">{esc(p.summary)}</p>
  </div>
</article>"""

    thumb = figure(img, rel, "thumbnail", "thumb", "84px")
    thumb = thumb or photo_placeholder(p, "thumb")
    return f"""<article class="card place-card">
  {thumb}
  <div class="card-body" style="padding:0">
    <h3 class="card-title"><a class="card-link" href="{url}">{esc(p.name)}</a></h3>
    <p class="card-dek">{esc(p.summary)}</p>
    <div class="metarow">{grade_html}{type_badge}</div>
  </div>
</article>"""


def day_card(d: Day, rel: str, region: Region | None = None) -> str:
    transfer = badge("caution", "거점 이동") if d.is_transfer else ""
    fatigue = f'<span>{ic("gauge")}피로 {esc(d.fatigue)}</span>' if d.fatigue else ""
    return f"""<article class="card day-card">
  <div class="card-body">
    <div class="day-card-head">
      <span class="day-date">{esc(d.date_label)}</span>
      <span class="day-num">DAY {d.n}</span>
    </div>
    <div class="day-route"><a class="card-link" href="{rel}/{d.url}">{esc(d.city)}</a></div>
    <p class="card-dek">{esc(d.title)}</p>
    <div class="metarow">{transfer}{fatigue}</div>
  </div>
</article>"""


def stop_map_url(s: Stop) -> str:
    """MapCard와 Action/Timeline이 공유하는 단일 지도 URL 생성 경로."""
    return maps_url(
        s.lat, s.lng, s.address or "",
        s.map_query or (s.place.map_query if s.place else ""),
        s.route_origin or "", s.route_destination or "", s.route_mode or "",
    )


def stop_official_url(s: Stop) -> str:
    if s.official_url:
        return s.official_url
    if not s.place:
        return ""
    url_fact = s.place.fact("url")
    if url_fact and url_fact.value.startswith("http"):
        return url_fact.value
    return first_source_url(s.place)


def stop_actions(s: Stop, rel: str, *, context: str) -> str:
    """링크가 있는 행동만, 지도 → Place → 공식 링크 순서로 만든다."""
    actions = []
    map_href = stop_map_url(s)
    if map_href:
        actions.append(
            f'<a class="btn btn-primary" href="{esc(map_href)}" '
            f'rel="nofollow noopener">{ic("map")}길찾기</a>')
    if context != "tl" and s.place:
        actions.append(
            f'<a class="btn btn-secondary" href="{rel}/places/{s.place.slug}.html">'
            f'{ic("pin")}장소 정보</a>')
    official = stop_official_url(s)
    if context != "tl" and official:
        actions.append(
            f'<a class="btn btn-secondary" href="{esc(official)}" '
            f'rel="nofollow noopener">{ic("link")}티켓·공식</a>')
    if not actions:
        return ""
    return f'<div class="{context}-actions btn-row">{"".join(actions)}</div>'


def stop_status_markup(s: Stop) -> tuple[str, str]:
    """(배지, 설명). Pilot 상태가 없으면 기존 reservation 표시를 보존한다."""
    marks, notes = [], []
    if s.execution_statuses:
        for status in s.execution_statuses:
            kind, default_label, icon = EXECUTION_STATUS_UI[status.type]
            label = status.label or default_label
            marks.append(badge(kind, label))
            if status.detail:
                notes.append(
                    f'<p class="tl-note tl-status tl-status-{status.type}">'
                    f'{ic(icon)}<strong>{esc(label)}</strong> · '
                    f'{esc(plain_inline(status.detail))}</p>')
    else:
        if s.optional:
            marks.append(badge("neutral", "선택"))
        if s.reservation:
            marks.append(badge("caution", "예약"))
            notes.append(
                f'<p class="tl-note tl-status">{ic("ticket")}'
                f'{esc(plain_inline(s.reservation))}</p>')
    if s.execution_note:
        notes.append(f'<p class="tl-note tl-execution-note">'
                     f'{esc(plain_inline(s.execution_note))}</p>')
    return "".join(marks), "".join(notes)


def timeline(d: Day, rel: str) -> str:
    """하루의 뼈대. stop 과 leg 를 순서대로 엮는다."""
    legs = {(l.frm, l.to): l for l in d.legs}
    rows, stops = [], d.stops
    for i, s in enumerate(stops):
        icon = CAT_ICON.get(s.category, "pin")
        name = esc(s.name)
        if s.place is not None:
            name = f'<a href="{rel}/places/{s.place.slug}.html">{name}</a>'
        marks, status_notes = stop_status_markup(s)
        summary = plain_inline(s.summary)
        note = f'<p class="tl-note tl-summary">{esc(summary)}</p>' if summary else ""
        # 한 stop 이 두 장소를 담을 때 보조 장소를 명시한다. 시간표는 한 줄로
        # 두되 장소 연결은 숨기지 않는다 — 그러지 않으면 그 장소가 어느
        # 날에도 걸리지 않는다.
        if s.related_places:
            links = " · ".join(
                f'<a href="{rel}/places/{x.slug}.html">{esc(x.name)}</a>'
                for x in s.related_places)
            note += f'<p class="tl-note">함께 보는 곳 — {links}</p>'
        actions = stop_actions(s, rel, context="tl")
        action_template = stop_actions(s, rel, context="action")
        if action_template:
            action_template = (f'<template class="tl-action-template">'
                               f'{action_template}</template>')
        rows.append(f"""<li class="tl-item tl-category-{esc(s.category)}" data-start="{esc(s.start or '')}" data-end="{esc(s.end or '')}">
  <div class="tl-time">{esc(s.start or '')}</div>
  <div class="tl-body">
    <div class="tl-name">{ic(icon)} <span class="tl-title">{name}</span>
      <span class="tl-marks">{marks}</span></div>
    {note}{status_notes}{actions}{action_template}
  </div>
</li>""")
        nxt = stops[i + 1] if i + 1 < len(stops) else None
        leg = legs.get((s.id, nxt.id)) if nxt else None
        if leg:
            bits = [MODE_LABEL.get(leg.mode, leg.mode)]
            unconfirmed = leg.mode == "unconfirmed"
            if leg.duration:
                bits.append(leg.duration)
            if leg.distance:
                bits.append(leg.distance)
            if leg.line:
                bits.append(leg.line)
            major = (
                (d.day_type == "driving" and leg.mode in {"car", "drive"})
                or (d.day_type == "transfer" and leg.mode not in {"walk", "unconfirmed"})
            )
            cls = "tl-leg tl-leg-open" if unconfirmed else "tl-leg"
            if major:
                cls += " tl-leg-major"
            route = (f'<strong class="tl-leg-route">{esc(s.name)} → '
                     f'{esc(nxt.name)}</strong>' if major else "")
            # Route queries are attached to the departure stop because they
            # describe this exact leg (not the destination's next movement).
            next_map = stop_map_url(s)
            next_action = (f'<a class="tl-leg-action" href="{esc(next_map)}" '
                           f'rel="nofollow noopener">{ic("map")}다음 목적지</a>'
                           if major and next_map else "")
            rows.append(f"""<li class="{cls}">
  <div></div>
  <div class="tl-body">{route}<span class="tl-leg-line">{ic(MODE_ICON.get(leg.mode, 'pin'))}
    {esc(' · '.join(bits))}</span>{next_action}</div>
</li>""")
    return f'<ol class="timeline">{"".join(rows)}</ol>'


def map_card(stops, rel: str, center=None, zoom: int = 14,
             label: str = "지도", region_groups=None, numbered: bool = True) -> str:
    """MapCard. 핀은 Place DB 에서만 온다 — HTML 에 좌표를 따로 박지 않는다.

    지도는 눌렀을 때만 불러온다. 43일 내내 열리는 화면마다 지도 SDK 를
    받으면 데이터가 약한 곳에서 첫 화면이 늦는다.

    JS 나 네트워크가 없어도 목록과 Google Maps 링크는 남는다. 현장에서
    스크립트가 안 뜨는 상황이 실제로 있고, 그때 좌표 링크만이라도 손에
    있어야 한다.
    """
    def _pin_dict(s):
        name_text = getattr(s, "route_title", None) or s.name
        return {
            "id": s.id, "name": name_text, "lat": s.lat, "lng": s.lng,
            "cat": s.category,
            "time": getattr(s, "formatted_when", s.start or ""),
            "address": s.address,
            "place": s.place.slug if s.place else None,
            "query": getattr(s, "map_query", None) or (s.place.map_query if s.place else None),
            "map_type": getattr(s, "map_type", "place"),
            "origin": getattr(s, "route_origin", None),
            "destination": getattr(s, "route_destination", None),
            "travel_mode": getattr(s, "route_mode", None),
        }

    valid_stops = [s for s in stops
                   if getattr(s, "map_type", "place") != "non_map"
                   and s.id not in {"cdg-departure", "inflight", "icn", "bcn-airport", "paris-return"}
                   and (s.category != "hotel" or getattr(s, "map_type", "place") == "route")
                   and ((s.lat and s.lng) or s.address)]
    if not valid_stops:
        return ""

    pins = [_pin_dict(s) for s in valid_stops]
    located = [p for p in pins if p["lat"] and p["lng"]]
    if center is None:
        if not located:
            return ""
        center = [sum(p["lat"] for p in located) / len(located),
                  sum(p["lng"] for p in located) / len(located)]
    payload = json.dumps({"center": center, "zoom": zoom, "pins": located},
                         ensure_ascii=False, separators=(",", ":")) \
        .replace("<", "\\u003c").replace(">", "\\u003e").replace("&", "\\u0026")

    def _render_stop_li(s, idx: int | None = None) -> str:
        name_text = getattr(s, "route_title", None) or s.name
        name_esc = esc(name_text)
        if s.place:
            name_html = f'<a href="{rel}/places/{s.place.slug}.html">{name_esc}</a>'
        else:
            name_html = name_esc
        when_text = getattr(s, "formatted_when", s.start or "")
        when_html = f'<span class="meta">{esc(when_text)}</span>' if when_text else ""
        num_html = f'<span class="map-num">{idx}.</span>' if idx is not None else ""
        href = maps_url(s.lat, s.lng, s.address or "",
                        getattr(s, "map_query", None) or (s.place.map_query if s.place else ""),
                        getattr(s, "route_origin", None) or "",
                        getattr(s, "route_destination", None) or "",
                        getattr(s, "route_mode", None) or "")
        return (f'<li data-pin="{esc(s.id)}">{num_html}{when_html}'
                f'<span class="map-name">{name_html}</span>'
                f'<a class="map-open" rel="nofollow noopener" href="{esc(href)}">'
                f'{ic("map")}'
                f'<span class="visually-hidden">{name_esc} </span>열기</a></li>')

    if region_groups:
        groups_html = []
        for r_name, r_stops in region_groups:
            r_valid = [s for s in r_stops
                       if getattr(s, "map_type", "place") != "non_map"
                       and s.id not in {"cdg-departure", "inflight", "icn", "bcn-airport", "paris-return"}
                       and (s.category != "hotel" or getattr(s, "map_type", "place") == "route")
                       and ((s.lat and s.lng) or s.address)]
            r_items = [_render_stop_li(s, idx=i + 1 if numbered else None)
                       for i, s in enumerate(r_valid)]
            if r_items:
                groups_html.append(
                    f'<section class="map-region-group">'
                    f'<h3 class="map-region-head">{esc(r_name)}</h3>'
                    f'<ol class="map-region-list">{"".join(r_items)}</ol>'
                    f'</section>')
        list_html = f'<div class="map-list">{"".join(groups_html)}</div>'
    else:
        items = [_render_stop_li(s, idx=i + 1 if numbered else None)
                 for i, s in enumerate(valid_stops)]
        list_html = f'<ol class="map-list">{"".join(items)}</ol>'

    return f"""<div class="map-card">
  <div class="map-canvas" hidden></div>
  <p class="map-status meta" role="status" aria-live="polite"></p>
  <script type="application/json" class="map-data-script">{payload}</script>
  <div class="map-card-foot">
    <span class="label">{esc(label)} · {len(pins)}곳</span>
    <div class="map-toggle" role="group" aria-label="지도와 목록 전환">
      <button type="button" data-view="map" aria-pressed="false">지도</button>
      <button type="button" data-view="list" aria-pressed="true">목록</button>
    </div>
  </div>
  {list_html}
</div>"""


def maps_url(lat=None, lng=None, address: str = "", query: str = "",
             origin: str = "", destination: str = "", travel_mode: str = "") -> str:
    """지도 링크. **경로면 Directions URL, 이름이 있으면 이름으로**, 없으면 주소, 그다음 좌표.

    좌표는 틀리면 조용히 틀린다. 실제로 숙소 좌표가 식당 9곳에 복사돼
    파리 Bouillon Chartier 는 2.5km, 리옹 Café Comptoir Abel 은 3km
    어긋나 있었다. 이름은 틀리면 검색이 실패해서 눈에 보인다.

    이름 검색어는 data/map-queries.json 이 정본이고, 이름으로 단일하게
    특정되는 것만 들어 있다.
    """
    if origin and destination:
        mode = travel_mode or "transit"
        return (f"https://www.google.com/maps/dir/?api=1"
                f"&origin={quote(origin)}&destination={quote(destination)}"
                f"&travelmode={mode}")
    if query:
        return ("https://www.google.com/maps/search/?api=1&query="
                + quote(query))
    if address:
        return ("https://www.google.com/maps/search/?api=1&query="
                + quote(address))
    if lat and lng:
        return f"https://www.google.com/maps/search/?api=1&query={lat},{lng}"
    return ""


def index_search(title: str, url: str, kind: str, extra: str = "") -> None:
    SEARCH_INDEX.append({"t": title, "u": url, "k": kind, "x": extra})


# ================================================================ 페이지

def build_place(p: Place, trip: Trip) -> str:
    """Place — Action → Experience → Practical → Deep Guide.

    길 찾는 손이 먼저 닿아야 하는 것은 서술이 아니라 지도와 시각이다.
    긴 역사·건축 서술은 지우지 않고 아래로 내린다.
    """
    rel = ".."
    region = trip.region(p.region)
    img = IMAGES["by_place"].get(p.slug)
    grade = GRADE_BADGE.get(p.grade or "")

    # --- Action -----------------------------------------------------------
    meta = []
    if grade:
        meta.append(badge(*grade))
    dur = p.fact("duration")
    if dur and dur.is_confirmed:
        meta.append(f'<span>{ic("clock")}{esc(dur.value)}</span>')
    for n in sorted(p.days):
        d = trip.day(n)
        if d:
            meta.append(f'<a href="{rel}/{d.url}">{esc(d.date_label)} · Day {n}</a>')

    actions = []
    # 목적지도 이름을 먼저 쓴다. 좌표가 없어 길찾기 버튼 자체가 없던 장소가
    # 36곳 있었는데, 이름이 있으면 그 장소들도 길을 열 수 있다.
    dest = quote(p.map_query) if p.map_query else (
        f"{p.lat},{p.lng}" if p.lat and p.lng else "")
    if dest:
        actions.append(
            f'<a class="btn btn-primary" rel="nofollow noopener" '
            f'href="https://www.google.com/maps/dir/?api=1&destination={dest}">'
            f'{ic("map")}길찾기</a>')
    url_fact = p.fact("url")
    official = (url_fact.value if url_fact and url_fact.value.startswith("http")
                else first_source_url(p))
    if official:
        actions.append(f'<a class="btn btn-secondary" href="{esc(official)}" '
                       f'rel="nofollow noopener">{ic("link")}공식 정보</a>')
    if p.wiki:
        wiki_url = (f"https://{p.wiki_lang}.wikipedia.org/wiki/"
                    f"{p.wiki.replace(' ', '_')}")
        actions.append(f'<a class="btn btn-secondary" href="{esc(wiki_url)}" '
                       f'rel="nofollow noopener">{ic("book")}위키백과</a>')

    hero = ""
    if img:
        src, srcset = img_src(img, "hero", rel)
        if src:
            hero = (f'<img class="hero-img" src="{src}" srcset="{srcset}" '
                    f'sizes="100vw" alt="{esc(img.get("altKo") or p.name)}" '
                    f'fetchpriority="high" decoding="async">')

    parts = [f"""<div class="hero">
  {hero}
  <div class="hero-body"><div class="hero-card">
    <div class="hero-eyebrow"><span class="label">{esc(region.name if region else '')}</span></div>
    <h1>{esc(p.name)}</h1>
    <div class="metarow">{''.join(meta)}</div>
    {f'<p class="hero-dek">{esc(p.summary)}</p>' if p.summary else ''}
    {f'<div class="btn-row" style="margin-top:1rem">{"".join(actions)}</div>' if actions else ''}
  </div></div>
</div>
<div class="wrap-read">
{credit_line(img)}"""]

    # --- Experience -------------------------------------------------------
    if p.why_go:
        parts.append(sec_head("WHY GO", "왜 가는가", rule=True))
        parts.append(f'<div class="prose">{md(strip_tokens(p.why_go))}</div>')
    if p.dont_miss:
        parts.append(sec_head("DON'T MISS", "놓치지 말 것", rule=True))
        items = "".join(f"<li>{esc(x)}</li>" for x in p.dont_miss)
        parts.append(f'<div class="prose"><ol>{items}</ol></div>')

    # --- Practical --------------------------------------------------------
    facts = [(k, f) for k, f in p.facts.items() if f.value or f.blocked_reason]
    if facts:
        parts.append(sec_head("PRACTICAL", "실용", rule=True))
        rows = []
        for key, f in sorted(facts, key=lambda kv: list(FACT_LABEL).index(kv[0])
                             if kv[0] in FACT_LABEL else 99):
            label = FACT_LABEL.get(key, key)
            if f.is_confirmed:
                mark = ""
                value = esc(f.value)
            else:
                # 확정처럼 보이면 안 된다. 이걸 믿고 움직이는 것이 최악의 사고다.
                mark = badge("caution", "재확인")
                value = esc(f.value) if f.value else \
                    f'<span class="meta">{esc(f.blocked_reason or "미확인")}</span>'
            src = (f'<br><span class="meta">출처 {linkify(esc(f.source))}</span>'
                   if f.source else "")
            rows.append(f"<tr><th scope=\"row\">{esc(label)}</th>"
                        f"<td>{value} {mark}{src}</td></tr>")
        parts.append(f'<div class="table-wrap"><table><tbody>{"".join(rows)}'
                     f"</tbody></table></div>")
    if p.practical_md.strip():
        if not facts:
            parts.append(sec_head("PRACTICAL", "실용", rule=True))
        parts.append(f'<div class="prose">{md(strip_tokens(p.practical_md))}</div>')

    # --- Deep Guide -------------------------------------------------------
    body = strip_tokens(p.body_md)
    if body.strip():
        parts.append(sec_head("DEEP GUIDE", "더 깊이", rule=True))
        parts.append(f'<div class="prose">{md(body)}</div>')

    # --- 다른 사진 --------------------------------------------------------
    # 한 장소에 사진이 여러 장 있으면 예전에는 첫 장만 쓰고 나머지는 저장소에
    # 남아 화면에 영영 안 나왔다. 별칭을 정리하고 나니 22장이 그 상태가 됐다.
    # 사진마다 제 설명과 저작자를 달고 나온다.
    extras = IMAGES.get("extras", {}).get(p.slug) or []
    if extras:
        figs = []
        for img in extras:
            fig = figure(img, rel, "content", "gallery-shot",
                         "(min-width:600px) 50vw, 100vw")
            if not fig:
                continue
            caption = esc(img.get("captionKo") or img.get("titleKo") or "")
            figs.append(f'<figure class="gallery-item">{fig}'
                        f'<figcaption>{caption}{credit_line(img)}</figcaption>'
                        "</figure>")
        if figs:
            parts.append(sec_head("PHOTOS", "다른 사진"))
            parts.append(f'<div class="grid grid-2">{"".join(figs)}</div>')

    # --- 현장 프랑스어 (Quick French) ------------------------------------
    pqf = place_quick_french(p, trip, rel)
    if pqf:
        parts.append(pqf)

    # 같은 지역의 다른 장소 — 길이 끊기지 않게 옆으로 나가는 문을 둔다
    if region:
        sibs = [x for x in region.places if x.slug != p.slug and x.summary][:6]
        if sibs:
            parts.append(sec_head("", f"{region.name} 의 다른 장소",
                                   more=("전체", f"{rel}/guide/{region.slug}.html")))
            parts.append('<div class="scroll-x place-siblings">'
                         + "".join(place_card(x, rel) for x in sibs) + "</div>")

    parts.append("</div>")

    index_search(p.name, f"places/{p.slug}.html", "place",
                 region.name if region else "")

    return page(
        title=p.name, body="\n".join(parts), rel=rel, tab="guide",
        region=p.region, country=region.country if region else "",
        description=p.summary,
        trail=[("홈", "index.html"), ("가이드", "guide/index.html"),
               (region.name if region else "", f"guide/{p.region}.html"),
               (p.name, None)],
    )


def sibling_days(d: Day, trip: Trip, region) -> list[Day]:
    """이 날의 형제 — 옆으로 움직일 수 있는 날들.

    보통은 그 지역의 날들이다. 그런데 정규 지역에 속하지 않는 날이 있다 —
    귀국일(`region: return`)이 그렇다. 그 날은 지역이 없으니 형제도 없고,
    스트립이 통째로 비어 **옆으로 갈 길이 화면에서 사라졌다.**

    지역으로 승격하지 않고 푼다. 같은 `region` 값을 가진 연속된 날들을 묶고,
    그 앞의 하루를 붙여 여정이 끊기지 않게 한다. Day 43 을 이름으로 부르지
    않으므로, 앞뒤에 비슷한 날(출국일·경유일)이 생겨도 같은 규칙이 적용된다.
    """
    if region:
        return list(region.days)
    kin = [x for x in trip.days if x.region == d.region]
    first = min(x.n for x in kin)
    lead = trip.day(first - 1)
    return ([lead] if lead else []) + kin


def build_day(d: Day, trip: Trip) -> str:
    """Day — 실행 화면. 첫 1~2 스크린에서 다음이 보여야 한다.
    지금 어디로 · 다음 일정 · 예약 · 주의 · 지도."""
    rel = ".."
    region = trip.region(d.region)
    prev_d, next_d = trip.day(d.n - 1), trip.day(d.n + 1)

    head_marks = []
    day_type = d.day_type or ("transfer" if d.is_transfer else None)
    if day_type in DAY_TYPE_LABEL:
        head_marks.append(badge(*DAY_TYPE_LABEL[day_type]))
    elif d.is_transfer:
        head_marks.append(badge("caution", "거점 이동"))
    if not d.is_authoritative:
        head_marks.append(badge("neutral", "검토 중"))
    if d.fatigue:
        head_marks.append(f'<span>{ic("gauge")}피로도 {esc(d.fatigue)}</span>')
    if d.total_distance:
        head_marks.append(f'<span>{esc(d.total_distance)}</span>')

    type_class = f" day-type-{day_type}" if day_type else ""
    parts = [f"""<div class="wrap{type_class}">
<div class="stack-lg" style="padding-top:1.5rem">
<header>
  <div class="metarow"><span class="day-date">{esc(d.date_label)}</span>
    <span class="day-num">DAY {d.n}</span></div>
  <h1>{esc(d.city)}</h1>
  <p class="day-summary">{esc(d.title)}</p>
  <div class="metarow">{''.join(head_marks)}</div>
</header>"""]

    # --- NEXT — 지금 시각 기준. 오늘이 아니면 첫 일정을 보여준다 ------------
    real = [s for s in d.stops if s.category != "hotel" and s.start]
    if real:
        first = real[0]
        name = esc(first.name)
        if first.place:
            name = f'<a href="{rel}/places/{first.place.slug}.html">{name}</a>'
        action_links = stop_actions(first, rel, context="action")
        parts.append(f"""<section class="action-card" id="next-action"
    data-day="{d.date.isoformat()}">
  <span class="label">NEXT</span>
  <div class="action-when">{esc(first.start)}</div>
  <div class="action-what">{name}</div>
  {f'<p class="card-dek action-summary">{esc(plain_inline(first.summary))}</p>' if first.summary else ''}
  {action_links}
</section>""")

    # --- 예약 — 당일에 잠긴 것 -------------------------------------------
    reserved = d.reserved_stops
    if reserved:
        def booking_text(s: Stop) -> str:
            selected = [x for x in s.execution_statuses
                        if x.type in {"confirmed", "book", "ticket"}]
            if selected:
                return " · ".join(
                    f"{x.label or EXECUTION_STATUS_UI[x.type][1]}: {plain_inline(x.detail)}"
                    for x in selected)
            return plain_inline(s.reservation or "")

        rows = "".join(
            f'<li><strong>{esc(s.start or "")} {esc(s.name)}</strong> — '
            f"{esc(booking_text(s))}</li>" for s in reserved)
        parts.append(sec_head("BOOKING", "오늘 예약"))
        parts.append(f'<div class="prose"><ul>{rows}</ul></div>')

    # --- 시간표 -----------------------------------------------------------
    parts.append(sec_head("TODAY", "오늘 일정"))
    parts.append(timeline(d, rel))

    # --- 현장 프랑스어 (Quick French) ------------------------------------
    dqf = day_quick_french(d, trip, rel)
    if dqf:
        parts.append(dqf)

    # --- 주의 ------------------------------------------------------------
    if d.backup:
        parts.append(sec_head("PLAN B", "일정 조정 기준", rule=True))
        parts.append(alert("caution", esc(plain_inline(d.backup))))
    checks = [esc(plain_inline(x)) for x in d.needs_review]
    if checks:
        parts.append(sec_head("PRE-TRIP CHECK", "출발 전 확인", rule=True))
        parts.append("".join(
            alert("caution", c) for c in checks))

    # --- 지도 ------------------------------------------------------------
    m = d.map or {}
    mc = map_card(d.stops, rel, center=m.get("center"), zoom=m.get("zoom", 14),
                  label=f"Day {d.n} 동선")
    if mc:
        parts.append(sec_head("MAP", "오늘 지도"))
        parts.append(mc)

    # --- 보조 정보 --------------------------------------------------------
    aside = []
    if d.transport:
        aside.append(("이동", d.transport))
    if d.food:
        aside.append(("식사", d.food))
    if d.highlights:
        aside.append(("오늘의 핵심", d.highlights))
    if aside:
        blocks = "".join(
            f'<details class="acc"><summary>{esc(name)}</summary>'
            f'<div class="acc-body"><ul>'
            + "".join(f"<li>{esc(x)}</li>" for x in items)
            + "</ul></div></details>" for name, items in aside)
        parts.append(f'<div class="stack">{blocks}</div>')

    # --- 이전/다음 --------------------------------------------------------
    # 라벨은 날짜다 — 현장에서 찾는 것은 "며칠"이다. Day 번호는 title 로 남는다.
    def day_hint(x) -> str:
        # 보이는 순서 그대로 읽히게 한다 — 날짜가 먼저, Day 번호가 뒤.
        # ISO 날짜까지 붙여 툴팁만 봐도 연도를 안다.
        return f"{x.date_label} · Day {x.n} · {x.date.isoformat()}"

    nav = []
    if prev_d:
        nav.append(f'<a class="btn btn-secondary" href="{rel}/{prev_d.url}" '
                   f'title="{esc(day_hint(prev_d))}">'
                   f"← {esc(prev_d.date_label)}</a>")
    if next_d:
        nav.append(f'<a class="btn btn-secondary" href="{rel}/{next_d.url}" '
                   f'title="{esc(day_hint(next_d))}">'
                   f"{esc(next_d.date_label)} →</a>")
    parts.append(f'<div class="btn-row" style="justify-content:space-between">'
                 f'{"".join(nav)}</div>')
    parts.append("</div></div>")

    # 형제 이동 — 그 지역의 날들
    sib = tabs_strip([
        (x.date_label, f"{rel}/{x.url}", x.n == d.n, f"Day {x.n}", day_hint(x))
        for x in sibling_days(d, trip, region)])

    index_search(f"{d.date_label} · Day {d.n} {d.city}", d.url, "day", f"Day {d.n} · {d.title}")

    return page(
        title=f"{d.date_label} · Day {d.n} · {d.city}", body="\n".join(parts), rel=rel,
        tab="today", region=d.region, country=d.country, subnav=sib,
        description=d.title,
        trail=[("홈", "index.html"), ("전체 일정", "schedule.html"),
               (f"{d.date_label} · Day {d.n}", None)],
    )


def link_food_text(text: str, rel: str, trip: Trip) -> str:
    """문장 안의 식당·시장 이름을 장소 페이지로 잇는다.

    정식 이름은 명부에서 그대로 잡고, 원고가 쓰는 줄임말('La Paradeta' ·
    'Bouillon Chartier')만 `data/place-name-aliases.json` 에서 읽는다.
    예전에는 이 별칭이 렌더러 안의 slug 분기 30여 개로 있었다 — 지역이
    늘 때마다 렌더러를 고치게 되는 구조라 지역 전용 우회로의 씨앗이었다.
    """
    escaped = esc(text)
    name_map = getattr(trip, "_food_name_map", None)
    if name_map is None:
        name_to_place = {}
        for p in trip.places.values():
            name_to_place[p.name] = p
        for slug, names in load_name_aliases().items():
            p = trip.places.get(slug)
            if p is None:
                continue
            for name in names:
                name_to_place.setdefault(name, p)
        # 긴 이름을 먼저 맞춘다 — 'Les Halles' 가 "Les Halles d'Avignon" 을
        # 반쪽만 잡아 링크가 이름 가운데서 끊기던 일이 있었다.
        name_map = sorted(name_to_place.items(), key=lambda x: len(x[0]),
                          reverse=True)
        trip._food_name_map = name_map

    for name, p in name_map:
        esc_name = esc(name)
        pattern = re.compile(rf"(?<![\">])({re.escape(esc_name)})(?![^<]*</a>)")
        if pattern.search(escaped):
            escaped = pattern.sub(f'<a href="{rel}/{p.url}">\\1</a>', escaped, count=1)
    return escaped


# '먹거리' 목록에서 걸러내는 실행 메모. 업소도 요리도 아닌 하루의 운영
# 지시라 지역 페이지가 아니라 Day 페이지가 맡는다.
GENERIC_FOOD_NOTES = [
    "기내", "편의점", "물만", "이동용 물", "출발 시각", "숙소 간단식", "숙소 저녁",
    "숙소식", "숙소 점심", "숙소권 간단", "숙소권 저녁", "숙소식 또는",
    "이동 중 간단식", "숙소 주변 가벼운 저녁", "가벼운 저녁", "가벼운 점심",
    "이른 저녁", "저녁 무예약", "동네 저녁", "가까운 저녁",
    "첫 장보기", "필수품만", "점심·휴식", "브런치·숙소", "숙소권 가벼운",
    "도착 점심은 가볍게", "점심 — 가볍게", "저녁은 가볍게", "마지막 저녁",
    "농가 첫 저녁", "농가 저녁", "플랫폼 대기", "경기장 식사", "축제권 점심",
    "동부 파리 점심",
]


def region_dishes(r: Region) -> list[str]:
    """'무엇을 먹는가'. Day 의 식사 슬롯에서 뽑는다.

    **통폐합을 끝낸 지역에서는 만들지 않는다.** Day 의 식사 슬롯은 본래
    '언제 어디서 먹는가' 라서, 지역 페이지에 목록으로 올리면 바로 위 식당
    카드가 말한 것을 업소 이름째 다시 말하게 된다. Nice 에서는 거기에
    '점심:' 같은 끼니 라벨과 `WISH-01` 같은 내부 코드까지 함께 새어 나왔다.

    통폐합한 지역은 챕터가 '이 지역에서 먹어볼 것' 을 따로 갖는다 —
    먹을 것은 거기가, 먹을 곳은 카드가, 언제 먹는지는 Day 가 맡는다.
    """
    if is_consolidated(r.slug):
        return []
    out = []
    for d in r.days:
        for item in d.food:
            item = item.strip()
            if any(g in item for g in GENERIC_FOOD_NOTES):
                continue
            if item not in out:
                out.append(item)
    return out[:12]


def attraction_card(p: Place, rel: str, large: bool = False) -> str:
    """볼거리 카드 = PlaceCard + 방문일 배지.

    날짜를 템플릿에 박지 않는다. Day SOT 에서 계산한 배지라 일정이 바뀌면
    카드도 같이 바뀐다.
    """
    card = place_card(p, rel, large=large)
    badges = visit_badges(p, rel)
    if not badges:
        return card
    return card.replace('<div class="metarow">',
                        f'<div class="metarow">{badges}', 1)


MEAL_LABEL = [(0, 4, "야식"), (4, 11, "아침"), (11, 16, "점심"), (16, 24, "저녁")]

FOOD_TYPE_LABEL = {
    "restaurant": "식당", "cafe": "카페", "bakery": "빵집",
    "market": "시장", "food-hall": "푸드홀", "wine-bar": "와인바",
}


def meal_of(start: str | None) -> str:
    """시각 → 끼니. 카드에 '점심'·'저녁' 을 붙이는 데 쓴다."""
    if not start or ":" not in start:
        return ""
    try:
        hour = int(start.split(":", 1)[0])
    except ValueError:
        return ""
    return next((label for lo, hi, label in MEAL_LABEL if lo <= hour < hi), "")


def place_visits(trip: Trip) -> dict[str, list[tuple[Day, Stop]]]:
    """장소 → 그 장소를 실제로 들르는 (Day, Stop).

    방문일 배지의 정본이다. 날짜 문자열을 지역 템플릿에 박지 않는다 —
    일정이 바뀌면 배지도 같이 바뀌어야 한다.
    """
    out: dict[str, list[tuple[Day, Stop]]] = {}
    for d in trip.days:
        for s in d.stops:
            if s.place is not None:
                out.setdefault(s.place.slug, []).append((d, s))
            # 한 stop 이 두 장소를 담을 때 보조 장소도 그날을 갖는다.
            # Day 13 의 시장 stop 이 Pâtisserie Weibel 을 함께 담는 것처럼.
            for extra in s.related_places:
                out.setdefault(extra.slug, []).append((d, s))
    for rows in out.values():
        rows.sort(key=lambda x: (x[0].n, x[1].order))
    return out


VISITS: dict[str, list[tuple[Day, Stop]]] = {}


def visit_badges(p: Place, rel: str, *, meals: bool = False) -> str:
    """[9.2(수) · Day 5] 배지. Day SOT 에서 계산한다."""
    out = []
    for d, s in VISITS.get(p.slug, [])[:3]:
        meal = f" · {meal_of(s.start)}" if meals and meal_of(s.start) else ""
        out.append(f'<a class="badge badge-day" href="{rel}/{d.url}">'
                   f'{esc(d.date_label)} · Day {d.n}{esc(meal)}</a>')
    if not out and p.days:
        for n in sorted(p.days)[:3]:
            out.append(f'<span class="badge">Day {n}</span>')
    return "".join(out)


def official_url(p: Place) -> str:
    """공식 페이지 하나. 없으면 빈 문자열 — 지어내지 않는다."""
    fact = p.fact("url")
    if fact and fact.value.startswith("http"):
        return fact.value
    m = URL_IN_TEXT.search(p.practical_md or "")
    if m and "wikipedia.org" not in m.group(0):
        return m.group(0).rstrip(".,·;)")
    return first_source_url(p)


def price_line(p: Place) -> str:
    """가격과 확인일. **확인하지 못한 값은 추정하지 않는다.**"""
    fact = p.fact("price_range") or p.fact("price_adult")
    if fact is None or not fact.value:
        return f'<dt>가격</dt><dd>{badge("caution", "미확인")}</dd>'
    mark = "" if fact.is_confirmed else " " + badge("caution", "재확인")
    when = f' · 확인 {esc(fact.verified_at)}' if fact.verified_at else ""
    return f'<dt>가격</dt><dd>{esc(fact.value)}{mark}{when}</dd>'


def food_card(p: Place, rel: str, trip: Trip) -> str:
    """식당·카페 카드. 사진 · 소개 · 방문일 · 추천 메뉴 · 가격 · 지도 · 공홈.

    없는 것은 숨기지 않고 자리도 만들지 않는다. 가격만 예외다 — 모르면
    '미확인' 이라고 쓴다. 빈칸은 '싸다' 로 읽히기 때문이다.
    """
    img = IMAGES["by_place"].get(p.slug)
    thumb = figure(img, rel, "content", "thumb",
                   "(min-width:600px) 50vw, 100vw") or photo_placeholder(p, "thumb")
    grade = GRADE_BADGE.get(p.grade or "")
    marks = [badge(*grade) if grade else ""]
    marks.append(f'<span class="badge">{ic("food")}'
                 f'{esc(FOOD_TYPE_LABEL.get(p.entity_type, "식당"))}</span>')
    marks.append(visit_badges(p, rel, meals=True))

    # 추천 메뉴는 그날의 stop 이 들고 있다 (daily-card 의 menu).
    menus = []
    for d, s in VISITS.get(p.slug, []):
        if s.menu and s.menu not in menus:
            menus.append(s.menu)
    menu_html = ""
    if menus:
        menu_html = ('<dt>추천 메뉴</dt><dd>'
                     + "<br>".join(esc(m) for m in menus) + "</dd>")

    rows = [menu_html, price_line(p)]
    for key, label in (("hours", "운영시간"), ("closed", "휴무"),
                       ("booking", "예약")):
        f = p.fact(key)
        if f and f.value:
            mark = "" if f.is_confirmed else " " + badge("caution", "재확인")
            rows.append(f"<dt>{label}</dt><dd>{esc(f.value)}{mark}</dd>")

    actions = []
    href = maps_url(p.lat, p.lng, "", p.map_query or p.name)
    if href:
        actions.append(f'<a class="btn btn-secondary" rel="nofollow noopener" '
                       f'href="{esc(href)}">{ic("map")}지도</a>')
    site = official_url(p)
    if site:
        actions.append(f'<a class="btn btn-secondary" rel="nofollow noopener" '
                       f'href="{esc(site)}">{ic("link")}공식</a>')

    return f"""<article class="card place-card-lg food-card">
  {thumb}
  <div class="card-body">
    <div class="metarow">{''.join(marks)}</div>
    <h3 class="card-title"><a class="card-link" href="{rel}/{p.url}">{esc(p.name)}</a></h3>
    <p class="card-dek">{esc(p.summary)}</p>
    <dl>{''.join(rows)}</dl>
    <div class="btn-row">{''.join(actions)}</div>
  </div>
</article>"""


def acc(title: str, body_md: str) -> str:
    """접이식 한 덩어리. 원고를 버리지 않으면서 화면 밀도를 낮춘다."""
    if not body_md or not body_md.strip():
        return ""
    return ('<details class="acc"><summary>' + esc(title) + '</summary>'
            f'<div class="acc-body prose">{md(strip_tokens(body_md))}'
            "</div></details>")


SCOPE_LABEL = {"region": "광역", "city": "도시", "site": "유적"}


def tourist_map_block(r, rel: str) -> str:
    """관광청 조망지도 — '이 고장에서 뭐가 어디 있나' 를 한 장으로 답한다.

    구글 지도는 목적지를 하나 찍어야 답을 준다. 조망지도는 반대다 — 무엇이
    유명하고 서로 얼마나 떨어져 있는지를 먼저 보여 준다. 현지 관광안내소에서
    받아야 하는 종이를 미리 받아 넣은 것이므로 오프라인에서 열려야 한다.

    이미지를 통째로 링크로 감싼다. 눌러 새 탭에서 열면 브라우저의 확대가
    그대로 듣는다 — 지도에 자바스크립트 확대를 붙이지 않는 이유다.
    """
    if not r.tourist_maps:
        return ""
    cards = []
    for m in r.tourist_maps:
        src = f'{rel}/{m["localPath"]}'
        scope = SCOPE_LABEL.get(m["scope"], "")
        cards.append(
            '<figure class="map-sheet">'
            f'<a href="{src}" target="_blank" rel="noopener">'
            f'<img src="{src}" alt="{esc(m["title"])}" loading="lazy" decoding="async"></a>'
            '<figcaption>'
            f'<div class="metarow"><span class="label">{esc(scope)}</span>'
            f'<strong>{esc(m["place"])}</strong>'
            f'<span>{esc(m["edition"])}</span></div>'
            f'<h3>{esc(m["title"])}</h3>'
            f'<p>{esc(m["usage"])}</p>'
            f'<p class="fine-print"><strong>저작권</strong> · {esc(m["license"])} '
            f'권리자: {esc(m["rightsHolder"])}<br>{esc(m["redistributionBasis"])}</p>'
            '<div class="actions">'
            f'<a class="btn btn-secondary" href="{src}" target="_blank" rel="noopener">크게 보기</a>'
            f'<a class="btn btn-secondary" href="{esc(m["sourceUrl"])}" target="_blank" rel="noopener">원본 내려받기</a>'
            '</div></figcaption></figure>')
    intro = ('<div class="prose"><p>현지 관광안내소가 종이로 나눠 주는 지도다. '
             '목적지를 찍어 찾아가는 지도가 아니라, 이 고장에서 무엇이 유명하고 '
             '서로 얼마나 떨어져 있는지를 한눈에 보여 주는 쪽이다. 눌러서 크게 '
             '열면 확대된다 — 인터넷이 끊겨도 열린다.</p></div>')
    return ('<div class="stack-lg" id="maps">'
            + sec_head("OVERVIEW MAPS", "조망지도", rule=True)
            + intro + "".join(cards) + "</div>")


def build_region(r: Region, trip: Trip) -> str:
    """Region — 지역을 이해하고 고르는 페이지. 상위 섹션은 여섯 개다.

        개요 · 볼거리 · 식당과 카페 · 숙소 · 생활권 · 교통

    FCR-02 에서 바뀐 것:
      · 볼거리와 먹을거리를 **엔티티로** 가른다. 이름에 '점심' 이 들어갔다고
        식당이 아니고, 등급이 '필수' 라고 관광지가 아니다.
      · 일정 섹션을 없앴다. 하루의 정본은 Day 페이지 하나다 — 지역 페이지는
        카드의 방문일 배지와 개요의 날짜 칩으로 그리로 보낸다.
      · '한눈에 보기' · '여행 전체에서의 역할' · '추천 체류 리듬' 은 개요
        안으로 접어 넣었다. 지우지 않았다 — 예상 체류·확정 일정·추천 이유는
        다른 어디에도 없다.
      · 숙박·생활을 숙소와 생활권 두 섹션으로 갈랐다.
      · 교통은 도착·출발 → 도시 교통 → 참고자료 순서 하나로 합쳤다.
    """
    global VISITS
    if not VISITS:
        VISITS = place_visits(trip)
    rel = ".."
    ed = r.editorial
    hero_img = IMAGES["heroes"].get(r.slug) or IMAGES["by_place"].get(r.slug)
    hero = ""
    if hero_img:
        src, srcset = img_src(hero_img, "hero", rel)
        if src:
            hero = (f'<img class="hero-img" src="{src}" srcset="{srcset}" '
                    f'sizes="100vw" alt="{esc(hero_img.get("altKo") or r.name)}" '
                    f'fetchpriority="high" decoding="async">')

    sections = [("overview", "개요")]
    if r.tourist_maps:
        sections.append(("maps", "조망지도"))
    sections += [("attractions", "볼거리"),
                 ("food", "식당·카페"), ("stay", "숙소"),
                 ("life", "생활권"), ("transport", "교통")]
    subnav = tabs_strip([(label, f"#{key}", False) for key, label in sections])

    parts = [f"""<div class="hero">
  {hero}
  <div class="hero-body"><div class="hero-card">
    <div class="hero-eyebrow"><span class="label">{esc(r.tagline)}</span></div>
    <h1>{esc(r.name)}</h1>
    <p class="hero-sub">{esc(r.date_range)} · {r.nights}박 · {esc(r.day_range)}
      · 거점 {esc(r.base)}</p>
    <p class="hero-dek">{esc(r.dek)}</p>
    <div class="btn-row" style="margin-top:1rem">
      <a class="btn btn-primary" href="{rel}/map/{r.slug}.html">{ic("map")}지역 지도</a>
      <a class="btn btn-secondary" href="{rel}/{r.days[0].url}">{ic("today")}첫날 열기</a>
    </div>
  </div></div>
</div>
<div class="wrap">
{credit_line(hero_img)}
<div class="stack-lg" id="overview">"""]

    # ================================================== 1 · 개요
    parts.append(sec_head("OVERVIEW", "개요", rule=True))

    # 날짜 칩 — 일정 섹션을 없앤 자리를 목록이 아니라 색인으로 잇는다.
    # Day 로 가는 길을 끊지 않는다. 시간표는 복제하지 않는다.
    chips = "".join(
        f'<a class="badge badge-day" href="{rel}/{d.url}">'
        f'{esc(d.date_label)} · Day {d.n}</a>' for d in r.days)
    parts.append(f'<div class="day-chips">{chips}</div>')

    if ed.get("verdict"):
        parts.append(f'<div class="prose">{md(strip_tokens(ed["verdict"]))}</div>')
    if ed.get("scenes"):
        parts.append(sec_head("EXPERIENCE", "꼭 경험할 세 장면"))
        parts.append(f'<div class="prose">{md(strip_tokens(ed["scenes"]))}</div>')
    if r.rain_plan:
        parts.append(alert("caution",
                           f"<strong>우천 전환</strong> — {esc(r.rain_plan)}"))
    # 접어 넣은 층. 섹션을 하나씩 세우지 않는다 — 개요 상단이 무거워진다.
    # 지역 사진 중 히어로로 쓰이지 않은 것. 예전에는 카탈로그에 있으면서
    # 화면에 영영 안 나왔다 (barcelona-city-aerial-01 · luberon-valley-01).
    region_photos = []
    other = IMAGES["by_place"].get(r.slug)
    if other and other is not hero_img:
        region_photos.append(other)
    region_photos += IMAGES.get("extras", {}).get(r.slug) or []
    figs = []
    for img in region_photos:
        fig = figure(img, rel, "content", "gallery-shot",
                     "(min-width:600px) 50vw, 100vw")
        if not fig:
            continue
        caption = esc(img.get("captionKo") or img.get("titleKo") or "")
        figs.append(f'<figure class="gallery-item">{fig}<figcaption>'
                    f'{caption}{credit_line(img)}</figcaption></figure>')
    if figs:
        parts.append(f'<div class="grid grid-2">{"".join(figs)}</div>')

    for title, key in (("생략해도 되는 것", "skip"),
                       ("한눈에 보기", "overview"),
                       ("여행 전체에서의 역할", "role"),
                       ("추천 체류 리듬", "rhythm"),
                       ("이 지역을 이해하는 층", "context")):
        parts.append(acc(layer_title(r.slug, key, title), ed.get(key, "")))
    parts.append("</div>")

    # ================================================== 1.5 · 조망지도
    parts.append(tourist_map_block(r, rel))

    # ================================================== 2 · 볼거리
    parts.append('<div class="stack-lg" id="attractions">')
    must, rec = r.must_visit, r.recommended
    if must:
        parts.append(sec_head("ATTRACTIONS · MUST VISIT", "꼭 가야 할 곳",
                              rule=True))
        big, small = must[:6], must[6:]
        parts.append('<div class="grid grid-2">'
                     + "".join(attraction_card(p, rel, large=True) for p in big)
                     + "</div>")
        if small:
            parts.append('<div class="grid grid-2">'
                         + "".join(attraction_card(p, rel) for p in small)
                         + "</div>")
    if rec:
        parts.append(sec_head("ATTRACTIONS · RECOMMENDED", "권할 만한 곳"))
        parts.append('<div class="grid grid-2">'
                     + "".join(attraction_card(p, rel) for p in rec) + "</div>")
    if not (must or rec):
        parts.append(alert("info", "이 지역의 장소 카드는 아직 준비 중이다.", "pin"))
    parts.append("</div>")

    # ================================================== 3 · 식당과 카페
    parts.append('<div class="stack-lg" id="food">')
    parts.append(sec_head("RESTAURANTS & CAFÉS", "식당과 카페", rule=True))
    food = r.food_places
    if food:
        # 하위 묶음 셋. 명부의 entity_type 을 그대로 쓴다 — 화면에서 나누되
        # 정본 분류를 새로 만들지 않는다. 한 묶음뿐이면 제목을 달지 않는다.
        groups = [
            ("RESTAURANTS", "식당", ("restaurant", "wine-bar")),
            ("CAFÉS", "카페", ("cafe",)),
            ("BAKERY · MARKET · FOOD HALL", "빵집·시장·푸드홀",
             ("bakery", "market", "food-hall")),
        ]
        filled = [(label, title, [p for p in food if p.entity_type in kinds])
                  for label, title, kinds in groups]
        filled = [g for g in filled if g[2]]
        for label, title, places in filled:
            if len(filled) > 1:
                parts.append(sec_head(label, title))
            parts.append('<div class="grid grid-2">'
                         + "".join(food_card(p, rel, trip) for p in places)
                         + "</div>")
    else:
        parts.append(alert("info",
                           "이 구간에는 예약·확정된 식당이 없다. 끼니는 그날의 "
                           "Day 페이지가 정본이다.", "food"))
    # 요리 사진. 장소가 아니라 요리를 찍은 것이라 어느 장소 카드에도 붙일 수
    # 없다 — 붙이면 그 가게 사진처럼 읽힌다. 지역의 음식 자리에 제 이름으로
    # 나온다.
    dish_photos = IMAGES.get("dishes", {}).get(r.slug) or []
    if dish_photos:
        figs = []
        for img in dish_photos:
            fig = figure(img, rel, "content", "gallery-shot",
                         "(min-width:600px) 33vw, 50vw")
            if not fig:
                continue
            figs.append(f'<figure class="gallery-item">{fig}<figcaption>'
                        f'<strong>{esc(img.get("dishLabel") or "")}</strong>'
                        f'{credit_line(img)}</figcaption></figure>')
        if figs:
            parts.append(f'<div class="grid grid-3">{"".join(figs)}</div>')

    dishes = region_dishes(r)
    if dishes:
        parts.append('<details class="acc"><summary>이 지역에서 먹는 것</summary>'
                     '<div class="acc-body prose"><ul>'
                     + "".join(f"<li>{link_food_text(x, rel, trip)}</li>"
                               for x in dishes)
                     + "</ul></div></details>")
    parts.append(acc(layer_title(r.slug, "food_culture", "이 지역의 음식과 시장"),
                     ed.get("food_culture", "")))
    parts.append("</div>")

    # ================================================== 4 · 숙소
    # 그 지역에서 **자는** 날의 숙소만 싣는다. 이동일은 두 지역에 걸쳐 있어
    # 그냥 모으면 다음 거점의 숙소가 이 지역 날짜를 달고 나타난다.
    parts.append('<div class="stack-lg" id="stay">')
    parts.append(sec_head("ACCOMMODATION", "숙소", rule=True))
    hotels = {d.hotel.get("name"): d.hotel for d in r.days
              if d.region == r.slug and d.hotel.get("name")}
    if hotels:
        cards = []
        for name, h in hotels.items():
            confirmed = h.get("status") == "confirmed"
            mark = badge("ok", "확정") if confirmed else badge("caution", "미확정")
            href = maps_url(h.get("lat"), h.get("lng"), h.get("address") or "",
                            HOTEL_QUERIES.get(name, ""))
            link = (f'<a class="btn btn-secondary" rel="nofollow noopener" '
                    f'href="{esc(href)}">{ic("map")}지도</a>') if href else ""
            addr = (f'<dt>주소</dt><dd>{esc(h["address"])}</dd>'
                    if h.get("address") else "")
            cards.append(f"""<article class="card booking-card">
  <div class="booking-head"><span class="booking-name">{esc(name)}</span>{mark}</div>
  <dl><dt>체크인</dt><dd>{esc(r.checkin.isoformat())}</dd>
      <dt>체크아웃</dt><dd>{esc(r.checkout.isoformat())}</dd>
      <dt>박수</dt><dd>{r.nights}박</dd>{addr}</dl>
  <div class="btn-row">{link}</div>
</article>""")
        parts.append(f'<div class="grid grid-2">{"".join(cards)}</div>')
    if not r.has_confirmed_stay:
        parts.append(alert("caution",
                           "<strong>숙소 미확정</strong> — 확정되면 여기에 표시된다. "
                           "확정 전 주소를 믿고 이동하지 않는다. 아래는 "
                           "'어느 동네에 묵을 것인가' 까지만 말한다.", "stay"))
    if r.essentials.get("staySummary"):
        parts.append(f'<div class="prose"><p>'
                     f'{esc(r.essentials["staySummary"])}</p></div>')
    parts.append(acc(layer_title(r.slug, "neighborhoods", "동네와 생활권"),
                     ed.get("neighborhoods", "")))
    parts.append(acc(layer_title(r.slug, "stay_budget", "숙소 예산과 확인 기준"),
                     ed.get("stay_budget", "")))
    parts.append("</div>")

    # ================================================== 5 · 생활권
    parts.append('<div class="stack-lg" id="life">')
    parts.append(sec_head("LOCAL LIFE", "생활권", rule=True))
    life = r.essentials.get("lifeEssentials") or []
    if life or r.essentials.get("lateReturnRule"):
        parts.append('<div class="prose"><ul>'
                     + "".join(f"<li>{esc(x)}</li>" for x in life) + "</ul>")
        if r.essentials.get("lateReturnRule"):
            parts.append(f'<p><strong>늦은 귀가</strong> — '
                         f'{esc(r.essentials["lateReturnRule"])}</p>')
        parts.append("</div>")
    else:
        parts.append(alert("info", "이 지역의 생활 정보는 아직 정리하지 않았다.",
                           "check"))
    parts.append("</div>")

    # ================================================== 6 · 교통
    parts.append('<div class="stack-lg" id="transport">')
    parts.append(sec_head("TRANSPORT", "교통", rule=True))
    arrive, leave = r.days[0], r.days[-1]
    parts.append(sec_head("ARRIVAL · DEPARTURE", "도착과 출발"))
    ess = r.essentials
    arr_body = (f'<p>{esc(ess["arrivalStrategy"])}</p>'
                if ess.get("arrivalStrategy") else "")
    dep_body = (f'<p>{esc(ess["departureStrategy"])}</p>'
                if ess.get("departureStrategy") else "")
    parts.append(f'''<div class="grid grid-2">
<article class="card"><div class="card-body"><h3>도착</h3>
  {arr_body}
  <a class="btn btn-secondary" href="{rel}/{arrive.url}">
    {esc(arrive.date_label)} · Day {arrive.n} · {esc(arrive.city)}</a>
</div></article>
<article class="card"><div class="card-body"><h3>출발</h3>
  {dep_body}
  <a class="btn btn-secondary" href="{rel}/{leave.url}">
    {esc(leave.date_label)} · Day {leave.n} · {esc(leave.city)}</a>
</div></article></div>''')

    transit = r.transit
    if transit:
        rec_t = transit["recommendation"]
        # 라벨이 '도시 교통' 이면 Girona 처럼 대중교통이 아예 없는 구간에서
        # 거짓말이 된다. 슬롯은 같고 이름만 사실에 맞춘다.
        parts.append(sec_head("GETTING AROUND", "구간 내 이동"))
        parts.append(alert("info", f'<strong>{esc(rec_t["title"])}</strong> — '
                           f'{esc(rec_t["summary"])}', "train"))
        products = transit.get("products") or []
        if products:
            product_cards = []
            for product in products:
                shared = "공동 사용" if product.get("shared") else "1인용"
                product_cards.append(f'''<article class="card"><div class="card-body">
  <h3>{esc(product.get("name"))}</h3>
  <div class="metarow"><strong>{esc(product.get("price"))}</strong><span>{shared}</span><span>{esc(product.get("accessNote"))}</span></div>
  <p>{esc(product.get("fit"))}</p>
</div></article>''')
            parts.append(f'<div class="grid grid-3">{"".join(product_cards)}</div>')
        parts.append('<div class="grid grid-2"><div class="prose"><h3>이용법</h3><ul>'
                     + "".join(f'<li>{esc(x)}</li>' for x in transit["howToUse"])
                     + '</ul></div><div class="prose"><h3>적용되지 않는 이동·예외</h3><ul>'
                     + "".join(f'<li>{esc(x)}</li>' for x in transit["exceptions"])
                     + '</ul></div></div>')
        uses = transit.get("itineraryUses") or []
        if uses:
            day_by_number = {day.n: day for day in trip.days}
            parts.append('<div class="prose"><h3>이 일정에서 쓰는 교통</h3><ul>'
                         + "".join(
                             f'<li><a href="{rel}/{day_by_number[x["day"]].url}">{esc(day_by_number[x["day"]].date_label)} · Day {x["day"]}</a> — {esc(x["label"])}</li>'
                             for x in uses) + '</ul></div>')
    parts.append(acc(layer_title(r.slug, "transport_deep", "이 지역에서 이동하기"),
                     ed.get("transport_deep", "")))

    # --- References — 누를 것만 모은다. 긴 설명은 위에서 이미 했다. -------
    refs = []
    for src in (transit.get("sources") or []) if transit else []:
        refs.append(f'<li><a href="{esc(src["url"])}" target="_blank" '
                    f'rel="noopener">{esc(src["label"])}</a> · 확인 '
                    f'{esc(src["verifiedAt"])} · 재확인 {esc(src["recheckBy"])}</li>')
    resource_cards = []
    for resource in r.transport_resources:
        local_path = resource.get("localPath")
        primary = ""
        if local_path:
            asset_rel = local_path.removeprefix("source/ASSETS/")
            primary = (f'<a class="btn btn-secondary" href="{rel}/assets/'
                       f'{esc(asset_rel)}" target="_blank">PDF 열기</a>')
        rights = ""
        if resource.get("rightsHolder"):
            rights = (f'<p class="fine-print"><strong>저작권</strong> · '
                      f'{esc(resource["license"])} 권리자: '
                      f'{esc(resource["rightsHolder"])}<br>'
                      f'{esc(resource["redistributionBasis"])}</p>')
        resource_cards.append(f'''<article class="card"><div class="card-body">
  <h3>{esc(resource["title"])}</h3>
  <div class="metarow"><span>{esc(resource["edition"])}</span></div>
  <p>{esc(resource["usage"])}</p>
  {rights}
  <div class="actions">{primary}<a class="btn btn-secondary" href="{esc(resource["officialUrl"])}" target="_blank" rel="noopener">권리자 사이트</a></div>
</div></article>''')
    if refs or resource_cards:
        parts.append(sec_head("REFERENCES", "공식 자료와 재확인"))
        if refs:
            parts.append(f'<div class="prose"><ul>{"".join(refs)}</ul></div>')
        if resource_cards:
            parts.append(f'<div class="grid grid-2">{"".join(resource_cards)}</div>')
    parts.append("</div>")

    parts.append("</div>")

    index_search(r.name, f"guide/{r.slug}.html", "region", r.tagline)

    return page(
        title=r.name, body="\n".join(parts), rel=rel, tab="guide",
        region=r.slug, country=r.country, subnav=subnav, description=r.dek,
        trail=[("홈", "index.html"), ("가이드", "guide/index.html"),
               (r.name, None)],
    )


def build_guide_index(trip: Trip) -> str:
    """가이드 — 8개 지역을 이동 순서대로. 축은 목록을 맡고 본문을 갖지 않는다."""
    rel = ".."   # guide/index.html 은 하위 디렉터리에 있다
    cards = []
    for r in trip.regions:
        img = IMAGES["heroes"].get(r.slug)
        thumb = figure(img, rel, "content", "thumb",
                       "(min-width:600px) 50vw, 100vw")
        cards.append(f"""<article class="card place-card-lg" data-region="{r.slug}">
  {thumb}
  <div class="card-body">
    <div class="metarow"><span class="label" style="color:var(--accent)">
      {esc(r.tagline)}</span></div>
    <h3 class="card-title"><a class="card-link" href="{r.slug}.html">
      {esc(r.name)}</a></h3>
    <p class="card-dek">{esc(r.dek)}</p>
    <div class="metarow"><span>{esc(r.date_range)}</span><span class="sep">·</span>
      <span>{r.nights}박</span><span class="sep">·</span>
      <span>{esc(r.day_range)}</span></div>
  </div>
</article>""")
    body = f"""<div class="wrap"><div class="stack-lg" style="padding-top:1.5rem">
<header>
  <h1>가이드</h1>
  <p class="hero-dek">8개 거점을 이동 순서대로 놓았다. 각 지역에서 무엇을 보고
    먹고 경험할지는 지역 페이지가 맡는다.</p>
</header>
<div class="grid grid-2">{''.join(cards)}</div>
</div></div>"""
    return page(title="가이드", body=body, rel=rel, tab="guide",
                description="8개 지역 가이드",
                trail=[("홈", "index.html"), ("가이드", None)])


def build_schedule(trip: Trip) -> str:
    """전체 일정 — 43일. 하루의 본문은 Day 가 갖는다, 여기는 목록이다."""
    rel = "."
    blocks = []
    for r in trip.regions:
        own = [d for d in r.days if d.region == r.slug]
        if not own:
            continue
        blocks.append(f'<div id="{r.slug}" data-region="{r.slug}">'
                      + sec_head(r.tagline, r.name,
                                 more=("지역 가이드", f"guide/{r.slug}.html"),
                                 rule=False) + "</div>")
        blocks.append('<div class="grid grid-2">'
                      + "".join(day_card(d, rel, r) for d in own) + "</div>")
    last = trip.day(43)
    if last and last.region == "return":
        blocks.append(sec_head("", "귀국"))
        blocks.append(f'<div class="grid grid-2">{day_card(last, rel)}</div>')

    jump = tabs_strip([(r.name, f"#{r.slug}", False) for r in trip.regions])
    regions_payload = json.dumps([{
        "slug": r.slug, "name": r.name,
        "start": r.days[0].date.isoformat(),
        "end": r.days[-1].date.isoformat()
    } for r in trip.regions], ensure_ascii=False)

    body = f"""<div class="wrap"><div class="stack-lg" style="padding-top:1.5rem">
<header>
  <h1>전체 일정</h1>
  <p class="hero-dek">{trip.start.isoformat()} — {trip.end.isoformat()} ·
    43일 42박 · 8개 거점</p>
</header>
{''.join(blocks)}
<script type="application/json" id="schedule-regions-data">{regions_payload}</script>
</div></div>"""
    return page(title="전체 일정", body=body, rel=rel, tab="schedule",
                subnav=jump, description="43일 전체 일정",
                bar_title="전체 일정",
                trail=[("홈", "index.html"), ("전체 일정", None)])


def build_home(trip: Trip, res: dict) -> str:
    """홈 — 중심은 Today.

    여행 전에는 준비, 여행 중에는 오늘이 첫 화면이다. 정적 사이트라
    빌드 시각에 모드를 굳히지 않고 브라우저가 오늘 날짜로 고른다 —
    출발 전에 빌드한 페이지가 여행 중에도 맞아야 한다.
    """
    rel = "."
    days_payload = [{
        "n": d.n, "date": d.date.isoformat(), "date_label": d.date_label,
        "city": d.city, "title": d.title,
        "url": d.url, "region": d.region,
        "next": [{"t": s.start, "n": s.name,
                  "u": f"places/{s.place.slug}.html" if s.place else None}
                 for s in d.stops if s.start and s.category != "hotel"][:4],
    } for d in trip.days]

    # 여행 전 화면 — 아직 예약하지 않은 것부터 보여준다.
    # 예약목표일이 가까운 순으로 다섯 개만 — 홈은 요약이고, 전체는 준비 화면이다.
    todo = res.get("todo", [])
    done = res.get("confirmed", [])
    pre_items = "".join(
        f'<li>{esc(r["예약항목"])}'
        + (f' <span class="meta">목표 {esc(r["예약목표일"])}</span>'
           if r["예약목표일"] else "") + "</li>"
        for r in todo[:5])

    region_cards = "".join(f"""<article class="card" data-region="{r.slug}"
    style="min-width:0">
  <div class="card-body">
    <span class="label" style="color:var(--accent)">{esc(r.day_range)}</span>
    <h3 class="card-title"><a class="card-link" href="guide/{r.slug}.html">
      {esc(r.name)}</a></h3>
    <p class="card-dek">{esc(r.tagline)}</p>
    <div class="metarow"><span>{esc(r.date_range)}</span><span class="sep">·</span>
      <span>{r.nights}박</span></div>
  </div>
</article>""" for r in trip.regions)

    body = f"""<div class="wrap"><div class="stack-lg" style="padding-top:1.5rem">

<section id="today-panel" class="stack" aria-live="polite">
  <noscript><p class="meta">오늘 화면은 기기의 날짜로 고릅니다.
    <a href="schedule.html">전체 일정</a>을 여세요.</p></noscript>
</section>

<section>
  {sec_head("JOURNEY", "여정", more=("전체 일정", "schedule.html"), rule=True)}
  <div class="scroll-x journey-grid">{region_cards}</div>
</section>

<section>
  {sec_head("PREPARE", "준비", more=("준비 화면", "prepare/index.html"), rule=True)}
  <div class="card"><div class="card-body">
    <div class="metarow">
      <span>{ic("check")}확정 {len(done)}건</span>
      <span class="sep">·</span>
      <span>{badge('caution', f'미예약 {len(todo)}건')}</span>
    </div>
    {f'<div class="prose"><ul>{pre_items}</ul></div>' if pre_items else ''}
  </div></div>
</section>

<section>
  {sec_head("TOOLS", "빠른 도구", rule=True)}
  <nav class="quick" aria-label="빠른 도구">
    <a href="#" id="quick-search">{ic("search")}<span>검색</span></a>
    <a href="map/index.html">{ic("map")}<span>지도</span></a>
    <a href="schedule.html">{ic("list")}<span>전체 일정</span></a>
    <a href="prepare/emergency.html">{ic("alert")}<span>긴급</span></a>
  </nav>
</section>

</div></div>
<script type="application/json" id="trip-data">{json.dumps(
    {"start": trip.start.isoformat(), "end": trip.end.isoformat(),
     "days": days_payload}, ensure_ascii=False)}</script>"""

    return page(title="오늘", body=body, rel=rel, tab="today",
                bar_title="2026년 유럽여행 가이드",
                description=f"{SITE_TITLE} — 43일 여행 가이드")


def build_map_pages(trip: Trip) -> dict[str, str]:
    """지도 — Trip · Region · Day 세 수준. 핀은 Place DB 에서만 온다."""
    out = {}
    rel = ".."

    def stop_region_slug(d, s):
        if s.place:
            return s.place.region
        # Geographic and semantic routing for non-place stops and routes
        if "bcn" in s.id or "barcelona" in s.id or "sitges" in s.id or "can-robert" in s.id:
            return "barcelona"
        if "girona" in s.id or "cadaques" in s.id or "tossa" in s.id or "sant-feliu" in s.id:
            return "girona"
        if "nice" in s.id or "cannes" in s.id or "antibes" in s.id or "monaco" in s.id or "menton" in s.id or "eze" in s.id or "villefranche" in s.id:
            return "nice"
        if "moustiers" in s.id or "verdon" in s.id or "palud" in s.id or "galetas" in s.id or "valensole" in s.id or "castellane" in s.id or "sublime" in s.id or "cretes" in s.id:
            return "verdon"
        if "aix" in s.id or "marseille" in s.id or "cassis" in s.id or "calanques" in s.id or "vallon" in s.id:
            return "aix"
        if "luberon" in s.id or "gordes" in s.id or "roussillon" in s.id or "farm" in s.id or "bories" in s.id:
            return "luberon"
        if "avignon" in s.id or "arles" in s.id or "uzes" in s.id or "gard" in s.id:
            return "avignon"
        if "lyon" in s.id or "annecy" in s.id or "funicular" in s.id or "saone" in s.id or "rosaire" in s.id or "part-dieu" in s.id:
            return "lyon"
        if ("paris" in s.id or "versailles" in s.id or "longchamp" in s.id or "cdg" in s.id or
                "tuileries" in s.id or "palais-royal" in s.id or "opera" in s.id or "invalides" in s.id or
                "champs-elysees" in s.id or "cour-carree" in s.id or "ranelagh" in s.id or "prix-de-l-arc" in s.id or
                "parc-monceau" in s.id or "iena" in s.id or "trocadero" in s.id or "first-grocery" in s.id or
                "city-bus-tour" in s.id or "rue-du-bac" in s.id):
            return "paris"
        if len(d.regions) > 1:
            intercity_stops = {"vy1521": "nice", "tgv-to-lyon": "lyon", "part-dieu": "lyon", "tgv-to-paris": "paris"}
            if s.id in intercity_stops:
                return intercity_stops[s.id]
        return d.region

    region_zoom = {
        "barcelona": 12,
        "girona": 9,
        "nice": 11,
        "aix": 10,
        "luberon": 11,
        "avignon": 10,
        "lyon": 12,
        "paris": 12,
    }

    # 1. Whole trip map page: 8 sequential region sections (Region name -> Map -> Numbered List)
    seen = set()
    region_sections = []
    total_stops_count = 0

    for r in trip.regions:
        r_stops = []
        for d in trip.days:
            for s in d.stops:
                reg_slug = stop_region_slug(d, s)
                if reg_slug != r.slug:
                    continue
                item_key = s.place.slug if s.place else s.id
                if item_key in seen:
                    continue
                if getattr(s, "map_type", "place") == "non_map":
                    continue
                if s.id in {"cdg-departure", "inflight", "icn", "bcn-airport", "paris-return"}:
                    continue
                if s.category == "hotel" and getattr(s, "map_type", "place") != "route":
                    continue
                if not ((s.lat and s.lng) or s.address):
                    continue
                seen.add(item_key)
                r_stops.append((d.date, s.start or "99:99", s.order, s))

        r_stops_sorted = [x[3] for x in sorted(r_stops, key=lambda x: (x[0], x[1], x[2]))]
        total_stops_count += len(r_stops_sorted)
        zoom = region_zoom.get(r.slug, 12)
        card_html = map_card(r_stops_sorted, rel, zoom=zoom, label=f"{r.name} 지도", numbered=True)
        region_sections.append(f"""<section class="region-map-section" id="section-{r.slug}">
{sec_head("", r.name)}
{card_html}
</section>""")

    out["index.html"] = page(
        title="지도", rel=rel, tab="map",
        trail=[("홈", "index.html"), ("지도", None)],
        body=f"""<div class="wrap"><div class="stack-lg" style="padding-top:1.5rem">
<header><h1>지도</h1>
<p class="hero-dek">전체 여정 8개 지역 {total_stops_count}곳. 각 지역 지도와 번호 매겨진 장소 목록.</p></header>
{"".join(region_sections)}
</div></div>""")

    # 2. Regional maps (map/{r.slug}.html)
    for r in trip.regions:
        r_seen = set()
        r_stops = []
        for d in trip.days:
            for s in d.stops:
                reg_slug = stop_region_slug(d, s)
                if reg_slug != r.slug:
                    continue
                item_key = s.place.slug if s.place else s.id
                if item_key in r_seen:
                    continue
                if getattr(s, "map_type", "place") == "non_map":
                    continue
                if s.id in {"cdg-departure", "inflight", "icn", "bcn-airport", "paris-return"}:
                    continue
                if s.category == "hotel" and getattr(s, "map_type", "place") != "route":
                    continue
                if not ((s.lat and s.lng) or s.address):
                    continue
                r_seen.add(item_key)
                r_stops.append((d.date, s.start or "99:99", s.order, s))

        r_stops_sorted = [x[3] for x in sorted(r_stops, key=lambda x: (x[0], x[1], x[2]))]
        zoom = region_zoom.get(r.slug, 12)
        day_links = "".join(
            f'<li><a href="../{d.url}">{esc(d.date_label)} · Day {d.n}</a> — '
            f"{esc(d.title)}</li>" for d in r.days)
        out[f"{r.slug}.html"] = page(
            title=f"{r.name} 지도", rel=rel, tab="map", region=r.slug,
            country=r.country,
            trail=[("홈", "index.html"), ("지도", "map/index.html"),
                   (r.name, None)],
            body=f"""<div class="wrap"><div class="stack-lg" style="padding-top:1.5rem">
<header><h1>{esc(r.name)} 지도</h1>
<p class="hero-dek">{esc(r.date_range)} · 장소 {len(r_stops_sorted)}곳</p></header>
{map_card(r_stops_sorted, rel, zoom=zoom, label=f"{r.name} 지도", numbered=True)}
{sec_head("", "날짜별 동선")}
<div class="prose"><ul>{day_links}</ul></div>
<div class="btn-row"><a class="btn btn-secondary" href="../guide/{r.slug}.html">
  {ic("region")}{esc(r.name)} 가이드</a></div>
</div></div>""")
    return out


def res_card(rec: dict, *, todo: bool = False) -> str:
    """예약 하나. 목록이 아니라 **현장에서 쓰는 카드**다.

    렌터카 카운터에서 필요한 것은 항목 이름이 아니라 예약번호이고, 체크인
    때 필요한 것은 주소다. 그래서 접지 않고 펼쳐 둔다.
    """
    mark = badge("caution", "미예약") if todo else badge("ok", "확정")
    when = rec["날짜"] + (f" {rec['시간']}" if rec["시간"] else "")
    money = (f"{rec['총액']} {rec['통화']}"
             if rec["총액"] and rec["총액"] != "0" else "")

    facts = []
    if when:
        facts.append(("일시", when))
    if rec["사업자"]:
        facts.append(("사업자", rec["사업자"]))
    if rec["예약번호"] and not todo:
        facts.append(("예약번호", rec["예약번호"]))
    if money:
        facts.append(("금액", money))
    if rec["주소/역"]:
        facts.append(("장소", rec["주소/역"]))
    if todo and rec["예약목표일"]:
        facts.append(("예약 목표", rec["예약목표일"]))
    if rec["무료취소기한"]:
        facts.append(("무료취소", rec["무료취소기한"]))

    rows = "".join(f"<dt>{esc(k)}</dt><dd>{linkify(esc(v))}</dd>" for k, v in facts)
    # 리스크는 접어 둔다 — 평소엔 방해가 되고, 문제가 생겼을 때만 필요하다
    extra = ""
    detail = " · ".join(x for x in (rec["리스크/대체안"], rec["비고"]) if x)
    if detail:
        extra = (f'<details class="acc"><summary>주의·메모</summary>'
                 f'<div class="acc-body"><p class="card-dek">{linkify(esc(detail))}'
                 f"</p></div></details>")
    region = f'<span class="meta">{esc(rec["지역"])}</span>' if rec["지역"] else ""
    return f"""<article class="card booking-card">
  <div class="booking-head">
    <span class="booking-name">{esc(rec["예약항목"])}</span>{mark}</div>
  <div class="metarow">{esc(rec["카테고리"])}{region}</div>
  <dl>{rows}</dl>
  {extra}
</article>"""


# ================================================================ 여행 프랑스어

DAY_FRENCH_MAP: dict[int, list[str]] = {
    7: ["fr_transport_003", "fr_market_013", "fr_hotel_001"],
    8: ["fr_market_001", "fr_market_011", "fr_restaurant_018"],
    9: ["fr_transport_006", "fr_sightseeing_002", "fr_restaurant_008"],
    10: ["fr_transport_005", "fr_restaurant_004", "fr_shopping_005"],
    11: ["fr_transport_008", "fr_hotel_001", "fr_restaurant_001"],
    12: ["fr_market_003", "fr_sightseeing_009", "fr_restaurant_013"],
    13: ["fr_transport_002", "fr_sightseeing_007", "fr_restaurant_009"],
    14: ["fr_driving_006", "fr_sightseeing_003", "fr_restaurant_016"],
    15: ["fr_market_002", "fr_restaurant_017", "fr_shopping_001"],
    16: ["fr_driving_001", "fr_market_010", "fr_driving_008"],
    17: ["fr_driving_009", "fr_sightseeing_001", "fr_market_005"],
    18: ["fr_driving_005", "fr_hotel_001", "fr_restaurant_015"],
    19: ["fr_sightseeing_002", "fr_sightseeing_010", "fr_restaurant_006"],
    20: ["fr_driving_010", "fr_driving_012", "fr_sightseeing_006"],
    21: ["fr_driving_007", "fr_sightseeing_008", "fr_market_012"],
    22: ["fr_driving_014", "fr_sightseeing_001", "fr_restaurant_007"],
    23: ["fr_driving_004", "fr_transport_005", "fr_hotel_010"],
    24: ["fr_transport_002", "fr_restaurant_010", "fr_market_004"],
    25: ["fr_market_003", "fr_market_009", "fr_restaurant_020"],
    26: ["fr_transport_011", "fr_sightseeing_007", "fr_restaurant_004"],
    27: ["fr_hotel_004", "fr_transport_013", "fr_hotel_001"],
    28: ["fr_transport_002", "fr_sightseeing_002", "fr_restaurant_001"],
    29: ["fr_restaurant_017", "fr_sightseeing_009", "fr_shopping_003"],
    30: ["fr_market_001", "fr_shopping_006", "fr_restaurant_018"],
    31: ["fr_sightseeing_004", "fr_restaurant_013", "fr_essential_018"],
    32: ["fr_sightseeing_002", "fr_sightseeing_008", "fr_restaurant_015"],
    33: ["fr_sightseeing_002", "fr_sightseeing_010", "fr_essential_014"],
    34: ["fr_transport_007", "fr_sightseeing_006", "fr_restaurant_008"],
    35: ["fr_sightseeing_007", "fr_restaurant_004", "fr_market_014"],
    36: ["fr_sightseeing_001", "fr_restaurant_006", "fr_shopping_005"],
    37: ["fr_transport_014", "fr_sightseeing_002", "fr_restaurant_012"],
    38: ["fr_market_002", "fr_market_015", "fr_restaurant_019"],
    39: ["fr_sightseeing_009", "fr_shopping_007", "fr_restaurant_018"],
    40: ["fr_sightseeing_002", "fr_restaurant_008", "fr_restaurant_020"],
    41: ["fr_sightseeing_002", "fr_restaurant_002", "fr_restaurant_020"],
    42: ["fr_hotel_004", "fr_shopping_007", "fr_transport_014"],
    43: ["fr_essential_004", "fr_essential_020"],
}

PLACE_CATEGORY_FRENCH_MAP: dict[str, list[str]] = {
    "restaurant": ["fr_restaurant_001", "fr_restaurant_013", "fr_restaurant_018"],
    "cafe": ["fr_restaurant_017", "fr_market_014", "fr_restaurant_019"],
    "bakery": ["fr_market_001", "fr_market_002", "fr_market_013"],
    "market": ["fr_market_003", "fr_market_011", "fr_market_012"],
    "food-hall": ["fr_market_009", "fr_market_014", "fr_restaurant_019"],
    "wine-bar": ["fr_restaurant_015", "fr_restaurant_016", "fr_restaurant_018"],
    "attraction": ["fr_sightseeing_002", "fr_sightseeing_007", "fr_sightseeing_009"],
    "walk": ["fr_essential_018", "fr_essential_019", "fr_sightseeing_007"],
    "transport-node": ["fr_transport_003", "fr_transport_005", "fr_transport_014"],
}

FRENCH_CATEGORY_LABEL: dict[str, str] = {
    "essential": "기본표현",
    "restaurant": "식당·카페",
    "market": "빵집·시장",
    "hotel": "숙소",
    "transport": "기차·교통",
    "driving": "렌터카·주차",
    "sightseeing": "관광·미술관",
    "shopping": "쇼핑",
    "emergency": "긴급상황",
}


def phrase_card(p: FrenchPhrase, compact: bool = False) -> str:
    cat_label = FRENCH_CATEGORY_LABEL.get(p.category, p.category)
    priority_badge = f'<span class="badge badge-must">P0</span>' if p.priority == "P0" else ""
    cat_badge = f'<span class="badge badge-neutral">{cat_label}</span>'
    hint_html = f'<p class="phrase-hint">{esc(p.pronunciation_hint)}</p>' if p.pronunciation_hint else ""
    note_html = f'<p class="phrase-note">{esc(p.usage_note)}</p>' if p.usage_note and not compact else ""
    search_data = f'{p.fr} {p.ko} {p.pronunciation_hint} {" ".join(p.tags)}'.lower()

    return f"""<article class="phrase-card" data-phrase-id="{p.id}" data-category="{p.category}" data-priority="{p.priority}" data-search="{esc(search_data)}">
  <div class="phrase-head">
    <div>{cat_badge} {priority_badge}</div>
  </div>
  <p class="phrase-fr">{esc(p.fr)}</p>
  <p class="phrase-ko">{esc(p.ko)}</p>
  {hint_html}
  {note_html}
  <div class="phrase-actions">
    <button type="button" class="phrase-btn btn-phrase-audio" data-audio="{esc(p.audio_text or p.fr)}" aria-label="발음 듣기" title="발음 듣기">
      {ic('sound')}<span>듣기</span>
    </button>
    <button type="button" class="phrase-btn btn-phrase-copy" data-copy="{esc(p.fr)}" aria-label="문구 복사" title="문구 복사">
      {ic('copy')}<span>복사</span>
    </button>
    <button type="button" class="phrase-btn btn-phrase-fav" data-fav-id="{p.id}" aria-label="즐겨찾기" title="즐겨찾기">
      {ic('star')}<span>저장</span>
    </button>
  </div>
</article>"""


def day_quick_french(d: Day, trip: Trip, rel: str) -> str:
    if d.n < 7 or (d.region in ("barcelona", "girona") and d.n <= 6):
        return ""
    phrase_ids = DAY_FRENCH_MAP.get(d.n, [])
    if not phrase_ids:
        return ""
    selected = [trip.french_phrases[pid] for pid in phrase_ids if pid in trip.french_phrases]
    if not selected:
        return ""
    cards = "".join(phrase_card(p, compact=True) for p in selected)
    return f"""<section class="quick-french-box">
  <div class="quick-french-head">
    <h3 class="quick-french-title">{ic('chat')} 오늘 현장 프랑스어 (Quick French)</h3>
    <a href="{rel}/prepare/french.html" class="meta">전체 120개 회화 →</a>
  </div>
  <div class="grid grid-2">
    {cards}
  </div>
</section>"""


def place_quick_french(p: Place, trip: Trip, rel: str) -> str:
    if p.region in ("barcelona", "girona"):
        return ""
    cat = p.entity_type
    phrase_ids = PLACE_CATEGORY_FRENCH_MAP.get(cat) or ["fr_essential_001", "fr_essential_003", "fr_essential_005"]
    selected = [trip.french_phrases[pid] for pid in phrase_ids if pid in trip.french_phrases]
    if not selected:
        return ""
    cards = "".join(phrase_card(ph, compact=True) for ph in selected[:3])
    return f"""<section class="quick-french-box">
  <div class="quick-french-head">
    <h3 class="quick-french-title">{ic('chat')} 현장 프랑스어 (Quick French)</h3>
    <a href="{rel}/prepare/french.html" class="meta">전체 120개 회화 →</a>
  </div>
  <div class="grid grid-2">
    {cards}
  </div>
</section>"""


def build_travel_french(trip: Trip) -> str:
    rel = ".."
    phrases = list(trip.french_phrases.values())
    guide = trip.french_guide or {}

    pron_blocks = []
    for section in guide.get("pronunciation_rules", []):
        rows = []
        for r in section.get("rules", []):
            rows.append(f"""<tr>
  <td><strong>{esc(r['pattern'])}</strong></td>
  <td><span class="sound">{esc(r['sound'])}</span></td>
  <td><span class="ex">{esc(r['example'])}</span></td>
  <td><span class="meta">{esc(r['note'])}</span></td>
</tr>""")
        pron_blocks.append(f"""<div class="card" style="margin-bottom:var(--s3)">
  <div class="card-body">
    <h3 style="margin-top:0">{esc(section['title'])}</h3>
    <div class="table-wrap"><table class="rule-table">
      <thead><tr><th>철자 패턴</th><th>한국어 발음 근사</th><th>대표 단어</th><th>발음 포인트</th></tr></thead>
      <tbody>{''.join(rows)}</tbody>
    </table></div>
  </div>
</div>""")
    pron_html = "".join(pron_blocks)

    signs_blocks = []
    for cat in guide.get("signs_and_menu", []):
        rows = []
        for w in cat.get("words", []):
            rows.append(f"""<tr>
  <td><strong style="color:var(--primary)">{esc(w['fr'])}</strong></td>
  <td>{esc(w['ko'])}</td>
  <td><span class="meta">{esc(w.get('pronunciation', ''))}</span></td>
</tr>""")
        signs_blocks.append(f"""<div class="card" style="margin-bottom:var(--s3)">
  <div class="card-body">
    <h3 style="margin-top:0">{esc(cat['title'])}</h3>
    <div class="table-wrap"><table>
      <thead><tr><th>프랑스어 표기</th><th>한국어 의미</th><th>발음 도움</th></tr></thead>
      <tbody>{''.join(rows)}</tbody>
    </table></div>
  </div>
</div>""")
    signs_html = "".join(signs_blocks)

    phrase_cards_html = "".join(phrase_card(p) for p in phrases)

    index_search("여행 프랑스어 (Travel French)", "prepare/french.html", "prepare", "120개 필수 문구 · 10분 발음 · 현장 표지판")

    return page(
        title="여행 프랑스어", rel=rel, tab="prepare",
        description="현장에서 즉시 쓰는 120개 필수 프랑스어 문구와 10분 발음·표지판 사전",
        trail=[("홈", "index.html"), ("준비", "prepare/index.html"),
               ("여행 프랑스어", None)],
        body=f"""<div class="wrap"><div class="stack-lg" style="padding-top:1.5rem">
<header>
  <h1>여행 프랑스어 (Travel French)</h1>
  <p class="hero-dek">현장에서 즉시 쓰는 120개 필수 회화 문구와 10분 발음·표지판 사전입니다. 발음 듣기(TTS)와 복사, 즐겨찾기를 지원합니다.</p>
</header>

<div class="card" style="background:var(--surface-2);border:1px solid var(--line-strong)"><div class="card-body stack">
  <div class="search-box" style="position:relative">
    <input type="search" id="french-search" class="form-control" style="width:100%;min-height:44px;padding:var(--s2) var(--s4);border:1px solid var(--line-strong);border-radius:var(--r-full);background:var(--surface);font-size:var(--t-body)" placeholder="프랑스어 / 한국어 / 발음 / 태그 검색 (예: 계산, 주차, addition, merci)..." aria-label="프랑스어 문구 검색">
  </div>
  <div class="chips" id="french-filter-chips" role="toolbar" aria-label="프랑스어 카테고리 필터">
    <button type="button" class="chip" data-category="all" aria-pressed="true">전체 (120)</button>
    <button type="button" class="chip" data-category="fav" aria-pressed="false">⭐ 즐겨찾기</button>
    <button type="button" class="chip" data-category="essential" aria-pressed="false">기본 20선 (P0)</button>
    <button type="button" class="chip" data-category="restaurant" aria-pressed="false">식당·카페</button>
    <button type="button" class="chip" data-category="market" aria-pressed="false">빵집·시장</button>
    <button type="button" class="chip" data-category="hotel" aria-pressed="false">숙소</button>
    <button type="button" class="chip" data-category="transport" aria-pressed="false">기차·교통</button>
    <button type="button" class="chip" data-category="driving" aria-pressed="false">렌터카·주차</button>
    <button type="button" class="chip" data-category="sightseeing" aria-pressed="false">관광·미술관</button>
    <button type="button" class="chip" data-category="shopping" aria-pressed="false">쇼핑</button>
    <button type="button" class="chip" data-category="emergency" aria-pressed="false">긴급상황</button>
  </div>
</div></div>

<section id="french-phrases-section">
  <div class="sec-head"><div class="sec-title-group"><span class="sec-eyebrow">PHRASES</span><h2 class="sec-title" id="french-list-title">상황별 회화 (120문구)</h2></div></div>
  <div id="french-no-results" class="alert-card alert-card-caution" style="display:none">
    {ic('alert')} <span>검색 결과가 없습니다.</span>
    <button type="button" class="btn btn-quiet" id="french-reset-btn" style="margin-inline-start:var(--s2)">전체 보기</button>
  </div>
  <div class="grid grid-2" id="french-phrase-grid">
    {phrase_cards_html}
  </div>
</section>

<section id="french-pronunciation-section" style="margin-top:var(--s5)">
  <details class="acc"><summary><h2 style="display:inline;font-size:var(--t-h3)">{ic('tip')} 10분 발음 & 읽기 규칙 (French in 10 Minutes)</h2></summary>
    <div class="acc-body stack" style="margin-top:var(--s3)">
      <p class="meta">정확한 음성학 학습이 아니라 간판과 메뉴를 읽기 위한 최소한의 발음 규칙입니다.</p>
      {pron_html}
    </div>
  </details>
</section>

<section id="french-signs-section" style="margin-top:var(--s3)">
  <details class="acc"><summary><h2 style="display:inline;font-size:var(--t-h3)">{ic('book')} 현장 표지판 & 메뉴 필수 어휘 사전</h2></summary>
    <div class="acc-body stack" style="margin-top:var(--s3)">
      <p class="meta">거리 표지, 역 안내판, 메뉴판에서 가장 빈번히 마주치는 단어들입니다.</p>
      {signs_html}
    </div>
  </details>
</section>

<div class="btn-row" style="margin-top:var(--s4)">
  <a class="btn btn-secondary" href="french.html">{ic('chat')}여행 프랑스어</a>
  <a class="btn btn-secondary" href="emergency.html">{ic('alert')}긴급 연락처</a>
  <a class="btn btn-secondary" href="index.html">{ic('check')}준비 메인</a>
  <a class="btn btn-secondary" href="../schedule.html">{ic('today')}전체 일정</a>
</div>

</div></div>"""
    )



def build_paris_museum_booking() -> str:
    """준비 — 파리 박물관·전시 예약 실행표 (RS02, Jason 2026-08-28 지시).

    파리 15박(9/24~10/9)의 미술관·전시 예약을 3단계 파도로 나눈 실행표다.
    행의 방문일은 Day 페이지로, 장소는 장소 정본으로 연결한다.
    """
    rel = ".."

    P1 = '<span class="badge badge-must">1차 · 지금</span>'
    P2 = '<span class="badge badge-caution">2차 · 9월 초</span>'
    P3 = '<span class="badge badge-neutral">3차 · 직전</span>'
    NOB = '<span class="badge badge-ok">예약 불필요</span>'

    # (우선순위, 방문일, day번호, 장소표기, place슬러그, 일정, 권장 예약 시점, 권장 실행일, 비고)
    rows = [
        (badge('ok', '예약 확정'), "9/25", 28, "Grand Palais — Cézanne et nous", "grand-palais",
         "17:00 특별전", "예약 완료", "티켓 저장", "사용자 예약 확정 · 16:45 보안검색 도착"),
        (P1, "9/26", 29, "Musée du Luxembourg — Warhol", "musee-du-luxembourg",
         "특별전", "3~4주 전", "지금", "원하는 시간대 확보"),
        (badge('ok', '예약 확정'), "9/30", 33, "Musée de l'Orangerie", "musee-de-l-orangerie",
         "10:00 · 상설 중심", "예약 완료", "티켓 저장", "수련 우선 90분 집중 관람 · 12:15 Chez Savy 연결"),
        (P3, "9/28", 31, "Musée Gustave Moreau", "musee-gustave-moreau",
         "일반관람", "1~2주 전", "9/14 전후", "예약 급하지 않음"),
        (P1, "9/29", 32, "Musée d'Orsay", "musee-d-orsay",
         "09:30 지정시간", "3~4주 전", "지금", "지정시간 일정 — 조기 확보"),
        (P3, "9/29", 32, "Musée Rodin", "musee-rodin",
         "14:30", "1~2주 전", "9/14 전후", "오전 Orsay 일정에 종속"),
        (P1, "10/1", 34, "Versailles", "versailles",
         "종일", "1~2개월 전", "지금", "Passport + 시간지정 입장"),
        (P1, "10/2", 35, "Louvre", "musee-du-louvre",
         "14:00 지정시간", "3~4주 전", "지금~9/2", "14:00 슬롯 유지 권장"),
        (P3, "10/3", 36, "Marmottan Monet", "musee-marmottan-monet",
         "14:00", "1~2주 전", "9/18 전후", "일정 유연성 유지"),
        (P2, "10/5", 38, "Jacquemart-André", "musee-jacquemart-andre",
         "미술관·전시", "2~4주 전", "9/7~15", "특별전이면 조금 빨리"),
        (P1, "10/6", 39, "Orsay — Mary Cassatt 특별전", "musee-d-orsay",
         "특별전", "판매 가능 즉시", "지금 확인·예약", "일반 Orsay 입장(9/29)과 별도로 판단 · "
         "회차 판매 여부 재확인"),
        (P3, "10/6", 39, "Musée Picasso Paris", "musee-picasso-paris",
         "13:00", "1~2주 전", "9/20 전후", "현장구매도 가능"),
        (P1, "10/7", 40, "Bourse de Commerce — Remember Me", "bourse-de-commerce-pinault-collection",
         "개막일", "판매 시작 즉시", "지금", "개막일 방문 — 조기 예약 필요"),
        (P3, "10/8", 41, "Musée Guimet", "musee-guimet",
         "10:00", "1~2주 전", "9/24 이후", "서두를 필요 없음"),
        (NOB, "10/8", 41, "MAM Paris (파리 시립 현대미술관)", "musee-d-art-moderne-de-paris",
         "상설전", "예약 불필요", "직전 확인", "상설 컬렉션 무료"),
    ]

    body_rows = "".join(
        f"<tr><td>{pr}</td>"
        f'<td><a href="{rel}/daily/day-{day:02d}.html">{esc(date)}</a></td>'
        f'<td><a href="{rel}/places/{slug}.html">{esc(name)}</a></td>'
        f"<td>{esc(sched)}</td><td>{esc(when)}</td><td>{esc(act)}</td>"
        f"<td>{esc(note)}</td></tr>"
        for pr, date, day, name, slug, sched, when, act, note in rows)

    wave = """
<div class="prose">
<p><strong>예약 확정:</strong> Grand Palais 9/25 17:00 · Orangerie 9/30 10:00.</p>
<p><strong>1차 — 지금 (8/28~8/31):</strong> Versailles → Orsay 9/29 →
Louvre 10/2 → Bourse de Commerce → Luxembourg → Mary Cassatt(판매 확인).
단순히 유명한 곳이라서가 아니라, <strong>날짜·시간이 고정되어 있거나 특별전 개막
직후에 방문</strong>하는 곳들이다. 특히 Versailles 10/1 · Louvre 10/2 · Orsay 9/29는
파리 일정 전체의 동선을 잡는 기준점이므로 먼저 확정한다.</p>
<p><strong>2차 — 9/1~9/10:</strong> Jacquemart-André. 1차가 끝난 뒤 처리한다.</p>
<p><strong>3차 — 9/12~9/20 (여행 중):</strong> Gustave Moreau · Rodin ·
Marmottan · Picasso. 너무 일찍 예약하면 파리 체류 중 날씨·피로도·공연 일정에
따른 변경 여지가 줄어든다 — 기다리는 것이 오히려 좋다.</p>
<p><strong>파리 도착 전후:</strong> Guimet 예약과 무료관(MAM) 운영시간 최종 확인.</p>
</div>"""

    principles = """
<div class="prose">
<p>예약 시간까지 가이드북 일정과 맞춰 <strong>고정해야 하는 곳</strong>은
Orsay 09:30 · Louvre 14:00 · Versailles 오전 입장이다. 반면 Rodin이나 Moreau처럼
유연한 곳은 앞뒤 일정에 여유를 두고 잡는다.</p>
<p>티켓을 살 때는 구매 완료만 관리하지 말고 <strong>① 날짜 ② 시간
③ 변경·취소 가능 여부 ④ QR 티켓 저장 위치</strong>까지 함께 기록한다.
파리 일정에는 특별전과 일반관람권이 섞여 있어 이 구분이 특히 중요하다.</p>
</div>"""

    index_search("파리 뮤지엄 예약 실행표", "prepare/paris-museums.html", "prepare",
                 "파리 15박 미술관·전시 예약 3단계 실행표")

    return page(
        title="파리 뮤지엄 예약", rel=rel, tab="prepare",
        description="파리 박물관·전시 예약을 3단계로 나눈 실행표",
        trail=[("홈", "index.html"), ("준비", "prepare/index.html"),
               ("파리 뮤지엄 예약", None)],
        body=f"""<div class="wrap"><div class="stack-lg" style="padding-top:1.5rem">
<header><h1>파리 뮤지엄 예약 실행표</h1>
<p class="hero-dek">파리 15박(9/24~10/9)의 미술관·전시 예약을 3단계 파도로 나눈다.
예약 실패 위험이 큰 것부터 지금 처리한다.</p></header>

{alert('caution',
       '<strong>오늘 처리 6건:</strong> Versailles · Orsay 9/29 · '
       'Louvre 10/2 · Bourse de Commerce · Luxembourg · Mary Cassatt(판매 확인). '
       '나머지는 단계별 시점에 맞춰 처리한다.')}

{sec_head('PLAN', '예약 실행표 — 15건', rule=True)}
<div class="table-wrap"><table>
<thead><tr><th>단계</th><th>방문일</th><th>장소 / 전시</th><th>일정</th>
<th>권장 예약 시점</th><th>권장 실행일</th><th>비고</th></tr></thead>
<tbody>{body_rows}</tbody>
</table></div>

{sec_head('WAVES', '3단계 파도 — 왜 이 순서인가', rule=True)}
{wave}

{sec_head('PRINCIPLES', '실제 예약할 때의 원칙', rule=True)}
{principles}

{sec_head('CHECKLIST', '압축 체크리스트', rule=True)}
<div class="prose"><ul>
<li><strong>예약 확정:</strong> Grand Palais 9/25 17:00 · Orangerie 9/30 10:00</li>
<li><strong>오늘:</strong> Versailles · Orsay · Louvre · Bourse de Commerce · Luxembourg · Mary Cassatt</li>
<li><strong>9월 초:</strong> Jacquemart-André</li>
<li><strong>9월 중순:</strong> Moreau · Rodin · Marmottan · Picasso</li>
<li><strong>파리 출발 직전:</strong> Guimet 예약 · 무료관(MAM) 운영시간 재확인</li>
</ul>
<p>이 순서대로 처리하면 가이드북 일정을 거의 그대로 유지하면서 예약 실패
위험을 크게 낮출 수 있다. 확정된 예약은 <a href="index.html">준비 현황</a>의
확정 목록과 트래커에 날짜·시간·취소조건·QR 저장 위치를 함께 기록한다.</p></div>

<div class="btn-row"><a class="btn btn-secondary" href="index.html">{ic('check')}준비 현황</a>
  <a class="btn btn-secondary" href="{rel}/guide/paris.html">{ic('region')}파리 가이드</a></div>
</div></div>""")

def build_prepare(trip: Trip, res: dict) -> dict[str, str]:
    """준비 — 무엇을 예약·확인해야 하는가.

    상태는 셋뿐이다: 확정 · 미예약 · 제외. 중간 상태를 두면 무엇을 해야
    하는지 알 수 없다.
    """
    rel = ".."
    out = {}
    todo, done, dropped = res["todo"], res["confirmed"], res["dropped"]

    def group(records, is_todo):
        by_cat = {}
        for r in records:
            by_cat.setdefault(r["카테고리"] or "기타", []).append(r)
        blocks = []
        for cat, items in by_cat.items():
            blocks.append(sec_head("", f"{cat} {len(items)}건"))
            blocks.append('<div class="grid grid-2">'
                          + "".join(res_card(r, todo=is_todo) for r in items)
                          + "</div>")
        return "".join(blocks)

    dropped_html = ""
    if dropped:
        rows = "".join(
            f'<li>{esc(r["예약항목"])}'
            + (f' — {esc(r["비고"][:80])}' if r["비고"] else "") + "</li>"
            for r in dropped)
        dropped_html = (f'<details class="acc"><summary>이번 일정에서 뺀 것 '
                        f'{len(dropped)}건</summary><div class="acc-body prose">'
                        f"<ul>{rows}</ul></div></details>")

    out["index.html"] = page(
        title="준비", rel=rel, tab="prepare",
        description="여행 준비 상태를 점검한다",
        trail=[("홈", "index.html"), ("준비", None)],
        body=f"""<div class="wrap"><div class="stack-lg" style="padding-top:1.5rem">
<header><h1>준비</h1>
<p class="hero-dek">확정 {len(done)}건 · 미예약 {len(todo)}건.
  상태는 셋뿐이다 — 확정 · 미예약 · 제외.</p></header>

<div class="btn-row"><a class="btn btn-primary" href="paris-museums.html">
  {ic('ticket')}파리 뮤지엄 예약</a></div>

{alert('caution',
       f'<strong>아직 {len(todo)}건이 예약되지 않았다.</strong> 예약이 없는 항목은 '
       f'주소·시각이 정해진 것이 아니다. 확정된 것만 확정으로 표시된다.')
 if todo else alert('ok', '<strong>모든 예약이 확정됐다.</strong>', 'check')}

{alert('ok', '<strong>예약번호는 뒤 4자리를 가렸다.</strong> 이 사이트는 누구나 '
       '열 수 있어서, 번호가 그대로 있으면 남이 그 예약을 조회할 수 있다. '
       '온전한 번호는 예약 확인 메일과 트래커 파일에 있다.', 'lock')}

{sec_head('TO BOOK', f'아직 예약하지 않은 것 — {len(todo)}건', rule=True) if todo else ''}
{group(todo, True)}

{sec_head('BOOKED', f'예약을 마친 것 — {len(done)}건', rule=True) if done else ''}
{group(done, False)}

{dropped_html}

<div class="btn-row"><a class="btn btn-secondary" href="french.html">
  {ic('chat')}여행 프랑스어</a>
  <a class="btn btn-secondary" href="emergency.html">
  {ic('alert')}긴급 연락처</a>
  <a class="btn btn-secondary" href="../offline.html">
  {ic('download')}오프라인 준비</a></div>
</div></div>""")

    out["paris-museums.html"] = build_paris_museum_booking()

    out["french.html"] = build_travel_french(trip)

    out["emergency.html"] = page(
        title="긴급", rel=rel, tab="prepare",
        trail=[("홈", "index.html"), ("준비", "prepare/index.html"),
               ("긴급", None)],
        body=f"""<div class="wrap-read"><div class="stack-lg" style="padding-top:1.5rem">
<header><h1>긴급 연락처</h1>
<p class="hero-dek">국경을 넘으면 번호가 바뀐다. Day 7 에 스페인에서
  프랑스로 넘어간다.</p></header>
<div class="table-wrap"><table>
<thead><tr><th>국가</th><th>번호</th><th>용도</th></tr></thead>
<tbody>
<tr><td>공통 (EU)</td><td><strong>112</strong></td><td>모든 긴급 — 경찰·소방·구급</td></tr>
<tr><td>스페인</td><td>091 / 061</td><td>경찰 / 구급</td></tr>
<tr><td>프랑스</td><td>17 / 15</td><td>경찰 / 구급 (SAMU)</td></tr>
</tbody></table></div>
{alert('ok', '<strong>112 는 EU 전역에서 통한다.</strong> 어느 나라인지 '
       '헷갈리면 112 를 누른다. 휴대폰 잠금 상태에서도 걸린다.', 'check')}
</div></div>""")
    return out


# ================================================================ 자산·색인

def load_reservations() -> dict:
    """예약을 상세까지 읽는다.

    예전에는 상태와 건수만 꺼냈다. 그런데 목록만 있고 세부가 없으면 현장에서
    쓸 수 없다 — "숙소 4박" 이라고만 적힌 줄은 어느 숙소인지 알려 주지 않고,
    렌터카 카운터에서 필요한 것은 목록이 아니라 예약번호다.

    상태는 셋뿐이다.
        확정   예약이 있다 (예약번호가 있거나 본인이 확인했다)
        미예약 아직 없다 — 해야 할 일이다
        제외   이번 일정에서 뺐다
    '재확인' 은 없앴다. 예약번호가 있는 것과 아예 없는 것을 한 낱말로 묶고
    있어서, 무엇을 해야 하는지 알 수 없었다.
    """
    empty = {"confirmed": [], "todo": [], "dropped": [],
             "active": 0, "undone": 0, "items": [], "by_date": {}}
    try:
        from openpyxl import load_workbook
    except ImportError:
        return empty
    if not TRACKER_XLSX.exists():
        return empty
    wb = load_workbook(TRACKER_XLSX, data_only=True)
    if "Reservations" not in wb.sheetnames:
        return empty
    rows = list(wb["Reservations"].iter_rows(values_only=True))
    hdr_i = next((k for k, r in enumerate(rows) if r and r[0] == "ID"), None)
    if hdr_i is None:
        return empty
    hdr = list(rows[hdr_i])
    ix = {name: hdr.index(name) for name in hdr if name}

    def cell(row, key):
        if key not in ix:
            return ""
        v = row[ix[key]]
        if v is None:
            return ""
        if hasattr(v, "date"):
            return v.date().isoformat()
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    confirmed, todo, dropped, by_date, items = [], [], [], {}, []
    BOOKING_CODES.clear()
    for row in rows[hdr_i + 1:]:
        if not row or not row[ix["ID"]]:
            continue
        rec = {k: cell(row, k) for k in
               ("ID", "카테고리", "지역", "예약항목", "날짜", "시간", "상태",
                "총액", "통화", "예약번호", "사업자", "주소/역", "무료취소기한",
                "예약목표일", "리스크/대체안", "비고")}
        collect_codes(rec["예약번호"])

        status = rec["상태"]
        items.append((rec["ID"], rec["예약항목"], status))
        if status == "제외":
            dropped.append(rec)
            continue
        if rec["날짜"]:
            by_date.setdefault(rec["날짜"], []).append(status)
        (confirmed if status == "확정" else todo).append(rec)

    # 같은 예약이 Reservations 와 Accommodation 에 나뉘어 적힌 경우가 있다.
    # 한쪽만 훑으면 다른 쪽 코드가 화면으로 새어 나간다.
    if "Accommodation" in wb.sheetnames:
        acc = wb["Accommodation"]
        acc_hdr = [c.value for c in acc[3]]
        if "예약번호" in acc_hdr:
            k = acc_hdr.index("예약번호")
            for row in acc.iter_rows(min_row=4, values_only=True):
                if k < len(row) and row[k]:
                    collect_codes(str(row[k]))

    order = {"숙소": 0, "항공": 1, "철도": 2, "렌터카": 3, "입장권": 4,
             "공연": 5, "기타": 6}
    key = lambda r: (order.get(r["카테고리"], 9), r["날짜"] or "9999")
    confirmed.sort(key=key)
    todo.sort(key=lambda r: (r["예약목표일"] or "9999", key(r)))

    return {"confirmed": confirmed, "todo": todo, "dropped": dropped,
            "active": len(confirmed) + len(todo), "undone": len(todo),
            "items": items, "by_date": by_date}


def write_assets(trip: Trip) -> None:
    out = SITE / "assets"
    out.mkdir(parents=True, exist_ok=True)
    # 아이콘은 CSS 마스크로 붙인다 — 페이지마다 스프라이트를 인라인하면
    # 페이지 수만큼 무게가 붙는다. 마스크는 CSS 한 번이고 페이지 무게는 0 이다.
    (out / "style.css").write_text(
        (ASSETS / "style.css").read_text(encoding="utf-8") + "\n" + icons.css(),
        encoding="utf-8")
    for name in ("app.js", "pwa.js"):
        shutil.copy(ASSETS / name, out / name)
    # 글꼴은 번들하지 않는다. 기기의 기본 한글 글꼴을 쓰므로 내려받을 것이
    # 없고, 그만큼 오프라인 패키지가 가벼워진다.
    pwa = ROOT / "source" / "ASSETS" / "pwa"
    if pwa.exists():
        shutil.copytree(pwa, out / "pwa", dirs_exist_ok=True)
    transport_guides = ROOT / "source" / "ASSETS" / "transport-guides"
    if transport_guides.exists():
        shutil.copytree(transport_guides, out / "transport-guides", dirs_exist_ok=True)
    tourist_maps = ROOT / "source" / "ASSETS" / "tourist-maps"
    if tourist_maps.exists():
        shutil.copytree(tourist_maps, out / "tourist-maps", dirs_exist_ok=True)

    # 사진 — 매니페스트에 있는 것만 옮긴다. 카탈로그에 없으면 자리도 없다.
    raw = json.loads(IMAGE_MANIFEST.read_text(encoding="utf-8"))
    copied = 0
    missing = []
    for img in raw.get("images", []):
        for variants in (img.get("variants") or {}).values():
            for v in variants:
                src = ROOT / v["path"]
                dst = SITE / v["sitePath"]
                if not src.exists():
                    # 조용히 건너뛰면 사진 자리만 비고 아무도 모른다.
                    # 슬러그를 바꾸고 파일을 안 바꾼 적이 실제로 있었다.
                    missing.append(f"{img.get('imageId')} — {v['path']}")
                    continue
                dst.parent.mkdir(parents=True, exist_ok=True)
                for _ in range(3):
                    try:
                        if not dst.exists():
                            shutil.copy(src, dst)
                            copied += 1
                        break
                    except OSError:
                        dst.parent.mkdir(parents=True, exist_ok=True)
    if missing:
        raise SystemExit("매니페스트가 가리키는 사진 파일이 없다:\n  "
                         + "\n  ".join(missing))
    print(f"  사진 {copied}개 복사")

    (out / "search-index.js").write_text(
        "window.SEARCH_INDEX=" + json.dumps(SEARCH_INDEX, ensure_ascii=False)
        + ";", encoding="utf-8")


PWA_ICON_SPECS = (
    ("apple-touch-icon.png", "180x180", "any"),
    ("icon-192.png", "192x192", "any"),
    ("icon-512.png", "512x512", "any"),
    ("icon-maskable-512.png", "512x512", "maskable"),
)

# 연결이 끊겨도 반드시 열려야 하는 것들. 없으면 빌드를 세운다.
PWA_CORE_PATHS = (
    "index.html",
    "offline.html",
    "offline-fallback.html",
    "schedule.html",
    "guide/index.html",
    "map/index.html",
    "prepare/index.html",
    "prepare/emergency.html",
    "prepare/french.html",
    "assets/style.css",
    "assets/app.js",
    "assets/search-index.js",
)


def build_offline_page() -> str:
    """오프라인 준비 화면. 여행 전에 전체를 기기에 저장시킨다.

    43일 중 상당 구간이 데이터가 약하거나 로밍이 비싸다. 현장에서 페이지가
    안 열리는 것이 이 도구의 실패다.

    DOM 은 build/assets/pwa.js 의 계약이다 — id 를 바꾸면 저장 기능이
    조용히 죽는다. 스크립트가 없으면 안내 문구만 남는다.
    """
    return page(
        title="오프라인 준비", rel=".", tab="prepare",
        description="가이드북 전체를 기기에 저장한다",
        trail=[("홈", "index.html"), ("오프라인 준비", None)],
        body=f"""<div class="wrap-read"><div class="stack-lg" style="padding-top:1.5rem">
<header><h1>오프라인 준비</h1>
<p class="hero-dek">가이드북 전체를 기기에 저장한다. 연결이 없어도 일정 ·
장소 · 지도 목록이 열린다.</p></header>

<div class="card" id="pwa-panel"><div class="card-body stack">
  <dl class="pwa-facts">
    <dt>저장 상태</dt><dd id="pwa-status">확인 중</dd>
    <dt>설치 형태</dt><dd id="pwa-install-mode">확인 중</dd>
    <dt>저장 용량</dt><dd id="pwa-storage">—</dd>
    <dt>버전</dt><dd id="pwa-version">—</dd>
  </dl>
  <p class="meta" id="pwa-detail"></p>
  <div class="btn-row">
    <button class="btn btn-primary" id="pwa-save" type="button">
      {ic("download")}전체 저장</button>
    <button class="btn btn-secondary" id="pwa-clear" type="button">
      {ic("close")}저장 비우기</button>
  </div>
  <progress class="pwa-progress" id="pwa-progress" value="0" max="1"
    aria-label="저장 진행률"></progress>
  <div id="pwa-update-box" hidden>
    {alert("caution", '새 버전이 있다. <button class="btn btn-quiet" '
           'id="pwa-activate-update" type="button">지금 적용</button>')}
  </div>
</div></div>

<noscript><p class="meta">저장 기능은 자바스크립트가 필요하다.
연결이 되는 곳에서 이 화면을 다시 열어라.</p></noscript>

{alert("caution", "<strong>출발 전과 장거리 이동 전에 다시 확인한다.</strong> "
       "iOS 는 저장 공간이 부족하거나 앱을 오래 쓰지 않으면 웹 데이터를 "
       "정리할 수 있다.")}

<div class="prose">
<h2>담기는 것</h2>
<ul>
  <li>43일 일정과 하루별 시간표 · 이동 · 예약</li>
  <li>장소 페이지 전체와 사진</li>
  <li>지역 가이드 · 지도 목록과 좌표 링크</li>
  <li>준비 화면과 긴급 연락처</li>
</ul>
<p>지도 타일은 담기지 않는다. 좌표와 Google Maps 링크는 저장되므로
연결이 되는 곳에서 열 수 있다.</p>
</div>
</div></div>""")


def build_offline_fallback() -> str:
    """저장되지 않은 페이지를 오프라인에서 열었을 때. 막다른 화면을 만들지 않는다."""
    return page(
        title="저장되지 않은 페이지", rel=".", tab="today",
        trail=[("홈", "index.html"), ("오프라인", None)],
        body=f"""<div class="wrap-read"><div class="stack-lg" style="padding-top:1.5rem">
<header><h1>이 페이지는 아직 저장되지 않았다</h1>
<p class="hero-dek">연결이 없고, 요청한 페이지가 기기에 없다.</p></header>
{alert("caution", "연결이 되는 곳에서 <strong>오프라인 준비</strong> 화면을 열어 "
       "전체 저장을 마치면 이런 일이 생기지 않는다.")}
<div class="btn-row">
  <a class="btn btn-primary" href="index.html">{ic("today")}저장된 홈 열기</a>
  <a class="btn btn-secondary" href="offline.html">{ic("download")}오프라인 준비</a>
</div>
</div></div>""")


def write_pwa() -> None:
    """manifest · 오프라인 목록 · 서비스워커.

    버전은 파일 내용 해시다. 내용이 안 바뀌면 버전도 안 바뀌고, 그래서
    기기가 헛되이 다시 받지 않는다.
    """
    (SITE / "manifest.webmanifest").write_text(json.dumps({
        "id": "./",
        "name": SITE_TITLE,
        "short_name": "유럽 가이드북",
        "description": "Barcelona 에서 Paris 까지 43일 여행 현장 가이드북",
        "lang": "ko",
        "start_url": "./index.html",
        "scope": "./",
        "display": "standalone",
        "background_color": "#FAF6EF",
        "theme_color": "#FAF6EF",
        "icons": [
            {"src": f"./assets/pwa/{name}", "sizes": sizes,
             "type": "image/png", "purpose": purpose}
            for name, sizes, purpose in PWA_ICON_SPECS
        ],
    }, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    excluded = {"sw.js", "offline-files.json", ".nojekyll"}
    records, version_hash = [], hashlib.sha256()
    for path in sorted(p for p in SITE.rglob("*") if p.is_file()):
        rel = path.relative_to(SITE).as_posix()
        if rel in excluded:
            continue
        content = path.read_bytes()
        digest = hashlib.sha256(content).hexdigest()
        records.append({"path": rel, "size": len(content), "sha256": digest})
        version_hash.update(rel.encode("utf-8") + b"\0")
        version_hash.update(digest.encode("ascii") + b"\n")
    version = version_hash.hexdigest()

    (SITE / "offline-files.json").write_text(json.dumps({
        "version": version,
        "totalFiles": len(records),
        "totalBytes": sum(r["size"] for r in records),
        "files": records,
    }, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")

    missing = [p for p in PWA_CORE_PATHS if not (SITE / p).is_file()]
    if missing:
        sys.exit("PWA 핵심 파일 누락: " + ", ".join(missing))

    template = (ASSETS / "service-worker.js").read_text(encoding="utf-8")
    sw = template.replace("__PWA_VERSION__", version).replace(
        "__PWA_CORE_PATHS__", json.dumps(list(PWA_CORE_PATHS), ensure_ascii=False))
    if "__PWA_" in sw:
        sys.exit("Service Worker 템플릿 토큰이 남아 있다")
    (SITE / "sw.js").write_text(sw, encoding="utf-8")
    mib = sum(r["size"] for r in records) / 1048576
    print(f"  PWA: {len(records)}개 파일 · {mib:.1f} MiB · 버전 {version[:12]}")


# 옛 주소 → 새 주소. 404 를 늘리지 않는다.
def write_redirects(trip: Trip) -> int:
    n = 0
    def put(path: str, target: str, label: str):
        nonlocal n
        p = SITE / path
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(redirect(target, label), encoding="utf-8")
        n += 1

    put("regions.html", "guide/index.html", "가이드")
    put("daily/index.html", "../schedule.html", "전체 일정")
    put("maps/index.html", "../map/index.html", "지도")
    put("tracker/index.html", "../prepare/index.html", "준비")
    put("places/index.html", "../guide/index.html", "가이드")

    # 챕터 10개 카테고리 → 새 지역 페이지의 해당 섹션
    section = {
        "index": "", "about": "#overview", "schedule": "#days",
        "places": "#places", "food": "#food", "stay": "#stay",
        "transport": "#transport", "tips": "#overview",
        "booking": "", "cost": "", "sources": "",
    }
    for r in trip.regions:
        for name, anchor in section.items():
            target = (f"../../prepare/index.html"
                      if name in ("booking", "cost")
                      else f"../../about/sources.html" if name == "sources"
                      else f"../../guide/{r.slug}.html{anchor}")
            put(f"chapters/{r.slug}/{name}.html", target, r.name)
        for d in r.days:
            put(f"chapters/{r.slug}/day-{d.n:02d}.html",
                f"../../{d.url}", f"Day {d.n}")
    for name in ("index", "days", "places", "food", "stay", "transport",
                 "tips", "booking", "cost", "essential", "p0", "about",
                 "sources", "reverify"):
        put(f"topics/{name}.html", "../guide/index.html", "가이드")
    for name in ("dashboard", "itinerary", "reservations", "accommodation",
                 "transport", "locks"):
        put(f"tracker/{name}.html", "../prepare/index.html", "준비")

    # 슬러그가 바뀐 장소 — 주소창에 남은 옛 이름을 404 로 만들지 않는다
    moved = ROOT / "data" / "slug-redirects.json"
    if moved.exists():
        table = json.loads(moved.read_text(encoding="utf-8")).get("places", {})
        for old_slug, rule in table.items():
            new_slug = rule["to"]
            place = trip.places.get(new_slug)
            if place is None:
                raise SystemExit(
                    f"슬러그 리다이렉트가 없는 곳을 가리킨다 — {old_slug} → {new_slug}")
            put(f"places/{old_slug}.html", f"{new_slug}.html", place.name)
    return n


def build_credits(trip: Trip) -> str:
    """사진 저작자 표시. 표시가 필요한 라이선스는 화면에 남긴다."""
    raw = json.loads(IMAGE_MANIFEST.read_text(encoding="utf-8"))
    def cell_link(url: str, label: str) -> str:
        """주소가 없으면 링크를 만들지 않는다. 빈 href 는 눌러도 제자리다."""
        if not url:
            return esc(label) if label else "—"
        return (f'<a href="{esc(url)}" rel="nofollow noopener">'
                f'{esc(label) if label else domain_of(url)}</a>')

    rows = "".join(f"""<tr><td>{esc(i.get('titleKo') or i.get('title'))}</td>
<td>{linkify(esc(i.get('creator') or '—'))}</td>
<td>{cell_link(i.get('licenseUrl') or '', i.get('license') or '')}</td>
<td>{cell_link(i.get('sourcePage') or '', '출처')}</td></tr>"""
        for i in raw.get("images", []))
    return page(
        title="사진 저작자 표시", rel="..", tab="guide",
        trail=[("홈", "index.html"), ("사진 저작자 표시", None)],
        body=f"""<div class="wrap"><div class="stack-lg" style="padding-top:1.5rem">
<header><h1>사진 저작자 표시 · 라이선스</h1>
<p class="hero-dek">이 사이트는 공개 배포되므로 재배포가 허용된 이미지만
쓴다. 저작자 표시가 필요한 라이선스는 여기에 표시한다.</p></header>
<div class="table-wrap"><table>
<thead><tr><th>사진</th><th>저작자</th><th>라이선스</th><th>출처</th></tr></thead>
<tbody>{rows}</tbody></table></div>
</div></div>""")


def build_sources(trip: Trip) -> str:
    return page(
        title="출처와 검증", rel="..", tab="guide",
        trail=[("홈", "index.html"), ("출처와 검증", None)],
        body=f"""<div class="wrap-read"><div class="stack-lg" style="padding-top:1.5rem">
<header><h1>출처와 검증</h1></header>
<div class="prose">
<p>이 가이드북은 확인한 것만 싣는다. 확인하지 못한 운영시간·요금은
기술하지 않거나 <strong>재확인</strong> 표시를 단다. 추측으로 채우지 않는다.</p>
<ul>
  <li>역사·건축·인물 — 복수 출처를 확인한 뒤 기술한다.</li>
  <li>운영시간·요금 — 근거와 확인 날짜를 함께 들고 다닌다.
    장소 페이지의 <strong>실용</strong> 표에서 볼 수 있다.</li>
  <li>개인 해석 — 주어를 밝힌다.</li>
  <li>미확인 — 기술하지 않는다.</li>
</ul>
<p>예약은 아직 전부 잠기지 않았다. <strong>확정 전 주소를 확정으로 믿고
이동하지 않는다.</strong> 확정된 것만 화면에 확정으로 표시된다.</p>
</div></div></div>""")


# ================================================================ 가드

def check_vocabulary(trip: Trip) -> list[str]:
    """데이터가 쓰는 값을 렌더러가 전부 아는가.

    콘텐츠 편집은 이 개편과 나란히 계속된다. 편집자가 새 이동수단이나
    새 카테고리를 쓰면 렌더러는 그 값을 모른 채 **영어 코드를 그대로**
    화면에 흘린다. 실제로 그런 일이 있었다 — Nice 일정이 갱신되면서
    'tram' 과 'unconfirmed' 가 들어왔고, 특히 'unconfirmed' 는 아직
    이동수단이 안 정해졌다는 뜻인데 확정된 구간과 똑같이 보였다.

    그래서 빌드를 세운다. 현장에서 읽는 화면이 조용히 틀리는 것보다
    빌드가 시끄럽게 멈추는 편이 낫다.
    """
    problems = []
    bad_modes, bad_cats, bad_status, bad_day_types, bad_execution = {}, {}, {}, {}, {}
    for d in trip.days:
        for leg in d.legs:
            if leg.mode not in MODE_LABEL:
                bad_modes.setdefault(leg.mode, []).append(d.n)
        for s in d.stops:
            if s.category not in CAT_ICON:
                bad_cats.setdefault(s.category, []).append(d.n)
            for status in s.execution_statuses:
                if status.type not in EXECUTION_STATUS_UI:
                    bad_execution.setdefault(status.type, []).append(d.n)
        if d.source_status not in ("authoritative", "candidate-latest-needs-review",
                                   "prototype-reviewed"):
            bad_status.setdefault(d.source_status, []).append(d.n)
        if d.day_type is not None and d.day_type not in DAY_TYPE_LABEL:
            bad_day_types.setdefault(d.day_type, []).append(d.n)

    for mode, days in bad_modes.items():
        problems.append(
            f"모르는 이동수단 '{mode}' — Day {sorted(set(days))[:6]}. "
            f"render.py 의 MODE_LABEL·MODE_ICON 에 한국어 표기를 더한다.")
    for cat, days in bad_cats.items():
        problems.append(
            f"모르는 장소 분류 '{cat}' — Day {sorted(set(days))[:6]}. "
            f"render.py 의 CAT_ICON 에 아이콘을 지정한다.")
    for st, days in bad_status.items():
        problems.append(f"모르는 sourceStatus '{st}' — Day {sorted(set(days))[:6]}")
    for day_type, days in bad_day_types.items():
        problems.append(f"모르는 dayType '{day_type}' — Day {sorted(set(days))[:6]}")
    for status, days in bad_execution.items():
        problems.append(
            f"모르는 execution status '{status}' — Day {sorted(set(days))[:6]}")

    # fact 키도 마찬가지다. 라벨이 없으면 화면에 'address' 가 그대로 뜬다 —
    # 실제로 주소·전화를 채운 날 영어 키가 새어 나왔다.
    bad_keys = {}
    for pl in trip.places.values():
        for key in pl.facts:
            if key not in FACT_LABEL:
                bad_keys.setdefault(key, []).append(pl.slug)
    for key, slugs in bad_keys.items():
        problems.append(
            f"모르는 fact 키 '{key}' — {sorted(slugs)[:4]}. "
            f"render.py 의 FACT_LABEL 에 한국어 표기를 더한다.")

    # 등급도 마찬가지다. 배지가 안 붙으면 '필수' 가 그냥 사라진다.
    for pl in trip.places.values():
        if pl.grade_label and pl.grade is None:
            problems.append(
                f"모르는 등급 '{pl.grade_label}' — {pl.slug}. "
                f"model.py 의 grade_map 에 더한다.")
    return problems


def check_place_prose(trip: Trip, promoted: set[str]) -> list[str]:
    """장문을 갖고 있던 장소가 장문을 잃지 않았는가.

    챕터 원고에서 절 제목이 바뀌면 대조에 실패해 장소 페이지가 조용히
    빈 껍데기가 된다. 그게 이 파이프라인에서 가장 잘 일어나는 사고다.
    """
    problems = []
    for slug in sorted(promoted):
        pl = trip.places.get(slug)
        if pl is None:
            problems.append(f"승격된 장문의 주인이 명부에 없다: {slug}")
        elif not pl.has_deep_guide and not pl.why_go:
            problems.append(f"장문이 사라졌다: {slug} — 챕터 원고의 절 제목이 "
                            f"바뀌었을 수 있다")
    return problems
