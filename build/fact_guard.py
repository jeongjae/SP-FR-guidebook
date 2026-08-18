#!/usr/bin/env python3
"""확정 사실 토큰 가드.

예약번호 · 전화번호 · 결제금액처럼 **이미 확정된 사실**이 독자 정본에서
유실되지 않았는지 본다. 확정된 것이 화면에서 사라지면 현장에서 다시
찾아야 하고, 그때는 대개 찾을 수 없다.

출처 셋을 대조한다.
  1) 110 Lock Register 의 확정 레코드
  2) TP_Europe_Travel_Master_Tracker_v1.2.xlsx 의 확정/완료 레코드
  3) build/confirmed_fact_manifest.json 에 동결된 챕터 상주 확정 사실

지역명·고정 개수를 하드코딩하지 않는다 — 데이터에서 뽑아 검증한다.
2026-08-18 개편에서 build.py 를 은퇴시키며 여기로 들어냈다. 이 가드는
렌더러에 의존하지 않고 원고만 읽으므로 그대로 옮길 수 있었다.

    python3 build/fact_guard.py
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
SOURCE = ROOT / "source"
TRACKER_XLSX = SOURCE / "OPERATIONS" / "TP_Europe_Travel_Master_Tracker_v1.2.xlsx"
MANIFEST_JSON = ROOT / "build" / "confirmed_fact_manifest.json"


def check_confirmed_fact_token_guards():
    """확정 사실 토큰(예약번호·전화번호·결제금액)이 독자 정본에서 유실되지 않도록 잠근다.
    출처:
      1) 110 Lock Register의 확정 레코드
      2) TP_Europe_Travel_Master_Tracker_v1.2.xlsx 의 Reservations·Transport·Accommodation 시트 확정/완료 레코드
      3) build/confirmed_fact_manifest.json 에 동결된 챕터 상주 확정 사실 레코드
    지역명·고정 개수 하드코딩 없이 데이터 기반으로 추출하여 검증한다.
    """
    import json
    import openpyxl
    problems = []
    exceptions = []

    reader_files = []
    reader_files.extend((SOURCE / "CURRENT/20_Regional_Chapters").glob("*.md"))
    reader_files.extend((SOURCE / "CURRENT/10_Core").glob("*.md"))
    reader_files.append(SOURCE / "ASSETS/90_Regional_Context_and_Place_Dossier_Compendium_v1.0.md")
    reader_files.append(SOURCE / "ASSETS/91_Place_Registry_v1.0.md")
    if (SOURCE.parent / "data/daily-cards").exists():
        reader_files.extend((SOURCE.parent / "data/daily-cards").glob("*.json"))

    reader_corpus = {}
    for f in reader_files:
        if f.exists():
            reader_corpus[f] = f.read_text(encoding="utf-8")

    combined_reader_text = "\n".join(reader_corpus.values())
    norm_corpus = re.sub(r"\s+", "", combined_reader_text)
    checked_tokens_set = set()

    def verify_token(token, source_desc):
        token_str = str(token).strip()
        if not token_str or token_str in ("None", "재확인", "확정", "미확정", "완료", "-", "—"):
            return True
        checked_tokens_set.add((token_str, source_desc))
        if token_str.startswith("+"):
            norm_token = re.sub(r"\s+", "", token_str)
            found = (token_str in combined_reader_text or norm_token in norm_corpus)
        elif token_str.startswith("KRW") or token_str.startswith("₩"):
            raw_num = re.sub(r"[^\d]", "", token_str)
            found = (token_str in combined_reader_text or
                     f"KRW {int(raw_num):,}" in combined_reader_text or
                     f"₩{int(raw_num):,}" in combined_reader_text or
                     f"{int(raw_num):,}" in combined_reader_text)
        elif token_str.startswith("€"):
            found = (token_str in combined_reader_text)
            if not found and "." in token_str:
                val = float(token_str.replace("€", "").strip())
                opt1 = f"€{val}"
                opt2 = f"€{val:.2f}"
                found = (opt1 in combined_reader_text or opt2 in combined_reader_text)
        else:
            found = (token_str in combined_reader_text)

        if not found:
            problems.append(f"토큰 누락: '{token_str}' (출처: {source_desc})")
        return found

    # 1. 110 Lock Register
    lock_register_path = SOURCE / "OPERATIONS/110_Phase8_Reservation_and_Operations_Lock_Register_v1.0.md"
    if lock_register_path.exists():
        text = lock_register_path.read_text(encoding="utf-8")
        for section in re.split(r"\n(?=##\s+)", text):
            header = section.splitlines()[0].strip() if section.splitlines() else ""
            is_confirmed = ("확정" in header or "CONFIRMED" in header or
                            bool(re.search(r"-\s*상태:\s*.*(?:확정|CONFIRMED)", section)))
            if not is_confirmed:
                continue
            if "guard: operations-only" in section:
                exceptions.append(f"LockRegister: {header}")
                continue
            for m in re.finditer(r"(?:확인번호|예약번호|예약코드|PNR|바우처)\s*[:：]?\s*([A-Za-z0-9.]+)", section):
                code = m.group(1).strip()
                if code not in ("재확인", "확정", "None", "미확정", "완료"):
                    verify_token(code, f"LockRegister: {header}")
            for m in re.finditer(r"(\+(?:33|34)(?:\s*\d+){4,})", section):
                verify_token(m.group(1).strip(), f"LockRegister: {header}")
            for m in re.finditer(r"(€\s*\d+(?:\.\d+)?|KRW\s*[\d,]+|₩\s*[\d,]+)", section):
                verify_token(m.group(1).strip(), f"LockRegister: {header}")

    # 2. TRACKER_XLSX
    tracker_path = SOURCE / "OPERATIONS/TP_Europe_Travel_Master_Tracker_v1.2.xlsx"
    if tracker_path.exists():
        wb = openpyxl.load_workbook(tracker_path, data_only=True)
        for sheet_name in ["Reservations", "Transport", "Accommodation"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[3]]
            for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
                if not any(row):
                    continue
                d = dict(zip(headers, row))
                state = str(d.get("상태") or "").strip()
                if any(k in state for k in ["예약완료", "확정", "CONFIRMED"]):
                    row_id = d.get("ID") or d.get("거점") or d.get("구간") or f"Row{row_idx}"
                    desc = f"Tracker {sheet_name} {row_id}"
                    note = str(d.get("비고") or "")
                    if "guard: operations-only" in note:
                        exceptions.append(desc)
                        continue
                    res_code = str(d.get("예약번호") or "").strip()
                    if res_code and res_code not in ["None", "미표기", "-", "—"]:
                        for code in re.findall(r"[A-Za-z0-9.]+", res_code):
                            if code not in ("Trip.com", "Airbnb", "booking.com", "None", "확인", "발권메일", "미표기", "PNR"):
                                verify_token(code, desc)
                    for m in re.finditer(r"(\+(?:33|34)(?:\s*\d+){4,})", note):
                        verify_token(m.group(1).strip(), desc)
                    for amt_col in ["총액", "실제총액", "결제액"]:
                        val = d.get(amt_col)
                        if val is not None and isinstance(val, (int, float)) and val > 0:
                            currency = str(d.get("예산통화") or d.get("통화") or ("KRW" if val > 10000 else "EUR")).strip()
                            if currency in ("EUR", "€"):
                                amt_token = f"€{val}"
                            elif currency in ("KRW", "₩", "원"):
                                amt_token = f"KRW {int(val):,}"
                            else:
                                amt_token = f"{val}"
                            verify_token(amt_token, desc)

    # 3. MANIFEST_JSON
    manifest_path = SOURCE.parent / "build/confirmed_fact_manifest.json"
    if manifest_path.exists():
        mdata = json.loads(manifest_path.read_text(encoding="utf-8"))
        for rec in mdata.get("records", []):
            r_id = rec.get("id")
            desc = f"Manifest {r_id} ({rec.get('entity')})"
            for t in rec.get("tokens", []):
                verify_token(t, desc)

    if exceptions:
        print(f"확정 사실 토큰 생존 가드 예외 목록: {', '.join(exceptions)}")

    if problems:
        print("확정 사실 토큰 생존 가드 실패:")
        for problem in problems:
            print("  " + problem)
        sys.exit(1)

    print(f"확정 사실 토큰 생존 가드: 확정 토큰 {len(checked_tokens_set)}건(검증 항목) 독자 정본 생존 확인 이상 없음")


if __name__ == "__main__":
    check_confirmed_fact_token_guards()
    print("확정 사실 토큰 가드 통과")
    sys.exit(0)
